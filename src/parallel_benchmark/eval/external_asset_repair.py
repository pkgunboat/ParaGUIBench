"""ParaGUIBench 外部二进制资产的可复现修复与验证工具。"""

from __future__ import annotations

from copy import copy
import json
import math
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


EXCEL_WORKBOOK_NAMES = (
    "store1.xlsx",
    "store2.xlsx",
    "store3.xlsx",
    "store4.xlsx",
)


def _require_excel_workbooks(root: Path) -> list[Path]:
    """解析并检查四个标准 Excel 工作簿路径。

    功能：按稳定文件名构造任务资产路径，并在任何输入缺失时立即报错，
    避免生成不完整但看似成功的替换资产目录。
    输入参数：root，包含任务工作簿的目录。
    输出返回值：按 ``store1.xlsx`` 到 ``store4.xlsx`` 排序的路径列表。
    """
    paths = [root / name for name in EXCEL_WORKBOOK_NAMES]
    missing = [path.name for path in paths if not path.is_file()]
    if missing:
        raise FileNotFoundError(f"缺少 Excel 资产: {', '.join(missing)}")
    return paths


def validate_excel002_inputs(root: Path) -> list[str]:
    """验证 Excel-002 初始资产没有预完成任务要求。

    功能：检查四个工作簿的 A3:C3 表头均未加粗，并检查 B4:C15 数值区域
    均未预设为右对齐，从而保证空操作不能获得格式任务分数。
    输入参数：root，包含四个初始 xlsx 的目录。
    输出返回值：问题描述列表；空列表表示全部资产符合初始状态契约。
    """
    issues: list[str] = []
    for path in _require_excel_workbooks(root):
        workbook = load_workbook(path, data_only=False)
        worksheet = workbook.active
        for column in range(1, 4):
            cell = worksheet.cell(row=3, column=column)
            if cell.font.bold is True:
                issues.append(f"{path.name}!{cell.coordinate} 已预先加粗")
        for row in range(4, 16):
            for column in range(2, 4):
                cell = worksheet.cell(row=row, column=column)
                if (cell.alignment.horizontal or "").casefold() == "right":
                    issues.append(f"{path.name}!{cell.coordinate} 已预先右对齐")
        workbook.close()
    return issues


def repair_excel002_inputs(source_dir: Path, destination_dir: Path) -> list[Path]:
    """重制 Excel-002 初始资产并清除预完成格式。

    功能：复制四个源工作簿的全部数据与非目标格式，只把 A3:C3 的粗体
    设为关闭，并把 B4:C15 显式居中，使数值列在任务执行前不会呈现为
    已完成的右对齐状态。
    输入参数：
        source_dir: 原始 Hugging Face 初始资产目录。
        destination_dir: 写入修复后工作簿的目标目录。
    输出返回值：按标准文件名排序的四个新工作簿路径。
    """
    destination_dir.mkdir(parents=True, exist_ok=True)
    output_paths: list[Path] = []
    for source_path in _require_excel_workbooks(source_dir):
        workbook = load_workbook(source_path, data_only=False)
        worksheet = workbook.active
        for column in range(1, 4):
            cell = worksheet.cell(row=3, column=column)
            font = copy(cell.font)
            font.bold = False
            cell.font = font
        for row in range(4, 16):
            for column in range(2, 4):
                cell = worksheet.cell(row=row, column=column)
                alignment = copy(cell.alignment)
                alignment.horizontal = "center"
                cell.alignment = alignment
        destination_path = destination_dir / source_path.name
        workbook.save(destination_path)
        workbook.close()
        output_paths.append(destination_path)

    remaining_issues = validate_excel002_inputs(destination_dir)
    if remaining_issues:
        raise ValueError("Excel-002 修复后验证失败: " + "; ".join(remaining_issues))
    return output_paths


def _load_excel005_contract(
    task_path: Path,
) -> tuple[int, int, int, float, float, float, dict[str, list[float]]]:
    """从任务 JSON 解析 Excel-005 的倍率评价契约。

    功能：查找 ``check_values_scaled_from_source`` 规则，解析单列单元格范围、
    除数、容差和每个工作簿的原始值，并验证数组长度与范围一致。
    输入参数：task_path，Excel-005 真实任务 JSON 路径。
    输出返回值：起始行、结束行、列号、除数、相对容差、绝对容差及
    按文件名组织的源数值字典。
    """
    task = json.loads(task_path.read_text(encoding="utf-8"))
    matching_rules = [
        rule
        for rule in task.get("eval_rules", [])
        if rule.get("check") == "check_values_scaled_from_source"
    ]
    if len(matching_rules) != 1:
        raise ValueError("任务必须且只能包含一个倍率换算评价规则")
    params = matching_rules[0].get("params") or {}
    start_cell = str(params.get("start_cell") or "")
    end_cell = str(params.get("end_cell") or "")
    min_col, min_row, max_col, max_row = range_boundaries(
        f"{start_cell}:{end_cell}"
    )
    if min_col != max_col:
        raise ValueError("倍率换算范围必须是单列")
    divisor = float(params.get("divisor"))
    if divisor == 0:
        raise ValueError("倍率换算除数不能为零")
    relative_tolerance = float(params.get("relative_tolerance", 0.0))
    absolute_tolerance = float(params.get("absolute_tolerance", 0.0))
    source_values = params.get("source_values_by_file") or {}
    expected_length = max_row - min_row + 1
    for filename in EXCEL_WORKBOOK_NAMES:
        values = source_values.get(filename)
        if not isinstance(values, list) or len(values) != expected_length:
            raise ValueError(
                f"{filename} 的源数值数量应为 {expected_length}，实际为 "
                f"{len(values) if isinstance(values, list) else 0}"
            )
    return (
        min_row,
        max_row,
        min_col,
        divisor,
        relative_tolerance,
        absolute_tolerance,
        source_values,
    )


def validate_excel005_answers(root: Path, task_path: Path) -> list[str]:
    """验证 Excel-005 外部答案符合真实任务倍率。

    功能：依据任务 JSON 中的源值、范围、除数与容差逐单元格验证四个答案，
    防止历史答案中的小数点错位再次进入 gold。
    输入参数：
        root: 包含四个答案工作簿的目录。
        task_path: Excel-005 真实任务 JSON 路径。
    输出返回值：问题描述列表；空列表表示全部答案符合倍率契约。
    """
    (
        min_row,
        max_row,
        column,
        divisor,
        relative_tolerance,
        absolute_tolerance,
        source_values,
    ) = _load_excel005_contract(task_path)
    issues: list[str] = []
    for path in _require_excel_workbooks(root):
        workbook = load_workbook(path, data_only=True)
        worksheet = workbook.active
        for row, source_value in zip(
            range(min_row, max_row + 1),
            source_values[path.name],
            strict=True,
        ):
            cell = worksheet.cell(row=row, column=column)
            expected = float(source_value) / divisor
            actual = cell.value
            if not isinstance(actual, (int, float)) or not math.isclose(
                float(actual),
                expected,
                rel_tol=relative_tolerance,
                abs_tol=absolute_tolerance,
            ):
                issues.append(
                    f"{path.name}!{cell.coordinate} 应为 {expected:g}，实际为 "
                    f"{actual!r}"
                )
        workbook.close()
    return issues


def repair_excel005_answers(
    source_dir: Path,
    destination_dir: Path,
    task_path: Path,
) -> list[Path]:
    """按任务 JSON 重生成 Excel-005 的外部答案数值。

    功能：根据任务 JSON 重新计算每个目标单元格；只重写实际错误的工作簿，
    已完全正确的文件使用字节级复制，以避免无关 OOXML 元数据变化。
    输入参数：
        source_dir: 原始历史 answer 工作簿目录。
        destination_dir: 写入修复后工作簿的目标目录。
        task_path: Excel-005 真实任务 JSON 路径。
    输出返回值：按标准文件名排序的四个新答案工作簿路径。
    """
    (
        min_row,
        max_row,
        column,
        divisor,
        relative_tolerance,
        absolute_tolerance,
        source_values,
    ) = _load_excel005_contract(task_path)
    destination_dir.mkdir(parents=True, exist_ok=True)
    output_paths: list[Path] = []
    for source_path in _require_excel_workbooks(source_dir):
        workbook = load_workbook(source_path, data_only=False)
        worksheet = workbook.active
        changed = False
        for row, source_value in zip(
            range(min_row, max_row + 1),
            source_values[source_path.name],
            strict=True,
        ):
            cell = worksheet.cell(row=row, column=column)
            expected = float(source_value) / divisor
            actual = cell.value
            if not isinstance(actual, (int, float)) or not math.isclose(
                float(actual),
                expected,
                rel_tol=relative_tolerance,
                abs_tol=absolute_tolerance,
            ):
                cell.value = expected
                changed = True
        destination_path = destination_dir / source_path.name
        if changed:
            workbook.save(destination_path)
            workbook.close()
        else:
            workbook.close()
            copy2(source_path, destination_path)
        output_paths.append(destination_path)

    remaining_issues = validate_excel005_answers(destination_dir, task_path)
    if remaining_issues:
        raise ValueError("Excel-005 修复后验证失败: " + "; ".join(remaining_issues))
    return output_paths

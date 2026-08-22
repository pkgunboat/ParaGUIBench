"""
Excel 表格（.xlsx）属性检查原语。

每个函数接收文件路径和参数字典，返回标准化结果：
    {"pass": bool, "score": float 0.0~1.0, "reason": str}

依赖: openpyxl
"""

import glob
import hashlib
import json
import logging
import math
import os
from collections import Counter
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger("eval.operation_checks.xlsx")


# ------------------------------------------------------------------
# 工具函数
# ------------------------------------------------------------------

def _load_workbook(file_path: str, data_only: bool = False):
    """
    安全加载 xlsx 文件。

    输入:
        file_path: xlsx 文件路径
        data_only: 是否只读计算值（不加载公式）
    输出:
        Workbook 对象；加载失败返回 None
    """
    try:
        return openpyxl.load_workbook(file_path, data_only=data_only)
    except Exception as exc:
        logger.error("无法打开 xlsx 文件 %s: %s", file_path, exc)
        return None


def _get_sheet(wb, sheet_name: Optional[str] = None):
    """
    获取指定工作表，未指定则返回活动表。

    输入:
        wb: Workbook 对象
        sheet_name: 工作表名称（可选）
    输出:
        Worksheet 对象；找不到返回 None
    """
    if sheet_name:
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        logger.warning("工作表 '%s' 不存在", sheet_name)
        return None
    return wb.active


def _ok(reason: str = "通过") -> Dict[str, Any]:
    return {"pass": True, "score": 1.0, "reason": reason}


def _fail(reason: str) -> Dict[str, Any]:
    return {"pass": False, "score": 0.0, "reason": reason}


def _partial(score: float, reason: str) -> Dict[str, Any]:
    # 严格阈值：仅当得分等于 1.0 时算通过；保留 score 用于聚合时的部分得分加权
    return {"pass": score >= 1.0 - 1e-9, "score": round(score, 4), "reason": reason}


def _config_error(reason: str) -> Dict[str, Any]:
    """评价器配置错误（缺参数 / 参数无法解析等）：score=-1 哨兵，由上层冒泡为 evaluator_error。"""
    return {"pass": False, "score": -1.0, "status": "evaluator_error", "reason": reason}


def _is_numeric_cell(cell) -> bool:
    """判断单元格是否为数值类型。"""
    return isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)


def _resolve_column(col, ws) -> Optional[int]:
    """
    将列标识转换为列号（1-based）。

    输入:
        col: 列标识——字符串（"A", "B"）或整数（1, 2）
        ws: Worksheet 对象（用于按表头名查找）
    输出:
        int 列号；解析失败返回 None
    """
    if isinstance(col, int):
        return col
    if isinstance(col, str):
        if col.isalpha():
            return column_index_from_string(col)
        # 尝试按表头名匹配
        for cell in ws[1]:
            if cell.value and str(cell.value).strip() == col.strip():
                return cell.column
    return None


# ------------------------------------------------------------------
# 检查函数
# ------------------------------------------------------------------

def check_cell_value(file_path: str, params: dict) -> dict:
    """
    检查指定单元格的值是否等于期望值。

    输入:
        file_path: xlsx 文件路径
        params:
            cell (str): 单元格坐标，如 "A1", "B3"
            expected (any): 期望值
            sheet_name (str, 可选): 工作表名称
            tolerance (float, 可选): 数值容差比例，默认 0.01（±1%）
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    cell_ref = params.get("cell", "")
    expected = params.get("expected")
    sheet_name = params.get("sheet_name")
    tolerance = params.get("tolerance", 0.01)

    if not cell_ref:
        return _config_error("参数缺少 cell")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        actual = ws[cell_ref].value

        # 数值比较（容差）
        if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
            if expected == 0:
                passed = abs(actual) < 1e-9
            else:
                passed = abs(actual - expected) / abs(expected) <= tolerance
            if passed:
                return _ok(f"单元格 {cell_ref} 值 {actual}（期望 {expected}）")
            return _fail(f"单元格 {cell_ref} 值 {actual}（期望 {expected}，容差 ±{tolerance:.0%}）")

        # 字符串比较（忽略前后空格和大小写）
        actual_str = str(actual).strip() if actual is not None else ""
        expected_str = str(expected).strip() if expected is not None else ""
        if actual_str.lower() == expected_str.lower():
            return _ok(f"单元格 {cell_ref} 值匹配: {actual_str}")
        return _fail(f"单元格 {cell_ref} 值 '{actual_str}'（期望 '{expected_str}'）")
    finally:
        wb.close()


def check_header_bold(file_path: str, params: dict) -> dict:
    """
    检查表头行（默认第1行）的所有非空单元格是否加粗。

    输入:
        file_path: xlsx 文件路径
        params:
            header_row (int): 表头行号（1-based），默认 1
            sheet_name (str, 可选): 工作表名称
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    header_row = params.get("header_row", 1)
    sheet_name = params.get("sheet_name")

    wb = _load_workbook(file_path)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        total_cells = 0
        bold_cells = 0
        non_bold = []

        for cell in ws[header_row]:
            if cell.value is None or str(cell.value).strip() == "":
                continue
            total_cells += 1
            if cell.font and cell.font.bold:
                bold_cells += 1
            else:
                non_bold.append(f"{get_column_letter(cell.column)}{header_row}")

        if total_cells == 0:
            return _ok(f"第 {header_row} 行无非空单元格")

        ratio = bold_cells / total_cells
        if ratio >= 1.0:
            return _ok(f"表头行全部加粗（{bold_cells}/{total_cells}）")

        return _partial(
            ratio,
            f"表头加粗率 {ratio:.1%}（{bold_cells}/{total_cells}），"
            f"未加粗: {', '.join(non_bold[:5])}"
        )
    finally:
        wb.close()


def check_column_alignment(file_path: str, params: dict) -> dict:
    """
    检查指定类型列的对齐方式。

    输入:
        file_path: xlsx 文件路径
        params:
            alignment (str): 期望的对齐方式（"left", "center", "right"）
            column_type (str): 列类型筛选 —— "numeric"（数值列）、"text"（文本列）、"all"
            columns (list, 可选): 指定列列表（字母或表头名），覆盖 column_type
            sheet_name (str, 可选): 工作表名称
            skip_header (bool): 是否跳过表头行，默认 True
            threshold (float): 符合比例阈值，默认 0.8
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    expected_align = params.get("alignment", "")
    column_type = params.get("column_type", "all")
    specific_columns = params.get("columns")
    sheet_name = params.get("sheet_name")
    skip_header = params.get("skip_header", True)
    threshold = params.get("threshold", 0.8)

    if not expected_align:
        return _config_error("参数缺少 alignment")

    wb = _load_workbook(file_path)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        # 确定需要检查的列
        target_cols = set()
        if specific_columns:
            for c in specific_columns:
                col_idx = _resolve_column(c, ws)
                if col_idx:
                    target_cols.add(col_idx)
        else:
            # 按 column_type 自动识别
            start_row = 2 if skip_header else 1
            for col_idx in range(1, ws.max_column + 1):
                has_numeric = False
                has_text = False
                for row_idx in range(start_row, min(ws.max_row + 1, start_row + 20)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is None:
                        continue
                    if _is_numeric_cell(cell):
                        has_numeric = True
                    else:
                        has_text = True

                if column_type == "numeric" and has_numeric:
                    target_cols.add(col_idx)
                elif column_type == "text" and has_text and not has_numeric:
                    target_cols.add(col_idx)
                elif column_type == "all":
                    target_cols.add(col_idx)

        if not target_cols:
            return _ok(f"未找到类型为 '{column_type}' 的列")

        total_cells = 0
        aligned_cells = 0
        start_row = 2 if skip_header else 1

        for row_idx in range(start_row, ws.max_row + 1):
            for col_idx in target_cols:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                total_cells += 1
                actual_align = cell.alignment.horizontal if cell.alignment else None
                if actual_align and actual_align.lower() == expected_align.lower():
                    aligned_cells += 1

        if total_cells == 0:
            return _ok("目标列无非空数据单元格")

        ratio = aligned_cells / total_cells
        if ratio >= threshold:
            return _ok(
                f"对齐率 {ratio:.1%}（{aligned_cells}/{total_cells}），"
                f"期望 {expected_align}"
            )

        return _partial(
            ratio,
            f"对齐率 {ratio:.1%}（{aligned_cells}/{total_cells}），"
            f"期望 {expected_align}，阈值 ≥ {threshold:.0%}"
        )
    finally:
        wb.close()


def check_sort_order(file_path: str, params: dict) -> dict:
    """
    检查指定列的数据是否按升序或降序排列。

    输入:
        file_path: xlsx 文件路径
        params:
            column (str/int): 列标识（字母、表头名或数字）
            order (str): "asc"（升序）或 "desc"（降序）
            sheet_name (str, 可选): 工作表名称
            skip_header (bool): 是否跳过表头行，默认 True
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    column = params.get("column")
    order = params.get("order", "desc")
    sheet_name = params.get("sheet_name")
    skip_header = params.get("skip_header", True)

    if column is None:
        return _config_error("参数缺少 column")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        col_idx = _resolve_column(column, ws)
        if col_idx is None:
            return _config_error(f"无法解析列: {column}")

        start_row = 2 if skip_header else 1
        values = []
        for row_idx in range(start_row, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                values.append(val)

        if len(values) <= 1:
            return _ok("数据量不足，无需排序检查")

        # 混合数值/文本类型是 Agent 输出问题，应评为失败，
        # 不应将 TypeError 上抛为评价器故障。
        try:
            is_ascending = all(values[i] <= values[i + 1] for i in range(len(values) - 1))
            is_descending = all(values[i] >= values[i + 1] for i in range(len(values) - 1))
        except TypeError:
            return _fail(f"列 {column} 包含无法互相比较的混合类型")

        if order == "asc" and is_ascending:
            return _ok(f"列 {column} 升序排列正确（{len(values)} 行）")
        if order == "desc" and is_descending:
            return _ok(f"列 {column} 降序排列正确（{len(values)} 行）")

        # 计算乱序程度（逆序对比例）
        inversions = 0
        total_pairs = 0
        try:
            for i in range(len(values) - 1):
                total_pairs += 1
                if order == "asc" and values[i] > values[i + 1]:
                    inversions += 1
                elif order == "desc" and values[i] < values[i + 1]:
                    inversions += 1
        except TypeError:
            return _fail(f"列 {column} 包含无法互相比较的混合类型")

        ratio = 1.0 - (inversions / total_pairs) if total_pairs else 1.0
        return _partial(
            ratio,
            f"列 {column} 排序不完全（有序比例 {ratio:.1%}，期望 {order}）"
        )
    finally:
        wb.close()


def check_has_sum_row(file_path: str, params: dict) -> dict:
    """
    检查工作表是否在末尾存在汇总行（数值列的合计）。

    通过检查最后几行是否包含与上方数据列之和接近的值来判断。

    输入:
        file_path: xlsx 文件路径
        params:
            sum_columns (list, 可选): 需要检查汇总的列（字母或表头名列表）；
                为空则自动检测所有数值列
            sheet_name (str, 可选): 工作表名称
            tolerance (float): 合计值容差比例，默认 0.01（±1%）
            check_last_rows (int): 检查最后几行，默认 3
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    sum_columns = params.get("sum_columns")
    sheet_name = params.get("sheet_name")
    tolerance = params.get("tolerance", 0.01)
    check_last = params.get("check_last_rows", 3)

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        if ws.max_row < 3:
            return _fail("数据行数不足")

        # 确定需要检查的数值列
        target_cols = []
        if sum_columns:
            for c in sum_columns:
                col_idx = _resolve_column(c, ws)
                if col_idx:
                    target_cols.append(col_idx)
        else:
            # 自动检测：第2行为数值的列
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=2, column=col_idx)
                if _is_numeric_cell(cell):
                    target_cols.append(col_idx)

        if not target_cols:
            return _fail("未找到数值列")

        total_cols = len(target_cols)
        matched_cols = 0

        for col_idx in target_cols:
            # 计算数据区域之和（第2行到倒数第 check_last+1 行）
            data_sum = 0.0
            data_end_row = ws.max_row - check_last
            if data_end_row < 2:
                data_end_row = ws.max_row - 1

            for row_idx in range(2, data_end_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    data_sum += val

            # 检查最后几行是否存在接近 data_sum 的值
            found_sum = False
            for row_idx in range(max(2, ws.max_row - check_last + 1), ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    if data_sum == 0:
                        if abs(val) < 1e-9:
                            found_sum = True
                            break
                    elif abs(val - data_sum) / abs(data_sum) <= tolerance:
                        found_sum = True
                        break

            if found_sum:
                matched_cols += 1

        ratio = matched_cols / total_cols
        if ratio >= 0.8:
            return _ok(
                f"汇总行检测通过（{matched_cols}/{total_cols} 列有合计值）"
            )

        return _partial(
            ratio,
            f"汇总行匹配率 {ratio:.1%}（{matched_cols}/{total_cols} 列有合计值）"
        )
    finally:
        wb.close()


# ------------------------------------------------------------------
# 任务专用检查函数
# ------------------------------------------------------------------

def check_batchexcel001_annual_sum(file_path: str, params: dict) -> dict:
    """
    Batchoperationexcel-001 专用：检查年度汇总行是否为对应列数据区域的累加。

    检查 B16 是否等于 B4:B15 的累加值，C16 是否等于 C4:C15 的累加值。

    输入:
        file_path: xlsx 文件路径
        params: 忽略（检查逻辑已内置）
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    sum_checks = [
        {"sum_cell": "B16", "range_start": 4, "range_end": 15, "col": "B"},
        {"sum_cell": "C16", "range_start": 4, "range_end": 15, "col": "C"},
    ]
    tolerance = 0.01

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = wb.active
        total = len(sum_checks)
        matched = 0
        details = []

        for check in sum_checks:
            col = check["col"]
            sum_cell = check["sum_cell"]
            row_start = check["range_start"]
            row_end = check["range_end"]

            # 计算数据区域累加值
            expected_sum = 0.0
            has_numeric = False
            for row in range(row_start, row_end + 1):
                val = ws[f"{col}{row}"].value
                if isinstance(val, (int, float)):
                    expected_sum += val
                    has_numeric = True

            if not has_numeric:
                details.append(f"{col}{row_start}:{col}{row_end} 无数值数据")
                continue

            actual = ws[sum_cell].value
            if actual is None:
                details.append(f"{sum_cell}: 空值（期望 {expected_sum}）")
                continue

            if isinstance(actual, (int, float)):
                if expected_sum == 0:
                    passed = abs(actual) < 1e-9
                else:
                    passed = abs(actual - expected_sum) / abs(expected_sum) <= tolerance
                if passed:
                    matched += 1
                else:
                    details.append(
                        f"{sum_cell}: {actual}（期望 {col}{row_start}:{col}{row_end} 累加 = {expected_sum}）"
                    )
            else:
                details.append(f"{sum_cell}: 值非数字类型 ({actual})")

        if matched == total:
            return _ok(f"全部 {total} 列汇总值正确")

        ratio = matched / total
        return _partial(ratio, f"{matched}/{total} 匹配; " + "; ".join(details))
    finally:
        wb.close()


def check_batchexcel002_header_bold(file_path: str, params: dict) -> dict:
    """
    Batchoperationexcel-002 专用：检查 A3, B3, C3 单元格是否加粗。

    输入:
        file_path: xlsx 文件路径
        params: 忽略
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    target_cells = ["A3", "B3", "C3"]

    wb = _load_workbook(file_path)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = wb.active
        total = len(target_cells)
        bold_count = 0
        non_bold = []

        for ref in target_cells:
            cell = ws[ref]
            if cell.font and cell.font.bold:
                bold_count += 1
            else:
                non_bold.append(ref)

        if bold_count == total:
            return _ok(f"{', '.join(target_cells)} 全部加粗")

        ratio = bold_count / total
        return _partial(ratio, f"未加粗: {', '.join(non_bold)}")
    finally:
        wb.close()


def check_batchexcel002_range_right_align(file_path: str, params: dict) -> dict:
    """
    Batchoperationexcel-002 专用：检查 B4:C15 区域是否右对齐。

    输入:
        file_path: xlsx 文件路径
        params: 忽略
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    wb = _load_workbook(file_path)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = wb.active
        total = 0
        aligned = 0
        misaligned = []

        for row in range(4, 16):  # 4~15
            for col_letter in ("B", "C"):
                ref = f"{col_letter}{row}"
                cell = ws[ref]
                if cell.value is None:
                    continue
                total += 1
                h_align = cell.alignment.horizontal if cell.alignment else None
                if h_align and h_align.lower() == "right":
                    aligned += 1
                else:
                    misaligned.append(ref)

        if total == 0:
            return _ok("B4:C15 区域无数据")

        ratio = aligned / total
        if ratio >= 0.9:
            return _ok(f"右对齐率 {ratio:.1%}（{aligned}/{total}）")

        sample = ", ".join(misaligned[:5])
        return _partial(ratio, f"右对齐率 {ratio:.1%}; 未对齐: {sample}")
    finally:
        wb.close()


# ------------------------------------------------------------------
# 扩展检查函数
# ------------------------------------------------------------------

def check_cell_contains_string(file_path: str, params: dict) -> dict:
    """
    检查指定单元格的值是否包含特定字符串。

    输入:
        file_path: xlsx 文件路径
        params:
            cell (str): 单元格坐标，如 "B4"
            expected (str): 期望包含的字符串
            sheet_name (str, 可选): 工作表名称
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    cell_ref = params.get("cell", "")
    expected = params.get("expected", "")
    sheet_name = params.get("sheet_name")

    if not cell_ref:
        return _config_error("参数缺少 cell")
    if not expected:
        return _config_error("参数缺少 expected")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        actual = ws[cell_ref].value
        if actual is None:
            return _fail(f"单元格 {cell_ref} 值为空")

        actual_str = str(actual)
        if expected in actual_str:
            return _ok(f"单元格 {cell_ref} 包含 '{expected}'")

        return _fail(f"单元格 {cell_ref} 值 '{actual_str}' 不包含 '{expected}'")
    finally:
        wb.close()


def check_values_are_decimals(file_path: str, params: dict) -> dict:
    """
    检查指定单元格区域的值是否为小数。

    输入:
        file_path: xlsx 文件路径
        params:
            start_cell (str): 起始单元格，如 "B4"
            end_cell (str): 结束单元格，如 "B15"
            sheet_name (str, 可选): 工作表名称
            threshold (float): 符合比例阈值，默认 0.8
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    start_cell = params.get("start_cell", "")
    end_cell = params.get("end_cell", "")
    sheet_name = params.get("sheet_name")
    threshold = params.get("threshold", 0.8)

    if not start_cell or not end_cell:
        return _config_error("参数缺少 start_cell 或 end_cell")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
        end_row = int(''.join(filter(str.isdigit, end_cell)))

        total = 0
        decimal_count = 0

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    continue
                total += 1
                if isinstance(cell.value, float):
                    decimal_count += 1
                elif isinstance(cell.value, int):
                    # 整数也可能是小数形式，检查是否为小数格式
                    val_str = str(cell.value)
                    if '.' in val_str or (isinstance(cell.value, float)):
                        decimal_count += 1

        if total == 0:
            return _fail(f"区域 {start_cell}:{end_cell} 无数据")

        ratio = decimal_count / total
        if ratio >= threshold:
            return _ok(f"小数比例 {ratio:.1%}（{decimal_count}/{total}）")

        return _partial(ratio, f"小数比例 {ratio:.1%}（{decimal_count}/{total}）")
    finally:
        wb.close()


def check_values_scaled_from_source(file_path: str, params: dict) -> dict:
    """检查单元格值是否按指定倍率从已知源数据换算。

    输入:
        file_path: xlsx 文件路径。
        params:
            start_cell (str): 目标区域起始单元格。
            end_cell (str): 目标区域结束单元格。
            source_values_by_file (dict[str, list[number]]): 按文件名保存的
                源区域行优先数值。
            divisor (number): 换算除数，例如元转万元为 10000。
            relative_tolerance (float): 相对容差，默认 0.001。
            absolute_tolerance (float): 绝对容差，默认 0.005。
            sheet_name (str, 可选): 工作表名。
    输出:
        所有单元格符合 source/divisor 时返回通过。
    """
    start_cell = params.get("start_cell", "")
    end_cell = params.get("end_cell", "")
    source_by_file = params.get("source_values_by_file", {})
    divisor = params.get("divisor")
    relative_tolerance = float(params.get("relative_tolerance", 0.001))
    absolute_tolerance = float(params.get("absolute_tolerance", 0.005))
    sheet_name = params.get("sheet_name")
    if not start_cell or not end_cell or not source_by_file or not divisor:
        return _config_error(
            "参数缺少 start_cell/end_cell/source_values_by_file/divisor"
        )
    filename = os.path.basename(file_path)
    source_values = source_by_file.get(filename)
    if not source_values:
        return _config_error(f"未配置 {filename} 的源数据")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")
    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")
        start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        actual_values = [
            ws.cell(row=row, column=col).value
            for row in range(start_row, end_row + 1)
            for col in range(start_col, end_col + 1)
        ]
        if len(actual_values) != len(source_values):
            return _config_error(
                f"区域有 {len(actual_values)} 个单元格，"
                f"但配置了 {len(source_values)} 个源值"
            )
        matched = 0
        failures = []
        for index, (actual, source) in enumerate(zip(actual_values, source_values)):
            expected = float(source) / float(divisor)
            if (
                isinstance(actual, (int, float))
                and not isinstance(actual, bool)
                and math.isclose(
                    float(actual), expected,
                    rel_tol=relative_tolerance,
                    abs_tol=absolute_tolerance,
                )
            ):
                matched += 1
            else:
                failures.append(f"#{index + 1}: {actual!r} != {expected:g}")
        if matched == len(source_values):
            return _ok(f"全部 {matched} 个数值均按 1/{divisor:g} 换算")
        return _partial(
            matched / len(source_values),
            f"仅 {matched}/{len(source_values)} 个数值换算正确: "
            f"{'; '.join(failures[:3])}",
        )
    finally:
        wb.close()


def _rgb_is_red(rgb) -> bool:
    """
    判断一个 openpyxl 颜色 rgb 值是否属于“红色”。

    采用红色通道占优判据（而非固定枚举），可覆盖纯红/深红/浅红填充：
    要求 R 通道足够高，且明显高于 G、B 通道。这样能正确接受真实
    agent 常用的浅红底色（如 LibreOffice 的 FFFFD7D7、Excel 的
    “浅红填充”FFFFC7CE），同时拒绝黄/白/绿等非红色。

    输入:
        rgb: openpyxl 颜色对象的 .rgb 属性（可能为 None / 非字符串 / 主题标记）
    输出:
        bool: 判定为红色返回 True，否则 False
    """
    if not rgb:
        return False
    s = str(rgb).upper()
    if s == "RED":
        return True
    if len(s) == 8:  # 去掉 ARGB 的 alpha 前缀
        s = s[2:]
    if len(s) != 6:  # 主题色/索引色等无法解析为 RRGGBB
        return False
    try:
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
    except ValueError:
        return False
    return r >= 150 and (r - g) >= 40 and (r - b) >= 40


def _cell_is_red(cell) -> bool:
    """
    判断单元格是否被标记为红色：字体颜色为红，或背景填充（fgColor/start_color）为红。

    输入:
        cell: openpyxl 单元格对象
    输出:
        bool: 字体色或背景填充色任一为红色返回 True，否则 False
    """
    # 1) 字体颜色
    try:
        if cell.font and cell.font.color and _rgb_is_red(getattr(cell.font.color, "rgb", None)):
            return True
    except Exception:
        pass
    # 2) 背景填充：仅当存在实际填充图案时才检查 fgColor / start_color
    try:
        fill = cell.fill
        if fill is not None and getattr(fill, "patternType", None):
            for attr in ("fgColor", "start_color"):
                col = getattr(fill, attr, None)
                if col is not None and _rgb_is_red(getattr(col, "rgb", None)):
                    return True
    except Exception:
        pass
    return False



def check_negative_values_colored(file_path: str, params: dict) -> dict:
    """
    检查负值单元格是否被标记为红色。

    输入:
        file_path: xlsx 文件路径
        params:
            start_cell (str): 起始单元格，如 "D4"
            end_cell (str): 结束单元格，如 "D15"
            sheet_name (str, 可选): 工作表名称
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    start_cell = params.get("start_cell", "")
    end_cell = params.get("end_cell", "")
    sheet_name = params.get("sheet_name")

    if not start_cell or not end_cell:
        return _config_error("参数缺少 start_cell 或 end_cell")

    wb = _load_workbook(file_path)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
        end_row = int(''.join(filter(str.isdigit, end_cell)))

        negative_cells = 0
        red_negative_cells = 0
        details = []

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    continue
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    negative_cells += 1
                    # 红色可来自字体颜色或红色背景填充，任一命中即视为已标红
                    if _cell_is_red(cell):
                        red_negative_cells += 1
                    else:
                        details.append(f"{get_column_letter(col)}{row}: 负值但非红色")

        if negative_cells == 0:
            return _ok(f"区域 {start_cell}:{end_cell} 内无负值单元格，无需红色标记")

        if red_negative_cells == negative_cells:
            return _ok(f"全部 {negative_cells} 个负值单元格都标记为红色")

        ratio = red_negative_cells / negative_cells
        return _partial(ratio, f"红色负值比例 {ratio:.1%}（{red_negative_cells}/{negative_cells}）")
    finally:
        wb.close()


def check_sorted_columns_exist(file_path: str, params: dict) -> dict:
    """
    检查多个xlsx文件中是否存在排序列。

    输入:
        file_path: xlsx 文件路径
        params:
            column (str/int): 需要检查的列
            order (str): "asc" 或 "desc"
            sheet_name (str, 可选): 工作表名称
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    # 此函数实际需要检查多个文件，所以使用 check_sort_order
    # 这里复用 check_sort_order 的逻辑
    return check_sort_order(file_path, params)


def _worksheet_value_rows(ws, start_row: int, column_count: int) -> List[tuple]:
    """提取工作表的固定列数值行。

    输入:
        ws: openpyxl Worksheet 对象。
        start_row: 起始行号（1-based）。
        column_count: 需要提取的列数。
    输出:
        从 start_row 到末行的元组列表，全空行被忽略。
    """
    rows = []
    for row_index in range(start_row, ws.max_row + 1):
        row = tuple(
            ws.cell(row=row_index, column=column_index).value
            for column_index in range(1, column_count + 1)
        )
        if any(value is not None for value in row):
            rows.append(row)
    return rows


def _rows_digest(rows: List[tuple]) -> str:
    """计算工作表值行的稳定 SHA-256。

    输入:
        rows: 值行元组列表。
    输出:
        UTF-8 JSON 序列化后的六十四位 SHA-256 十六进制字符串。
    """
    payload = json.dumps(rows, ensure_ascii=False, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _is_monotonic_values(values: List[Any], order: str) -> bool:
    """判断值序列是否按指定方向单调。

    输入:
        values: 待比较的非空值列表。
        order: "asc" 或 "desc"。
    输出:
        序列可比且单调时返回 True；混合类型返回 False。
    """
    try:
        if order == "desc":
            return all(left >= right for left, right in zip(values, values[1:]))
        return all(left <= right for left, right in zip(values, values[1:]))
    except TypeError:
        return False


def _has_distinct_sort_assignment(candidates: List[set], requirement_count: int) -> bool:
    """检查排序要求能否一对一分配给不同文件。

    输入:
        candidates: 每个副本可覆盖的排序要求索引集合。
        requirement_count: 排序要求总数。
    输出:
        存在覆盖全部要求的不同文件匹配时返回 True。
    """
    matched_file_by_requirement = {}

    def assign(file_index: int, seen: set) -> bool:
        """尝试为单个副本分配一个尚可增广的排序要求。

        输入:
            file_index: candidates 中的副本索引。
            seen: 本轮增广已访问的要求索引。
        输出:
            成功新建或重排一条匹配时返回 True。
        """
        for requirement_index in candidates[file_index]:
            if requirement_index in seen:
                continue
            seen.add(requirement_index)
            previous_file = matched_file_by_requirement.get(requirement_index)
            if previous_file is None or assign(previous_file, seen):
                matched_file_by_requirement[requirement_index] = file_index
                return True
        return False

    for file_index in range(len(candidates)):
        assign(file_index, set())
    return len(matched_file_by_requirement) == requirement_count


def check_sorted_copies_preserve_rows(result_dir: str, params: dict) -> dict:
    """检查多个排序副本是否保留源表全部数据并覆盖指定列。

    输入:
        result_dir: Agent 产出根目录。
        params:
            source_filename (str): 应保留的源工作簿文件名。
            header_row (int): 表头行，默认 1。
            column_count (int): 参与完整性比较的列数。
            required_sorts (list[dict]): 每项包含 column 与 order。
            source_data_sha256 (str): 初始源值行的稳定摘要。
            expected_data_rows (int): 参与排序的初始数据行数；表尾
                汇总行不计入。
            expected_total_rows (int): 含表尾汇总行的源值行总数。
    输出:
        源数据未被替换、每个副本行多重集一致，且每个排序要求
        均由不同副本覆盖时返回通过。
    """
    source_filename = params.get("source_filename", "")
    header_row = int(params.get("header_row", 1))
    column_count = int(params.get("column_count", 0))
    required_sorts = params.get("required_sorts", [])
    expected_digest = params.get("source_data_sha256", "")
    expected_data_rows = int(params.get("expected_data_rows", 0))
    expected_total_rows = int(params.get("expected_total_rows", 0))
    if not source_filename or column_count <= 0 or not required_sorts or not expected_digest:
        return _config_error(
            "参数缺少 source_filename/column_count/required_sorts/source_data_sha256"
        )

    source_matches = sorted(
        path for path in glob.glob(os.path.join(result_dir, "**", source_filename), recursive=True)
        if os.path.isfile(path)
    )
    if len(source_matches) != 1:
        return _fail(f"源文件 {source_filename} 期望唯一，实际 {len(source_matches)} 个")

    source_wb = _load_workbook(source_matches[0], data_only=True)
    if source_wb is None:
        return _fail(f"无法打开源文件 {source_filename}")
    try:
        source_ws = source_wb.active
        source_rows = _worksheet_value_rows(source_ws, header_row + 1, column_count)
    finally:
        source_wb.close()
    if expected_total_rows and len(source_rows) != expected_total_rows:
        return _fail(f"源值行总数 {len(source_rows)} != {expected_total_rows}")
    if expected_data_rows <= 0 or expected_data_rows > len(source_rows):
        return _config_error("expected_data_rows 必须在源值行数范围内")
    if _rows_digest(source_rows) != expected_digest:
        return _fail("源工作簿数据已被替换或篡改")

    source_counter = Counter(source_rows)
    all_xlsx = sorted(
        path for path in glob.glob(os.path.join(result_dir, "**", "*.xlsx"), recursive=True)
        if os.path.isfile(path) and not os.path.basename(path).startswith("~$")
    )
    copy_paths = [path for path in all_xlsx if os.path.realpath(path) != os.path.realpath(source_matches[0])]
    if len(copy_paths) < len(required_sorts):
        return _fail(f"仅找到 {len(copy_paths)} 个排序副本，期望 {len(required_sorts)} 个")

    candidates = []
    invalid_copies = []
    for path in copy_paths:
        workbook = _load_workbook(path, data_only=True)
        if workbook is None:
            invalid_copies.append(os.path.basename(path))
            continue
        try:
            worksheet = workbook.active
            rows = _worksheet_value_rows(worksheet, header_row + 1, column_count)
        finally:
            workbook.close()
        if Counter(rows) != source_counter:
            invalid_copies.append(os.path.basename(path))
            continue
        coverage = set()
        sortable_rows = rows[:expected_data_rows]
        for requirement_index, requirement in enumerate(required_sorts):
            column = requirement.get("column")
            try:
                column_index = (
                    int(column) if isinstance(column, int)
                    else column_index_from_string(str(column))
                ) - 1
            except (TypeError, ValueError):
                return _config_error(f"无法解析排序列: {column}")
            values = [
                row[column_index] for row in sortable_rows
                if row[column_index] is not None
            ]
            if _is_monotonic_values(values, requirement.get("order", "asc")):
                coverage.add(requirement_index)
        candidates.append(coverage)

    if invalid_copies:
        return _fail(
            f"排序副本与源数据行多重集不一致: "
            f"{', '.join(invalid_copies[:5])}"
        )
    if not _has_distinct_sort_assignment(candidates, len(required_sorts)):
        return _fail("未由四个不同完整副本覆盖全部排序列")
    return _ok(
        f"{len(required_sorts)} 个排序要求均由保留全部 "
        f"{len(source_rows)} 行源数据的不同副本覆盖"
    )


def check_no_na_values(file_path: str, params: dict) -> dict:
    """
    检查xlsx文件中是否存在N/A值。

    输入:
        file_path: xlsx 文件路径
        params:
            sheet_name (str, 可选): 工作表名称
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    sheet_name = params.get("sheet_name")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        na_cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    val_str = str(cell.value).upper().strip()
                    if val_str in ("N/A", "NA", "#N/A", "#NA", "NAN", "NULL"):
                        na_cells.append(f"{get_column_letter(cell.column)}{cell.row}")

        if not na_cells:
            return _ok("文件中无N/A值")

        ratio = 0.0
        return _fail(f"发现 {len(na_cells)} 个N/A值: {', '.join(na_cells[:5])}")
    finally:
        wb.close()


def check_sequential_numbers(file_path: str, params: dict) -> dict:
    """
    检查序列号是否从0开始按顺序递增（步长为1）。

    输入:
        file_path: xlsx 文件路径
        params:
            start_after (str): 从该单元格下方开始检查，如 "A3"
            sheet_name (str, 可选): 工作表名称
            threshold (float): 符合比例阈值，默认 0.8
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    start_after = params.get("start_after", "")
    sheet_name = params.get("sheet_name")
    threshold = params.get("threshold", 0.8)
    expected_format = params.get("format")

    if not start_after:
        return _config_error("参数缺少 start_after")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        # 解析起始单元格
        start_col = column_index_from_string(''.join(filter(str.isalpha, start_after)))
        start_row = int(''.join(filter(str.isdigit, start_after))) + 1

        # 收集序列值
        values = []
        for row in range(start_row, ws.max_row + 1):
            val = ws.cell(row=row, column=start_col).value
            if val is not None:
                values.append(val)

        if len(values) < 2:
            return _ok("序列值数量不足")

        def _matches_expected(value, expected_num: int) -> bool:
            if expected_format:
                expected_text = str(expected_format).replace("#", str(expected_num))
                return str(value).strip() == expected_text
            if value == expected_num:
                return True
            if isinstance(value, str) and value.strip().isdigit():
                return int(value.strip()) == expected_num
            return False

        # 检查是否为从0开始的递增序列，可选格式如 No.#
        expected = 0
        correct_count = 0
        for val in values:
            if _matches_expected(val, expected):
                correct_count += 1
            expected += 1

        ratio = correct_count / len(values)
        if ratio >= threshold:
            return _ok(f"序列正确 {ratio:.1%}（{correct_count}/{len(values)}）")

        return _partial(ratio, f"序列正确率 {ratio:.1%}（{correct_count}/{len(values)}）")
    finally:
        wb.close()


def check_multi_cell_values(file_path: str, params: dict) -> dict:
    """
    检查多个单元格的值是否匹配期望值。

    输入:
        file_path: xlsx 文件路径
        params:
            cells (list[dict]): 单元格检查列表，格式：
                [
                    {"cell": "B4", "expected": 1096},
                    {"cell": "B5", "contains": "London"},
                    {"cell": "C3", "contains_any": ["San Diego", "Mexico City"]}
                ]
            tolerance (float): 数值容差，默认 0.01
            sheet_name (str, 可选): 工作表名称
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    cells = params.get("cells", [])
    tolerance = params.get("tolerance", 0.01)
    sheet_name = params.get("sheet_name")

    if not cells:
        return _config_error("参数缺少 cells")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        total = len(cells)
        matched = 0
        details = []

        for check in cells:
            cell_ref = check.get("cell", "")
            expected = check.get("expected")
            expected_str = check.get("contains")
            expected_any = check.get("contains_any")
            tolerance_val = check.get("tolerance", tolerance)

            if not cell_ref:
                continue

            actual = ws[cell_ref].value

            if expected_any is not None:
                # 检查是否包含列表中的任意一个字符串
                actual_str = str(actual) if actual is not None else ""
                matched_any = False
                for candidate in expected_any:
                    if str(candidate).lower() in actual_str.lower():
                        matched_any = True
                        break
                if matched_any:
                    matched += 1
                else:
                    details.append(f"{cell_ref}: '{actual_str}' 不包含任意 {expected_any}")
            elif expected_str is not None:
                # 检查是否包含字符串
                actual_str = str(actual) if actual is not None else ""
                if expected_str.lower() in actual_str.lower():
                    matched += 1
                else:
                    details.append(f"{cell_ref}: '{actual_str}' 不包含 '{expected_str}'")
            elif expected is not None:
                # 检查数值或字符串相等
                if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
                    if expected == 0:
                        passed = abs(actual) < 1e-9
                    else:
                        passed = abs(actual - expected) / abs(expected) <= tolerance_val
                    if passed:
                        matched += 1
                    else:
                        details.append(f"{cell_ref}: {actual} (期望 {expected})")
                else:
                    actual_str = str(actual).strip() if actual is not None else ""
                    expected_str_val = str(expected).strip()
                    if actual_str.lower() == expected_str_val.lower():
                        matched += 1
                    else:
                        details.append(f"{cell_ref}: '{actual_str}' (期望 '{expected_str_val}')")
            else:
                details.append(f"{cell_ref}: 未指定期望值")

        if matched == total:
            return _ok(f"全部 {total} 个单元格值正确")

        ratio = matched / total
        return _partial(ratio, f"{matched}/{total} 匹配: {', '.join(details[:5])}")
    finally:
        wb.close()


def check_cells_filled(file_path: str, params: dict) -> dict:
    """
    检查指定单元格区域是否都有值。

    输入:
        file_path: xlsx 文件路径
        params:
            start_cell (str): 起始单元格，如 "A4"
            end_cell (str): 结束单元格，如 "C6"
            sheet_name (str, 可选): 工作表名称
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    start_cell = params.get("start_cell", "")
    end_cell = params.get("end_cell", "")
    sheet_name = params.get("sheet_name")

    if not start_cell or not end_cell:
        return _config_error("参数缺少 start_cell 或 end_cell")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
        end_row = int(''.join(filter(str.isdigit, end_cell)))

        total = 0
        filled = 0
        empty_cells = []

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                total += 1
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    filled += 1
                else:
                    empty_cells.append(f"{get_column_letter(col)}{row}")

        ratio = filled / total if total > 0 else 0.0
        if filled == total:
            return _ok(f"全部 {total} 个单元格都有值")

        return _partial(ratio, f"{filled}/{total} 有值，空白: {', '.join(empty_cells[:5])}")
    finally:
        wb.close()


# ---- 以下一项自 2026-07 审计修复线补回 ----
# 线 C 的 Excel-003 仍按固定列号判排序，无法发现「B 列排好但真实销售额列未排序」
# 的情形；此函数按表头关键字定位列，为 BatchOperationExcel-003 所引用。


def check_sort_order_by_header_keywords(file_path: str, params: dict) -> dict:
    """
    按表头关键字定位列并检查排序。

    比固定列号更贴近“按 sales amount 排序”这类任务，避免 agent 只把
    B 列排好但真实销售额列未排序也通过。
    """
    keywords = [str(k).lower() for k in params.get("header_keywords", [])]
    order = params.get("order", "desc")
    sheet_name = params.get("sheet_name")
    header_rows = int(params.get("header_rows", 5))
    skip_header = params.get("skip_header", True)

    if not keywords:
        return _config_error("参数缺少 header_keywords")

    wb = _load_workbook(file_path, data_only=True)
    if wb is None:
        return _fail(f"无法打开文件: {file_path}")

    try:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            return _fail(f"工作表不存在: {sheet_name}")

        best = None
        for row_idx in range(1, min(header_rows, ws.max_row) + 1):
            for col_idx in range(1, ws.max_column + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                text = str(value).lower()
                if all(keyword in text for keyword in keywords):
                    best = (row_idx, col_idx, value)
                    break
            if best:
                break

        if not best:
            # 定位不到待检查的列属于规则配置与数据不匹配，不是 agent 失败；
            # 判 FAIL 会让该任务对所有模型系统性失败，故按评价器故障处理，
            # 使其退出成功率分母并在报告中显式暴露。
            return _config_error(
                f"未找到包含关键字 {keywords} 的表头（前 {header_rows} 行）"
            )

        header_row, col_idx, header_value = best
        start_row = header_row + 1 if skip_header else header_row
        values = []
        for row_idx in range(start_row, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            if isinstance(val, str):
                normalized = val.replace(",", "").strip()
                try:
                    val = float(normalized)
                except ValueError:
                    continue
            values.append(val)

        if len(values) <= 1:
            return _ok(f"列 {header_value} 数据量不足，无需排序检查")

        try:
            is_ascending = all(values[i] <= values[i + 1] for i in range(len(values) - 1))
            is_descending = all(values[i] >= values[i + 1] for i in range(len(values) - 1))
        except TypeError:
            return _fail(f"列 {header_value} 含不可比较值")

        if order == "asc" and is_ascending:
            return _ok(f"列 {header_value} 升序排列正确（{len(values)} 行）")
        if order == "desc" and is_descending:
            return _ok(f"列 {header_value} 降序排列正确（{len(values)} 行）")

        inversions = 0
        total_pairs = 0
        for i in range(len(values) - 1):
            total_pairs += 1
            if order == "asc" and values[i] > values[i + 1]:
                inversions += 1
            elif order == "desc" and values[i] < values[i + 1]:
                inversions += 1

        ratio = 1.0 - (inversions / total_pairs) if total_pairs else 1.0
        return _partial(
            ratio,
            f"列 {header_value} 排序不完全（有序比例 {ratio:.1%}，期望 {order}）"
        )
    finally:
        wb.close()



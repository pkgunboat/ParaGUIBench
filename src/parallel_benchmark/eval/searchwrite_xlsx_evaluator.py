#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Search&Write xlsx 任务评估器。

通过逐单元格模糊匹配，评估 Agent 在共享 xlsx 文档中填写的内容与 Ground Truth 的一致性。

评估逻辑:
    1. 加载模板 xlsx → GT xlsx → Agent 结果 xlsx
    2. 遍历 GT 所有单元格，筛选出"需评估单元格"（GT 有值 AND 模板对应位置为空）
    3. 对每个需评估单元格，根据值类型选择匹配策略
    4. 返回 score = 匹配数 / 总需评估单元格数

值类型匹配策略:
    - 年份 (1800-2100): 精确匹配
    - 数值 (可转 float): 允许 ±1% 误差
    - URL (含 .com/.org/.edu 等): normalize 后精确匹配
    - N/A: Agent 也为 N/A 或为空均算通过
    - 一般文本: 忽略大小写 + 去首尾空格后精确匹配
"""

from __future__ import annotations

import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ============================================================
# 值类型检测与匹配
# ============================================================


def _is_year(value: str) -> bool:
    """
    判断是否为年份（4 位数字，1800-2100）。

    输入:
        value: 字符串值

    输出:
        bool
    """
    return bool(re.fullmatch(r"\d{4}", value.strip())) and 1800 <= int(value.strip()) <= 2100


def _is_number(value: str) -> bool:
    """
    判断是否为数值（可转换为 float）。

    输入:
        value: 字符串值

    输出:
        bool
    """
    try:
        float(value.strip().replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False


def _parse_number(value: str) -> float:
    """
    解析数值字符串为 float。

    输入:
        value: 字符串值

    输出:
        float
    """
    return float(value.strip().replace(",", ""))


_URL_PATTERN = re.compile(
    r"(https?://|www\.)|"
    r"\.(com|org|edu|net|gov|io|co|uk|cn|jp|de|fr|au|ca|info|biz)(/|$)",
    re.IGNORECASE,
)


def _is_url(value: str) -> bool:
    """
    判断是否为 URL（含常见域名后缀或 http/www 前缀）。

    输入:
        value: 字符串值

    输出:
        bool
    """
    return bool(_URL_PATTERN.search(value.strip()))


def normalize_url(url: str) -> str:
    """
    URL 归一化：去 scheme、去 www.、去末尾斜杠、转小写。

    输入:
        url: 原始 URL 字符串

    输出:
        归一化后的 URL
    """
    url = url.strip().lower()
    url = re.sub(r"^https?://", "", url)
    url = re.sub(r"^www\.", "", url)
    url = url.rstrip("/")
    return url


def _is_na(value: str) -> bool:
    """
    判断是否为 N/A 标记。

    输入:
        value: 字符串值

    输出:
        bool
    """
    return value.strip().lower() in {"n/a", "na", "n.a.", "none", "-", "—", ""}


def _cell_to_str(value: Any) -> str:
    """
    将单元格值转为字符串（处理 None、数字等类型）。

    输入:
        value: openpyxl 单元格值（可能是 str/int/float/None）

    输出:
        字符串
    """
    if value is None:
        return ""
    if isinstance(value, float):
        # 整数值去掉 .0（如 2012.0 → "2012"）
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


# ============================================================
# 单元格匹配
# ============================================================


def match_cell(gt_val: str, result_val: str) -> Tuple[bool, str]:
    """
    根据 GT 值的类型选择匹配策略，判断 Agent 结果是否匹配。

    输入:
        gt_val: Ground Truth 单元格值（字符串）
        result_val: Agent 结果单元格值（字符串）

    输出:
        (matched: bool, match_type: str)
        match_type 为 "year"/"number"/"url"/"na"/"text"/"empty_result"
    """
    # N/A 特殊处理：GT 为 N/A 时，Agent 为 N/A 或空均通过
    if _is_na(gt_val):
        if _is_na(result_val) or result_val.strip() == "":
            return True, "na"
        return False, "na"

    # Agent 结果为空 → 未填写
    if result_val.strip() == "":
        return False, "empty_result"

    # 年份精确匹配
    if _is_year(gt_val):
        return gt_val.strip() == result_val.strip(), "year"

    # 数值匹配（±1% 误差）
    if _is_number(gt_val) and _is_number(result_val):
        gt_num = _parse_number(gt_val)
        result_num = _parse_number(result_val)
        if gt_num == 0:
            return result_num == 0, "number"
        return abs(gt_num - result_num) / abs(gt_num) <= 0.01, "number"

    # URL 归一化匹配
    if _is_url(gt_val):
        if _is_url(result_val):
            return normalize_url(gt_val) == normalize_url(result_val), "url"
        return False, "url"

    # 一般文本：忽略大小写 + 去首尾空格
    return gt_val.strip().lower() == result_val.strip().lower(), "text"


# ============================================================
# 需评估单元格检测
# ============================================================


def _find_evaluable_cells(
    template_ws,
    gt_ws,
) -> List[Tuple[int, int, str, str]]:
    """
    找出需要评估的单元格：GT 有值且模板对应位置为空。

    输入:
        template_ws: 模板 worksheet
        gt_ws: Ground Truth worksheet

    输出:
        [(row, col, cell_coord, gt_value_str), ...]
    """
    evaluable = []
    for row in range(1, gt_ws.max_row + 1):
        for col in range(1, gt_ws.max_column + 1):
            gt_cell = gt_ws.cell(row=row, column=col)
            gt_val = _cell_to_str(gt_cell.value)
            if not gt_val:
                continue

            # 检查模板中对应位置是否为空
            template_cell = template_ws.cell(row=row, column=col)
            template_val = _cell_to_str(template_cell.value)
            if template_val:
                # 模板已有值 → 这是表头或预填内容，不评估
                continue

            coord = gt_cell.coordinate  # 如 "B4"
            evaluable.append((row, col, coord, gt_val))

    return evaluable


# ============================================================
# 主评估函数
# ============================================================


def evaluate(
    template_path: str,
    gt_path: str,
    result_path: str,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    评估单个 xlsx 文件的 Search&Write 结果。

    输入:
        template_path: 模板 xlsx 文件路径
        gt_path: Ground Truth xlsx 文件路径
        result_path: Agent 结果 xlsx 文件路径
        sheet_name: 指定 sheet 名称（None 则使用第一个 sheet）

    输出:
        {
            "score": float,           # 匹配率 (0.0~1.0)
            "pass": bool,             # score >= 0.5
            "total_cells": int,       # 需评估的单元格总数
            "matched_cells": int,     # 匹配的单元格数
            "cell_details": {         # 逐单元格明细
                "B4": {"gt": "London", "result": "london", "matched": True, "type": "text"},
                ...
            }
        }
    """
    template_wb = load_workbook(template_path, data_only=True)
    gt_wb = load_workbook(gt_path, data_only=True)
    result_wb = load_workbook(result_path, data_only=True)

    # 选择 sheet
    if sheet_name:
        template_ws = template_wb[sheet_name]
        gt_ws = gt_wb[sheet_name]
        result_ws = result_wb[sheet_name]
    else:
        template_ws = template_wb.active
        gt_ws = gt_wb.active
        result_ws = result_wb.active

    # 找出需评估的单元格
    evaluable = _find_evaluable_cells(template_ws, gt_ws)

    if not evaluable:
        return {
            "score": 0.0,
            "pass": False,
            "total_cells": 0,
            "matched_cells": 0,
            "cell_details": {},
        }

    # 逐单元格评估
    matched_count = 0
    cell_details: Dict[str, Dict[str, Any]] = {}

    for row, col, coord, gt_val in evaluable:
        result_val = _cell_to_str(result_ws.cell(row=row, column=col).value)  # type: ignore[union-attr]

        matched, match_type = match_cell(gt_val, result_val)
        if matched:
            matched_count += 1

        cell_details[coord] = {
            "gt": gt_val,
            "result": result_val,
            "matched": matched,
            "type": match_type,
        }

    score = matched_count / len(evaluable)
    return {
        "score": score,
        "pass": score >= 1.0 - 1e-6,
        "total_cells": len(evaluable),
        "matched_cells": matched_count,
        "cell_details": cell_details,
    }


def evaluate_multi_file(
    file_pairs: List[Dict[str, str]],
) -> Dict[str, Any]:
    """
    评估多个 xlsx 文件（如 SAW-004 有 2 个文件），取平均分。

    输入:
        file_pairs: [
            {"template": "path", "gt": "path", "result": "path", "name": "filename"},
            ...
        ]

    输出:
        {
            "score": float,           # 各文件平均得分
            "pass": bool,             # score >= 0.5
            "total_cells": int,       # 所有文件需评估单元格总数
            "matched_cells": int,     # 所有文件匹配的单元格总数
            "file_results": {         # 按文件名分组的结果
                "filename.xlsx": { ... evaluate() 返回结构 ... },
                ...
            }
        }
    """
    file_results: Dict[str, Dict[str, Any]] = {}
    total_cells = 0
    total_matched = 0

    for pair in file_pairs:
        name = pair.get("name", "unknown")
        try:
            result = evaluate(
                template_path=pair["template"],
                gt_path=pair["gt"],
                result_path=pair["result"],
            )
        except Exception as exc:
            result = {
                "score": 0.0,
                "pass": False,
                "total_cells": 0,
                "matched_cells": 0,
                "cell_details": {},
                "error": str(exc),
            }

        file_results[name] = result
        total_cells += result["total_cells"]
        total_matched += result["matched_cells"]

    # 总分用单元格级汇总（而非文件级平均），对单元格数不均匀的情况更公平
    overall_score = total_matched / total_cells if total_cells > 0 else 0.0

    return {
        "score": overall_score,
        "pass": overall_score >= 1.0 - 1e-6,
        "total_cells": total_cells,
        "matched_cells": total_matched,
        "file_results": file_results,
    }

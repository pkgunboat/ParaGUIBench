"""
Master Table 报告生成器。

从 master.csv 全量读取后，渲染为：
    reports/master_summary.xlsx — 8 个 Sheet
    reports/master_summary.md   — Markdown 摘要

Sheet 清单（详见 spec §7）:
    1. Raw          — 全量原表
    2. Main         — mode=full AND condition∈{baseline, gui_only_seed18}
    3. Plan Ablation
    4. GUI Ablation
    5. GUI-Only Ablation
    6. Parallelism Ablation
    7. Oracle Ablation
    8. Coverage     — 每组 (mode, condition, pipeline) 的任务覆盖率
"""

import os
import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

import master_table as mt

MAIN_CONDITIONS = ("baseline", "gui_only_seed18")


# ============================================================
# 数据加载
# ============================================================

def _load_typed() -> List[Dict[str, Any]]:
    """从 master.csv 读取已做类型转换的行列表。"""
    return mt.load_master()


# ============================================================
# Sheet: Main
# ============================================================

def select_main(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """筛选 mode=full 且 condition ∈ MAIN_CONDITIONS 的行。"""
    return [r for r in rows
            if r.get("mode") == "full"
            and r.get("condition") in MAIN_CONDITIONS]


# ============================================================
# Sheet: Coverage
# ============================================================

def build_coverage(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    按 (mode, condition, pipeline) 聚合：应跑 / pass / empty / error / needs_rerun。

    输入:
        rows: master.csv 的所有行

    输出:
        list of dict，每个 dict 含:
            mode, condition, pipeline,
            total, pass, empty, error, needs_rerun, coverage (float 0-1)
    """
    by_group: Dict[tuple, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        key = (r["mode"], r["condition"], r["pipeline"])
        by_group[key].append(r)

    out = []
    for (mode, cond, pipeline), group in sorted(by_group.items()):
        total = len(group)
        outcomes = _summarize_master_outcomes(group)
        passed = outcomes["pass"]
        empty = sum(1 for g in group if g.get("empty") is True)
        error = sum(1 for g in group if g.get("error") is True)
        needs_rerun = sum(1 for g in group if g.get("needs_rerun") is True)
        eligible = total - outcomes["skip"]
        ran = max(0, eligible - empty)
        coverage = ran / eligible if eligible else 0.0
        out.append({
            "mode": mode, "condition": cond, "pipeline": pipeline,
            "total": total, "eligible": eligible,
            "ran": ran,
            "pass": passed,
            "skip": outcomes["skip"],
            "evaluator_error": outcomes["evaluator_error"],
            "interrupted": outcomes["interrupted"],
            "empty": empty,
            "error": error,
            "needs_rerun": needs_rerun,
            "coverage": round(coverage, 3),
        })
    return out


# ============================================================
# 消融子表筛选
# ============================================================

def _base_ablation_filter(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """所有消融子表的前置：mode=ablation。"""
    return [r for r in rows if r.get("mode") == "ablation"]


def select_plan_ablation(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Plan Ablation: mode=ablation AND agent_mode=plan AND gui_agent=seed18
                   AND vms_per_task=5 AND NOT oracle_plan_injected。
    行维度：plan_model。
    """
    return [r for r in _base_ablation_filter(rows)
            if r.get("agent_mode") == "plan"
            and r.get("gui_agent") == "seed18"
            and r.get("vms_per_task") == 5
            and not r.get("oracle_plan_injected", False)]


def select_gui_ablation(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    GUI Ablation: mode=ablation AND agent_mode=plan AND plan_model=gpt-5.4
                  AND vms_per_task=5 AND NOT oracle_plan_injected。
    行维度：gui_agent。
    """
    return [r for r in _base_ablation_filter(rows)
            if r.get("agent_mode") == "plan"
            and r.get("plan_model") == "gpt-5.4"
            and r.get("vms_per_task") == 5
            and not r.get("oracle_plan_injected", False)]


def select_gui_only_ablation(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """GUI-Only Ablation: mode=ablation AND agent_mode=gui_only AND vms_per_task=1。"""
    return [r for r in _base_ablation_filter(rows)
            if r.get("agent_mode") == "gui_only"
            and r.get("vms_per_task") == 1]


def select_parallelism_ablation(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Parallelism: mode=ablation AND agent_mode=plan AND plan_model=gpt-5.4
                 AND gui_agent=seed18 AND NOT oracle_plan_injected。
    行维度：vms_per_task。
    """
    return [r for r in _base_ablation_filter(rows)
            if r.get("agent_mode") == "plan"
            and r.get("plan_model") == "gpt-5.4"
            and r.get("gui_agent") == "seed18"
            and not r.get("oracle_plan_injected", False)]


def select_oracle_ablation(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Oracle Ablation: mode=ablation AND condition ∈ {baseline, oracle_plan}。"""
    return [r for r in _base_ablation_filter(rows)
            if r.get("condition") in ("baseline", "oracle_plan")]


# ============================================================
# 聚合指标（复用 report_generator 的思路）
# ============================================================

def _is_eval_error_score(score: Any) -> bool:
    """判断分数是否为评价器故障哨兵。

    输入参数：score 为记录中的分数字段，可能是数值、空串或 None。
    输出返回值：分数为负数时返回 True；代码库中没有任何合法模型结果
    使用负分，因此负分一律代表评价器自身故障或 GT 缺失。
    """
    return isinstance(score, (int, float)) and not isinstance(score, bool) and score < 0


def _classify_master_outcome(row: Dict[str, Any]) -> str:
    """按互斥优先级判定一条 master 记录的评价结果。

    功能：依次识别 SKIP、EVALUATOR_ERROR、INTERRUPTED、PASS、FAIL；空任务、
    执行错误及旧 CSV 中没有可用 pass/status 的记录归为 NOT_EVALUATED。
    输入参数：row 为 ``master.csv`` 的一条已类型化记录。
    输出返回值：上述六种大写状态之一。
    """
    status = str(row.get("status") or "").strip().lower()
    if status == "skip":
        return "SKIP"
    if status == "evaluator_error":
        return "EVALUATOR_ERROR"
    # 旧 CSV 没有 status 列，评价器故障只以 score=-1 哨兵体现；缺这一路会把
    # 评价器故障当成模型失败并计入成功率分母。
    if not status and _is_eval_error_score(row.get("score")):
        return "EVALUATOR_ERROR"
    if row.get("interrupted") is True:
        return "INTERRUPTED"
    if row.get("empty") is True or row.get("error") is True:
        return "NOT_EVALUATED"
    if row.get("pass") is True:
        return "PASS"
    if row.get("pass") is False:
        return "FAIL"
    return "NOT_EVALUATED"


def _summarize_master_outcomes(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """汇总 master 记录并只用 PASS+FAIL 计算有效评价通过率。

    功能：以 ``_classify_master_outcome`` 的互斥分类统计六种结果，保证
    SKIP、评价器错误、中断和未评价记录均不进入有效评价分母。
    输入参数：rows 为同一实验分组的 master 记录列表。
    输出返回值：含 total、valid、各状态计数和 rate 的字典；无有效评价时
    rate 返回 None。
    """
    counts = {
        "pass": 0,
        "fail": 0,
        "skip": 0,
        "evaluator_error": 0,
        "interrupted": 0,
        "not_evaluated": 0,
    }
    key_by_status = {
        "PASS": "pass",
        "FAIL": "fail",
        "SKIP": "skip",
        "EVALUATOR_ERROR": "evaluator_error",
        "INTERRUPTED": "interrupted",
        "NOT_EVALUATED": "not_evaluated",
    }
    for row in rows:
        counts[key_by_status[_classify_master_outcome(row)]] += 1
    valid = counts["pass"] + counts["fail"]
    return {
        "total": len(rows),
        "valid": valid,
        **counts,
        "rate": counts["pass"] / valid if valid else None,
    }


def _aggregate_by(rows: List[Dict[str, Any]], group_key: str) -> List[Dict[str, Any]]:
    """
    按 group_key 列聚合，计算 Pass/Rate/Token/Cost/Time 等合计。

    输入:
        rows: 已筛选的子表行
        group_key: 作为分组维度的列名（如 "plan_model"）

    输出:
        list of dict，每个 dict 含 group_key + 聚合指标。
    """
    grouped: Dict[Any, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        grouped[r.get(group_key)].append(r)

    out = []
    for key, group in sorted(grouped.items(), key=lambda kv: str(kv[0])):
        outcomes = _summarize_master_outcomes(group)
        sum_plan = _safe_sum(group, "plan_rounds")
        sum_gui_total = _safe_sum(group, "gui_rounds_total")
        sum_gui_seq = _safe_sum(group, "gui_steps_sequential")
        parallelism = (sum_gui_total / sum_gui_seq) if sum_gui_seq else 0.0
        out.append({
            group_key: key,
            "total": outcomes["total"],
            "valid": outcomes["valid"],
            "pass": outcomes["pass"],
            "fail": outcomes["fail"],
            "skip": outcomes["skip"],
            "evaluator_error": outcomes["evaluator_error"],
            "interrupted": outcomes["interrupted"],
            "not_evaluated": outcomes["not_evaluated"],
            "rate": (
                round(outcomes["rate"], 3)
                if outcomes["rate"] is not None
                else None
            ),
            "plan_rounds": sum_plan,
            "gui_rounds_total": sum_gui_total,
            "gui_steps_sequential": sum_gui_seq,
            "parallelism": round(parallelism, 2),
            "token_plan": _safe_sum(group, "token_plan"),
            "token_gui": _safe_sum(group, "token_gui"),
            "token_total": _safe_sum(group, "token_total"),
            "cost_usd": round(_safe_sum(group, "cost_usd"), 2),
            "elapsed_time_sec": round(_safe_sum(group, "elapsed_time_sec"), 1),
        })
    return out


def _safe_sum(group: List[Dict[str, Any]], col: str) -> float:
    """对列求和，空字符串 / None 视为 0。"""
    s = 0.0
    for r in group:
        v = r.get(col)
        if isinstance(v, (int, float)):
            s += v
    return s


# ============================================================
# xlsx / md 渲染
# ============================================================

def rebuild_reports() -> str:
    """
    从 master.csv 渲染 8 个 Sheet 到 reports/master_summary.xlsx + md 摘要。

    输出:
        报告目录绝对路径。
    """
    rows = _load_typed()
    out_dir = os.path.join(mt.MASTER_DIR, "reports")
    os.makedirs(out_dir, exist_ok=True)

    sheets = [
        ("Raw", rows, None),
        ("Main", select_main(rows), ("condition", "pipeline")),
        ("Plan Ablation", select_plan_ablation(rows), ("plan_model", "pipeline")),
        ("GUI Ablation", select_gui_ablation(rows), ("gui_agent", "pipeline")),
        ("GUI-Only Ablation", select_gui_only_ablation(rows),
         ("gui_agent", "pipeline")),
        ("Parallelism Ablation", select_parallelism_ablation(rows),
         ("vms_per_task", "pipeline")),
        ("Oracle Ablation", select_oracle_ablation(rows), ("condition", "pipeline")),
        ("Coverage", build_coverage(rows), None),
    ]

    xlsx_path = os.path.join(out_dir, "master_summary.xlsx")
    _write_xlsx(xlsx_path, sheets)

    md_path = os.path.join(out_dir, "master_summary.md")
    _write_md(md_path, sheets)

    return out_dir


def _write_xlsx(path: str, sheets) -> None:
    """用 openpyxl 写多 Sheet。sheets: [(name, rows, group_by_keys)]。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl 未安装，请 pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")

    for name, rows, group_keys in sheets:
        ws = wb.create_sheet(title=name[:31])
        if name in ("Raw", "Coverage"):
            records = rows
        elif name == "Main":
            records = _aggregate_main_by_pipeline(rows)
        else:
            records = _aggregate_by(rows, group_keys[0]) if group_keys else rows
        if not records:
            ws.append(["(no data)"])
            continue
        headers = list(records[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for rec in records:
            ws.append([rec.get(h, "") for h in headers])

    wb.save(path)


def _aggregate_main_by_pipeline(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Main 表按 (condition, pipeline) 双维度展开，聚合列与 Ablation 子表对齐。

    输出字段（与 _aggregate_by 返回的 15 个指标对齐）:
        condition, pipeline, total, pass, fail, interrupted, rate,
        plan_rounds, gui_rounds_total, gui_steps_sequential, parallelism,
        token_plan, token_gui, token_total, cost_usd, elapsed_time_sec
    """
    grouped: Dict[tuple, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        grouped[(r["condition"], r["pipeline"])].append(r)
    out = []
    for (cond, pipeline), group in sorted(grouped.items()):
        outcomes = _summarize_master_outcomes(group)
        sum_plan = _safe_sum(group, "plan_rounds")
        sum_gui_total = _safe_sum(group, "gui_rounds_total")
        sum_gui_seq = _safe_sum(group, "gui_steps_sequential")
        parallelism = (sum_gui_total / sum_gui_seq) if sum_gui_seq else 0.0
        out.append({
            "condition": cond,
            "pipeline": pipeline,
            "total": outcomes["total"],
            "valid": outcomes["valid"],
            "pass": outcomes["pass"],
            "fail": outcomes["fail"],
            "skip": outcomes["skip"],
            "evaluator_error": outcomes["evaluator_error"],
            "interrupted": outcomes["interrupted"],
            "not_evaluated": outcomes["not_evaluated"],
            "rate": (
                round(outcomes["rate"], 3)
                if outcomes["rate"] is not None
                else None
            ),
            "plan_rounds": sum_plan,
            "gui_rounds_total": sum_gui_total,
            "gui_steps_sequential": sum_gui_seq,
            "parallelism": round(parallelism, 2),
            "token_plan": _safe_sum(group, "token_plan"),
            "token_gui": _safe_sum(group, "token_gui"),
            "token_total": _safe_sum(group, "token_total"),
            "cost_usd": round(_safe_sum(group, "cost_usd"), 2),
            "elapsed_time_sec": round(_safe_sum(group, "elapsed_time_sec"), 1),
        })
    return out


def _write_md(path: str, sheets) -> None:
    """简要 md 摘要：每 Sheet 第一行指标。"""
    lines = ["# Master Table Summary", ""]
    for name, rows, _ in sheets:
        lines.append(f"## {name}")
        lines.append(f"rows: {len(rows)}")
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def import_run(ablation_dir: str) -> int:
    """
    从一个 ablation_<timestamp>/ 目录导入所有条件 / pipeline 的结果。

    假设目录结构:
        <dir>/
        ├── ablation_summary.json
        └── <condition>/
            ├── ablation_config.json
            └── <pipeline>_results.json

    输入:
        ablation_dir: 绝对路径或相对 cwd

    输出:
        导入行数
    """
    import json
    ablation_dir = os.path.abspath(ablation_dir)
    if not os.path.isdir(ablation_dir):
        raise FileNotFoundError(ablation_dir)

    run_timestamp = os.path.basename(ablation_dir).replace("ablation_", "")
    total_imported = 0

    for cond_name in sorted(os.listdir(ablation_dir)):
        cond_dir = os.path.join(ablation_dir, cond_name)
        config_path = os.path.join(cond_dir, "ablation_config.json")
        if not os.path.isfile(config_path):
            continue
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        env = cfg.get("env", {})
        context = {
            "mode": cfg.get("mode", "ablation"),
            "condition": cfg.get("condition", cond_name),
            "run_timestamp": run_timestamp,
            "run_dir": os.path.relpath(cond_dir,
                                      os.path.join(mt.UBUNTU_ENV_DIR, "logs")),
            "plan_model": env.get("ABLATION_PLAN_MODEL", ""),
            "gui_agent": env.get("ABLATION_GUI_AGENT", ""),
            "agent_mode": cfg.get("agent_mode") or "plan",
            "vms_per_task": cfg.get("vms_per_task", 5),
            "oracle_plan_injected":
                env.get("ABLATION_ORACLE_PLAN_INJECTED", "") == "1",
        }
        for fname in os.listdir(cond_dir):
            if not fname.endswith("_results.json"):
                continue
            pipeline = fname.replace("_results.json", "")
            fpath = os.path.join(cond_dir, fname)
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    results = json.load(f)
            except Exception:
                continue
            expected = sorted(r.get("task_id") for r in results.values()
                              if r.get("task_id"))
            mt.upsert_results(
                results=results,
                expected_task_ids=expected,
                pipeline=pipeline,
                context=context,
            )
            total_imported += len(expected)
    return total_imported

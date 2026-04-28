"""
实验结果统计报告生成器。

功能:
    从 pipeline 结果字典生成 Markdown 和 Excel 格式的统计报告。

输出文件:
    report/summary.md   — Markdown 总表 + 各 pipeline 子表
    report/summary.xlsx — Excel 版（Sheet "Summary" + 各 pipeline Sheet）
"""

import os
import glob
import json
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# 详情表列定义
DETAIL_COLUMNS = [
    ("Task ID", "task_id"),
    ("Score", "score"),
    ("Pass", "pass"),
    ("Plan Rounds", "plan_rounds"),
    ("GUI Steps(Total)", "gui_rounds_total"),
    ("GUI Steps(Seq)", "gui_steps_sequential"),
    ("Parallelism", None),  # 计算列
    ("Token(Plan)", "token_plan"),
    ("Token(GUI)", "token_gui"),
    ("Token(Total)", "token_total"),
    ("Cost($)", "cost_usd"),
    ("Time(s)", "elapsed_time_sec"),
]


def _calc_parallelism(total: int, seq: int) -> str:
    """
    计算并行度。

    输入:
        total: GUI 总步骤数
        seq: GUI 串行等效步骤数

    输出:
        并行度字符串（如 "3.0x"），seq=0 时返回 "-"
    """
    if seq == 0:
        return "-"
    return f"{total / seq:.1f}x"


def _as_int(value: Any) -> int:
    """宽松转换统计字段，空值/非法值按 0 处理。"""
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _extract_gui_only_step_count(record: Dict[str, Any]) -> int:
    """
    从 gui_only execution_record 中提取 GUI step 数。

    优先使用完整 steps，其次 rounds_timing，再回退 summary.total_rounds。
    """
    if not isinstance(record, dict):
        return 0

    summary = record.get("summary", {})
    if not isinstance(summary, dict) or summary.get("mode") != "gui_only":
        return 0

    steps = record.get("steps")
    if isinstance(steps, list) and steps:
        return len(steps)

    rounds_timing = record.get("rounds_timing")
    if isinstance(rounds_timing, list) and rounds_timing:
        return len(rounds_timing)

    return _as_int(summary.get("total_rounds"))


def _execution_record_candidates(result: Dict[str, Any], output_dir: str) -> List[str]:
    """按可信度列出可能的 execution_record.json 路径。"""
    task_id = result.get("task_id")
    if not task_id or not output_dir:
        return []

    candidates: List[str] = []
    for key in ("result_dir", "result_dir_path"):
        result_dir = result.get(key)
        if result_dir:
            candidates.append(os.path.join(result_dir, "execution_record.json"))

    condition = result.get("condition")
    if condition:
        candidates.append(os.path.join(output_dir, condition, task_id,
                                       "execution_record.json"))

    candidates.append(os.path.join(output_dir, task_id, "execution_record.json"))
    candidates.extend(glob.glob(os.path.join(output_dir, "*", task_id,
                                             "execution_record.json")))

    seen = set()
    out = []
    for path in candidates:
        norm = os.path.abspath(path)
        if norm in seen:
            continue
        seen.add(norm)
        if os.path.isfile(norm):
            out.append(norm)
    return out


def _backfill_gui_only_steps(result: Dict[str, Any], output_dir: str) -> Dict[str, Any]:
    """
    当历史 *_results.json 里 GUI step 为 0 时，从任务目录的
    execution_record.json 回填 gui_only step 统计。
    """
    if result.get("agent_mode") != "gui_only":
        return result
    if _as_int(result.get("gui_rounds_total")) > 0 \
            or _as_int(result.get("gui_steps_sequential")) > 0:
        return result

    for path in _execution_record_candidates(result, output_dir):
        try:
            with open(path, "r", encoding="utf-8") as f:
                record = json.load(f)
        except Exception:
            continue
        step_count = _extract_gui_only_step_count(record)
        if step_count <= 0:
            continue
        enriched = dict(result)
        enriched["gui_rounds_total"] = step_count
        enriched["gui_steps_sequential"] = step_count
        enriched.setdefault("result_dir", os.path.dirname(path))
        return enriched

    return result


def enrich_results_with_gui_step_metrics(
    results: Dict[str, Dict[str, Any]],
    output_dir: str,
) -> Dict[str, Dict[str, Any]]:
    """回填 GUI-only step 指标，供报告和汇总复用。"""
    return {
        key: _backfill_gui_only_steps(result, output_dir)
        for key, result in results.items()
    }


def compute_results_summary(
    results: Dict[str, Dict[str, Any]],
    output_dir: str = "",
) -> Dict[str, Any]:
    """计算一次 pipeline 结果的机器可读汇总。"""
    if output_dir:
        results = enrich_results_with_gui_step_metrics(results, output_dir)
    summary = _compute_pipeline_summary(list(results.values()))
    return {
        "tasks": summary["total"],
        "pass": summary["passed"],
        "fail": summary["failed"],
        "interrupted": summary["interrupted"],
        "pass_rate": summary["rate"],
        "plan_rounds": summary["plan_rounds"],
        "gui_rounds_total": summary["gui_total"],
        "gui_steps_sequential": summary["gui_seq"],
        "parallelism": summary["parallelism"],
        "token_plan": summary["token_plan"],
        "token_gui": summary["token_gui"],
        "token_total": summary["token_total"],
        "cost_usd": summary["cost_usd"],
        "elapsed_time_sec": summary["time_sec"],
    }


def generate_report(
    results: Dict[str, Dict[str, Any]],
    output_dir: str,
    log=None,
) -> str:
    """
    从结果字典生成统计报告（Markdown + Excel）。

    输入:
        results: {task_key: task_result_dict} 格式的结果数据
        output_dir: 报告输出目录（会在其下创建 report/ 子目录）
        log: logger（可选）

    输出:
        报告目录路径

    生成文件:
        <output_dir>/report/summary.md
        <output_dir>/report/summary.xlsx
    """
    report_dir = os.path.join(output_dir, "report")
    os.makedirs(report_dir, exist_ok=True)

    results = enrich_results_with_gui_step_metrics(results, output_dir)

    # 按 pipeline 分组
    by_pipeline = defaultdict(list)
    for key, result in results.items():
        pipeline = result.get("pipeline", "unknown")
        by_pipeline[pipeline].append(result)

    # 生成 Markdown
    md_path = os.path.join(report_dir, "summary.md")
    md_content = _generate_markdown(by_pipeline)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    if log:
        log.info("Markdown 报告已生成: %s", md_path)

    # 生成 Excel
    if HAS_OPENPYXL:
        xlsx_path = os.path.join(report_dir, "summary.xlsx")
        _generate_excel(by_pipeline, xlsx_path)
        if log:
            log.info("Excel 报告已生成: %s", xlsx_path)
    else:
        if log:
            log.warning("openpyxl 未安装，跳过 Excel 报告生成")

    return report_dir


def _compute_pipeline_summary(tasks: List[Dict]) -> Dict[str, Any]:
    """
    计算单个 pipeline 的汇总统计。

    输入:
        tasks: 该 pipeline 下所有任务的结果列表

    输出:
        汇总字典
    """
    total = len(tasks)
    passed = sum(1 for t in tasks if t.get("pass", False))
    interrupted = sum(1 for t in tasks if t.get("interrupted", False))
    failed = total - passed - interrupted

    sum_plan_rounds = sum(t.get("plan_rounds", 0) for t in tasks)
    sum_gui_total = sum(t.get("gui_rounds_total", 0) for t in tasks)
    sum_gui_seq = sum(t.get("gui_steps_sequential", 0) for t in tasks)
    sum_token_plan = sum(t.get("token_plan", 0) for t in tasks)
    sum_token_gui = sum(t.get("token_gui", 0) for t in tasks)
    sum_token_total = sum(t.get("token_total", 0) for t in tasks)
    sum_cost = sum(t.get("cost_usd", 0) for t in tasks)
    sum_time = sum(t.get("elapsed_time_sec", 0) for t in tasks)

    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "interrupted": interrupted,
        "rate": f"{passed/total*100:.1f}%" if total > 0 else "0.0%",
        "plan_rounds": sum_plan_rounds,
        "gui_total": sum_gui_total,
        "gui_seq": sum_gui_seq,
        "parallelism": _calc_parallelism(sum_gui_total, sum_gui_seq),
        "token_plan": sum_token_plan,
        "token_gui": sum_token_gui,
        "token_total": sum_token_total,
        "cost_usd": round(sum_cost, 2),
        "time_sec": round(sum_time, 1),
    }


def _generate_markdown(by_pipeline: Dict[str, List[Dict]]) -> str:
    """
    生成 Markdown 格式的报告。

    输入:
        by_pipeline: {pipeline_name: [task_results]} 分组数据

    输出:
        Markdown 字符串
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [f"# 实验结果报告", f"", f"生成时间: {now}", f""]

    # ── 总表 ──
    lines.append("## 总体统计")
    lines.append("")
    header = "| Pipeline | Tasks | Pass | Fail | Int. | Rate | Σ Plan Rounds | Σ GUI Steps(Total) | Σ GUI Steps(Seq) | Parallelism | Σ Token(Plan) | Σ Token(GUI) | Σ Token | Σ Cost($) | Σ Time(s) |"
    sep    = "|----------|-------|------|------|------|------|---------------|--------------------|--------------------|-------------|---------------|--------------|---------|-----------|-----------|"
    lines.append(header)
    lines.append(sep)

    grand = {"total": 0, "passed": 0, "failed": 0, "interrupted": 0,
             "plan_rounds": 0, "gui_total": 0, "gui_seq": 0,
             "token_plan": 0, "token_gui": 0, "token_total": 0,
             "cost_usd": 0.0, "time_sec": 0.0}

    for pipeline_name in sorted(by_pipeline.keys()):
        s = _compute_pipeline_summary(by_pipeline[pipeline_name])
        lines.append(
            f"| {pipeline_name} | {s['total']} | {s['passed']} | {s['failed']} | {s['interrupted']} "
            f"| {s['rate']} | {s['plan_rounds']} | {s['gui_total']} | {s['gui_seq']} "
            f"| {s['parallelism']} | {s['token_plan']} | {s['token_gui']} | {s['token_total']} "
            f"| {s['cost_usd']} | {s['time_sec']} |"
        )
        for k in grand:
            if k in s:
                grand[k] += s[k] if isinstance(s[k], (int, float)) else 0

    # 汇总行
    grand_rate = f"{grand['passed']/grand['total']*100:.1f}%" if grand["total"] > 0 else "0.0%"
    grand_par = _calc_parallelism(grand["gui_total"], grand["gui_seq"])
    lines.append(
        f"| **Total** | **{grand['total']}** | **{grand['passed']}** | **{grand['failed']}** "
        f"| **{grand['interrupted']}** | **{grand_rate}** | **{grand['plan_rounds']}** "
        f"| **{grand['gui_total']}** | **{grand['gui_seq']}** | **{grand_par}** "
        f"| **{grand['token_plan']}** | **{grand['token_gui']}** | **{grand['token_total']}** "
        f"| **{round(grand['cost_usd'], 2)}** | **{round(grand['time_sec'], 1)}** |"
    )
    lines.append("")

    # ── 各 pipeline 详情表 ──
    for pipeline_name in sorted(by_pipeline.keys()):
        tasks = by_pipeline[pipeline_name]
        lines.append(f"## {pipeline_name} Pipeline 详情")
        lines.append("")

        col_headers = [c[0] for c in DETAIL_COLUMNS]
        lines.append("| " + " | ".join(col_headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(col_headers)) + " |")

        for t in sorted(tasks, key=lambda x: x.get("task_id", "")):
            row = []
            for col_name, col_key in DETAIL_COLUMNS:
                if col_name == "Parallelism":
                    row.append(_calc_parallelism(
                        t.get("gui_rounds_total", 0),
                        t.get("gui_steps_sequential", 0)))
                elif col_name == "Pass":
                    row.append("PASS" if t.get("pass", False) else
                               ("INT" if t.get("interrupted", False) else "FAIL"))
                elif col_key:
                    val = t.get(col_key, "-")
                    if isinstance(val, float):
                        val = f"{val:.2f}" if col_name in ("Score", "Cost($)") else f"{val:.1f}"
                    row.append(str(val))
                else:
                    row.append("-")
            lines.append("| " + " | ".join(row) + " |")

        lines.append("")

    return "\n".join(lines)


def _generate_excel(by_pipeline: Dict[str, List[Dict]], xlsx_path: str):
    """
    生成 Excel 格式的报告。

    输入:
        by_pipeline: {pipeline_name: [task_results]} 分组数据
        xlsx_path: 输出文件路径
    """
    wb = Workbook()

    # ── Summary Sheet ──
    ws = wb.active
    ws.title = "Summary"

    summary_headers = ["Pipeline", "Tasks", "Pass", "Fail", "Interrupted", "Rate",
                       "Σ Plan Rounds", "Σ GUI Steps(Total)", "Σ GUI Steps(Seq)",
                       "Parallelism", "Σ Token(Plan)", "Σ Token(GUI)", "Σ Token",
                       "Σ Cost($)", "Σ Time(s)"]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for pipeline_name in sorted(by_pipeline.keys()):
        s = _compute_pipeline_summary(by_pipeline[pipeline_name])
        values = [pipeline_name, s["total"], s["passed"], s["failed"], s["interrupted"],
                  s["rate"], s["plan_rounds"], s["gui_total"], s["gui_seq"],
                  s["parallelism"], s["token_plan"], s["token_gui"], s["token_total"],
                  s["cost_usd"], s["time_sec"]]
        for col, v in enumerate(values, 1):
            ws.cell(row=row, column=col, value=v)
        row += 1

    # ── 各 Pipeline Detail Sheet ──
    for pipeline_name in sorted(by_pipeline.keys()):
        ws_detail = wb.create_sheet(title=pipeline_name[:31])  # Excel sheet 名最长 31 字符
        col_headers = [c[0] for c in DETAIL_COLUMNS]

        for col, h in enumerate(col_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for r, t in enumerate(sorted(by_pipeline[pipeline_name],
                                      key=lambda x: x.get("task_id", "")), 2):
            for c, (col_name, col_key) in enumerate(DETAIL_COLUMNS, 1):
                if col_name == "Parallelism":
                    val = _calc_parallelism(
                        t.get("gui_rounds_total", 0),
                        t.get("gui_steps_sequential", 0))
                elif col_name == "Pass":
                    val = "PASS" if t.get("pass") else ("INT" if t.get("interrupted") else "FAIL")
                elif col_key:
                    val = t.get(col_key, "")
                else:
                    val = ""
                ws_detail.cell(row=r, column=c, value=val)

    wb.save(xlsx_path)

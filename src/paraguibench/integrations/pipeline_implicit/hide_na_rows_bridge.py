"""Excel-008 generic artifact 闭集到隐藏行 typed 观测的转换边界。"""

from __future__ import annotations

import hashlib
from io import BytesIO
import json
import math
import multiprocessing
import os
import posixpath
import stat
import tempfile
import time
from types import MappingProxyType
from typing import Any
import xml.etree.ElementTree as ET
import zipfile

from paraguibench.evaluation.pipeline_implicit import (
    HIDE_NA_ROWS_PROTOCOL_ID,
    HIDE_NA_ROWS_TASK_ID,
    HideNARowsObservation,
    WorkbookHiddenRows,
)

from .artifact_evidence import (
    PipelineImplicitArtifactEvidenceError,
    PipelineImplicitArtifactObservation,
)
from .searchwrite_bridge import (
    _PARSER_POLL_INTERVAL_SECONDS,
    _PARSER_RSS_LIMIT_BYTES,
    _PARSER_WALL_TIMEOUT_SECONDS,
    _WorkbookParseRejected,
    _WorkbookParserInternalError,
    _cleanup_parser_sandbox,
    _decode_parser_message,
    _enter_parser_python_write_boundary,
    _install_parser_resource_limits,
    _parser_resident_bytes,
    _preflight_xlsx_bytes,
    _receive_parser_frame_after_final_rss_check,
    _send_parser_message,
    _silence_parser_worker_output,
    _terminate_parser_process,
    _validate_workbook_resource_bounds,
)


_VERIFIED_DEFAULT_THEME_SHA256 = frozenset(
    {
        "156137ac2d7fae74e0286df47c4d1c75e65d5ef1455ff74c4d46176aef06fe56",
        "d15e8ebf78ef7b9720839d7ae8fdc81a7df5bc24706d8e137df61a5683c358d9",
    }
)
PINNED_HIDE_NA_ROWS_BASELINE_SHA256 = MappingProxyType(
    {
        "KFC_Monthly_Data.xlsx": (
            "ef390c7fcfadccd6ebdde8b9aef381f0de781e67be3bd1a9e3924ef7dfc516c0"
        ),
        "McDonalds_Monthly_Data.xlsx": (
            "3928c55f6e1c4f03e5dc608b1b9ce33c0fd26c1b7df453b1abf9897c9344d0f1"
        ),
        "Mixue_Monthly_Data.xlsx": (
            "887a9e1649835c22b0b28d418d561df821abd4d449e3b56870ca8d253f835a6d"
        ),
        "PizzaHut_Monthly_Data.xlsx": (
            "d97ba8d1bf4deb5c3912a04616d4a01c7742e90a20767cc92d4c08b993acac07"
        ),
        "Subway_Monthly_Data.xlsx": (
            "00273dc972a5c365515cefce471cc4901bd9936ebf00096a6bdd56286822431f"
        ),
    }
)
_MAX_HIDDEN_ROWS = 256
_SHEET_PROTECTION_FIELDS = (
    "selectLockedCells",
    "selectUnlockedCells",
    "algorithmName",
    "sheet",
    "objects",
    "insertRows",
    "insertHyperlinks",
    "autoFilter",
    "scenarios",
    "formatColumns",
    "deleteColumns",
    "insertColumns",
    "pivotTables",
    "deleteRows",
    "formatCells",
    "saltValue",
    "formatRows",
    "sort",
    "spinCount",
    "password",
    "hashValue",
)
_WORKBOOK_PROTECTION_FIELDS = (
    "workbookPassword",
    "workbookPasswordCharacterSet",
    "revisionsPassword",
    "revisionsPasswordCharacterSet",
    "lockStructure",
    "lockWindows",
    "lockRevision",
    "revisionsAlgorithmName",
    "revisionsHashValue",
    "revisionsSaltValue",
    "revisionsSpinCount",
    "workbookAlgorithmName",
    "workbookHashValue",
    "workbookSaltValue",
    "workbookSpinCount",
)
_DEFINED_NAME_FIELDS = (
    "name",
    "comment",
    "customMenu",
    "description",
    "help",
    "statusBar",
    "localSheetId",
    "hidden",
    "function",
    "vbProcedure",
    "xlm",
    "functionGroupId",
    "shortcutKey",
    "publishToServer",
    "workbookParameter",
    "attr_text",
)
_PAGE_BREAK_FIELDS = ("id", "min", "max", "man", "pt")
_WORKBOOK_VIEW_FIELDS = (
    "visibility",
    "minimized",
    "showHorizontalScroll",
    "showVerticalScroll",
    "showSheetTabs",
    "xWindow",
    "yWindow",
    "windowWidth",
    "windowHeight",
    "tabRatio",
    "firstSheet",
    "activeTab",
    "autoFilterDateGrouping",
)
_WORKBOOK_VIEW_GEOMETRY_FIELDS = (
    "xWindow",
    "yWindow",
    "windowWidth",
    "windowHeight",
    "tabRatio",
)
_VERIFIED_WORKBOOK_VIEW_GEOMETRIES = frozenset(
    {
        (0, 780, 34_200, 19_860, 600),
        (0, 0, 16_384, 8_192, 500),
    }
)
_WORKSHEET_PROPERTY_FIELDS = (
    "codeName",
    "enableFormatConditionsCalculation",
    "filterMode",
    "published",
    "syncHorizontal",
    "syncRef",
    "syncVertical",
    "transitionEvaluation",
    "transitionEntry",
)
_OUTLINE_PROPERTY_FIELDS = (
    "applyStyles",
    "summaryBelow",
    "summaryRight",
    "showOutlineSymbols",
)
_OOXML_VISIBLE_PART_MAX_BYTES = 2 * 1024 * 1024
_SPREADSHEET_DRAWING_NAMESPACE = (
    "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
)
_DRAWINGML_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/main"
_OFFICE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_DRAWING_ALLOWED_NAMESPACES = frozenset(
    {
        _SPREADSHEET_DRAWING_NAMESPACE,
        _DRAWINGML_NAMESPACE,
        _OFFICE_RELATIONSHIP_NAMESPACE,
    }
)
_SPREADSHEETML_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_MARKUP_COMPATIBILITY_NAMESPACE = (
    "http://schemas.openxmlformats.org/markup-compatibility/2006"
)
_OFFICE_2014_REVISION_NAMESPACE = (
    "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
)
_WORKSHEET_ALLOWED_ATTRIBUTE_NAMESPACES = frozenset(
    {
        _MARKUP_COMPATIBILITY_NAMESPACE,
        _OFFICE_2014_REVISION_NAMESPACE,
        _OFFICE_RELATIONSHIP_NAMESPACE,
    }
)
_PACKAGE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/relationships"
)
_OFFICE_RELATIONSHIP_TYPE_PREFIX = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
)
_CORE_PROPERTIES_RELATIONSHIP_TYPE = (
    "http://schemas.openxmlformats.org/package/2006/relationships/metadata/"
    "core-properties"
)
_RELATIONSHIP_PART_OWNERS = MappingProxyType(
    {
        "_rels/.rels": "",
        "xl/_rels/workbook.xml.rels": "xl/workbook.xml",
        "xl/worksheets/_rels/sheet1.xml.rels": "xl/worksheets/sheet1.xml",
    }
)
_REQUIRED_OOXML_PACKAGE_MEMBERS = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "docProps/app.xml",
        "docProps/core.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/styles.xml",
        "xl/workbook.xml",
        "xl/worksheets/sheet1.xml",
    }
)
_OPTIONAL_OOXML_PACKAGE_MEMBERS = frozenset(
    {
        "docProps/custom.xml",
        "xl/drawings/drawing1.xml",
        "xl/sharedStrings.xml",
        "xl/theme/theme1.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
    }
)
_CONTENT_TYPES_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/content-types"
)
_RELATIONSHIPS_CONTENT_TYPE = "application/vnd.openxmlformats-package.relationships+xml"
_EXPECTED_CONTENT_TYPE_BY_MEMBER = MappingProxyType(
    {
        "_rels/.rels": _RELATIONSHIPS_CONTENT_TYPE,
        "docProps/app.xml": (
            "application/vnd.openxmlformats-officedocument.extended-properties+xml"
        ),
        "docProps/core.xml": (
            "application/vnd.openxmlformats-package.core-properties+xml"
        ),
        "docProps/custom.xml": (
            "application/vnd.openxmlformats-officedocument.custom-properties+xml"
        ),
        "xl/_rels/workbook.xml.rels": _RELATIONSHIPS_CONTENT_TYPE,
        "xl/drawings/drawing1.xml": (
            "application/vnd.openxmlformats-officedocument.drawing+xml"
        ),
        "xl/sharedStrings.xml": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml."
            "sharedStrings+xml"
        ),
        "xl/styles.xml": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"
        ),
        "xl/theme/theme1.xml": (
            "application/vnd.openxmlformats-officedocument.theme+xml"
        ),
        "xl/workbook.xml": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
        ),
        "xl/worksheets/_rels/sheet1.xml.rels": _RELATIONSHIPS_CONTENT_TYPE,
        "xl/worksheets/sheet1.xml": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
        ),
    }
)
_ALLOWED_DEFAULT_CONTENT_TYPES = MappingProxyType(
    {
        "rels": _RELATIONSHIPS_CONTENT_TYPE,
        "xml": "application/xml",
        "png": "image/png",
        "jpeg": "image/jpeg",
    }
)
_EXTENDED_PROPERTIES_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
)
_CUSTOM_PROPERTIES_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
)
_DOCUMENT_PROPERTY_VALUE_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"
)
_CORE_PROPERTIES_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
)
_DUBLIN_CORE_NAMESPACE = "http://purl.org/dc/elements/1.1/"
_DUBLIN_CORE_TERMS_NAMESPACE = "http://purl.org/dc/terms/"
_DUBLIN_CORE_TYPE_NAMESPACE = "http://purl.org/dc/dcmitype/"
_XML_SCHEMA_INSTANCE_NAMESPACE = "http://www.w3.org/2001/XMLSchema-instance"
_CALCULATION_FIELDS = (
    "calcId",
    "calcMode",
    "fullCalcOnLoad",
    "refMode",
    "iterate",
    "iterateCount",
    "iterateDelta",
    "fullPrecision",
    "calcCompleted",
    "calcOnSave",
    "concurrentCalc",
    "concurrentManualCount",
    "forceFullCalc",
)
_VERIFIED_CALCULATION_DEFAULTS = MappingProxyType(
    {
        "calcId": (0, 124_519),
        "refMode": (None, "A1"),
        "iterate": (None, False),
        "iterateCount": (None, 100),
        "iterateDelta": (None, 0.0001),
    }
)
_FONT_ELEMENT_FIELDS = (
    "name",
    "charset",
    "family",
    "b",
    "i",
    "strike",
    "outline",
    "shadow",
    "condense",
    "color",
    "extend",
    "sz",
    "u",
    "vertAlign",
    "scheme",
)
_CELL_PROTECTION_FIELDS = ("locked", "hidden")
_PRINT_OPTIONS_FIELDS = (
    "horizontalCentered",
    "verticalCentered",
    "headings",
    "gridLines",
    "gridLinesSet",
)
_PAGE_SETUP_FIELDS = (
    "orientation",
    "paperSize",
    "scale",
    "fitToHeight",
    "fitToWidth",
    "firstPageNumber",
    "useFirstPageNumber",
    "paperHeight",
    "paperWidth",
    "pageOrder",
    "usePrinterDefaults",
    "blackAndWhite",
    "draft",
    "cellComments",
    "errors",
    "horizontalDpi",
    "verticalDpi",
    "copies",
    "id",
)


def build_hide_na_rows_observation(
    artifact_observation: PipelineImplicitArtifactObservation,
) -> HideNARowsObservation:
    """把已冻结的五工作簿闭集投影为正式 typed observation。

    输入参数：
        artifact_observation：经 manifest—nofollow—manifest 和逐文件
            size/SHA-256 双重校验的 generic observation。
    输出返回值：
        仅保留精确文件身份、隐藏行集和基线匹配布尔值的
        ``HideNARowsObservation``。
    异常：
        PipelineImplicitArtifactEvidenceError：任务、协议或 parser
            可信边界自身失效；异常仅含固定脱敏码。
    """

    if (
        not isinstance(
            artifact_observation,
            PipelineImplicitArtifactObservation,
        )
        or artifact_observation.task_id != HIDE_NA_ROWS_TASK_ID
        or artifact_observation.protocol_id != HIDE_NA_ROWS_PROTOCOL_ID
        or artifact_observation.complete is not True
    ):
        raise PipelineImplicitArtifactEvidenceError("TYPED_OBSERVATION_INVALID")
    workbooks: list[WorkbookHiddenRows] = []
    unexpected_file_count = 0
    for artifact_file in artifact_observation.iter_files_for_evaluator():
        expected_baseline_sha256 = PINNED_HIDE_NA_ROWS_BASELINE_SHA256.get(
            artifact_file.relative_path
        )
        if expected_baseline_sha256 is None:
            unexpected_file_count += 1
            workbooks.append(
                WorkbookHiddenRows(
                    document_name=f"unexpected-{unexpected_file_count}.xlsx",
                    hidden_rows=(),
                    content_matches_baseline=False,
                )
            )
            continue
        payload = artifact_file.read_for_evaluator()
        try:
            _preflight_xlsx_bytes(payload)
        except (OSError, ValueError, zipfile.BadZipFile):
            workbooks.append(_failed_workbook(artifact_file.relative_path))
            continue
        try:
            hidden_rows, baseline_sha256 = _parse_xlsx_controlled(payload)
        except _WorkbookParseRejected:
            workbooks.append(_failed_workbook(artifact_file.relative_path))
            continue
        except _WorkbookParserInternalError:
            raise PipelineImplicitArtifactEvidenceError(
                "TYPED_OBSERVATION_INVALID"
            ) from None
        workbooks.append(
            WorkbookHiddenRows(
                document_name=artifact_file.relative_path,
                hidden_rows=hidden_rows,
                content_matches_baseline=(baseline_sha256 == expected_baseline_sha256),
            )
        )
    return HideNARowsObservation(complete=True, workbooks=tuple(workbooks))


def derive_hide_na_rows_baseline_sha256(content: bytes) -> str:
    """在与生产取证相同的受控边界内重算语义摘要。

    输入参数：
        content：已从固定 input 或 gold manifest 验证的 XLSX 字节。
    输出返回值：
        排除本任务允许的行 hidden 状态后的小写 SHA-256。
    异常：
        PipelineImplicitArtifactEvidenceError：OOXML 或 parser 边界无法
            形成可信指纹；异常不回显路径、单元格或内容。

    该函数是 formal asset 审计 builder 的窄入口；它不读取
    Agent final text，也不持久化工作簿投影。
    """

    try:
        _preflight_xlsx_bytes(content)
        _, baseline_sha256 = _parse_xlsx_controlled(content)
    except (
        OSError,
        ValueError,
        zipfile.BadZipFile,
        _WorkbookParseRejected,
        _WorkbookParserInternalError,
    ):
        raise PipelineImplicitArtifactEvidenceError(
            "TYPED_OBSERVATION_INVALID"
        ) from None
    return baseline_sha256


def _parse_xlsx_controlled(
    content: bytes,
) -> tuple[tuple[int, ...], str]:
    """在 wall/CPU/RSS/fd/禁写边界内解析一个 Excel-008 XLSX。

    输入参数：
        content：已通过流式 OOXML 预检的工作簿字节。
    输出返回值：
        一号 sheet 的有序隐藏行号与语义基线 SHA-256。
    异常：
        _WorkbookParseRejected：输入错误或资源超限。
        _WorkbookParserInternalError：spawn、监控或 IPC 边界失效。
    """

    if not isinstance(content, bytes) or not content:
        raise _WorkbookParseRejected("PARSER_INPUT_INVALID")
    sandbox: Any | None = None
    try:
        sandbox = tempfile.TemporaryDirectory(
            prefix="paraguibench-hide-na-rows-parser-"
        )
        os.chmod(sandbox.name, stat.S_IRUSR | stat.S_IXUSR)
    except OSError:
        if sandbox is not None:
            try:
                _cleanup_parser_sandbox(sandbox)
            except _WorkbookParserInternalError:
                pass
        raise _WorkbookParserInternalError("PARSER_SANDBOX_FAILED") from None
    try:
        return _run_parser_process(
            content,
            sandbox_cwd=sandbox.name,
        )
    finally:
        _cleanup_parser_sandbox(sandbox)


def _run_parser_process(
    content: bytes,
    *,
    sandbox_cwd: str,
) -> tuple[tuple[int, ...], str]:
    """启动并监控 Excel-008 专属 parser 子进程。

    输入参数：
        content：已由父层验证的工作簿字节。
        sandbox_cwd：父进程创建且移除写权限的空目录。
    输出返回值：
        经父进程二次验证的隐藏行 tuple 和指纹。
    """

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    parent_rss_ack = context.Event()
    process = context.Process(
        target=_workbook_parse_worker,
        args=(sender, content, None, sandbox_cwd, parent_rss_ack),
        daemon=True,
        name="paraguibench-hide-na-rows-parser",
    )
    try:
        process.start()
    except (OSError, RuntimeError):
        receiver.close()
        sender.close()
        raise _WorkbookParserInternalError("PARSER_START_FAILED") from None
    sender.close()
    raw_message: bytes | None = None
    rejected_by_resource_gate = False
    internal_monitor_error = False
    deadline = time.monotonic() + _PARSER_WALL_TIMEOUT_SECONDS
    try:
        while True:
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                rejected_by_resource_gate = True
                break
            if receiver.poll(min(_PARSER_POLL_INTERVAL_SECONDS, remaining)):
                try:
                    raw_message = _receive_parser_frame_after_final_rss_check(
                        receiver,
                        process,
                    )
                except _WorkbookParseRejected:
                    rejected_by_resource_gate = True
                except _WorkbookParserInternalError:
                    internal_monitor_error = True
                break
            if not process.is_alive():
                break
            resident_bytes = _parser_resident_bytes(process.pid)
            if resident_bytes is None:
                internal_monitor_error = True
                break
            if resident_bytes > _PARSER_RSS_LIMIT_BYTES:
                rejected_by_resource_gate = True
                break
    finally:
        parent_rss_ack.set()
        receiver.close()
        process.join(timeout=0.5)
        if process.is_alive():
            _terminate_parser_process(process)
    if rejected_by_resource_gate:
        process.close()
        raise _WorkbookParseRejected("PARSER_RESOURCE_LIMIT")
    if internal_monitor_error:
        process.close()
        raise _WorkbookParserInternalError("PARSER_MONITOR_FAILED")
    if process.exitcode != 0:
        exitcode = process.exitcode
        process.close()
        if isinstance(exitcode, int) and exitcode < 0:
            raise _WorkbookParseRejected("PARSER_PROCESS_LIMITED")
        raise _WorkbookParserInternalError("PARSER_PROCESS_FAILED")
    process.close()
    return _validate_parser_message(_decode_parser_message(raw_message))


def _workbook_parse_worker(
    sender: Any,
    content: bytes,
    materializer: Any | None = None,
    sandbox_cwd: str | None = None,
    parent_rss_ack: Any | None = None,
) -> None:
    """在受限子进程中物化工作簿并只返回小型 JSON 投影。

    输入参数：
        sender：只向父进程发送一个有界 frame 的单向 pipe。
        content：与受控入口一致的工作簿字节。
        materializer：默认使用生产解析器；仅供边界故障注入
            测试的可拾取顶层 callable。
        sandbox_cwd：只读空工作目录。
        parent_rss_ack：生产父进程在最终 RSS 采样后设置的事件；
            边界单测可省略。
    输出返回值：
        无；成功发送 ``ok``，已知输入错误发送 ``rejected``，
        未知错误只发送 ``internal_error``。
    """

    try:
        _silence_parser_worker_output()
        import sys

        sys.dont_write_bytecode = True
        _install_parser_resource_limits()
        _enter_parser_python_write_boundary(sandbox_cwd)
        parser = _materialize_workbook if materializer is None else materializer
        if not callable(parser):
            raise TypeError("parser materializer 无效")
    except BaseException:
        _send_safely(sender, ("internal_error",))
        sender.close()
        return
    try:
        hidden_rows, baseline_sha256 = parser(content)
    except (
        IndexError,
        KeyError,
        OSError,
        ValueError,
        zipfile.BadZipFile,
        ET.ParseError,
    ):
        _send_safely(sender, ("rejected",))
    except BaseException:
        _send_safely(sender, ("internal_error",))
    else:
        _send_safely(sender, ("ok", hidden_rows, baseline_sha256))
    finally:
        sender.close()
        if parent_rss_ack is not None:
            try:
                parent_rss_ack.wait(timeout=_PARSER_WALL_TIMEOUT_SECONDS)
            except BaseException:
                pass


def _send_safely(sender: Any, message: tuple[object, ...]) -> None:
    """发送脱敏 JSON frame，pipe 已断开时不外泄异常。

    输入参数：
        sender：单向字节 pipe 发送端。
        message：仅含固定状态、整数行号和 SHA-256 的 tuple。
    输出返回值：无。
    """

    try:
        _send_parser_message(sender, message)
    except (BrokenPipeError, EOFError, OSError):
        pass


def _validate_parser_message(message: object) -> tuple[tuple[int, ...], str]:
    """验证不可信 parser 返回的隐藏行投影。

    输入参数：
        message：从有界 JSON frame 解码的候选对象。
    输出返回值：
        严格递增的正整数行号 tuple 与小写 SHA-256。
    """

    if message == ("internal_error",):
        raise _WorkbookParserInternalError("PARSER_INTERNAL_ERROR")
    if message == ("rejected",):
        raise _WorkbookParseRejected("PARSER_INPUT_REJECTED")
    if (
        not isinstance(message, tuple)
        or len(message) != 3
        or message[0] != "ok"
        or not isinstance(message[1], tuple)
        or not isinstance(message[2], str)
        or len(message[2]) != 64
        or any(character not in "0123456789abcdef" for character in message[2])
        or len(message[1]) > _MAX_HIDDEN_ROWS
        or any(type(row) is not int or not 1 <= row <= 1_048_576 for row in message[1])
        or tuple(sorted(set(message[1]))) != message[1]
    ):
        raise _WorkbookParserInternalError("PARSER_RESULT_INVALID")
    return message[1], message[2]


def _materialize_workbook(
    content: bytes,
) -> tuple[tuple[int, ...], str]:
    """在子进程中用 openpyxl 物化并投影工作簿语义。

    输入参数：
        content：已预检的 XLSX 字节。
    输出返回值：
        一号 sheet 隐藏行集和忽略隐藏行后的语义指纹。
    """

    import openpyxl

    _validate_ooxml_package_members(content)
    _validate_ooxml_content_types(content)
    _validate_ooxml_relationship_closure(content)
    ooxml_visible_parts = [
        *_normalize_drawing_parts(content),
        *_normalize_worksheet_extension_parts(content),
        *_normalize_non_core_property_parts(content),
    ]
    stream = BytesIO(content)
    workbook = openpyxl.load_workbook(
        stream,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        _validate_workbook_resource_bounds(workbook)
        if len(workbook.worksheets) != 1:
            raise ValueError("Excel-008 sheet 数不匹配")
        worksheet = workbook.worksheets[0]
        hidden_rows = tuple(
            index
            for index, dimension in sorted(worksheet.row_dimensions.items())
            if bool(dimension.hidden)
        )
        return hidden_rows, _baseline_semantic_sha256(
            workbook,
            ooxml_visible_parts=ooxml_visible_parts,
        )
    finally:
        workbook.close()
        stream.close()


def _baseline_semantic_sha256(
    workbook: Any,
    *,
    ooxml_visible_parts: list[dict[str, str]],
) -> str:
    """生成排除隐藏行状态后的工作簿语义指纹。

    输入参数：
        workbook：已通过 OOXML 资源门并被物化的工作簿。
        ooxml_visible_parts：不依赖 openpyxl 识别能力的
            受控可见 OOXML part 投影。
    输出返回值：
        覆盖 sheet、值/公式、合并区、行高列宽、可见样式与
        新增可见对象计数的 SHA-256。指纹仅忽略本任务
        允许的行 ``hidden`` 状态和已证实的 Office 默认序列化差异。
    """

    sheets: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        cells: list[list[object]] = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                if cell.value is None and not cell.has_style:
                    continue
                cells.append(
                    [
                        cell.coordinate,
                        cell.data_type,
                        _normalize_baseline_scalar_or_none(cell.value),
                        _normalize_cell_visible_style(cell),
                    ]
                )
        row_dimensions = [
            [
                str(index),
                _normalize_visible_length(dimension.height, quantum=1.0),
                int(dimension.outlineLevel or 0),
                bool(dimension.collapsed),
                bool(dimension.thickTop),
                bool(dimension.thickBot),
            ]
            for index, dimension in sorted(worksheet.row_dimensions.items())
            if not _is_default_row_dimension_ignoring_hidden(
                dimension,
                default_height=worksheet.sheet_format.defaultRowHeight,
            )
        ]
        column_dimensions = [
            [
                str(index),
                int(dimension.min or 0),
                int(dimension.max or 0),
                _normalize_visible_length(dimension.width, quantum=0.1),
                bool(dimension.hidden),
                bool(dimension.bestFit),
                int(dimension.outlineLevel or 0),
                bool(dimension.collapsed),
            ]
            for index, dimension in sorted(worksheet.column_dimensions.items())
        ]
        freeze_panes = worksheet.freeze_panes
        if hasattr(freeze_panes, "coordinate"):
            freeze_panes = freeze_panes.coordinate
        if freeze_panes is not None and not isinstance(freeze_panes, str):
            raise ValueError("freeze pane 类型未固定")
        comments = 0
        hyperlinks = 0
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                comments += int(cell.comment is not None)
                hyperlinks += int(cell.hyperlink is not None)
        sheets.append(
            {
                "title": worksheet.title,
                "state": worksheet.sheet_state,
                "dimension": worksheet.calculate_dimension(),
                "merged_cells": sorted(
                    str(cell_range) for cell_range in worksheet.merged_cells.ranges
                ),
                "freeze_panes": freeze_panes,
                "sheet_format": {
                    "base_column_width": (
                        int(worksheet.sheet_format.baseColWidth or 0)
                        if worksheet.sheet_format.defaultColWidth is None
                        else None
                    ),
                    "default_column_width": _normalize_visible_length(
                        worksheet.sheet_format.defaultColWidth,
                        quantum=0.1,
                    ),
                    "default_row_height": _normalize_visible_length(
                        worksheet.sheet_format.defaultRowHeight,
                        quantum=1.0,
                    ),
                    "custom_height": _default_bool(
                        worksheet.sheet_format.customHeight,
                        False,
                    ),
                    "zero_height": bool(worksheet.sheet_format.zeroHeight),
                    "thick_top": bool(worksheet.sheet_format.thickTop),
                    "thick_bottom": bool(worksheet.sheet_format.thickBottom),
                    "outline_level_row": _default_int(
                        worksheet.sheet_format.outlineLevelRow,
                        0,
                    ),
                    "outline_level_column": _default_int(
                        worksheet.sheet_format.outlineLevelCol,
                        0,
                    ),
                },
                "row_dimensions": row_dimensions,
                "column_dimensions": column_dimensions,
                "cells": cells,
                "visible_object_counts": [
                    len(worksheet._images),
                    len(worksheet._charts),
                    len(worksheet.tables),
                    len(worksheet.data_validations.dataValidation),
                    len(worksheet.conditional_formatting),
                    comments,
                    hyperlinks,
                ],
                "sheet_view": _normalize_sheet_view(worksheet.sheet_view),
                "tab_color": _normalize_style_color(
                    worksheet.sheet_properties.tabColor
                ),
                "auto_filter": _normalize_auto_filter(worksheet.auto_filter.ref),
                "page_setup": _normalize_page_setup(worksheet),
                "header_footer": _normalize_header_footer(worksheet.HeaderFooter),
                "sheet_protection": _normalize_closed_attributes(
                    worksheet.protection,
                    _SHEET_PROTECTION_FIELDS,
                ),
                "print_ranges": _normalize_print_ranges(worksheet),
                "page_breaks": _normalize_page_breaks(worksheet),
                "sheet_properties": _normalize_sheet_properties(
                    worksheet.sheet_properties
                ),
            }
        )
    encoded = json.dumps(
        {
            "version": 4,
            "theme": _normalize_workbook_theme(workbook),
            "workbook_protection": _normalize_closed_attributes(
                workbook.security,
                _WORKBOOK_PROTECTION_FIELDS,
            ),
            "defined_names": _normalize_defined_names(workbook),
            "workbook_views": _normalize_workbook_views(workbook),
            "calculation": _normalize_workbook_calculation(workbook),
            "ooxml_visible_parts": ooxml_visible_parts,
            "sheets": sheets,
        },
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8", errors="strict")
    return hashlib.sha256(encoded).hexdigest()


def _normalize_drawing_parts(content: bytes) -> list[dict[str, str]]:
    """从原始 OOXML 闭集投影 openpyxl 可能丢弃的 DrawingML。

    输入参数：
        content：已通过成员、展开大小、XML 深度和元素数
            预检的 XLSX 字节。
    输出返回值：
        按 part 身份排序的非空 drawing 固定 payload SHA-256；
        Lee gold 中的精确空 ``xdr:wsDr`` 与不存在 part 等价。
    异常：
        ValueError：drawing part 身份、资源、根元素或 namespace
            脱离已审定闭集。

    该投影不读取 ``worksheet._images/_charts``，因此未受支持的
    shape/text/sparkline 即使被 openpyxl 静默丢弃，也不会
    以空对象计数 fail-open。
    """

    records: list[dict[str, str]] = []
    with zipfile.ZipFile(BytesIO(content), mode="r") as archive:
        for member in sorted(
            archive.infolist(),
            key=lambda item: item.filename.encode("utf-8", errors="strict"),
        ):
            name = member.filename
            if not name.startswith("xl/drawings/") or not name.endswith(".xml"):
                continue
            stem = name.removeprefix("xl/drawings/drawing").removesuffix(".xml")
            if not stem.isascii() or not stem.isdigit() or int(stem) <= 0:
                raise ValueError("DrawingML part 身份无效")
            if member.file_size > _OOXML_VISIBLE_PART_MAX_BYTES:
                raise ValueError("DrawingML part 资源超限")
            with archive.open(member, mode="r") as stream:
                payload = stream.read(_OOXML_VISIBLE_PART_MAX_BYTES + 1)
                tail = stream.read(1)
            if (
                len(payload) > _OOXML_VISIBLE_PART_MAX_BYTES
                or len(payload) != member.file_size
                or tail != b""
            ):
                raise ValueError("DrawingML part 读取不完整")
            try:
                root = ET.fromstring(payload)
            except ET.ParseError:
                raise ValueError("DrawingML XML 无效") from None
            if root.tag != f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}wsDr" or root.attrib:
                raise ValueError("DrawingML 根元素闭集失效")
            for element in root.iter():
                tag_namespace = _qualified_name_namespace(element.tag)
                if tag_namespace not in _DRAWING_ALLOWED_NAMESPACES:
                    raise ValueError("DrawingML 元素 namespace 未审定")
                for attribute in element.attrib:
                    attribute_namespace = _qualified_name_namespace(
                        attribute,
                        allow_unqualified=True,
                    )
                    if (
                        attribute_namespace is not None
                        and attribute_namespace not in _DRAWING_ALLOWED_NAMESPACES
                    ):
                        raise ValueError("DrawingML 属性 namespace 未审定")
            if len(root) == 0 and not (root.text or "").strip():
                continue
            records.append(
                {
                    "part": name,
                    "payload_sha256": hashlib.sha256(payload).hexdigest(),
                }
            )
    return records


def _normalize_worksheet_extension_parts(
    content: bytes,
) -> list[dict[str, str]]:
    """投影 openpyxl 会警告后删除的 worksheet 扩展。

    输入参数：
        content：已通过被动 OOXML 资源预检的 XLSX 字节。
    输出返回值：
        含 ``extLst``、非 SpreadsheetML 元素 namespace 或
        未审定属性 namespace 时，返回 worksheet 固定 payload
        摘要与 namespace-set 摘要；固定 input/gold 无此结构。
    异常：
        ValueError：worksheet part 资源、根元素或 QName 无效。

    这一原始 part 投影使 x14 sparkline 等未支持扩展在
    openpyxl 移除它们之前就进入基线。
    """

    member_name = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(BytesIO(content), mode="r") as archive:
        try:
            member = archive.getinfo(member_name)
        except KeyError:
            raise ValueError("worksheet part 缺失") from None
        if member.file_size > _OOXML_VISIBLE_PART_MAX_BYTES:
            raise ValueError("worksheet part 资源超限")
        with archive.open(member, mode="r") as stream:
            payload = stream.read(_OOXML_VISIBLE_PART_MAX_BYTES + 1)
            tail = stream.read(1)
    if (
        len(payload) > _OOXML_VISIBLE_PART_MAX_BYTES
        or len(payload) != member.file_size
        or tail != b""
    ):
        raise ValueError("worksheet part 读取不完整")
    try:
        root = ET.fromstring(payload)
    except ET.ParseError:
        raise ValueError("worksheet XML 无效") from None
    if root.tag != f"{{{_SPREADSHEETML_NAMESPACE}}}worksheet":
        raise ValueError("worksheet 根元素 namespace 无效")

    extension_detected = False
    observed_namespaces: set[str] = set()
    for element in root.iter():
        element_namespace = _qualified_name_namespace(element.tag)
        observed_namespaces.add(element_namespace)
        if (
            element_namespace != _SPREADSHEETML_NAMESPACE
            or element.tag == f"{{{_SPREADSHEETML_NAMESPACE}}}extLst"
        ):
            extension_detected = True
        for attribute in element.attrib:
            attribute_namespace = _qualified_name_namespace(
                attribute,
                allow_unqualified=True,
            )
            if attribute_namespace is not None:
                observed_namespaces.add(attribute_namespace)
                if attribute_namespace not in _WORKSHEET_ALLOWED_ATTRIBUTE_NAMESPACES:
                    extension_detected = True
    if not extension_detected:
        return []
    namespace_payload = json.dumps(
        sorted(observed_namespaces),
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8", errors="strict")
    return [
        {
            "part": member_name,
            "payload_sha256": hashlib.sha256(payload).hexdigest(),
            "namespace_set_sha256": hashlib.sha256(namespace_payload).hexdigest(),
        }
    ]


def _validate_ooxml_relationship_closure(content: bytes) -> None:
    """校验 Excel-008 固定 package 的全部 relationship 图。

    输入参数：
        content：已通过 ZIP/XML 资源预检的 XLSX 字节。
    输出返回值：
        无；package、workbook 与单 worksheet 的关系集在忽略
        无语义 ``rId`` 命名后精确等于已审定图时返回。
    异常：
        ValueError：出现未登记 relationship part/type/target、
            重复 ID、外部模式或悬空 target。
    """

    with zipfile.ZipFile(BytesIO(content), mode="r") as archive:
        members = frozenset(archive.namelist())
        relationship_parts = frozenset(
            name for name in members if name.casefold().endswith(".rels")
        )
        expected_relationship_parts = {
            "_rels/.rels",
            "xl/_rels/workbook.xml.rels",
        }
        drawing_present = "xl/drawings/drawing1.xml" in members
        if drawing_present:
            expected_relationship_parts.add("xl/worksheets/_rels/sheet1.xml.rels")
        if relationship_parts != expected_relationship_parts:
            raise ValueError("OOXML relationship part 闭集失效")

        root_expected = {
            (
                _OFFICE_RELATIONSHIP_TYPE_PREFIX + "officeDocument",
                "xl/workbook.xml",
            ),
            (_CORE_PROPERTIES_RELATIONSHIP_TYPE, "docProps/core.xml"),
            (
                _OFFICE_RELATIONSHIP_TYPE_PREFIX + "extended-properties",
                "docProps/app.xml",
            ),
        }
        if "docProps/custom.xml" in members:
            root_expected.add(
                (
                    _OFFICE_RELATIONSHIP_TYPE_PREFIX + "custom-properties",
                    "docProps/custom.xml",
                )
            )
        workbook_expected = {
            (
                _OFFICE_RELATIONSHIP_TYPE_PREFIX + "worksheet",
                "xl/worksheets/sheet1.xml",
            ),
            (
                _OFFICE_RELATIONSHIP_TYPE_PREFIX + "styles",
                "xl/styles.xml",
            ),
        }
        if "xl/sharedStrings.xml" in members:
            workbook_expected.add(
                (
                    _OFFICE_RELATIONSHIP_TYPE_PREFIX + "sharedStrings",
                    "xl/sharedStrings.xml",
                )
            )
        if "xl/theme/theme1.xml" in members:
            workbook_expected.add(
                (
                    _OFFICE_RELATIONSHIP_TYPE_PREFIX + "theme",
                    "xl/theme/theme1.xml",
                )
            )
        expected_by_part: dict[str, set[tuple[str, str]]] = {
            "_rels/.rels": root_expected,
            "xl/_rels/workbook.xml.rels": workbook_expected,
        }
        if drawing_present:
            expected_by_part["xl/worksheets/_rels/sheet1.xml.rels"] = {
                (
                    _OFFICE_RELATIONSHIP_TYPE_PREFIX + "drawing",
                    "xl/drawings/drawing1.xml",
                )
            }

        for relationship_part in sorted(expected_relationship_parts):
            payload = _read_ooxml_member_bounded(archive, relationship_part)
            try:
                root = ET.fromstring(payload)
            except ET.ParseError:
                raise ValueError("OOXML relationship XML 无效") from None
            if (
                root.tag != f"{{{_PACKAGE_RELATIONSHIP_NAMESPACE}}}Relationships"
                or root.attrib
                or (root.text or "").strip()
            ):
                raise ValueError("OOXML relationship 根闭集失效")
            owner = _RELATIONSHIP_PART_OWNERS[relationship_part]
            ids: set[str] = set()
            observed: list[tuple[str, str]] = []
            for relationship in root:
                if (
                    relationship.tag
                    != f"{{{_PACKAGE_RELATIONSHIP_NAMESPACE}}}Relationship"
                    or set(relationship.attrib) != {"Id", "Type", "Target"}
                    or len(relationship)
                    or (relationship.text or "").strip()
                ):
                    raise ValueError("OOXML relationship 元素闭集失效")
                relationship_id = relationship.attrib["Id"]
                relationship_type = relationship.attrib["Type"]
                target = relationship.attrib["Target"]
                if (
                    not relationship_id
                    or relationship_id in ids
                    or not relationship_type
                    or not target
                ):
                    raise ValueError("OOXML relationship 身份无效")
                ids.add(relationship_id)
                resolved_target = _resolve_relationship_target(owner, target)
                if resolved_target not in members:
                    raise ValueError("OOXML relationship target 悬空")
                observed.append((relationship_type, resolved_target))
            if (
                len(observed) != len(expected_by_part[relationship_part])
                or set(observed) != expected_by_part[relationship_part]
            ):
                raise ValueError("OOXML relationship 语义图闭集失效")


def _normalize_non_core_property_parts(content: bytes) -> list[dict[str, str]]:
    """校验文档属性 namespace 并摘要非空自定义载荷。

    输入参数：
        content：已通过 package 成员、content type 和关系图闭集
            的 XLSX 字节。
    输出返回值：
        固定 gold 的空 custom-properties part 与缺失等价；
        非空自定义属性返回仅含 part 身份与 payload SHA-256
        的脱敏记录。app/core 产生器 metadata 仅做精确
        namespace/QName 分类，不作为表格可见语义。
    异常：
        ValueError：属性 part 根、元素或属性 namespace 漂移。
    """

    records: list[dict[str, str]] = []
    specifications = (
        (
            "docProps/app.xml",
            f"{{{_EXTENDED_PROPERTIES_NAMESPACE}}}Properties",
            frozenset(
                {
                    _EXTENDED_PROPERTIES_NAMESPACE,
                    _DOCUMENT_PROPERTY_VALUE_NAMESPACE,
                }
            ),
            frozenset(),
        ),
        (
            "docProps/core.xml",
            f"{{{_CORE_PROPERTIES_NAMESPACE}}}coreProperties",
            frozenset(
                {
                    _CORE_PROPERTIES_NAMESPACE,
                    _DUBLIN_CORE_NAMESPACE,
                    _DUBLIN_CORE_TERMS_NAMESPACE,
                    _DUBLIN_CORE_TYPE_NAMESPACE,
                }
            ),
            frozenset({_XML_SCHEMA_INSTANCE_NAMESPACE}),
        ),
    )
    with zipfile.ZipFile(BytesIO(content), mode="r") as archive:
        for (
            member_name,
            root_tag,
            element_namespaces,
            attribute_namespaces,
        ) in specifications:
            payload = _read_ooxml_member_bounded(archive, member_name)
            _validate_property_part_namespaces(
                payload,
                root_tag=root_tag,
                element_namespaces=element_namespaces,
                attribute_namespaces=attribute_namespaces,
            )
        if "docProps/custom.xml" not in archive.namelist():
            return records
        custom_payload = _read_ooxml_member_bounded(
            archive,
            "docProps/custom.xml",
        )
    custom_root = _validate_property_part_namespaces(
        custom_payload,
        root_tag=f"{{{_CUSTOM_PROPERTIES_NAMESPACE}}}Properties",
        element_namespaces=frozenset(
            {
                _CUSTOM_PROPERTIES_NAMESPACE,
                _DOCUMENT_PROPERTY_VALUE_NAMESPACE,
            }
        ),
        attribute_namespaces=frozenset(),
    )
    if len(custom_root) == 0 and not (custom_root.text or "").strip():
        return records
    records.append(
        {
            "part": "docProps/custom.xml",
            "payload_sha256": hashlib.sha256(custom_payload).hexdigest(),
        }
    )
    return records


def _validate_property_part_namespaces(
    payload: bytes,
    *,
    root_tag: str,
    element_namespaces: frozenset[str],
    attribute_namespaces: frozenset[str],
) -> ET.Element:
    """验证一个 document-property part 的精确 namespace 边界。

    输入参数：
        payload：已按 2 MiB 上限完整读取的 XML。
        root_tag：预期 Clark notation 根元素。
        element_namespaces：已审定元素 namespace 闭集。
        attribute_namespaces：除无 namespace 属性外允许的闭集。
    输出返回值：
        通过根、QName 和 namespace 审计的 ElementTree 根。
    """

    try:
        root = ET.fromstring(payload)
    except ET.ParseError:
        raise ValueError("OOXML document properties XML 无效") from None
    if root.tag != root_tag or root.attrib:
        raise ValueError("OOXML document properties 根闭集失效")
    for element in root.iter():
        if _qualified_name_namespace(element.tag) not in element_namespaces:
            raise ValueError("OOXML document properties namespace 未审定")
        for attribute in element.attrib:
            namespace = _qualified_name_namespace(
                attribute,
                allow_unqualified=True,
            )
            if namespace is not None and namespace not in attribute_namespaces:
                raise ValueError("OOXML document properties 属性 namespace 未审定")
    return root


def _validate_ooxml_package_members(content: bytes) -> None:
    """校验 Excel-008 单工作表 OOXML part 身份闭集。

    输入参数：
        content：已通过 ZIP 便携身份、数量和展开预算门的
            XLSX 字节。
    输出返回值：
        无；固定核心 part 全部存在，可选产生器 part 仅来自
        主题、shared strings、空 drawing/自定义属性闭集时返回。
    异常：
        ValueError：必要 part 缺失或出现任何未登记成员。
    """

    with zipfile.ZipFile(BytesIO(content), mode="r") as archive:
        members = frozenset(archive.namelist())
    allowed_members = _REQUIRED_OOXML_PACKAGE_MEMBERS | _OPTIONAL_OOXML_PACKAGE_MEMBERS
    if not _REQUIRED_OOXML_PACKAGE_MEMBERS.issubset(members) or not members.issubset(
        allowed_members
    ):
        raise ValueError("OOXML part 身份闭集失效")


def _validate_ooxml_content_types(content: bytes) -> None:
    """校验每个已审定 package part 的精确 content type。

    输入参数：
        content：成员身份已通过 Excel-008 闭集门的 XLSX 字节。
    输出返回值：
        无；``Default``/``Override`` namespace、属性、唯一性
        与每个现存 part 的最终 MIME 映射全部精确时返回。
    异常：
        ValueError：声明重复、悬空、namespace/属性漂移，
            或任一 part 被声明为错误 MIME。
    """

    with zipfile.ZipFile(BytesIO(content), mode="r") as archive:
        members = frozenset(archive.namelist())
        payload = _read_ooxml_member_bounded(archive, "[Content_Types].xml")
    try:
        root = ET.fromstring(payload)
    except ET.ParseError:
        raise ValueError("OOXML content types XML 无效") from None
    if (
        root.tag != f"{{{_CONTENT_TYPES_NAMESPACE}}}Types"
        or root.attrib
        or (root.text or "").strip()
    ):
        raise ValueError("OOXML content types 根闭集失效")

    defaults: dict[str, str] = {}
    overrides: dict[str, str] = {}
    for declaration in root:
        if len(declaration) or (declaration.text or "").strip():
            raise ValueError("OOXML content type 子元素无效")
        if declaration.tag == f"{{{_CONTENT_TYPES_NAMESPACE}}}Default":
            if set(declaration.attrib) != {"Extension", "ContentType"}:
                raise ValueError("OOXML Default 字段闭集漂移")
            extension = declaration.attrib["Extension"]
            content_type = declaration.attrib["ContentType"]
            if (
                extension in defaults
                or _ALLOWED_DEFAULT_CONTENT_TYPES.get(extension) != content_type
            ):
                raise ValueError("OOXML Default content type 未审定")
            defaults[extension] = content_type
            continue
        if declaration.tag == f"{{{_CONTENT_TYPES_NAMESPACE}}}Override":
            if set(declaration.attrib) != {"PartName", "ContentType"}:
                raise ValueError("OOXML Override 字段闭集漂移")
            raw_part_name = declaration.attrib["PartName"]
            if not raw_part_name.startswith("/") or raw_part_name.startswith("//"):
                raise ValueError("OOXML Override part 身份无效")
            part_name = raw_part_name[1:]
            content_type = declaration.attrib["ContentType"]
            if (
                part_name in overrides
                or part_name not in members
                or _EXPECTED_CONTENT_TYPE_BY_MEMBER.get(part_name) != content_type
            ):
                raise ValueError("OOXML Override content type 未审定")
            overrides[part_name] = content_type
            continue
        raise ValueError("OOXML content type 元素 namespace 未审定")

    if defaults.get("rels") != _RELATIONSHIPS_CONTENT_TYPE:
        raise ValueError("OOXML rels 默认 content type 缺失")
    if defaults.get("xml") != "application/xml":
        raise ValueError("OOXML xml 默认 content type 缺失")
    for member_name in members - {"[Content_Types].xml"}:
        extension = member_name.rsplit(".", 1)[-1]
        resolved_content_type = overrides.get(member_name, defaults.get(extension))
        if resolved_content_type != _EXPECTED_CONTENT_TYPE_BY_MEMBER.get(member_name):
            raise ValueError("OOXML part content type 不匹配")


def _resolve_relationship_target(owner_part: str, target: str) -> str:
    """将内部 OOXML relationship target 解析为包成员身份。

    输入参数：
        owner_part：relationship 归属 part；package 根关系为空串。
        target：经预检证明为内部的 Target 属性。
    输出返回值：
        无前导斜线、已解析 ``..`` 的规范 POSIX 成员名。
    """

    if (
        not isinstance(owner_part, str)
        or not isinstance(target, str)
        or not target
        or "\\" in target
        or "\x00" in target
        or "?" in target
        or "#" in target
    ):
        raise ValueError("OOXML relationship target 无效")
    if target.startswith("/"):
        candidate = target.removeprefix("/")
    else:
        candidate = posixpath.join(posixpath.dirname(owner_part), target)
    normalized = posixpath.normpath(candidate)
    if (
        not normalized
        or normalized in {".", ".."}
        or normalized.startswith("../")
        or normalized.startswith("/")
        or ":" in normalized.split("/", 1)[0]
    ):
        raise ValueError("OOXML relationship target 越界")
    return normalized


def _read_ooxml_member_bounded(
    archive: zipfile.ZipFile,
    member_name: str,
) -> bytes:
    """在公用 XML 预算内读取一个已定位 OOXML 成员。

    输入参数：
        archive：已通过被动 ZIP 预检的工作簿。
        member_name：由当前闭集决定的成员身份。
    输出返回值：
        声明大小与实际流一致且不超过 2 MiB 的字节。
    """

    try:
        member = archive.getinfo(member_name)
    except KeyError:
        raise ValueError("OOXML 成员缺失") from None
    if member.file_size > _OOXML_VISIBLE_PART_MAX_BYTES:
        raise ValueError("OOXML 成员资源超限")
    with archive.open(member, mode="r") as stream:
        payload = stream.read(_OOXML_VISIBLE_PART_MAX_BYTES + 1)
        tail = stream.read(1)
    if (
        len(payload) > _OOXML_VISIBLE_PART_MAX_BYTES
        or len(payload) != member.file_size
        or tail != b""
    ):
        raise ValueError("OOXML 成员读取不完整")
    return payload


def _qualified_name_namespace(
    name: object,
    *,
    allow_unqualified: bool = False,
) -> str | None:
    """验证 ElementTree QName 并返回精确 namespace。

    输入参数：
        name：ElementTree 解析后的 tag 或 attribute 名。
        allow_unqualified：是否允许 OOXML 中的无 namespace 属性。
    输出返回值：
        Clark notation 中的 namespace；允许的无限定名返回
        ``None``。
    """

    if not isinstance(name, str) or not name:
        raise ValueError("OOXML QName 无效")
    if name.startswith("{") and "}" in name:
        namespace, local_name = name[1:].split("}", 1)
        if namespace and local_name:
            return namespace
    elif allow_unqualified and "{" not in name and "}" not in name:
        return None
    raise ValueError("OOXML QName 未限定")


def _normalize_workbook_theme(workbook: Any) -> list[str]:
    """将固定 Excel 默认 theme 与 LibreOffice 缺省表示归一。

    输入参数：
        workbook：已由受控 openpyxl parser 物化的工作簿。
    输出返回值：
        已验证默认 theme 或其他 theme 的脱敏摘要标记。
    """

    theme = workbook.loaded_theme
    if theme is None:
        return ["verified-default"]
    if not isinstance(theme, bytes):
        raise ValueError("工作簿 theme 类型无效")
    digest = hashlib.sha256(theme).hexdigest()
    if digest in _VERIFIED_DEFAULT_THEME_SHA256:
        return ["verified-default"]
    return ["custom-sha256", digest]


def _normalize_sheet_view(sheet_view: Any) -> dict[str, object]:
    """将工作表的可见 view 字段规范化为默认等价闭集。

    输入参数：
        sheet_view：openpyxl ``SheetView``。
    输出返回值：
        window protection、gridlines、ruler、whitespace、zoom、headers、
        zeros、RTL、color/workbook view ID 及已验证默认字段。

    ``pane`` 和 ``selection`` 是当前 GUI 会话的滚动/光标导航状态；
    固定 input/gold 的 activeCell/sqref 本就不同，因此不纳入
    artifact 语义指纹。会影响表格布局的 frozen pane 由上层
    ``freeze_panes`` 字段独立锁定。
    """

    return {
        "window_protection": _default_bool(
            sheet_view.windowProtection,
            False,
        ),
        "show_formulas": _default_bool(sheet_view.showFormulas, False),
        "show_gridlines": _default_bool(sheet_view.showGridLines, True),
        "show_headers": _default_bool(
            sheet_view.showRowColHeaders,
            True,
        ),
        "show_zeros": _default_bool(sheet_view.showZeros, True),
        "right_to_left": _default_bool(sheet_view.rightToLeft, False),
        "tab_selected": _default_bool(sheet_view.tabSelected, False),
        "show_outline_symbols": _default_bool(
            sheet_view.showOutlineSymbols,
            True,
        ),
        "show_ruler": _default_bool(sheet_view.showRuler, True),
        "show_white_space": _default_bool(
            sheet_view.showWhiteSpace,
            True,
        ),
        "default_grid_color": _default_bool(
            sheet_view.defaultGridColor,
            True,
        ),
        "view": _default_string(sheet_view.view, "normal"),
        "top_left_cell": _default_string(
            sheet_view.topLeftCell,
            "A1",
        ),
        "zoom": _default_int(sheet_view.zoomScale, 100),
        "zoom_normal": _default_int(
            sheet_view.zoomScaleNormal,
            100,
        ),
        "zoom_page_layout": _default_int(
            sheet_view.zoomScalePageLayoutView,
            100,
        ),
        "zoom_sheet_layout": _default_int(
            sheet_view.zoomScaleSheetLayoutView,
            100,
        ),
        "zoom_to_fit": _default_bool(sheet_view.zoomToFit, False),
        "color_id": _default_int(sheet_view.colorId, 64),
        "workbook_view_id": _default_int(
            sheet_view.workbookViewId,
            0,
        ),
    }


def _normalize_auto_filter(value: object) -> list[str]:
    """归一固定 gold 中 LibreOffice 显式写入的等价筛选范围。

    输入参数：
        value：``worksheet.auto_filter.ref`` 候选值。
    输出返回值：
        input 的 ``None`` 与固定 gold ``B3:F16`` 统一为
        ``verified-default``；任何第三范围保留精确值。
    """

    if value is None or value == "B3:F16":
        return ["verified-default"]
    if not isinstance(value, str) or not value:
        raise ValueError("auto-filter 范围无效")
    return ["range", value]


def _normalize_header_footer(header_footer: Any) -> dict[str, object]:
    """投影完整页眉页脚标志与六组三区可见内容。

    输入参数：
        header_footer：openpyxl ``HeaderFooter``，包含 odd/even/first
            的 header/footer 项和四个显式标志。
    输出返回值：
        固定顺序的标志及 left/center/right 文本、字体、
        字号和颜色闭集。
    """

    items: dict[str, object] = {}
    for item_name in (
        "oddHeader",
        "evenHeader",
        "firstHeader",
        "oddFooter",
        "evenFooter",
        "firstFooter",
    ):
        item = getattr(header_footer, item_name)
        items[item_name] = {
            part_name: {
                "text": _optional_style_string(part.text),
                "font": _optional_style_string(part.font),
                "size": _optional_style_number(part.size),
                "color": _optional_style_string(part.color),
            }
            for part_name, part in (
                ("left", item.left),
                ("center", item.center),
                ("right", item.right),
            )
        }
    return {
        "different_odd_even": _default_bool(
            header_footer.differentOddEven,
            False,
        ),
        "different_first": _default_bool(
            header_footer.differentFirst,
            False,
        ),
        "scale_with_document": _default_bool(
            header_footer.scaleWithDoc,
            True,
        ),
        "align_with_margins": _default_bool(
            header_footer.alignWithMargins,
            True,
        ),
        "items": items,
    }


def _normalize_closed_attributes(
    instance: Any | None,
    expected_fields: tuple[str, ...],
) -> dict[str, list[object]]:
    """以显式字段闭集投影 openpyxl 序列化对象。

    输入参数：
        instance：SheetProtection、WorkbookProtection 或 ``None``。
        expected_fields：当前依赖版本经审计的全字段顺序。
    输出返回值：
        每个字段都带类型标记的固定字典；``None`` 对象
        与所有字段均为 ``None`` 的显式序列化等价。

    异常：
        ValueError：三方序列化字段集漂移或字段类型未封闭。
    """

    if instance is None:
        return {field: ["none"] for field in expected_fields}
    if tuple(getattr(instance, "__attrs__", ())) != expected_fields:
        raise ValueError("Office 序列化字段闭集漂移")
    return {
        field: _normalize_metadata_scalar_or_none(getattr(instance, field))
        for field in expected_fields
    }


def _normalize_metadata_scalar_or_none(value: object) -> list[object]:
    """把可选 protection/定义名元数据投影为类型标量。

    输入参数：
        value：``None`` 或 bool/int/finite-float/str 之一。
    输出返回值：
        不丢失 ``None`` 与显式默认差异的类型标记列表。
    """

    if value is None:
        return ["none"]
    return _normalize_baseline_scalar(value)


def _normalize_defined_names(workbook: Any) -> list[dict[str, list[object]]]:
    """按全字段内容排序投影 workbook defined names。

    输入参数：
        workbook：已由受控 parser 物化的 openpyxl 工作簿。
    输出返回值：
        name、scope、公式与所有可选属性均在内的有序记录；
        XML/字典插入顺序不影响语义指纹。
    """

    records = [
        _normalize_closed_attributes(item, _DEFINED_NAME_FIELDS)
        for item in workbook.defined_names.values()
    ]
    return sorted(
        records,
        key=lambda record: json.dumps(
            record,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8", errors="strict"),
    )


def _normalize_workbook_views(workbook: Any) -> list[dict[str, object]]:
    """投影工作簿窗口可见状态并只合并已审定几何对。

    输入参数：
        workbook：已由受控 parser 物化的 openpyxl 工作簿。
    输出返回值：
        按 workbook XML 顺序的窗口观测；包含可见性、最小化、
        滚动条/标签、firstSheet/activeTab、日期分组和几何。

    Lee 固定 input 的 ``0/780/34200/19860/600`` 与 gold 的
    ``0/0/16384/8192/500`` 是已审定的 LibreOffice/Excel 窗口差异；
    两者仅在这里合并，其他任何几何值保留并使基线失效。
    """

    records: list[dict[str, object]] = []
    for view in workbook.views:
        if tuple(getattr(view, "__attrs__", ())) != _WORKBOOK_VIEW_FIELDS:
            raise ValueError("workbook view 字段闭集漂移")
        geometry = tuple(
            getattr(view, field) for field in _WORKBOOK_VIEW_GEOMETRY_FIELDS
        )
        if geometry in _VERIFIED_WORKBOOK_VIEW_GEOMETRIES:
            normalized_geometry: object = ["verified-default"]
        else:
            normalized_geometry = [
                "exact",
                {
                    field: _normalize_metadata_scalar_or_none(getattr(view, field))
                    for field in _WORKBOOK_VIEW_GEOMETRY_FIELDS
                },
            ]
        records.append(
            {
                field: _normalize_metadata_scalar_or_none(getattr(view, field))
                for field in _WORKBOOK_VIEW_FIELDS
                if field not in _WORKBOOK_VIEW_GEOMETRY_FIELDS
            }
            | {"geometry": normalized_geometry}
        )
    return records


def _normalize_workbook_calculation(
    workbook: Any,
) -> dict[str, list[object]]:
    """投影 CalcProperties 全字段并只合并真实产生器差异。

    输入参数：
        workbook：已由受控 openpyxl 边界物化的工作簿。
    输出返回值：
        ``CalcProperties.__attrs__`` 13 字段的类型明确字典。

    Lee input/gold 仅存在 calcId、A1 refMode 和迭代默认值的
    显式/缺省差异；这些精确已审定值合并，任何第三值
    仍以原类型进入基线。
    """

    calculation = workbook.calculation
    if tuple(getattr(calculation, "__attrs__", ())) != _CALCULATION_FIELDS:
        raise ValueError("workbook calculation 字段闭集漂移")
    result: dict[str, list[object]] = {}
    for field in _CALCULATION_FIELDS:
        value = getattr(calculation, field)
        verified_values = _VERIFIED_CALCULATION_DEFAULTS.get(field)
        if verified_values is not None and _matches_typed_value(
            value,
            verified_values,
        ):
            result[field] = ["verified-default"]
        else:
            result[field] = _normalize_metadata_scalar_or_none(value)
    return result


def _matches_typed_value(value: object, candidates: tuple[object, ...]) -> bool:
    """按精确 Python 类型匹配已审定默认值。

    输入参数：
        value：当前 openpyxl 字段值。
        candidates：允许合并的精确值 tuple。
    输出返回值：
        类型和值同时相等时为真，防止 ``False == 0``
        或 ``True == 1`` 扩大已审定等价类。
    """

    return any(
        type(value) is type(candidate) and value == candidate
        for candidate in candidates
    )


def _normalize_print_ranges(worksheet: Any) -> dict[str, list[str]]:
    """投影 openpyxl 从内建 defined names 拆出的打印区域。

    输入参数：
        worksheet：当前受控工作表。
    输出返回值：
        重复行、重复列和 print area 的显式字符串投影。

    openpyxl 加载 ``_xlnm.Print_Titles/_xlnm.Print_Area`` 时会将它们
    从 ``workbook.defined_names`` 移入 worksheet 属性，因此必须
    在此显式收集，才能完成定义名语义闭集。
    """

    return {
        "title_rows": _normalize_optional_reference(worksheet.print_title_rows),
        "title_columns": _normalize_optional_reference(worksheet.print_title_cols),
        "area": _normalize_optional_reference(worksheet.print_area),
    }


def _normalize_optional_reference(value: object) -> list[str]:
    """验证可选 defined-name 引用字符串。

    输入参数：
        value：``None``、空串或 openpyxl 解析后的引用字符串。
    输出返回值：
        空引用为空列表，否则为单元字符串列表。
    """

    if value is None or value == "":
        return []
    if not isinstance(value, str):
        raise ValueError("打印定义名引用类型无效")
    return [value]


def _normalize_page_setup(worksheet: Any) -> dict[str, object]:
    """归一 input/gold 显式默认差异并锁定打印可见语义。

    输入参数：
        worksheet：当前受控工作表。
    输出返回值：
        页面方向、纸张、缩放、拟合、DPI、页边距和打印选项。
    """

    page_setup = worksheet.page_setup
    margins = worksheet.page_margins
    print_options = worksheet.print_options
    page_setup_properties = worksheet.sheet_properties.pageSetUpPr
    if tuple(getattr(page_setup, "__attrs__", ())) != _PAGE_SETUP_FIELDS:
        raise ValueError("page setup 字段闭集漂移")
    if tuple(getattr(print_options, "__attrs__", ())) != _PRINT_OPTIONS_FIELDS:
        raise ValueError("print options 字段闭集漂移")
    grid_lines_set = print_options.gridLinesSet
    if _matches_typed_value(grid_lines_set, (None, True)):
        normalized_grid_lines_set: list[object] = ["verified-default"]
    else:
        normalized_grid_lines_set = _normalize_metadata_scalar_or_none(grid_lines_set)
    return {
        "orientation": _default_string(
            page_setup.orientation,
            "portrait",
        ),
        "paper_size": _default_int(page_setup.paperSize, 9),
        "scale": _default_int(page_setup.scale, 100),
        "fit_to_height": _default_int(page_setup.fitToHeight, 1),
        "fit_to_width": _default_int(page_setup.fitToWidth, 1),
        "first_page_number": _normalize_metadata_scalar_or_none(
            page_setup.firstPageNumber
        ),
        "use_first_page_number": _normalize_metadata_scalar_or_none(
            page_setup.useFirstPageNumber
        ),
        "paper_height": _normalize_metadata_scalar_or_none(page_setup.paperHeight),
        "paper_width": _normalize_metadata_scalar_or_none(page_setup.paperWidth),
        "page_order": _default_string(
            page_setup.pageOrder,
            "downThenOver",
        ),
        "use_printer_defaults": _normalize_metadata_scalar_or_none(
            page_setup.usePrinterDefaults
        ),
        "black_and_white": _default_bool(
            page_setup.blackAndWhite,
            False,
        ),
        "draft": _default_bool(page_setup.draft, False),
        "cell_comments": _normalize_metadata_scalar_or_none(page_setup.cellComments),
        "errors": _normalize_metadata_scalar_or_none(page_setup.errors),
        "horizontal_dpi": _default_int(
            page_setup.horizontalDpi,
            300,
        ),
        "vertical_dpi": _default_int(
            page_setup.verticalDpi,
            300,
        ),
        "copies": _default_int(page_setup.copies, 1),
        "relationship_id": _normalize_metadata_scalar_or_none(page_setup.id),
        "fit_to_page": _default_bool(
            page_setup_properties.fitToPage,
            False,
        ),
        "auto_page_breaks": _default_bool(
            page_setup_properties.autoPageBreaks,
            False,
        ),
        "margins": {
            "left": _normalize_visible_length(margins.left, quantum=0.05),
            "right": _normalize_visible_length(margins.right, quantum=0.05),
            "top": _normalize_visible_length(margins.top, quantum=0.05),
            "bottom": _normalize_visible_length(
                margins.bottom,
                quantum=0.05,
            ),
            "header": _normalize_visible_length(
                margins.header,
                quantum=0.05,
            ),
            "footer": _normalize_visible_length(
                margins.footer,
                quantum=0.05,
            ),
        },
        "print_options": {
            "horizontal_centered": _default_bool(
                print_options.horizontalCentered,
                False,
            ),
            "vertical_centered": _default_bool(
                print_options.verticalCentered,
                False,
            ),
            "headings": _default_bool(print_options.headings, False),
            "gridlines": _default_bool(print_options.gridLines, False),
            "gridlines_set": normalized_grid_lines_set,
        },
    }


def _normalize_page_breaks(worksheet: Any) -> dict[str, list[dict[str, list[object]]]]:
    """按完整 Break 字段投影手动行/列分页符。

    输入参数：
        worksheet：当前受控工作表。
    输出返回值：
        分别按 id/min/max/man/pt 全字段内容排序的
        row breaks 和 column breaks。
    """

    def normalize(items: list[Any]) -> list[dict[str, list[object]]]:
        """投影并排序一类 Break 序列。

        输入参数：
            items：openpyxl ``Break`` 列表。
        输出返回值：
            完整字段的规范化有序记录。
        """

        records = [
            _normalize_closed_attributes(item, _PAGE_BREAK_FIELDS) for item in items
        ]
        return sorted(
            records,
            key=lambda record: json.dumps(
                record,
                ensure_ascii=False,
                separators=(",", ":"),
                sort_keys=True,
            ).encode("utf-8", errors="strict"),
        )

    return {
        "rows": normalize(list(worksheet.row_breaks.brk)),
        "columns": normalize(list(worksheet.col_breaks.brk)),
    }


def _normalize_sheet_properties(properties: Any) -> dict[str, object]:
    """投影 worksheet properties 和 outlinePr 的完整字段闭集。

    输入参数：
        properties：openpyxl ``WorksheetProperties``。
    输出返回值：
        codeName、格式计算、filter/published、sync、transition
        以及 outlinePr 四字段的固定投影。

    固定 input 的 ``filterMode=None`` 与 gold 的 ``True`` 由
    已验证的 auto-filter 序列化差异产生；仅这两值合并，
    显式 ``False`` 作为第三值保留并导致基线失效。
    """

    if tuple(getattr(properties, "__attrs__", ())) != _WORKSHEET_PROPERTY_FIELDS:
        raise ValueError("worksheet properties 字段闭集漂移")
    result = {
        field: _normalize_metadata_scalar_or_none(getattr(properties, field))
        for field in _WORKSHEET_PROPERTY_FIELDS
        if field != "filterMode"
    }
    filter_mode = properties.filterMode
    if filter_mode is None or filter_mode is True:
        result["filterMode"] = ["verified-default"]
    elif filter_mode is False:
        result["filterMode"] = ["bool", False]
    else:
        raise ValueError("worksheet filterMode 类型无效")
    result["outlinePr"] = _normalize_closed_attributes(
        properties.outlinePr,
        _OUTLINE_PROPERTY_FIELDS,
    )
    return result


def _default_bool(value: object, default: bool) -> bool:
    """把可选布尔字段归一为固定默认值。

    输入参数：
        value：``None`` 或真实 bool。
        default：经 input/gold 审定的默认值。
    输出返回值：归一后的 bool。
    """

    if value is None:
        return default
    if type(value) is not bool or type(default) is not bool:
        raise ValueError("布尔默认字段无效")
    return value


def _default_int(value: object, default: int) -> int:
    """把可选整数字段归一为固定默认值。

    输入参数：
        value：``None`` 或真实 int。
        default：经 input/gold 审定的默认整数。
    输出返回值：归一后的 int。
    """

    if value is None:
        return default
    if type(value) is not int or type(default) is not int:
        raise ValueError("整数默认字段无效")
    return value


def _default_string(value: object, default: str) -> str:
    """把可选字符串字段归一为固定默认值。

    输入参数：
        value：``None`` 或字符串。
        default：经 input/gold 审定的默认字符串。
    输出返回值：归一后的字符串。
    """

    if value is None:
        return default
    if not isinstance(value, str) or not isinstance(default, str):
        raise ValueError("字符串默认字段无效")
    return value


def _normalize_baseline_scalar_or_none(value: object) -> list[object]:
    """规范化单元格值，并为有样式空格提供显式标记。

    输入参数：
        value：openpyxl 返回的单元格值。
    输出返回值：
        ``None`` 的固定标记或通用基线标量投影。
    """

    if value is None:
        return ["none"]
    return _normalize_baseline_scalar(value)


def _normalize_cell_visible_style(cell: Any) -> dict[str, object]:
    """将 Excel-008 单元格的可见样式规范化为 JSON 闭集。

    输入参数：
        cell：已受资源门限制的 ``Cell`` 或 ``MergedCell``。
    输出返回值：
        字体、填充、边框、对齐、保护与数字格式的
        稳定全字段字典。
    """

    font = cell.font
    alignment = cell.alignment
    border = cell.border
    protection = cell.protection
    if tuple(getattr(font, "__elements__", ())) != _FONT_ELEMENT_FIELDS:
        raise ValueError("单元格字体字段闭集漂移")
    if tuple(getattr(protection, "__attrs__", ())) != _CELL_PROTECTION_FIELDS:
        raise ValueError("单元格 protection 字段闭集漂移")
    charset = font.charset
    if _matches_typed_value(charset, (None, 1)):
        normalized_charset: list[object] = ["verified-default"]
    else:
        normalized_charset = _normalize_metadata_scalar_or_none(charset)
    scheme = font.scheme
    if _matches_typed_value(scheme, (None, "minor")):
        normalized_scheme: list[object] = ["verified-default"]
    else:
        normalized_scheme = _normalize_metadata_scalar_or_none(scheme)
    return {
        "font": {
            "name": _optional_style_string(font.name),
            "charset": normalized_charset,
            "family": _optional_style_number(font.family),
            "size": _optional_style_number(font.sz),
            "bold": bool(font.bold),
            "italic": bool(font.italic),
            "strike": bool(font.strike),
            "shadow": bool(font.shadow),
            "outline": bool(font.outline),
            "condense": bool(font.condense),
            "extend": bool(font.extend),
            "underline": _optional_style_string(font.underline),
            "vertical_alignment": _optional_style_string(font.vertAlign),
            "scheme": normalized_scheme,
            "color": _normalize_style_color(font.color),
        },
        "fill": _normalize_cell_fill(cell.fill),
        "border": {
            side_name: _normalize_border_side(getattr(border, side_name))
            for side_name in (
                "start",
                "end",
                "left",
                "right",
                "top",
                "bottom",
                "diagonal",
                "vertical",
                "horizontal",
            )
        }
        | {
            "diagonal_up": bool(border.diagonalUp),
            "diagonal_down": bool(border.diagonalDown),
            "outline": border.outline is not False,
        },
        "alignment": {
            "horizontal": alignment.horizontal or "general",
            "vertical": alignment.vertical or "bottom",
            "rotation": int(alignment.textRotation or 0),
            "wrap_text": bool(alignment.wrapText),
            "shrink_to_fit": bool(alignment.shrinkToFit),
            "indent": _normalize_style_number(alignment.indent or 0),
            "relative_indent": _normalize_style_number(alignment.relativeIndent or 0),
            "justify_last_line": bool(alignment.justifyLastLine),
            "reading_order": _normalize_style_number(alignment.readingOrder or 0),
        },
        "number_format": _optional_style_string(cell.number_format),
        "protection": _normalize_closed_attributes(
            protection,
            _CELL_PROTECTION_FIELDS,
        ),
        "quote_prefix": bool(cell.quotePrefix),
        "pivot_button": bool(cell.pivotButton),
    }


def _normalize_baseline_scalar(value: object) -> list[object]:
    """把单元格值规范化为类型明确的 JSON 标量。

    输入参数：
        value：openpyxl 返回的 bool/int/finite-float/str。
    输出返回值：
        含类型标记的稳定两元列表。
    """

    if type(value) is bool:
        return ["bool", value]
    if type(value) is int:
        return ["int", str(value)]
    if type(value) is float and math.isfinite(value):
        return ["float", value.hex()]
    if type(value) is str:
        return ["str", value]
    raise ValueError("基线单元格标量类型未固定")


def _normalize_cell_fill(fill: Any) -> list[object]:
    """规范化单元格 pattern fill 并拒绝未固定渐变填充。

    输入参数：
        fill：openpyxl 单元格填充对象。
    输出返回值：
        无填充、solid 或其他 pattern 的固定列表。
    """

    fill_type = getattr(fill, "fill_type", None)
    is_pattern_fill = hasattr(fill, "fgColor") and hasattr(fill, "bgColor")
    if fill_type is None:
        if not is_pattern_fill:
            raise ValueError("单元格填充类型未固定")
        return ["none"]
    if not isinstance(fill_type, str) or not is_pattern_fill:
        raise ValueError("单元格填充类型无效")
    foreground = _normalize_style_color(fill.fgColor)
    if fill_type == "solid":
        return ["solid", foreground]
    return [
        "pattern",
        fill_type,
        foreground,
        _normalize_style_color(fill.bgColor),
    ]


def _normalize_border_side(side: Any | None) -> list[object]:
    """将一个可选边线归一为线型和颜色。

    输入参数：
        side：openpyxl ``Side`` 或 ``None``。
    输出返回值：
        无线型的缺省列表，或线型+规范化颜色。
    """

    if side is None or side.style is None:
        return [None, None]
    if not isinstance(side.style, str):
        raise ValueError("单元格边线类型无效")
    return [side.style, _normalize_style_color(side.color)]


def _normalize_style_color(color: Any | None) -> list[object]:
    """规范化 Office 可见颜色的 RGB/indexed/theme 表示。

    输入参数：
        color：openpyxl ``Color`` 或缺省值。
    输出返回值：
        自动色、ARGB、indexed 或 theme+tint 的固定列表。
    """

    if color is None:
        return ["automatic"]
    color_type = getattr(color, "type", None)
    if color_type == "auto":
        return ["automatic"]
    tint = _normalize_style_number(getattr(color, "tint", 0) or 0)
    if color_type == "rgb":
        rgb = getattr(color, "rgb", None)
        if not isinstance(rgb, str):
            raise ValueError("单元格 RGB 颜色无效")
        normalized_rgb = rgb.upper()
        if len(normalized_rgb) == 6:
            normalized_rgb = "FF" + normalized_rgb
        if len(normalized_rgb) != 8 or any(
            character not in "0123456789ABCDEF" for character in normalized_rgb
        ):
            raise ValueError("单元格 RGB 颜色无效")
        if normalized_rgb == "FF000000" and tint == "0x0.0p+0":
            return ["opaque-black"]
        return ["argb", normalized_rgb, tint]
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        if type(theme) is not int or theme < 0:
            raise ValueError("单元格 theme 颜色无效")
        if theme == 1 and tint == "0x0.0p+0":
            return ["opaque-black"]
        return ["theme", theme, tint]
    if color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        if type(indexed) is not int or indexed < 0:
            raise ValueError("单元格 indexed 颜色无效")
        if indexed == 64:
            return ["automatic"]
        return ["indexed", indexed, tint]
    raise ValueError("单元格颜色类型未固定")


def _normalize_style_number(value: object) -> str:
    """把样式有限数值规范化为十六进制浮点字符串。

    输入参数：
        value：int/float 样式字段。
    输出返回值：稳定 ``float.hex`` 字符串。
    """

    if (
        isinstance(value, bool)
        or not isinstance(value, (int, float))
        or not math.isfinite(float(value))
    ):
        raise ValueError("单元格样式数值无效")
    return float(value).hex()


def _optional_style_string(value: object) -> str | None:
    """验证一个可选样式字段。

    输入参数：
        value：字体名、下划线或垂直对齐候选值。
    输出返回值：``None`` 或原字符串。
    """

    if value is None or isinstance(value, str):
        return value
    raise ValueError("单元格样式字符串无效")


def _normalize_visible_length(
    value: object,
    *,
    quantum: float,
) -> str | None:
    """将 Office 可见尺寸按固定半入粒度归一。

    输入参数：
        value：可选行高、列宽或页边距。
        quantum：任务协议固定的最小可见粒度。
    输出返回值：``None`` 或整数桶字符串。
    """

    if value is None:
        return None
    if (
        isinstance(value, bool)
        or not isinstance(value, (int, float))
        or isinstance(quantum, bool)
        or not isinstance(quantum, (int, float))
    ):
        raise ValueError("可见尺寸类型无效")
    normalized = float(value)
    normalized_quantum = float(quantum)
    if (
        not math.isfinite(normalized)
        or normalized < 0
        or not math.isfinite(normalized_quantum)
        or normalized_quantum <= 0
    ):
        raise ValueError("可见尺寸无效")
    return str(math.floor(normalized / normalized_quantum + 0.5))


def _optional_style_number(value: object) -> str | None:
    """验证并规范化一个可选样式数值。

    输入参数：
        value：字体大小等可选 int/float 字段。
    输出返回值：
        ``None`` 或稳定的十六进制浮点字符串。
    """

    if value is None:
        return None
    return _normalize_style_number(value)


def _is_default_row_dimension_ignoring_hidden(
    dimension: Any,
    *,
    default_height: object,
) -> bool:
    """判断行维度除本任务允许的 hidden 外是否仅重申默认值。

    输入参数：
        dimension：openpyxl ``RowDimension``。
        default_height：当前 sheet 默认行高。
    输出返回值：
        行高处于同一 1 pt 桶，且其余可见结构缺省时为真。
    """

    return (
        _normalize_visible_length(dimension.height, quantum=1.0)
        == _normalize_visible_length(default_height, quantum=1.0)
        and not dimension.outlineLevel
        and not dimension.collapsed
        and not dimension.thickTop
        and not dimension.thickBot
        and not dimension.has_style
    )


def _failed_workbook(document_name: str) -> WorkbookHiddenRows:
    """为存在但无法安全解析的期望文件生成零分占位。

    输入参数：
        document_name：由精确固定文件名映射的身份。
    输出返回值：
        无隐藏行且基线失效的工作簿观测。
    """

    return WorkbookHiddenRows(
        document_name=document_name,
        hidden_rows=(),
        content_matches_baseline=False,
    )


__all__ = [
    "PINNED_HIDE_NA_ROWS_BASELINE_SHA256",
    "build_hide_na_rows_observation",
    "derive_hide_na_rows_baseline_sha256",
]

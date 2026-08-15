"""SearchAndWrite-008 generic artifact 闭集到强类型工作簿观测的转换边界。"""

from __future__ import annotations

import hashlib
from io import BytesIO
import json
import math
import multiprocessing
import os
from pathlib import PurePosixPath
import stat
import sys
import tempfile
import time
from types import MappingProxyType
from typing import Any
import unicodedata
import xml.etree.ElementTree as ET
import zipfile

from paraguibench.evaluation.pipeline_implicit import (
    SEARCHWRITE_DOCUMENT_CONTRACTS,
    SEARCHWRITE_XLSX_PROTOCOL_ID,
    SEARCHWRITE_XLSX_TASK_ID,
    SearchWriteCell,
    SearchWriteObservation,
    SearchWriteWorkbook,
)

from .artifact_evidence import (
    PipelineImplicitArtifactEvidenceError,
    PipelineImplicitArtifactObservation,
)


_WORKBOOK_CONTRACTS = MappingProxyType(
    {contract.relative_path: contract for contract in SEARCHWRITE_DOCUMENT_CONTRACTS}
)
_MAX_ARCHIVE_MEMBERS = 64
_MAX_ARCHIVE_MEMBER_BYTES = 2 * 1024 * 1024
_MAX_ARCHIVE_EXPANDED_BYTES = 8 * 1024 * 1024
_MAX_ARCHIVE_COMPRESSION_RATIO = 100
_MAX_XML_MEMBER_BYTES = 2 * 1024 * 1024
_MAX_XML_ELEMENTS = 8192
_MAX_XML_DEPTH = 64
_MAX_WORKSHEET_XML_ROWS = 128
_MAX_WORKSHEET_XML_CELLS = 1024
_MAX_WORKSHEET_XML_COLUMNS = 128
_MAX_WORKSHEET_XML_MERGES = 128
_MAX_RELATIONSHIP_ELEMENTS = 256
_MAX_WORKBOOK_SHEETS = 2
_MAX_WORKSHEET_ROWS = 128
_MAX_WORKSHEET_COLUMNS = 64
_MAX_WORKSHEET_CELLS = 4096
_MAX_WORKBOOK_CELLS = 4096
_MAX_POPULATED_CELLS = 1024
_MAX_CELL_TEXT_BYTES = 4096
_MAX_WORKBOOK_TEXT_BYTES = 128 * 1024
_MAX_FORMULAS = 256
_MAX_MERGED_RANGES = 128
_MAX_DIMENSION_ENTRIES = 256
_PARSER_WALL_TIMEOUT_SECONDS = 5.0
_PARSER_CPU_SECONDS = 2
_PARSER_ADDRESS_SPACE_BYTES = 768 * 1024 * 1024
_PARSER_RSS_LIMIT_BYTES = 256 * 1024 * 1024
_PARSER_OPEN_FILE_LIMIT = 64
_PARSER_POLL_INTERVAL_SECONDS = 0.05
_PARSER_RESULT_MAX_BYTES = 64 * 1024
_VERIFIED_DEFAULT_THEME_SHA256 = (
    "156137ac2d7fae74e0286df47c4d1c75e65d5ef1455ff74c4d46176aef06fe56"
)
_PARSER_WRITE_OPEN_FLAGS = (
    os.O_WRONLY | os.O_RDWR | os.O_CREAT | os.O_TRUNC | os.O_APPEND
)
_PARSER_DENIED_AUDIT_EVENTS = frozenset(
    {
        "ctypes.call_function",
        "ctypes.dlopen",
        "ctypes.dlsym",
        "ctypes.dlsym/handle",
        "os.chdir",
        "os.chmod",
        "os.chown",
        "os.fork",
        "os.forkpty",
        "os.link",
        "os.mkfifo",
        "os.mkdir",
        "os.mknod",
        "os.posix_spawn",
        "os.remove",
        "os.rename",
        "os.rmdir",
        "os.symlink",
        "os.system",
        "os.truncate",
        "os.utime",
        "os.setxattr",
        "os.removexattr",
        "subprocess.Popen",
    }
)
_REQUIRED_XLSX_MEMBERS = frozenset({"[Content_Types].xml", "xl/workbook.xml"})


class _WorkbookParseRejected(RuntimeError):
    """表示不可信 Office parser 未在受控边界内产生合法结果。"""


class _WorkbookParserInternalError(RuntimeError):
    """表示 parser 边界自身故障，必须记为 evaluator ERROR。"""


def build_searchwrite_observation(
    artifact_observation: PipelineImplicitArtifactObservation,
) -> SearchWriteObservation:
    """把已冻结的两工作簿闭集投影为正式 typed observation。

    输入参数：
        artifact_observation：经 manifest—nofollow—manifest 和逐文件
            size/SHA-256 双重校验的 production generic observation。
    输出返回值：
        只保留固定九个目标单元格和基线完整性结论的
        ``SearchWriteObservation``。
    异常：
        PipelineImplicitArtifactEvidenceError：任务、协议、依赖或
            XLSX 解析边界无效；异常仅含固定脱敏码。
    """

    if (
        not isinstance(
            artifact_observation,
            PipelineImplicitArtifactObservation,
        )
        or artifact_observation.task_id != SEARCHWRITE_XLSX_TASK_ID
        or artifact_observation.protocol_id != SEARCHWRITE_XLSX_PROTOCOL_ID
        or artifact_observation.complete is not True
    ):
        raise PipelineImplicitArtifactEvidenceError("TYPED_OBSERVATION_INVALID")
    workbooks: list[SearchWriteWorkbook] = []
    unexpected_file_count = 0
    for artifact_file in artifact_observation.iter_files_for_evaluator():
        contract = _WORKBOOK_CONTRACTS.get(artifact_file.relative_path)
        if contract is None:
            unexpected_file_count += 1
            workbooks.append(_failed_workbook(f"unexpected-{unexpected_file_count}"))
            continue
        document_id = contract.document_id
        coordinates = contract.target_coordinates
        expected_baseline_sha256 = contract.baseline_sha256
        payload = artifact_file.read_for_evaluator()
        try:
            _preflight_xlsx_bytes(payload)
        except (OSError, ValueError, zipfile.BadZipFile):
            workbooks.append(_failed_workbook(document_id))
            continue
        try:
            parsed_cells, baseline_unchanged = _parse_xlsx_controlled(
                payload,
                coordinates=coordinates,
                expected_baseline_sha256=expected_baseline_sha256,
            )
        except _WorkbookParseRejected:
            workbooks.append(_failed_workbook(document_id))
            continue
        except _WorkbookParserInternalError:
            raise PipelineImplicitArtifactEvidenceError(
                "TYPED_OBSERVATION_INVALID"
            ) from None
        workbooks.append(
            SearchWriteWorkbook(
                document_id=document_id,
                cells=tuple(
                    SearchWriteCell(coordinate, value)
                    for coordinate, value in parsed_cells
                ),
                baseline_unchanged=baseline_unchanged,
            )
        )
    return SearchWriteObservation(complete=True, workbooks=tuple(workbooks))


def _parse_xlsx_controlled(
    content: bytes,
    *,
    coordinates: tuple[str, ...],
    expected_baseline_sha256: str,
) -> tuple[tuple[tuple[str, object], ...], bool]:
    """在受控 spawn 子进程中解析 XLSX，并失败关闭越界结果。

    输入参数：
        content：已通过 ZIP/XML 流式预检的完整工作簿字节。
        coordinates：当前固定文档允许读取的目标坐标闭集。
        expected_baseline_sha256：固定 input 基线语义指纹。
    输出返回值：
        已再次校验的 ``(coordinate, scalar)`` tuple 和基线布尔结论。
    异常：
        _WorkbookParseRejected：输入解析被已知文档错误、5 秒
            wall-clock、CPU、Linux 地址空间硬限制或父进程 RSS
            软监控预算拒绝。
        _WorkbookParserInternalError：spawn/监控边界失效、未知
            parser 异常或子进程返回非法协议消息。
    """

    if (
        not isinstance(content, bytes)
        or not isinstance(coordinates, tuple)
        or not coordinates
        or not isinstance(expected_baseline_sha256, str)
        or len(expected_baseline_sha256) != 64
    ):
        raise _WorkbookParseRejected("PARSER_INPUT_INVALID")
    sandbox: Any | None = None
    try:
        sandbox = tempfile.TemporaryDirectory(prefix="paraguibench-searchwrite-parser-")
        os.chmod(sandbox.name, stat.S_IRUSR | stat.S_IXUSR)
    except OSError:
        if sandbox is not None:
            try:
                _cleanup_parser_sandbox(sandbox)
            except _WorkbookParserInternalError:
                pass
        raise _WorkbookParserInternalError("PARSER_SANDBOX_FAILED") from None
    try:
        return _run_parser_process(
            content,
            coordinates=coordinates,
            expected_baseline_sha256=expected_baseline_sha256,
            sandbox_cwd=sandbox.name,
        )
    finally:
        _cleanup_parser_sandbox(sandbox)


def _cleanup_parser_sandbox(sandbox: Any) -> None:
    """恢复并删除 parser 临时 cwd，清理故障使用固定内部错误。

    输入参数：
        sandbox：父进程创建的 ``TemporaryDirectory`` 对象。
    输出返回值：
        清理成功时无返回值。
    异常：
        _WorkbookParserInternalError：恢复权限或递归清理产生 OSError；
            不传播可能包含绝对路径的底层异常文本。
    """

    cleanup_failed = False
    try:
        os.chmod(sandbox.name, stat.S_IRWXU)
    except FileNotFoundError:
        pass
    except OSError:
        cleanup_failed = True
    try:
        sandbox.cleanup()
    except OSError:
        cleanup_failed = True
    if cleanup_failed:
        raise _WorkbookParserInternalError("PARSER_SANDBOX_FAILED") from None


def _run_parser_process(
    content: bytes,
    *,
    coordinates: tuple[str, ...],
    expected_baseline_sha256: str,
    sandbox_cwd: str,
) -> tuple[tuple[tuple[str, object], ...], bool]:
    """启动并监控一个已分配只读空工作目录的 parser 子进程。

    输入参数：
        content/coordinates/expected_baseline_sha256：已通过父层校验的
            工作簿字节、目标坐标闭集和基线指纹。
        sandbox_cwd：父进程新建且移除写权限的空目录绝对路径。
    输出返回值：
        经父进程再次验证的目标标量 tuple 与基线布尔结论。
    异常：
        _WorkbookParseRejected：资源门或已知输入错误拒绝解析。
        _WorkbookParserInternalError：spawn、监控或 IPC 边界故障。
    """

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    process = context.Process(
        target=_workbook_parse_worker,
        args=(
            sender,
            content,
            coordinates,
            expected_baseline_sha256,
            None,
            sandbox_cwd,
        ),
        daemon=True,
        name="paraguibench-searchwrite-parser",
    )
    try:
        process.start()
    except (OSError, RuntimeError):
        receiver.close()
        sender.close()
        raise _WorkbookParserInternalError("PARSER_START_FAILED") from None
    sender.close()
    raw_message: bytes | None = None
    rejected_by_resource_gate = False
    internal_monitor_error = False
    deadline = time.monotonic() + _PARSER_WALL_TIMEOUT_SECONDS
    try:
        while True:
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                rejected_by_resource_gate = True
                break
            if receiver.poll(min(_PARSER_POLL_INTERVAL_SECONDS, remaining)):
                try:
                    raw_message = _receive_parser_frame_after_final_rss_check(
                        receiver,
                        process,
                    )
                except _WorkbookParseRejected:
                    rejected_by_resource_gate = True
                except _WorkbookParserInternalError:
                    internal_monitor_error = True
                break
            if not process.is_alive():
                break
            resident_bytes = _parser_resident_bytes(process.pid)
            if resident_bytes is None:
                internal_monitor_error = True
                break
            if resident_bytes > _PARSER_RSS_LIMIT_BYTES:
                rejected_by_resource_gate = True
                break
    finally:
        receiver.close()
        process.join(timeout=0.5)
        if process.is_alive():
            _terminate_parser_process(process)
    if rejected_by_resource_gate:
        process.close()
        raise _WorkbookParseRejected("PARSER_RESOURCE_LIMIT")
    if internal_monitor_error:
        process.close()
        raise _WorkbookParserInternalError("PARSER_MONITOR_FAILED")
    if process.exitcode != 0:
        exitcode = process.exitcode
        process.close()
        if isinstance(exitcode, int) and exitcode < 0:
            raise _WorkbookParseRejected("PARSER_PROCESS_LIMITED")
        raise _WorkbookParserInternalError("PARSER_PROCESS_FAILED")
    process.close()
    message = _decode_parser_message(raw_message)
    return _validate_parser_message(message, coordinates=coordinates)


def _receive_parser_frame_after_final_rss_check(
    receiver: Any,
    process: Any,
) -> bytes | None:
    """在消费已就绪 frame 前完成最后一次父进程 RSS 采样。

    输入参数：
        receiver：已由 ``poll`` 报告 frame 就绪的单向 pipe 接收端。
        process：提供正整数 pid 的受控 parser 子进程。
    输出返回值：
        不超过 64 KiB 的原始 frame；pipe 在边界处关闭时返回 ``None``，
        后续固定协议验证会将其归类为 evaluator 内部错误。
    异常：
        _WorkbookParseRejected：最后一次采样超过 256 MiB 预算。
        _WorkbookParserInternalError：当前平台无法证明该次 RSS 状态。

    本检查补齐“frame 快速就绪即跳过监控”的窗口，但 macOS RSS
    仍是父进程轮询的软监控，不能替代 OS 级硬内存上限；Linux 另有
    ``RLIMIT_AS`` 硬限制。
    """

    resident_bytes = _parser_resident_bytes(process.pid)
    if resident_bytes is None:
        raise _WorkbookParserInternalError("PARSER_MONITOR_FAILED")
    if resident_bytes > _PARSER_RSS_LIMIT_BYTES:
        raise _WorkbookParseRejected("PARSER_RESOURCE_LIMIT")
    try:
        return receiver.recv_bytes(_PARSER_RESULT_MAX_BYTES)
    except (EOFError, OSError):
        return None


def _workbook_parse_worker(
    sender: Any,
    content: bytes,
    coordinates: tuple[str, ...],
    expected_baseline_sha256: str,
    materializer: Any | None = None,
    sandbox_cwd: str | None = None,
) -> None:
    """在资源限额子进程中物化工作簿并只返回受限标量。

    输入参数：
        sender：只向父进程发送一个小型结果的单向 pipe。
        content/coordinates/expected_baseline_sha256：与父进程受控入口相同。
        materializer：默认为生产 ``_materialize_workbook``；仅用于
            边界故障注入测试的可拾取顶层 callable。
        sandbox_cwd：父进程新建且移除写权限的空工作目录。
    输出返回值：
        无；成功发送固定 ``ok`` tuple，可预期解析错误只发送
        ``rejected``，不传播 exception 文本、路径或单元格值。
    """

    try:
        _silence_parser_worker_output()
        sys.dont_write_bytecode = True
        _install_parser_resource_limits()
        _enter_parser_python_write_boundary(sandbox_cwd)
        parser = _materialize_workbook if materializer is None else materializer
        if not callable(parser):
            raise TypeError("parser materializer 无效")
    except BaseException:
        try:
            _send_parser_message(sender, ("internal_error",))
        except (BrokenPipeError, EOFError, OSError):
            pass
        sender.close()
        return
    try:
        parsed_cells, baseline_unchanged = parser(
            content,
            coordinates=coordinates,
            expected_baseline_sha256=expected_baseline_sha256,
        )
    except (
        IndexError,
        KeyError,
        OSError,
        ValueError,
        zipfile.BadZipFile,
        ET.ParseError,
    ):
        try:
            _send_parser_message(sender, ("rejected",))
        except (BrokenPipeError, EOFError, OSError):
            pass
    except BaseException:
        try:
            _send_parser_message(sender, ("internal_error",))
        except (BrokenPipeError, EOFError, OSError):
            pass
    else:
        try:
            _send_parser_message(
                sender,
                ("ok", parsed_cells, baseline_unchanged),
            )
        except BaseException:
            try:
                _send_parser_message(sender, ("internal_error",))
            except (BrokenPipeError, EOFError, OSError):
                pass
    finally:
        sender.close()


def _send_parser_message(sender: Any, message: tuple[object, ...]) -> None:
    """以有界 UTF-8 JSON frame 发送 parser 结果，禁止 pickle。

    输入参数：
        sender：multiprocessing 单向字节 pipe 的发送端。
        message：只含固定状态、有界目标标量和 bool 的 tuple。
    输出返回值：
        无；紧凑 JSON 不超过 64 KiB 时通过 ``send_bytes`` 发送。
    异常：
        TypeError/ValueError：含非 JSON 类型、NaN/Infinity 或超长结果。
        OSError：pipe 无法完整写入。
    """

    encoded = json.dumps(
        message,
        ensure_ascii=False,
        allow_nan=False,
        separators=(",", ":"),
    ).encode("utf-8", errors="strict")
    if len(encoded) > _PARSER_RESULT_MAX_BYTES:
        raise ValueError("parser JSON 结果超限")
    sender.send_bytes(encoded)


def _decode_parser_message(content: bytes | None) -> object:
    """从有界 UTF-8 JSON 解码 parser 结果，不执行对象构造。

    输入参数：
        content：``recv_bytes`` 在 64 KiB 上限下收到的完整 frame。
    输出返回值：
        固定状态 tuple；``ok`` 的 cell 列表也转为 tuple，
        交给下一层做坐标和标量白名单验证。
    异常：
        _WorkbookParserInternalError：frame 缺失、超长、非 UTF-8、
            非严格 JSON 或顶层容器不是列表。
    """

    if not isinstance(content, bytes) or not (
        0 < len(content) <= _PARSER_RESULT_MAX_BYTES
    ):
        raise _WorkbookParserInternalError("PARSER_RESULT_INVALID")
    try:
        decoded = json.loads(
            content.decode("utf-8", errors="strict"),
            parse_constant=_reject_nonstandard_json_constant,
        )
    except (UnicodeDecodeError, ValueError):
        raise _WorkbookParserInternalError("PARSER_RESULT_INVALID") from None
    if not isinstance(decoded, list):
        raise _WorkbookParserInternalError("PARSER_RESULT_INVALID")
    if len(decoded) == 1:
        return (decoded[0],)
    if len(decoded) == 3 and isinstance(decoded[1], list):
        cells = tuple(
            tuple(entry) if isinstance(entry, list) else entry for entry in decoded[1]
        )
        return decoded[0], cells, decoded[2]
    return tuple(decoded)


def _reject_nonstandard_json_constant(value: str) -> object:
    """拒绝 JSON 解析器默认接受的 NaN/Infinity 扩展。

    输入参数：
        value：解析器遇到的非标准数值 token，不回显。
    输出返回值：不返回，始终抛出 ``ValueError``。
    """

    del value
    raise ValueError("parser JSON 数值非标准")


def _silence_parser_worker_output() -> None:
    """在不可信 parser 执行前丢弃子进程 stdout/stderr。

    输入参数：无。
    输出返回值：
        无；将文件描述符 1/2 指向 ``os.devnull``，因此未知
        Python/C-extension 异常也不能把 traceback、路径或值写入日志。
    异常：
        OSError：重定向无法完成时由 worker 的固定
            ``internal_error`` 通道处理。
    """

    sink = os.open(os.devnull, os.O_WRONLY)
    try:
        os.dup2(sink, 1)
        os.dup2(sink, 2)
    finally:
        os.close(sink)


def _install_parser_resource_limits() -> None:
    """在导入 openpyxl 前安装 OS 级 parser 资源上限。

    输入参数：无。
    输出返回值：
        无；关闭 core dump，CPU 硬限制 2 秒，Linux 地址空间
        硬限制 768 MiB，最多 64 个文件句柄，且输出文件大小
        硬限制为零。macOS 不设置不可靠的 RLIMIT_AS/DATA；
        其 256 MiB RSS 约束是父进程轮询与接收前复查的软监控，
        无法排除采样间的短暂峰值。
    异常：
        ImportError/OSError/ValueError：当前平台不能完整安装限额时
            失败关闭，不在无保护模式下继续解析。
    """

    import resource

    _set_resource_limit(resource, resource.RLIMIT_CORE, 0)
    _set_resource_limit(resource, resource.RLIMIT_CPU, _PARSER_CPU_SECONDS)
    if sys.platform != "darwin":
        _set_resource_limit(
            resource,
            resource.RLIMIT_AS,
            _PARSER_ADDRESS_SPACE_BYTES,
        )
    _set_resource_limit(
        resource,
        resource.RLIMIT_NOFILE,
        _PARSER_OPEN_FILE_LIMIT,
    )
    _set_resource_limit(resource, resource.RLIMIT_FSIZE, 0)


def _enter_parser_python_write_boundary(sandbox_cwd: str | None) -> None:
    """进入只读空 cwd，并安装不可移除的 Python audit 写入门。

    输入参数：
        sandbox_cwd：父进程创建、确认为空并移除写权限的绝对目录。
    输出返回值：
        无；完成 ``chdir`` 后安装进程级 audit hook，拒绝 Python
        API 发起的写模式 open、截断、删除、重命名、建删目录、
        改权限、链接、再次切换 cwd、拉起子进程和 ``ctypes`` 绕行。
    异常：
        OSError/TypeError/ValueError：目录身份、状态或 audit hook
        安装失败；worker 以固定 ``internal_error`` 失败关闭。

    该边界只承诺受信任 Python parser 的 Python 层写隔离；它不是
    针对任意本地机器码的 OS 沙箱。CPU、文件大小和句柄在
    macOS/Linux 均由 RLIMIT 硬限制，Linux 地址空间另有
    RLIMIT_AS；macOS 内存只有父进程 RSS 软监控。
    """

    if (
        not isinstance(sandbox_cwd, str)
        or not sandbox_cwd
        or not os.path.isabs(sandbox_cwd)
        or os.path.islink(sandbox_cwd)
        or not os.path.isdir(sandbox_cwd)
        or os.listdir(sandbox_cwd)
    ):
        raise ValueError("parser 只读 cwd 无效")
    mode = stat.S_IMODE(os.stat(sandbox_cwd, follow_symlinks=False).st_mode)
    if mode & (stat.S_IWUSR | stat.S_IWGRP | stat.S_IWOTH):
        raise ValueError("parser cwd 仍可写")
    os.chdir(sandbox_cwd)
    sys.addaudithook(_deny_parser_python_write)
    _install_parser_unaudited_write_guards()


def _deny_parser_python_write(event: str, arguments: tuple[object, ...]) -> None:
    """拒绝 parser 在 Python audit 层发起文件系统变更。

    输入参数：
        event：CPython audit 事件名称。
        arguments：事件参数；只检查 ``open`` 的模式/flags，不回显路径。
    输出返回值：
        只读事件正常返回；任何已列入的变更事件抛出固定
        ``PermissionError``，不包含文件路径、单元格或 parser 文本。
    """

    if event == "open":
        mode = arguments[1] if len(arguments) > 1 else None
        flags = arguments[2] if len(arguments) > 2 else None
        mode_writes = isinstance(mode, str) and any(marker in mode for marker in "wax+")
        flags_write = (
            isinstance(flags, int)
            and not isinstance(flags, bool)
            and bool(flags & _PARSER_WRITE_OPEN_FLAGS)
        )
        if mode_writes or flags_write:
            raise PermissionError("PARSER_PYTHON_WRITE_DENIED")
        return
    if event in _PARSER_DENIED_AUDIT_EVENTS:
        raise PermissionError("PARSER_PYTHON_WRITE_DENIED")


def _install_parser_unaudited_write_guards() -> None:
    """禁用 CPython 当前未可靠发出 audit event 的路径创建 API。

    输入参数：无。
    输出返回值：
        无；在 ``os`` 与其 Unix 底层 ``posix`` 模块中，将可用的
        ``mkfifo``/``mknod`` 替换为固定拒绝函数。这样受信任 parser
        经常规模块属性调用时不能在只读 cwd 外创建特殊文件。

    该补丁不撤销安装前已被任意第三方代码保存的 C 函数引用，故仍
    属于受信任 Python parser 合同，不构成任意代码的 OS 沙箱。
    """

    import posix

    for module in (os, posix):
        for function_name in ("mkfifo", "mknod"):
            if hasattr(module, function_name):
                setattr(
                    module,
                    function_name,
                    _reject_parser_unaudited_write,
                )


def _reject_parser_unaudited_write(
    *arguments: object,
    **keywords: object,
) -> None:
    """以固定脱敏错误拒绝未受 audit 覆盖的文件系统变更。

    输入参数：
        arguments/keywords：被禁用 API 的任意调用参数；全部丢弃，
            不格式化或回显其中路径。
    输出返回值：
        不返回，始终抛出固定 ``PermissionError``。
    """

    del arguments, keywords
    raise PermissionError("PARSER_PYTHON_WRITE_DENIED")


def _set_resource_limit(resource_module: Any, kind: int, requested: int) -> None:
    """将单个软/硬资源限额收紧到当前硬上限内。

    输入参数：
        resource_module：子进程本地导入的 ``resource`` 模块。
        kind：RLIMIT 类型。
        requested：协议要求的非负软/硬限额。
    输出返回值：无；成功后软硬限额均不高于请求值。
    """

    _, hard = resource_module.getrlimit(kind)
    target = requested
    if hard != resource_module.RLIM_INFINITY:
        target = min(target, hard)
    if target < 0:
        raise ValueError("parser 资源限额无效")
    resource_module.setrlimit(kind, (target, target))


def _materialize_workbook(
    content: bytes,
    *,
    coordinates: tuple[str, ...],
    expected_baseline_sha256: str,
) -> tuple[tuple[tuple[str, object], ...], bool]:
    """在子进程中使用 openpyxl 物化并投影一个工作簿。

    输入参数：
        content：已预检 XLSX 字节。
        coordinates：允许投影的目标坐标。
        expected_baseline_sha256：预期基线指纹。
    输出返回值：不含路径的目标标量与基线布尔结论。
    """

    import openpyxl

    stream = BytesIO(content)
    workbook = openpyxl.load_workbook(
        stream,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        _validate_workbook_resource_bounds(workbook)
        if len(workbook.worksheets) != 1:
            raise ValueError("工作簿 sheet 数不匹配")
        worksheet = workbook.worksheets[0]
        cells: list[tuple[str, object]] = []
        for coordinate in coordinates:
            value = worksheet[coordinate].value
            if value is None:
                continue
            _validate_target_scalar(value)
            cells.append((coordinate, value))
        baseline_unchanged = (
            _baseline_semantic_sha256(
                workbook,
                target_coordinates=frozenset(coordinates),
            )
            == expected_baseline_sha256
        )
        return tuple(cells), baseline_unchanged
    finally:
        workbook.close()
        stream.close()


def _validate_parser_message(
    message: object,
    *,
    coordinates: tuple[str, ...],
) -> tuple[tuple[tuple[str, object], ...], bool]:
    """验证不可信 parser 子进程返回的小型投影。

    输入参数：
        message：从单向 pipe 反序列化的候选对象。
        coordinates：固定允许坐标闭集。
    输出返回值：唯一坐标的有界标量 tuple 与 bool。
    异常：
        _WorkbookParseRejected：子进程显式返回已知输入拒绝。
        _WorkbookParserInternalError：状态、容器、坐标、值或
            基线类型不符合父子进程固定协议。
    """

    if message == ("internal_error",):
        raise _WorkbookParserInternalError("PARSER_INTERNAL_ERROR")
    if message == ("rejected",):
        raise _WorkbookParseRejected("PARSER_INPUT_REJECTED")
    if (
        not isinstance(message, tuple)
        or len(message) != 3
        or message[0] != "ok"
        or not isinstance(message[1], tuple)
        or type(message[2]) is not bool
        or len(message[1]) > len(coordinates)
    ):
        raise _WorkbookParserInternalError("PARSER_RESULT_INVALID")
    allowed = set(coordinates)
    observed: set[str] = set()
    cells: list[tuple[str, object]] = []
    for entry in message[1]:
        if (
            not isinstance(entry, tuple)
            or len(entry) != 2
            or not isinstance(entry[0], str)
            or entry[0] not in allowed
            or entry[0] in observed
        ):
            raise _WorkbookParserInternalError("PARSER_RESULT_INVALID")
        try:
            _validate_target_scalar(entry[1])
        except ValueError:
            raise _WorkbookParserInternalError("PARSER_RESULT_INVALID") from None
        observed.add(entry[0])
        cells.append((entry[0], entry[1]))
    return tuple(cells), message[2]


def _parser_resident_bytes(pid: int | None) -> int | None:
    """读取 parser 子进程当前常驻内存，不读取命令行。

    输入参数：
        pid：由 multiprocessing 创建的正整数子进程 ID。
    输出返回值：
        Linux 从 ``/proc/<pid>/statm``、macOS 从 ``libproc``
        的 task info 返回 RSS 字节数；进程已退出时返回 0，
        监控边界无法证明时返回 ``None`` 使父进程失败关闭。
    """

    if not isinstance(pid, int) or isinstance(pid, bool) or pid <= 0:
        return None
    if sys.platform.startswith("linux"):
        try:
            with open(f"/proc/{pid}/statm", encoding="ascii") as stream:
                fields = stream.read(128).split()
            if len(fields) < 2:
                return None
            resident_pages = int(fields[1])
            page_size = os.sysconf("SC_PAGE_SIZE")
            if resident_pages < 0 or not isinstance(page_size, int):
                return None
            return resident_pages * page_size
        except FileNotFoundError:
            return 0
        except (OSError, ValueError):
            return None
    if sys.platform == "darwin":
        import ctypes

        class ProcTaskInfo(ctypes.Structure):
            """镜像 macOS ``proc_taskinfo`` 中的固定宽度字段。"""

            _fields_ = [
                ("virtual_size", ctypes.c_uint64),
                ("resident_size", ctypes.c_uint64),
                ("total_user", ctypes.c_uint64),
                ("total_system", ctypes.c_uint64),
                ("threads_user", ctypes.c_uint64),
                ("threads_system", ctypes.c_uint64),
                ("policy", ctypes.c_int32),
                ("faults", ctypes.c_int32),
                ("pageins", ctypes.c_int32),
                ("cow_faults", ctypes.c_int32),
                ("messages_sent", ctypes.c_int32),
                ("messages_received", ctypes.c_int32),
                ("syscalls_mach", ctypes.c_int32),
                ("syscalls_unix", ctypes.c_int32),
                ("context_switches", ctypes.c_int32),
                ("thread_count", ctypes.c_int32),
                ("running_thread_count", ctypes.c_int32),
                ("priority", ctypes.c_int32),
            ]

        try:
            libproc = ctypes.CDLL(
                "/usr/lib/libproc.dylib",
                use_errno=True,
            )
            proc_pidinfo = libproc.proc_pidinfo
            proc_pidinfo.argtypes = [
                ctypes.c_int,
                ctypes.c_int,
                ctypes.c_uint64,
                ctypes.c_void_p,
                ctypes.c_int,
            ]
            proc_pidinfo.restype = ctypes.c_int
            task_info = ProcTaskInfo()
            result = proc_pidinfo(
                pid,
                4,
                0,
                ctypes.byref(task_info),
                ctypes.sizeof(task_info),
            )
        except (AttributeError, OSError):
            return None
        if result == ctypes.sizeof(task_info):
            return int(task_info.resident_size)
        try:
            os.kill(pid, 0)
        except ProcessLookupError:
            return 0
        except OSError:
            return None
        return None
    return None


def _terminate_parser_process(process: Any) -> None:
    """终止超时或未干净退出的 parser 子进程。

    输入参数：
        process：已成功 ``start`` 的 multiprocessing Process。
    输出返回值：无；先 terminate，0.5 秒后仍存活则 kill。
    """

    process.terminate()
    process.join(timeout=0.5)
    if process.is_alive():
        process.kill()
        process.join(timeout=0.5)


def _failed_workbook(document_id: str) -> SearchWriteWorkbook:
    """为存在但无法安全解析的期望文件生成零分占位。

    输入参数：
        document_id：由精确固定文件名映射的逻辑文档身份。
    输出返回值：
        无目标单元格且基线失效的工作簿观测；评价器
        因此仍会将该文件的固定单元格全部计入分母。
    """

    return SearchWriteWorkbook(
        document_id=document_id,
        cells=(),
        baseline_unchanged=False,
    )


def _preflight_xlsx_bytes(content: bytes) -> None:
    """在 Office parser 导入内容前校验 XLSX ZIP 的被动资源门。

    输入参数：
        content：已由 production capture 校验外层 size/SHA-256
            且不超过单文件上限的原始 XLSX 字节。
    输出返回值：
        无；必要成员、路径、便携唯一性、文件类型、加密、
        宏、单项/总展开大小、压缩比、XML 主动实体、
        流式元素/深度/工作表结构计数及外部关系均通过时
        正常返回。资源预算以固定四个真实工作簿的
        10 成员、28 个 cell XML 和不超过 427 个总元素为基线收紧。
    异常：
        ValueError/zipfile.BadZipFile：任一闭集或安全约束失败。
        上层将该文件转为脱敏零分占位，不回显成员身份。
    """

    if not isinstance(content, bytes) or not content:
        raise ValueError("XLSX 字节无效")
    with zipfile.ZipFile(BytesIO(content), mode="r") as archive:
        infos = archive.infolist()
        if not infos or len(infos) > _MAX_ARCHIVE_MEMBERS:
            raise ValueError("XLSX 成员数超限")
        members: dict[str, zipfile.ZipInfo] = {}
        portable_names: set[tuple[str, ...]] = set()
        expanded_bytes = 0
        xml_members: list[zipfile.ZipInfo] = []
        for member in infos:
            normalized_name = (
                member.filename[:-1]
                if member.filename.endswith("/")
                else member.filename
            )
            path = PurePosixPath(normalized_name)
            if (
                not normalized_name
                or member.filename.startswith("/")
                or "\\" in member.filename
                or "\x00" in member.filename
                or path.is_absolute()
                or any(part in {"", ".", ".."} for part in path.parts)
                or (path.parts and ":" in path.parts[0])
                or any(
                    not unicodedata.is_normalized("NFC", part) for part in path.parts
                )
            ):
                raise ValueError("XLSX 成员路径无效")
            portable_name = tuple(
                unicodedata.normalize("NFC", part).casefold() for part in path.parts
            )
            if portable_name in portable_names:
                raise ValueError("XLSX 成员身份碰撞")
            portable_names.add(portable_name)
            mode = (member.external_attr >> 16) & 0o170000
            if (
                member.flag_bits & 0x1
                or member.file_size < 0
                or member.compress_size < 0
                or member.file_size > _MAX_ARCHIVE_MEMBER_BYTES
                or (mode and mode not in {stat.S_IFREG, stat.S_IFDIR})
                or normalized_name.casefold().endswith("vbaproject.bin")
                or normalized_name.casefold().startswith("xl/externallinks/")
            ):
                raise ValueError("XLSX 成员 metadata 无效")
            if member.file_size > 0 and (
                member.compress_size <= 0
                or member.file_size / member.compress_size
                > _MAX_ARCHIVE_COMPRESSION_RATIO
            ):
                raise ValueError("XLSX 成员压缩比超限")
            expanded_bytes += member.file_size
            if expanded_bytes > _MAX_ARCHIVE_EXPANDED_BYTES:
                raise ValueError("XLSX 展开大小超限")
            members[normalized_name] = member
            if normalized_name.lower().endswith((".xml", ".rels")):
                xml_members.append(member)
        if not _REQUIRED_XLSX_MEMBERS.issubset(members):
            raise ValueError("XLSX 必要成员缺失")
        xml_counts = {
            "elements": 0,
            "worksheet_rows": 0,
            "worksheet_cells": 0,
            "worksheet_columns": 0,
            "worksheet_merges": 0,
            "relationships": 0,
        }
        for member in xml_members:
            payload = _read_bounded_zip_member(
                archive,
                member,
                max_bytes=_MAX_XML_MEMBER_BYTES,
            )
            lowered = payload.replace(b"\x00", b"").lower()
            if b"<!doctype" in lowered or b"<!entity" in lowered:
                raise ValueError("XLSX XML 含主动实体声明")
            _validate_streaming_xml_structure(
                payload,
                member_name=member.filename,
                counts=xml_counts,
            )
            if member.filename.lower().endswith(".rels"):
                _validate_relationship_xml(payload)


def _validate_streaming_xml_structure(
    content: bytes,
    *,
    member_name: str,
    counts: dict[str, int],
) -> None:
    """在 Office 物化前流式计数 OOXML 结构。

    输入参数：
        content：已通过单成员字节上限与 DTD/entity 门的 XML。
        member_name：经 ZIP 便携身份门确认的内部成员名，
            只用于区分 worksheet 和 relationship 预算。
        counts：当前工作簿所有 XML 成员共享的聚合计数器。
    输出返回值：
        无；通过 ``iterparse`` 在 ``end`` 事件立即清理元素，
        不构建整棵 XML DOM。
    异常：
        ValueError：XML 不可解释，或总元素、嵌套深度、worksheet
            row/cell/column/merge 或 relationship 计数超出固定预算。
    """

    lowered_name = member_name.casefold()
    is_worksheet = lowered_name.startswith("xl/worksheets/") and lowered_name.endswith(
        ".xml"
    )
    is_relationship = lowered_name.endswith(".rels")
    depth = 0
    try:
        events = ET.iterparse(BytesIO(content), events=("start", "end"))
        for event, element in events:
            if event == "start":
                depth += 1
                if depth > _MAX_XML_DEPTH:
                    raise ValueError("XLSX XML 嵌套超限")
                counts["elements"] += 1
                if counts["elements"] > _MAX_XML_ELEMENTS:
                    raise ValueError("XLSX XML 元素数超限")
                local_name = element.tag.rsplit("}", 1)[-1]
                if is_worksheet:
                    counter_name: str | None = None
                    limit = 0
                    if local_name == "row":
                        counter_name = "worksheet_rows"
                        limit = _MAX_WORKSHEET_XML_ROWS
                    elif local_name == "c":
                        counter_name = "worksheet_cells"
                        limit = _MAX_WORKSHEET_XML_CELLS
                    elif local_name == "col":
                        counter_name = "worksheet_columns"
                        limit = _MAX_WORKSHEET_XML_COLUMNS
                    elif local_name == "mergeCell":
                        counter_name = "worksheet_merges"
                        limit = _MAX_WORKSHEET_XML_MERGES
                    if counter_name is not None:
                        counts[counter_name] += 1
                        if counts[counter_name] > limit:
                            raise ValueError("XLSX worksheet 结构数超限")
                if is_relationship and local_name == "Relationship":
                    counts["relationships"] += 1
                    if counts["relationships"] > _MAX_RELATIONSHIP_ELEMENTS:
                        raise ValueError("XLSX relationship 数超限")
            else:
                element.clear()
                depth -= 1
    except ET.ParseError:
        raise ValueError("XLSX XML 无效") from None
    if depth != 0:
        raise ValueError("XLSX XML 嵌套无效")


def _read_bounded_zip_member(
    archive: zipfile.ZipFile,
    member: zipfile.ZipInfo,
    *,
    max_bytes: int,
) -> bytes:
    """以声明大小和实际流双重上限读取一个 XML 成员。

    输入参数：
        archive/member：已完成 central-directory 元数据门的 ZIP
            及当前 XML/关系成员。
        max_bytes：当前成员允许的最大展开字节数。
    输出返回值：
        声明长度与实际流长度一致的完整字节。
    异常：
        ValueError：成员超限、截断或含超额尾部。
    """

    if member.file_size > max_bytes:
        raise ValueError("XLSX XML 成员超限")
    with archive.open(member, mode="r") as stream:
        content = stream.read(max_bytes + 1)
        tail = stream.read(1)
    if len(content) > max_bytes or len(content) != member.file_size or tail != b"":
        raise ValueError("XLSX XML 成员读取不完整")
    return content


def _validate_relationship_xml(content: bytes) -> None:
    """拒绝 SearchWrite 协议不需要的 OOXML 外部关系。

    输入参数：
        content：已受 XML 单成员大小上限约束，且已拒绝
            DTD/entity 声明的 ``.rels`` 字节。
    输出返回值：
        无；所有 Relationship 均为内部 target 时正常返回。
    异常：
        ValueError：XML 无效、声明 external TargetMode，或 target
            呈现绝对网络/邮件 URI 身份。
    """

    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        raise ValueError("XLSX relationship XML 无效") from None
    for element in root.iter():
        if element.tag.rsplit("}", 1)[-1] != "Relationship":
            continue
        target_mode = element.attrib.get("TargetMode")
        target = element.attrib.get("Target")
        if target_mode is not None and target_mode.casefold() != "internal":
            raise ValueError("XLSX 外部 relationship 被拒绝")
        if isinstance(target, str) and (
            "://" in target
            or target.startswith("//")
            or target.casefold().startswith(("mailto:", "file:"))
        ):
            raise ValueError("XLSX 外部 relationship 被拒绝")


def _validate_workbook_resource_bounds(workbook: Any) -> None:
    """在语义投影前给工作簿结构和有效单元格设定上限。

    输入参数：
        workbook：已通过 ZIP/XML 预检并由 openpyxl 被动解析
            的工作簿。
    输出返回值：
        无；sheet 数、行/列/矩形单元格、非空单元格、公式、
        字符串字节、合并区和行列维度记录均在固定预算内时返回。
    异常：
        ValueError：任一资源边界超限或形状不可解释。

    本函数在任何 ``iter_rows`` 之前先验证矩形形状，
    避免恶意 dimension 把一个小 OOXML 扩张为无界迭代。
    """

    worksheets = workbook.worksheets
    if not isinstance(worksheets, list) or not (
        1 <= len(worksheets) <= _MAX_WORKBOOK_SHEETS
    ):
        raise ValueError("工作簿 sheet 数超限")
    total_rectangular_cells = 0
    populated_cells = 0
    formula_count = 0
    text_bytes = 0
    for worksheet in worksheets:
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        if (
            type(max_row) is not int
            or type(max_column) is not int
            or not 1 <= max_row <= _MAX_WORKSHEET_ROWS
            or not 1 <= max_column <= _MAX_WORKSHEET_COLUMNS
            or max_row * max_column > _MAX_WORKSHEET_CELLS
            or len(worksheet.merged_cells.ranges) > _MAX_MERGED_RANGES
            or len(worksheet.row_dimensions) > _MAX_DIMENSION_ENTRIES
            or len(worksheet.column_dimensions) > _MAX_DIMENSION_ENTRIES
        ):
            raise ValueError("工作表形状超限")
        total_rectangular_cells += max_row * max_column
        if total_rectangular_cells > _MAX_WORKBOOK_CELLS:
            raise ValueError("工作簿单元格矩形超限")
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_column,
        ):
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                populated_cells += 1
                if populated_cells > _MAX_POPULATED_CELLS:
                    raise ValueError("工作簿非空单元格超限")
                if cell.data_type == "f":
                    formula_count += 1
                    if formula_count > _MAX_FORMULAS:
                        raise ValueError("工作簿公式数超限")
                if isinstance(value, str):
                    try:
                        encoded = value.encode("utf-8", errors="strict")
                    except UnicodeEncodeError:
                        raise ValueError("单元格字符串无效") from None
                    if len(encoded) > _MAX_CELL_TEXT_BYTES:
                        raise ValueError("单元格字符串超限")
                    text_bytes += len(encoded)
                    if text_bytes > _MAX_WORKBOOK_TEXT_BYTES:
                        raise ValueError("工作簿字符串总量超限")


def _validate_target_scalar(value: object) -> None:
    """验证目标单元格可被正式九格协议安全消费。

    输入参数：
        value：固定目标坐标的 openpyxl 值。
    输出返回值：
        无；类型为非 bool 整数、有限浮点数或有界字符串
        时正常返回。
    异常：
        ValueError：值类型、有限性或 UTF-8 长度不符合协议。
    """

    if isinstance(value, bool) or not isinstance(value, (str, int, float)):
        raise ValueError("目标单元格类型无效")
    if isinstance(value, float) and not math.isfinite(value):
        raise ValueError("目标单元格非有限")
    if isinstance(value, str):
        try:
            encoded = value.encode("utf-8", errors="strict")
        except UnicodeEncodeError:
            raise ValueError("目标单元格字符串无效") from None
        if len(encoded) > _MAX_CELL_TEXT_BYTES:
            raise ValueError("目标单元格字符串超限")


def _baseline_semantic_sha256(
    workbook: Any,
    *,
    target_coordinates: frozenset[str],
) -> str:
    """生成排除目标填空格后的工作簿语义指纹。

    输入参数：
        workbook：已经 OOXML 资源门校验并由 openpyxl
            被动解析的工作簿。
        target_coordinates：首工作表中允许 Agent 填入的固定
            坐标闭集。
    输出返回值：
        对 sheet 顺序/名称/状态/维度/合并区、freeze pane、
        默认行列尺寸、显式行高/列宽/隐藏/折叠结构、所有有效格
        的可见样式、图片/图表/表格/批注等对象库存，以及
        非目标单元格值进行规范化后的 SHA-256。只排除九个目标
        坐标的内容与 data type；目标格样式和其余可见结构仍固定。
        行高按 1 pt、列宽按 0.1 字符宽度做半入规范化，
        用于合并固定 input 与 LibreOffice gold 中已验证的
        16.5/17 pt 及二进制列宽序列化差异。
        Office ZIP 序列化元数据不参与指纹。
    异常：
        ValueError：工作簿包含未固定的单元格标量类型。
    """

    sheets: list[dict[str, object]] = []
    for sheet_index, worksheet in enumerate(workbook.worksheets):
        cells: list[list[object]] = []
        comment_count = 0
        hyperlink_count = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    comment_count += 1
                if cell.hyperlink is not None:
                    hyperlink_count += 1
                is_target = sheet_index == 0 and cell.coordinate in target_coordinates
                if is_target:
                    cells.append(
                        [
                            cell.coordinate,
                            "target-content-excluded",
                            _normalize_cell_visible_style(cell),
                        ]
                    )
                    continue
                if cell.value is None:
                    if cell.has_style:
                        cells.append(
                            [
                                cell.coordinate,
                                "empty",
                                _normalize_cell_visible_style(cell),
                            ]
                        )
                    continue
                cells.append(
                    [
                        cell.coordinate,
                        cell.data_type,
                        _normalize_baseline_scalar(cell.value),
                        _normalize_cell_visible_style(cell),
                    ]
                )
        row_dimensions = [
            [
                str(index),
                _normalize_visible_length(dimension.height, quantum=1.0),
                bool(dimension.hidden),
                int(dimension.outlineLevel or 0),
                bool(dimension.collapsed),
                bool(dimension.thickTop),
                bool(dimension.thickBot),
            ]
            for index, dimension in sorted(worksheet.row_dimensions.items())
            if not _is_default_row_dimension(
                dimension,
                default_height=worksheet.sheet_format.defaultRowHeight,
            )
        ]
        column_dimensions = [
            [
                str(index),
                int(dimension.min or 0),
                int(dimension.max or 0),
                _normalize_visible_length(dimension.width, quantum=0.1),
                bool(dimension.hidden),
                bool(dimension.bestFit),
                int(dimension.outlineLevel or 0),
                bool(dimension.collapsed),
            ]
            for index, dimension in sorted(worksheet.column_dimensions.items())
        ]
        freeze_panes = worksheet.freeze_panes
        if hasattr(freeze_panes, "coordinate"):
            freeze_panes = freeze_panes.coordinate
        if freeze_panes is not None and not isinstance(freeze_panes, str):
            raise ValueError("freeze pane 类型未固定")
        sheets.append(
            {
                "title": worksheet.title,
                "state": worksheet.sheet_state,
                "dimension": worksheet.calculate_dimension(),
                "merged_cells": sorted(
                    str(cell_range) for cell_range in worksheet.merged_cells.ranges
                ),
                "freeze_panes": freeze_panes,
                "sheet_format": {
                    "base_column_width": (
                        int(worksheet.sheet_format.baseColWidth or 0)
                        if worksheet.sheet_format.defaultColWidth is None
                        else None
                    ),
                    "default_column_width": _normalize_visible_length(
                        worksheet.sheet_format.defaultColWidth,
                        quantum=0.1,
                    ),
                    "default_row_height": _normalize_visible_length(
                        worksheet.sheet_format.defaultRowHeight,
                        quantum=1.0,
                    ),
                    "zero_height": bool(worksheet.sheet_format.zeroHeight),
                    "thick_top": bool(worksheet.sheet_format.thickTop),
                    "thick_bottom": bool(worksheet.sheet_format.thickBottom),
                },
                "row_dimensions": row_dimensions,
                "column_dimensions": column_dimensions,
                "visible_objects": _normalize_visible_object_inventory(
                    worksheet,
                    comment_count=comment_count,
                    hyperlink_count=hyperlink_count,
                ),
                "cells": cells,
            }
        )
    encoded = json.dumps(
        {
            "version": 6,
            "theme": _normalize_workbook_theme(workbook),
            "sheets": sheets,
        },
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8", errors="strict")
    return hashlib.sha256(encoded).hexdigest()


def _normalize_workbook_theme(workbook: Any) -> list[str]:
    """把固定 input theme 与 LibreOffice 缺省 theme 合并为同一语义。

    输入参数：
        workbook：已由受控 openpyxl parser 物化的工作簿。
    输出返回值：
        真实 input 的已核验 theme SHA-256 和 gold 的缺省 theme 均返回
        ``["verified-default"]``；其他 theme 只返回脱敏 SHA-256 身份。
    异常：
        ValueError：``loaded_theme`` 不是 ``None`` 或不可变字节。

    该字段约束 theme-1 等颜色编号的实际语义，防止修改 theme 后仍被
    单元格颜色规范化误判为固定黑色；原始 theme XML 不进入结果或日志。
    """

    theme = workbook.loaded_theme
    if theme is None:
        return ["verified-default"]
    if not isinstance(theme, bytes):
        raise ValueError("工作簿 theme 类型无效")
    digest = hashlib.sha256(theme).hexdigest()
    if digest == _VERIFIED_DEFAULT_THEME_SHA256:
        return ["verified-default"]
    return ["custom-sha256", digest]


def _normalize_visible_object_inventory(
    worksheet: Any,
    *,
    comment_count: int,
    hyperlink_count: int,
) -> dict[str, object]:
    """把工作表非单元格可见对象纳入固定基线闭集。

    输入参数：
        worksheet：已受资源门约束的 openpyxl 工作表。
        comment_count：有批注单元格数，由受限矩形迭代得到。
        hyperlink_count：有超链接单元格数，由同一矩形迭代得到。
    输出返回值：
        图片、图表、表格、批注、超链接、数据验证、条件格式数量，
        以及自动筛选范围的稳定字典。固定真实基线均为零对象，
        因此不读取或持久化对象内可能含私密内容的表示。
    异常：
        ValueError：计数或自动筛选范围类型不符合闭集。
    """

    counts = {
        "images": len(worksheet._images),
        "charts": len(worksheet._charts),
        "tables": len(worksheet.tables),
        "comments": comment_count,
        "hyperlinks": hyperlink_count,
        "data_validations": len(worksheet.data_validations.dataValidation),
        "conditional_formatting": len(worksheet.conditional_formatting),
    }
    if any(type(value) is not int or value < 0 for value in counts.values()):
        raise ValueError("工作表可见对象计数无效")
    auto_filter_reference = worksheet.auto_filter.ref
    if auto_filter_reference is not None and not isinstance(auto_filter_reference, str):
        raise ValueError("工作表自动筛选范围无效")
    return counts | {"auto_filter": auto_filter_reference}


def _normalize_cell_visible_style(cell: Any) -> dict[str, object]:
    """把单元格可见样式规范化为固定 JSON 闭集。

    输入参数：
        cell：已由受控 openpyxl parser 物化的 Cell/MergedCell。
    输出返回值：
        字体、填充、边框、对齐、数字格式、保护和 quote/pivot
        标志的稳定字典。规范化只合并经真实 input/gold 验证的
        Office 默认等价项，不使用内部 style id 作为语义证据。
    异常：
        ValueError：样式字段包含未固定类型、非有限数值或未知填充。
    """

    font = cell.font
    alignment = cell.alignment
    protection = cell.protection
    font_name = _optional_style_string(font.name)
    font_size = _normalize_style_number(font.sz)
    font_family = _normalize_style_number(font.family)
    font_color = _normalize_style_color(font.color)
    return {
        "font": {
            "name": font_name,
            "size": font_size,
            "family": font_family,
            "bold": bool(font.bold),
            "italic": bool(font.italic),
            "strike": bool(font.strike),
            "outline": bool(font.outline),
            "shadow": bool(font.shadow),
            "condense": bool(font.condense),
            "extend": bool(font.extend),
            "underline": _optional_style_string(font.underline),
            "vertical_alignment": _optional_style_string(font.vertAlign),
            "color": font_color,
            "serialization_variant": _normalize_font_serialization_variant(
                font,
                normalized_name=font_name,
                normalized_size=font_size,
                normalized_family=font_family,
                normalized_color=font_color,
            ),
        },
        "fill": _normalize_cell_fill(cell.fill),
        "border": _normalize_cell_border(cell.border),
        "alignment": {
            "horizontal": alignment.horizontal or "general",
            "vertical": alignment.vertical or "bottom",
            "rotation": int(alignment.textRotation or 0),
            "wrap_text": bool(alignment.wrapText),
            "shrink_to_fit": bool(alignment.shrinkToFit),
            "indent": _normalize_style_number(alignment.indent or 0),
            "relative_indent": _normalize_style_number(alignment.relativeIndent or 0),
            "justify_last_line": bool(alignment.justifyLastLine),
            "reading_order": _normalize_style_number(alignment.readingOrder or 0),
        },
        "number_format": _optional_style_string(cell.number_format),
        "protection": {
            "locked": protection.locked is not False,
            "hidden": bool(protection.hidden),
        },
        "quote_prefix": bool(cell.quotePrefix),
        "pivot_button": bool(cell.pivotButton),
    }


def _normalize_font_serialization_variant(
    font: Any,
    *,
    normalized_name: str | None,
    normalized_size: str,
    normalized_family: str,
    normalized_color: list[object],
) -> list[object]:
    """只合并真实 input/gold 已证明等价的 CJK 字体 metadata。

    输入参数：
        font：openpyxl 字体对象。
        normalized_name/size/family/color：同一字体已验证的规范化字段。
    输出返回值：
        固定 11pt 宋体黑色的 ``(charset=None, scheme=minor)`` 与
        LibreOffice ``(charset=1, scheme=None)`` 返回共同标记；其余
        组合保留显式 charset/scheme，避免兼容规则被扩大成绕过。
    异常：
        ValueError：charset 或 scheme 类型不在固定闭集。
    """

    charset = font.charset
    if charset is not None and (type(charset) is not int or charset < 0):
        raise ValueError("单元格字体 charset 无效")
    scheme = font.scheme
    if scheme is not None and not isinstance(scheme, str):
        raise ValueError("单元格字体 scheme 无效")
    if (
        normalized_name == "宋体"
        and normalized_size == float(11).hex()
        and normalized_family == float(2).hex()
        and normalized_color == ["opaque-black"]
        and (charset, scheme) in {(None, "minor"), (1, None)}
    ):
        return ["verified-cjk-default"]
    return ["explicit", charset, scheme]


def _normalize_cell_fill(fill: Any) -> list[object]:
    """规范化 PatternFill，并显式拒绝未冻结的渐变样式。

    输入参数：
        fill：openpyxl 单元格填充对象。
    输出返回值：
        无填充、solid 前景色或其他 pattern 前/背景色的固定列表；
        solid 的背景色不会渲染，故不参与语义指纹。
    异常：
        ValueError：出现当前任务基线未声明的 GradientFill/未知对象。
    """

    fill_type = getattr(fill, "fill_type", None)
    is_pattern_fill = hasattr(fill, "fgColor") and hasattr(fill, "bgColor")
    if fill_type is None:
        if not is_pattern_fill:
            raise ValueError("单元格填充类型未固定")
        return ["none"]
    if not isinstance(fill_type, str):
        raise ValueError("单元格填充类型无效")
    if not is_pattern_fill:
        raise ValueError("单元格渐变填充未固定")
    foreground = _normalize_style_color(fill.fgColor)
    if fill_type == "solid":
        return ["solid", foreground]
    return [
        "pattern",
        fill_type,
        foreground,
        _normalize_style_color(fill.bgColor),
    ]


def _normalize_cell_border(border: Any) -> dict[str, object]:
    """规范化单元格边框的可见线型和颜色。

    输入参数：
        border：openpyxl ``Border`` 对象。
    输出返回值：
        八个方向边线以及 diagonal/outline 标志的稳定字典；无边线
        时忽略无效颜色，缺省色与 ``auto`` 统一为自动颜色。
    """

    return {
        side_name: _normalize_border_side(getattr(border, side_name))
        for side_name in (
            "start",
            "end",
            "left",
            "right",
            "top",
            "bottom",
            "diagonal",
            "vertical",
            "horizontal",
        )
    } | {
        "diagonal_up": bool(border.diagonalUp),
        "diagonal_down": bool(border.diagonalDown),
        "outline": border.outline is not False,
    }


def _normalize_border_side(side: Any | None) -> list[object]:
    """把一个可选 Border Side 投影为线型和规范化颜色。

    输入参数：
        side：openpyxl ``Side`` 或 ``None``。
    输出返回值：
        无边线统一为 ``[None, None]``；有边线返回类型和颜色。
    """

    if side is None or side.style is None:
        return [None, None]
    if not isinstance(side.style, str):
        raise ValueError("单元格边框线型无效")
    return [side.style, _normalize_style_color(side.color)]


def _normalize_style_color(color: Any | None) -> list[object]:
    """规范化 Office 样式颜色，并合并已验证的默认黑色表示。

    输入参数：
        color：openpyxl ``Color`` 或缺省值。
    输出返回值：
        自动色、ARGB、indexed 或 theme+tint 的固定列表。真实 input
        的 theme-1 黑与 LibreOffice gold 的不透明 RGB 黑归一化。
    异常：
        ValueError：颜色类型、索引、tint 或 RGB 值不可解释。
    """

    if color is None or getattr(color, "auto", False) is True:
        return ["automatic"]
    color_type = getattr(color, "type", None)
    tint = _normalize_style_number(getattr(color, "tint", 0) or 0)
    if color_type == "rgb":
        rgb = getattr(color, "rgb", None)
        if not isinstance(rgb, str):
            raise ValueError("单元格 RGB 颜色无效")
        normalized_rgb = rgb.upper()
        if len(normalized_rgb) == 6:
            normalized_rgb = "FF" + normalized_rgb
        if len(normalized_rgb) != 8 or any(
            character not in "0123456789ABCDEF" for character in normalized_rgb
        ):
            raise ValueError("单元格 RGB 颜色无效")
        if normalized_rgb == "FF000000" and tint == "0x0.0p+0":
            return ["opaque-black"]
        return ["argb", normalized_rgb, tint]
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        if type(theme) is not int or theme < 0:
            raise ValueError("单元格 theme 颜色无效")
        if theme == 1 and tint == "0x0.0p+0":
            return ["opaque-black"]
        return ["theme", theme, tint]
    if color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        if type(indexed) is not int or indexed < 0:
            raise ValueError("单元格 indexed 颜色无效")
        if indexed == 64:
            return ["automatic"]
        return ["indexed", indexed, tint]
    raise ValueError("单元格颜色类型未固定")


def _normalize_style_number(value: object) -> str:
    """把样式有限数值规范化为稳定十六进制浮点字符串。

    输入参数：
        value：openpyxl 样式字段中的 int/float。
    输出返回值：
        对应有限 ``float`` 的 ``hex`` 表示。
    异常：
        ValueError：布尔、非数值或非有限值。
    """

    if (
        isinstance(value, bool)
        or not isinstance(value, (int, float))
        or not math.isfinite(float(value))
    ):
        raise ValueError("单元格样式数值无效")
    return float(value).hex()


def _optional_style_string(value: object) -> str | None:
    """验证一个可选样式字段为字符串。

    输入参数：
        value：字体名称、下划线、垂直对齐或数字格式候选值。
    输出返回值：
        ``None`` 或原字符串。
    异常：
        ValueError：非空值不是字符串。
    """

    if value is None or isinstance(value, str):
        return value
    raise ValueError("单元格样式字符串无效")


def _normalize_visible_length(
    value: object,
    *,
    quantum: float,
) -> str | None:
    """将 Office 可见尺寸按固定半入粒度规范化。

    输入参数：
        value：可选行高或列宽。
        quantum：任务协议固定的最小可见粒度。
    输出返回值：
        ``None`` 或按粒度折算的稳定整数桶字符串。
    异常：
        ValueError：值、粒度类型非法或不是有限正数。
    """

    if value is None:
        return None
    if (
        isinstance(value, bool)
        or not isinstance(value, (int, float))
        or isinstance(quantum, bool)
        or not isinstance(quantum, (int, float))
    ):
        raise ValueError("工作表可见尺寸类型无效")
    normalized = float(value)
    normalized_quantum = float(quantum)
    if (
        not math.isfinite(normalized)
        or not math.isfinite(normalized_quantum)
        or normalized_quantum <= 0
    ):
        raise ValueError("工作表可见尺寸非法")
    bucket = math.floor(normalized / normalized_quantum + 0.5)
    return str(bucket)


def _is_default_row_dimension(
    dimension: Any,
    *,
    default_height: object,
) -> bool:
    """判断显式行维度是否仅重申工作表默认值。

    输入参数：
        dimension：openpyxl ``RowDimension``。
        default_height：当前 sheet 的默认行高。
    输出返回值：
        行高与默认值处于同一 1 pt 桶，且没有
        隐藏、折叠、加粗边或自定义样式时返回 ``True``。
    """

    return (
        _normalize_visible_length(dimension.height, quantum=1.0)
        == _normalize_visible_length(default_height, quantum=1.0)
        and not dimension.hidden
        and not dimension.outlineLevel
        and not dimension.collapsed
        and not dimension.thickTop
        and not dimension.thickBot
        and not dimension.has_style
    )


def _normalize_baseline_scalar(value: object) -> list[object]:
    """把非目标单元格值规范化为稳定 JSON 标量。

    输入参数：
        value：openpyxl 返回的候选单元格值。
    输出返回值：
        含显式类型标记的 ``[kind, value]``；整数和浮点数
        分别用十进制字符串和 ``float.hex`` 避免序列化漂移。
    异常：
        ValueError：值不是 bool/int/有限 float/str。
    """

    if type(value) is bool:
        return ["bool", value]
    if type(value) is int:
        return ["int", str(value)]
    if type(value) is float and math.isfinite(value):
        return ["float", value.hex()]
    if type(value) is str:
        return ["str", value]
    raise ValueError("基线单元格标量类型未固定")


__all__ = ["build_searchwrite_observation"]

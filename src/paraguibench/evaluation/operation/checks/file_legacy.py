"""从最终旧源机械提取的 Operation 检查最小传递闭包。"""

from __future__ import annotations
import glob
import html
import logging
import os
import re
import unicodedata
from html.parser import HTMLParser
from typing import Any, Dict, List

logger = logging.getLogger("eval.operation_checks.file")


class _HTMLTextAndTagCollector(HTMLParser):
    """提取 HTML 的标签与可见文本。

    输入:
        HTMLParser.feed 接收的 HTML 字符串。
    输出:
        tags: 小写标签名集合。
        text_parts: 可见文本片段列表。
    """

    def __init__(self) -> None:
        """初始化 HTML 标签与可见文本收集器。

        输入参数：无。
        输出返回值：无；初始化空的 ``tags`` 与 ``text_parts`` 容器。
        """
        super().__init__(convert_charrefs=True)
        self.tags = set()
        self.text_parts: List[str] = []

    def handle_starttag(self, tag: str, attrs: list) -> None:
        """记录开始标签。

        输入:
            tag: HTML 标签名。
            attrs: 标签属性列表（本检查无需使用）。
        输出:
            无；结果写入 tags。
        """
        del attrs
        self.tags.add(tag.lower())

    def handle_data(self, data: str) -> None:
        """记录可见文本片段。

        输入:
            data: 标签之间的文本。
        输出:
            无；非空文本写入 text_parts。
        """
        if data.strip():
            self.text_parts.append(data)


def _ok(reason: str = "通过") -> Dict[str, Any]:
    """功能：执行传递闭包内的 _ok 检查步骤。

    输入参数：
        由函数签名定义；调用值仅在 evaluator 内存中使用。
    输出返回值：
        由函数签名定义；公开聚合层会删除路径、内容、gold 与原始 reason。"""
    return {"pass": True, "score": 1.0, "reason": reason}


def _fail(reason: str) -> Dict[str, Any]:
    """功能：执行传递闭包内的 _fail 检查步骤。

    输入参数：
        由函数签名定义；调用值仅在 evaluator 内存中使用。
    输出返回值：
        由函数签名定义；公开聚合层会删除路径、内容、gold 与原始 reason。"""
    return {"pass": False, "score": 0.0, "reason": reason}


def _partial(score: float, reason: str) -> Dict[str, Any]:
    """功能：执行传递闭包内的 _partial 检查步骤。

    输入参数：
        由函数签名定义；调用值仅在 evaluator 内存中使用。
    输出返回值：
        由函数签名定义；公开聚合层会删除路径、内容、gold 与原始 reason。"""
    return {"pass": score >= 1.0 - 1e-09, "score": round(score, 4), "reason": reason}


def _config_error(reason: str) -> Dict[str, Any]:
    """功能：执行传递闭包内的 _config_error 检查步骤。

    输入参数：
        由函数签名定义；调用值仅在 evaluator 内存中使用。
    输出返回值：
        由函数签名定义；公开聚合层会删除路径、内容、gold 与原始 reason。

    旧源语义说明：
    评价器配置错误（缺参数等）：score=-1 哨兵，由上层冒泡为 evaluator_error。"""
    return {"pass": False, "score": -1.0, "status": "evaluator_error", "reason": reason}


def check_files_in_same_folder(result_dir: str, params: dict) -> dict:
    """
    检查指定文件是否在同一个子文件夹中。

    输入:
        result_dir: 搜索的根目录
        params:
            file_groups (list[list[str]]): 文件分组列表，每组内的文件应在同一文件夹
                例如: [["file1.xlsx", "file2.xlsx"], ["file3.pptx", "file4.pptx"]]
            require_subfolder (bool): 是否禁止将分组文件全部留在结果根目录，
                默认 True。
            require_distinct_group_folders (bool): 不同有效分组是否必须位于
                不同子目录，默认 False。
            min_distinct_folders (int): 所有通过分组最少使用的不同子目录数，
                默认 1。
    输出:
        {"pass": bool, "score": float, "reason": str}

    说明:
        - 组内文件数 < 2 时跳过该组（单文件组无法评估"同一文件夹"语义，
          否则 len(unique_dirs)==1 恒 true 会误判通过）
        - 组内任一文件缺失 → 该组计为 fail（不再仅对已找到的文件比较目录）
        - 若所有组都被跳过（全部为空组或单文件组）→ 返回 _fail
    """
    file_groups = params.get("file_groups", [])
    require_subfolder = params.get("require_subfolder", True)
    require_distinct = params.get("require_distinct_group_folders", False)
    min_distinct_folders = int(params.get("min_distinct_folders", 1))
    if not file_groups:
        return _config_error("参数缺少 file_groups")
    effective_groups = 0
    passed_groups = 0
    details: List[str] = []
    skipped_details: List[str] = []
    passed_folder_paths: List[str] = []
    root_dir = os.path.realpath(result_dir)
    for i, group in enumerate(file_groups):
        if not group:
            skipped_details.append(f"组{i + 1}: 空组，已跳过")
            continue
        if len(group) < 2:
            skipped_details.append(
                f"组{i + 1}: 单文件组({group[0]})，无法评估同一文件夹语义，已跳过"
            )
            continue
        file_dirs: List[str] = []
        missing: List[str] = []
        duplicates: List[str] = []
        for filename in group:
            pattern = os.path.join(result_dir, "**", filename)
            matches = sorted(
                (
                    path
                    for path in glob.glob(pattern, recursive=True)
                    if os.path.isfile(path)
                )
            )
            if len(matches) == 1:
                file_dirs.append(os.path.realpath(os.path.dirname(matches[0])))
            elif len(matches) > 1:
                duplicates.append(filename)
            else:
                missing.append(filename)
        effective_groups += 1
        if missing:
            details.append(f"组{i + 1}: 缺失文件 {', '.join(missing[:3])}")
            continue
        if duplicates:
            details.append(f"组{i + 1}: 文件出现多份副本 {', '.join(duplicates[:3])}")
            continue
        unique_dirs = set(file_dirs)
        if len(unique_dirs) == 1:
            common_dir = next(iter(unique_dirs))
            if require_subfolder and common_dir == root_dir:
                details.append(f"组{i + 1}: 文件仍在结果根目录，未分类到子目录")
                continue
            passed_groups += 1
            passed_folder_paths.append(common_dir)
            folder_name = os.path.basename(common_dir) or common_dir
            details.append(f"组{i + 1}: {len(group)} 个文件全部在 {folder_name}")
        else:
            folder_names = [os.path.basename(d) or d for d in unique_dirs]
            details.append(
                f"组{i + 1}: 文件分散在 {len(unique_dirs)} 个文件夹: {', '.join(folder_names[:3])}"
            )
    if effective_groups == 0:
        reason = "无有效分组可评估"
        if skipped_details:
            reason += "：" + "; ".join(skipped_details[:5])
        return _fail(reason)
    distinct_folder_count = len(set(passed_folder_paths))
    folder_separation_ok = distinct_folder_count >= min_distinct_folders
    if require_distinct and len(passed_folder_paths) != distinct_folder_count:
        folder_separation_ok = False
        details.append("不同主题分组被放入了同一子目录")
    if distinct_folder_count < min_distinct_folders:
        details.append(
            f"仅使用 {distinct_folder_count} 个子目录，期望至少 {min_distinct_folders} 个"
        )
    ratio = passed_groups / effective_groups
    if not folder_separation_ok:
        ratio = min(ratio, distinct_folder_count / max(min_distinct_folders, 1))
    combined = details + skipped_details
    if passed_groups == effective_groups and folder_separation_ok:
        ok_reason = f"{effective_groups} 组文件全部在同一文件夹"
        if skipped_details:
            ok_reason += (
                f"（跳过 {len(skipped_details)} 组: {'; '.join(skipped_details[:3])}）"
            )
        return _ok(ok_reason)
    return _partial(
        ratio,
        f"{passed_groups}/{effective_groups} 组在同一文件夹: {'; '.join(combined[:5])}",
    )


def check_html_files_for_xlsx(result_dir: str, params: dict) -> dict:
    """
    检查是否存在与 xlsx 文件名相同的 html 文件。

    输入:
        result_dir: 搜索的根目录
        params:
            search_subdirs (bool): 是否搜索子目录，默认 True
            validate_content (bool): 是否校验 HTML 表格内容来自对应 xlsx，
                默认 True。
            min_match_ratio (float): xlsx 抽样单元格值在 HTML 中的最小
                命中比例，默认 0.8。
            sample_cell_limit (int): 每个工作簿最多抽样的非空值数，
                默认 24。
    输出:
        {"pass": bool, "score": float, "reason": str}
    """
    search_subdirs = params.get("search_subdirs", True)
    validate_content = params.get("validate_content", True)
    min_match_ratio = float(params.get("min_match_ratio", 0.8))
    sample_cell_limit = int(params.get("sample_cell_limit", 24))
    if search_subdirs:
        xlsx_pattern = os.path.join(result_dir, "**", "*.xlsx")
        xlsx_files = glob.glob(xlsx_pattern, recursive=True)
    else:
        xlsx_pattern = os.path.join(result_dir, "*.xlsx")
        xlsx_files = glob.glob(xlsx_pattern)
    if not xlsx_files:
        return _fail("未找到 xlsx 文件")
    total = len(xlsx_files)
    matched = 0
    details = []
    for xlsx_path in xlsx_files:
        xlsx_name = os.path.basename(xlsx_path)
        xlsx_base = os.path.splitext(xlsx_name)[0]
        html_name = xlsx_base + ".html"
        html_path = os.path.join(os.path.dirname(xlsx_path), html_name)
        html_matches = [html_path] if os.path.isfile(html_path) else []
        if html_matches and (
            not validate_content
            or _html_contains_xlsx_content(
                xlsx_path, html_matches[0], min_match_ratio, sample_cell_limit
            )
        ):
            matched += 1
            details.append(f"{xlsx_name} -> {html_name}")
        elif html_matches:
            details.append(f"{xlsx_name} -> {html_name} 内容与 xlsx 不匹配")
        else:
            details.append(f"{xlsx_name} -> 缺失")
    ratio = matched / total
    if matched == total:
        return _ok(f"全部 {total} 个 xlsx 都有对应的 html 文件")
    return _partial(
        ratio, f"{matched}/{total} 个 xlsx 有对应 html: {', '.join(details[:5])}"
    )


def _normalize_exported_text(value: Any) -> str:
    """归一化 xlsx 单元格值与 HTML 导出文本。

    输入:
        value: 单元格值或 HTML 文本。
    输出:
        小写、去空白与常见数字分组符后的比较字符串。
    """
    text = unicodedata.normalize("NFKC", html.unescape(str(value))).lower()
    text = re.sub("(?<=\\d)[,\\u00a0](?=\\d{3}\\b)", "", text)
    return re.sub("\\s+", "", text)


def _html_contains_xlsx_content(
    xlsx_path: str, html_path: str, min_match_ratio: float, sample_cell_limit: int
) -> bool:
    """检查 HTML 是否是对应 xlsx 的实质表格导出。

    输入:
        xlsx_path: 源工作簿路径。
        html_path: 待检查 HTML 路径。
        min_match_ratio: 抽样值的最小命中比例。
        sample_cell_limit: 最多抽样的非空值数。
    输出:
        HTML 包含 table 且达到内容命中阈值时返回 True。
    """
    try:
        from openpyxl import load_workbook

        with open(html_path, "r", encoding="utf-8", errors="ignore") as stream:
            exported_html = stream.read()
        parser = _HTMLTextAndTagCollector()
        parser.feed(exported_html)
        if "table" not in parser.tags:
            return False
        exported_text = _normalize_exported_text(" ".join(parser.text_parts))
        workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
        try:
            samples: List[str] = []
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(values_only=True):
                    for value in row:
                        if value is None:
                            continue
                        normalized = _normalize_exported_text(value)
                        if normalized and normalized not in samples:
                            samples.append(normalized)
                        if len(samples) >= sample_cell_limit:
                            break
                    if len(samples) >= sample_cell_limit:
                        break
                if len(samples) >= sample_cell_limit:
                    break
        finally:
            workbook.close()
    except Exception:
        logger.warning("Operation 检查内部事件")
        return False
    if not samples:
        return False
    matched = sum((1 for sample in samples if sample in exported_text))
    return matched / len(samples) >= min_match_ratio


FILE_LEGACY_CHECKS = {
    "check_files_in_same_folder": check_files_in_same_folder,
    "check_html_files_for_xlsx": check_html_files_for_xlsx,
}

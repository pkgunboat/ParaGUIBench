"""OSWorld artifact runtime evidence source 的公共纵向契约测试。"""

from __future__ import annotations

from contextlib import contextmanager
from dataclasses import replace
import hashlib
from io import BytesIO
import json
import sys
from types import ModuleType

from PIL import Image
import pytest

from paraguibench.runtime.osworld_artifact_evidence import (
    OSWorldArtifactEvidenceError,
    OSWorldArtifactEvidenceSource,
)
from paraguibench.runtime.osworld_environment import (
    OSWORLD_ARTIFACT_RUNTIME_FINALIZE_TASK_IDS,
)
from paraguibench.integrations.osworld.artifact_evidence_specs import (
    OSWORLD_ARTIFACT_EVIDENCE_SPECS,
    canonical_artifact_evidence_spec_json,
)
from paraguibench.integrations.osworld.controller import (
    OSWorldGuestPathMissingError,
)


_TASK_ID = "Operation-FileOperate-BatchOperation-001"
_BIBTEX_TASK_ID = "Operation-FileOperate-CombinationDocs-015"
_FIRST_SHEET_TASK_ID = "Operation-FileOperate-CombinationDocs-010"
_NAMED_SHEET_TASK_ID = "Operation-FileOperate-SearchAndWrite-009"
_PDF_DIRECTORY_TASK_ID = "Operation-FileOperate-CombinationDocs-011"
_FINALIZED_FIRST_SHEET_TASK_ID = "Operation-FileOperate-SearchAndWrite-001"
_FIRST_SHEET_GOLD_KEY = (
    "osworld-gold:aceb0368-56b8-4073-b70e-3dc9aee184e0:expected:0:v1"
)
_BIBTEX_GOLD_KEY = "osworld-gold:df67aebb-fb3a-44fd-b75b-51b6012df509:expected:0:v1"


def test_runtime_finalized_family_task_can_capture_with_wired_capability() -> None:
    """验证完成 runtime finalizer 接线的任务不再被旧 source 门禁拒绝。

    输入参数：
        无；使用 runtime 导出的精确能力集，以及等价的合成 actual/gold XLSX。
    输出返回值：
        无；环境已负责在本 source 前执行 finalizer，证据源随后应正常捕获并评分。
    """

    assert _FINALIZED_FIRST_SHEET_TASK_ID in OSWORLD_ARTIFACT_RUNTIME_FINALIZE_TASK_IDS
    workbook = _xlsx_bytes("PRIVATE_FINALIZED_FIRST_SHEET_VALUE")
    resolver = _GoldResolver(workbook)
    controller = _SingleFileController(workbook)

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=resolver,
    ).capture(
        _FINALIZED_FIRST_SHEET_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert slot.status == "available"
    assert tuple(metric.score for metric in slot.metric_scores) == (1.0,)
    assert controller.collect_calls[0]["guest_path"] == (
        "/srv/paraguibench-test/user/Desktop/Professor_Contact.xlsx"
    )
    assert "PRIVATE_FINALIZED_FIRST_SHEET_VALUE" not in repr(observation)


class _ImageDirectoryController:
    """模拟真实 controller 边界，不替换内部 metric 或证据逻辑。"""

    def __init__(self, records: tuple[tuple[str, str], ...]) -> None:
        """保存 guest helper 返回的图片像素摘要记录及调用记录。

        输入参数：
            records：按 guest 观察顺序排列的 ``(pixel_sha256, name)``。
        输出返回值：
            无；实例将在窄证据 getter 返回该摘要。
        """

        self._records = records
        self.collect_calls: list[dict[str, object]] = []

    def get_desktop_path(self) -> str:
        """返回用于动态推导 guest home 的可信 Desktop 路径。

        输入参数：
            无。
        输出返回值：
            固定测试 guest 的 Desktop 绝对路径。
        """

        return "/home/user/Desktop"

    def collect_image_pixel_hashes(
        self,
        guest_directory: str,
        *,
        max_entries: int,
        max_name_bytes: int,
        max_compressed_item_bytes: int,
        max_total_compressed_bytes: int,
        max_pixels_per_image: int,
        max_decoded_item_bytes: int,
        max_total_decoded_bytes: int,
        max_response_bytes: int,
        timeout_seconds: float,
    ) -> tuple[tuple[str, str], ...]:
        """模拟单次 nofollow/Pillow guest helper 的安全输出。

        输入参数：
            guest_directory：证据源根据 spec 推导的目录。
            其余关键字参数：固定的压缩、解码、成员及响应资源上限。
        输出返回值：
            按 guest 观察顺序排列的 ``(pixel_sha256, name)`` tuple。
        """

        self.collect_calls.append(
            {
                "guest_directory": guest_directory,
                "max_entries": max_entries,
                "max_name_bytes": max_name_bytes,
                "max_compressed_item_bytes": max_compressed_item_bytes,
                "max_total_compressed_bytes": max_total_compressed_bytes,
                "max_pixels_per_image": max_pixels_per_image,
                "max_decoded_item_bytes": max_decoded_item_bytes,
                "max_total_decoded_bytes": max_total_decoded_bytes,
                "max_response_bytes": max_response_bytes,
                "timeout_seconds": timeout_seconds,
            }
        )
        return self._records


class _SingleFileController:
    """模拟 production 单文件 getter，记录脱敏前的受信调用参数。"""

    def __init__(
        self,
        content: bytes = b"",
        *,
        error: Exception | None = None,
        events: list[str] | None = None,
    ) -> None:
        """保存待返回 actual 字节或受控异常。

        输入参数：
            content：getter 成功时返回的原始字节。
            error：非 ``None`` 时由 getter 原样抛出的模拟边界异常。
            events：可选的跨 resolver/controller 顺序记录列表。
        输出返回值：
            无；初始化空调用记录。
        """

        self._content = content
        self._error = error
        self._events = events
        self.collect_calls: list[dict[str, object]] = []

    def collect_file_bytes(
        self,
        guest_path: str,
        *,
        max_bytes: int,
        max_response_bytes: int,
        timeout_seconds: float,
    ) -> bytes:
        """模拟 controller 的 nofollow、regular-only 单文件读取。

        输入参数：
            guest_path：source 由冻结 guest home 与 locator 推导的绝对路径。
            max_bytes/max_response_bytes/timeout_seconds：版本化资源边界。
        输出返回值：
            构造时保存的原始字节；配置异常时不返回。
        """

        self.collect_calls.append(
            {
                "guest_path": guest_path,
                "max_bytes": max_bytes,
                "max_response_bytes": max_response_bytes,
                "timeout_seconds": timeout_seconds,
            }
        )
        if self._events is not None:
            self._events.append("guest-file")
        if self._error is not None:
            raise self._error
        return self._content


class _GoldResolver:
    """只在可信 host 内存暴露 gold 流的 evaluator 边界 fake。"""

    def __init__(
        self,
        content: bytes,
        *,
        error: Exception | None = None,
        events: list[str] | None = None,
    ) -> None:
        """保存 gold 字节或固定失败，并初始化调用记录。

        输入参数：
            content：成功时由 context manager 交付的只读字节流内容。
            error：非 ``None`` 时模拟未 provision、完整性或媒体错误。
            events：可选的跨 resolver/controller 顺序记录列表。
        输出返回值：
            无。
        """

        self._content = content
        self._error = error
        self._events = events
        self.open_calls: list[dict[str, object]] = []

    @contextmanager
    def open_verified(
        self,
        logical_key: str,
        *,
        max_bytes: int,
        expected_media_types: frozenset[str],
    ):
        """模拟已完成 fd/nofollow/大小/摘要/媒体校验的 gold lease。

        输入参数：
            logical_key：spec 固定的 evaluator-only gold 身份。
            max_bytes：source 允许交给纯 metric 的最大字节数。
            expected_media_types：该 artifact family 的媒体 allowlist。
        输出返回值：
            context manager 内交付 seekable ``BytesIO``；失败时不交付。
        """

        self.open_calls.append(
            {
                "logical_key": logical_key,
                "max_bytes": max_bytes,
                "expected_media_types": expected_media_types,
            }
        )
        if self._events is not None:
            self._events.append("gold-open")
        if self._error is not None:
            raise self._error
        with BytesIO(self._content) as stream:
            yield stream


class _PDFDirectoryController(_SingleFileController):
    """模拟同一 guest 中的 PDF getter 与目录成员 getter。"""

    def __init__(self, content: bytes, *, events: list[str]) -> None:
        """保存 PDF 字节、共享顺序记录并初始化目录调用。

        输入参数：
            content：单文件 getter 返回的受控 PDF 字节。
            events：gold/file/directory 边界共享的顺序列表。
        输出返回值：
            无；初始化父类和空目录调用记录。
        """

        super().__init__(content, events=events)
        self.directory_calls: list[dict[str, object]] = []

    def list_directory(
        self,
        guest_path: str,
        *,
        max_entries: int,
        max_name_bytes: int,
        max_response_bytes: int,
    ) -> tuple[str, ...]:
        """返回只含目标发票的排序目录成员闭集。

        输入参数：
            guest_path：spec 推导的 guest 目录绝对路径。
            max_entries/max_name_bytes/max_response_bytes：版本化的资源边界。
        输出返回值：
            含一个固定目标文件名的 tuple。
        """

        self.directory_calls.append(
            {
                "guest_path": guest_path,
                "max_entries": max_entries,
                "max_name_bytes": max_name_bytes,
                "max_response_bytes": max_response_bytes,
            }
        )
        if self._events is not None:
            self._events.append("guest-directory")
        return ("Invoice # 243729.pdf",)


class _FakePdfPage:
    """提供受控页面文本的 pypdf 页面替身。"""

    def __init__(self, text: str) -> None:
        """保存后续 ``extract_text`` 返回的文本。

        输入参数：
            text：受控 PDF 页面文本。
        输出返回值：
            无。
        """

        self._text = text

    def extract_text(self) -> str:
        """返回构造时保存的 PDF 页面文本。

        输入参数：
            无。
        输出返回值：
            受控文本字符串。
        """

        return self._text


class _FakePdfReader:
    """从受控 PDF marker 解析单页文本的 pypdf 替身。"""

    def __init__(self, stream: BytesIO, *, strict: bool = True) -> None:
        """解析 ``TEXT`` 段并暴露未加密单页集合。

        输入参数：
            stream：含受控 PDF 字节的内存流。
            strict：兼容 production pypdf 严格解析参数。
        输出返回值：
            无；设置 ``is_encrypted`` 和 ``pages``。
        """

        del strict
        content = stream.read()
        prefix = b"\nTEXT:"
        suffix = b"\n%%EOF"
        start = content.index(prefix) + len(prefix)
        end = content.index(suffix, start)
        self.is_encrypted = False
        self.pages = (_FakePdfPage(content[start:end].decode("utf-8")),)


def _install_fake_pypdf(monkeypatch: pytest.MonkeyPatch) -> None:
    """将受控 pypdf 模块注入当前测试进程。

    输入参数：
        monkeypatch：pytest 提供的模块表替换器。
    输出返回值：
        无；``sys.modules['pypdf']`` 暂时指向 fake reader。
    """

    module = ModuleType("pypdf")
    module.PdfReader = _FakePdfReader
    monkeypatch.setitem(sys.modules, "pypdf", module)


def _pdf_bytes(*, text: str, marker: str) -> bytes:
    """构造含可提取文本与独立 marker 的受控 PDF 字节。

    输入参数：
        text：评价应提取的页面文本。
        marker：不属于评价语义的私有文件标记。
    输出返回值：
        符合 projection 边界最小签名要求的 UTF-8 PDF bytes。
    """

    return f"%PDF-1.7\n%{marker}\nTEXT:{text}\n%%EOF\n".encode("utf-8")


def _png_and_pixel_digest() -> tuple[bytes, str]:
    """构造文件字节哈希与像素字节哈希不同的最小 PNG。

    输入参数：
        无。
    输出返回值：
        PNG 编码字节，以及旧 OSWorld ``Image.tobytes()`` 语义对应的
        SHA-256。
    """

    image = Image.new("RGB", (2, 1), color=(17, 29, 43))
    pixel_digest = hashlib.sha256(image.tobytes()).hexdigest()
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    encoded = buffer.getvalue()
    assert hashlib.sha256(encoded).hexdigest() != pixel_digest
    return encoded, pixel_digest


def _xlsx_bytes(cell_value: str) -> bytes:
    """构造 first-sheet runtime tracer 使用的最小 XLSX。

    输入参数：
        cell_value：写入活动工作表 A1 的测试值。
    输出返回值：
        由 openpyxl 保存到内存的 OOXML 字节。
    """

    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active["A1"] = cell_value
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _named_xlsx_bytes(*, noise: str) -> bytes:
    """构造含噪声首页和 ``unseen_movies`` 目标页的 XLSX。

    输入参数：
        noise：写入首页的私有哨兵文本。
    输出返回值：
        仅目标页语义固定的内存 OOXML 字节。
    """

    from openpyxl import Workbook

    workbook = Workbook()
    first = workbook.active
    first.title = "noise"
    first.append(("value",))
    first.append((noise,))
    target = workbook.create_sheet("unseen_movies")
    target.append(("title",))
    target.append(("Arrival",))
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_first_sheet_family_capture_projects_verified_gold_before_guest_io() -> None:
    """验证新 13 项首个 first-sheet 任务贯通真实 production adapter。

    输入参数：
        无；使用等价的合成 actual/gold XLSX 与带调用记录的窄边界。
    输出返回值：
        无；gold 必须先被验证，随后只捕获一次 guest 文件并返回脱敏满分。
    """

    actual = _xlsx_bytes("PRIVATE_FIRST_SHEET_VALUE")
    resolver = _GoldResolver(actual)
    controller = _SingleFileController(actual)

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=resolver,
    ).capture(
        _FIRST_SHEET_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    assert len(observation.artifact_slots) == 1
    slot = observation.artifact_slots[0]
    assert slot.status == "available"
    assert tuple(score.score for score in slot.metric_scores) == (1.0,)
    assert resolver.open_calls == [
        {
            "logical_key": _FIRST_SHEET_GOLD_KEY,
            "max_bytes": controller.collect_calls[0]["max_bytes"],
            "expected_media_types": frozenset(
                {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
            ),
        }
    ]
    assert controller.collect_calls[0]["guest_path"] == (
        "/srv/paraguibench-test/user/exam/grades.xlsx"
    )
    assert controller.collect_calls[0]["max_bytes"] == 12_579_840
    for private_value in (
        "PRIVATE_FIRST_SHEET_VALUE",
        "grades.xlsx",
        "/srv/",
    ):
        assert private_value not in repr(observation)


def test_named_sheet_family_capture_ignores_noise_and_uses_verified_gold() -> None:
    """验证 named-sheet 任务经 production source 只评价固定目标页。

    输入参数：
        无；actual/gold 首页噪声不同，``unseen_movies`` 页相同。
    输出返回值：
        无；断言受控 gold 媒体、guest locator、满分与脱敏 repr。
    """

    actual_noise = "PRIVATE_ACTUAL_NOISE"
    gold_noise = "PRIVATE_GOLD_NOISE"
    actual = _named_xlsx_bytes(noise=actual_noise)
    gold = _named_xlsx_bytes(noise=gold_noise)
    resolver = _GoldResolver(gold)
    controller = _SingleFileController(actual)

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=resolver,
    ).capture(
        _NAMED_SHEET_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert slot.status == "available"
    assert tuple(metric.score for metric in slot.metric_scores) == (1.0,)
    assert resolver.open_calls[0]["expected_media_types"] == frozenset(
        {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )
    assert controller.collect_calls[0]["guest_path"] == (
        "/srv/paraguibench-test/user/Desktop/movies.xlsx"
    )
    for private_value in (
        actual_noise,
        gold_noise,
        "movies.xlsx",
        "/srv/",
    ):
        assert private_value not in repr(observation)


def test_pdf_and_directory_family_capture_scores_after_gold_preflight(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 PDF external gold 与目录 inline metric 在同一任务中闭环。

    输入参数：
        monkeypatch：pytest 提供的受控 pypdf 模块替换器。
    输出返回值：
        无；gold 必须先于任何 guest getter 验证，随后 PDF 与目录
        两个槽位均以脱敏分数通过。
    """

    _install_fake_pypdf(monkeypatch)
    private_text = "PRIVATE_INVOICE_TEXT"
    actual_marker = "PRIVATE_ACTUAL_PDF_MARKER"
    gold_marker = "PRIVATE_GOLD_PDF_MARKER"
    actual = _pdf_bytes(text=private_text, marker=actual_marker)
    gold = _pdf_bytes(text=private_text, marker=gold_marker)
    events: list[str] = []
    resolver = _GoldResolver(gold, events=events)
    controller = _PDFDirectoryController(actual, events=events)

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=resolver,
    ).capture(
        _PDF_DIRECTORY_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    assert tuple(slot.status for slot in observation.artifact_slots) == (
        "available",
        "available",
    )
    assert tuple(
        metric.score
        for slot in observation.artifact_slots
        for metric in slot.metric_scores
    ) == (1.0, 1.0)
    assert resolver.open_calls[0]["expected_media_types"] == frozenset(
        {"application/pdf"}
    )
    assert events == ["gold-open", "guest-file", "guest-directory"]
    for private_value in (
        private_text,
        actual_marker,
        gold_marker,
        "Invoice # 243729.pdf",
        "/srv/",
    ):
        assert private_value not in repr(observation)


def test_family_source_rejects_resigned_spec_drift_before_trusted_io() -> None:
    """验证外层 spec 与 family/projection catalog 不会形成分裂身份。

    输入参数：
        无；仅修改 first-sheet locator 并重签 canonical 摘要，其余
        task、slot 与 metric 身份保持不变。
    输出返回值：
        无；source 必须在 gold resolver 与 guest getter 首次 I/O 前拒绝。
    """

    source = OSWORLD_ARTIFACT_EVIDENCE_SPECS[_FIRST_SHEET_TASK_ID]
    drifted_slot = replace(
        source.artifact_slots[0],
        source_locator_relative_paths=("exam/drifted-grades.xlsx",),
        locator_relative_paths=("exam/drifted-grades.xlsx",),
    )
    unsigned = replace(
        source,
        artifact_slots=(drifted_slot,),
        evidence_spec_sha256="",
    )
    drifted_spec = replace(
        unsigned,
        evidence_spec_sha256=hashlib.sha256(
            canonical_artifact_evidence_spec_json(unsigned).encode("utf-8")
        ).hexdigest(),
    )
    content = _xlsx_bytes("PRIVATE_SPLIT_BRAIN_SENTINEL")
    resolver = _GoldResolver(content)
    controller = _SingleFileController(content)

    with pytest.raises(OSWorldArtifactEvidenceError):
        OSWorldArtifactEvidenceSource(
            specs={_FIRST_SHEET_TASK_ID: drifted_spec},
            gold_resolver=resolver,
        ).capture(
            _FIRST_SHEET_TASK_ID,
            controller,
            guest_shared_dir="/srv/paraguibench-test/user/shared",
        )

    assert resolver.open_calls == []
    assert controller.collect_calls == []


def _spec_with_inline_pixel_digest(pixel_digest: str):
    """为测试像素构造仍通过公共 canonical digest 绑定的可信 spec。

    输入参数：
        pixel_digest：测试图片解码后像素字节的 SHA-256。
    输出返回值：
        仅替换内联 gold、随后重新计算 evidence spec SHA 的不可变 spec。
    """

    source = OSWORLD_ARTIFACT_EVIDENCE_SPECS[_TASK_ID]
    metric = source.artifact_slots[0].metrics[0]
    inline_rule = {
        "expected": {pixel_digest: ["Mount Test"]},
        "expect_in_result": True,
        "result_not_list": True,
    }
    replaced_metric = replace(
        metric,
        expected_options_json=json.dumps(
            inline_rule,
            ensure_ascii=False,
            allow_nan=False,
            sort_keys=True,
            separators=(",", ":"),
        ),
    )
    replaced_slot = replace(source.artifact_slots[0], metrics=(replaced_metric,))
    unsigned = replace(
        source,
        artifact_slots=(replaced_slot,),
        evidence_spec_sha256="",
    )
    digest = hashlib.sha256(
        canonical_artifact_evidence_spec_json(unsigned).encode("utf-8")
    ).hexdigest()
    return replace(unsigned, evidence_spec_sha256=digest)


def test_capture_scores_decoded_image_pixels_from_adapted_shared_directory() -> None:
    """验证首条 production evidence 链忠实复现 Pillow 像素哈希。

    输入参数：
        无；使用真实 Pillow 编解码与 fake controller 外部边界。
    输出返回值：
        无；断言动态 home、Pictures→shared 适配、有界 I/O、spec SHA
        绑定和最终 metric observation 同时成立。
    """

    _image_bytes, pixel_digest = _png_and_pixel_digest()
    spec = _spec_with_inline_pixel_digest(pixel_digest)
    controller = _ImageDirectoryController(((pixel_digest, "Mount Test"),))

    observation = OSWorldArtifactEvidenceSource(specs={_TASK_ID: spec}).capture(
        _TASK_ID,
        controller,
        guest_shared_dir="/home/user/shared",
    )

    assert observation.rule_id == spec.rule_id
    assert observation.source_contract_sha256 == spec.source_contract_sha256
    assert observation.evidence_spec_sha256 == spec.evidence_spec_sha256
    assert len(observation.artifact_slots) == 1
    slot = observation.artifact_slots[0]
    assert (slot.slot_id, slot.status) == ("renamed_picture_set", "available")
    assert tuple((metric.metric_id, metric.score) for metric in slot.metric_scores) == (
        ("check_direct_json_object", 1.0),
    )
    assert controller.collect_calls == [
        {
            "guest_directory": "/home/user/shared",
            "max_entries": spec.limits.max_items,
            "max_name_bytes": 255,
            "max_compressed_item_bytes": spec.limits.max_single_item_bytes,
            "max_total_compressed_bytes": spec.limits.max_total_bytes,
            "max_pixels_per_image": (spec.limits.max_container_expanded_bytes // 16),
            "max_decoded_item_bytes": (spec.limits.max_container_expanded_bytes),
            "max_total_decoded_bytes": (spec.limits.max_container_expanded_bytes),
            "max_response_bytes": spec.limits.max_text_bytes,
            "timeout_seconds": spec.limits.getter_timeout_seconds,
        }
    ]
    assert "Mount Test" not in repr(observation)
    assert pixel_digest not in repr(observation)


def test_capture_rejects_controller_records_above_the_spec_item_limit() -> None:
    """验证 source 不把底层 controller 视为无需复核的可信内存。

    输入参数：
        无；fake controller 返回比 spec 上限多一条的结构合法记录。
    输出返回值：
        无；槽位必须标记 ``schema_error`` 且不执行任何 metric，避免异常
        controller 绕过版本化资源上限后被当成普通 Agent 零分。
    """

    spec = OSWORLD_ARTIFACT_EVIDENCE_SPECS[_TASK_ID]
    records = tuple(
        (hashlib.sha256(str(index).encode("ascii")).hexdigest(), f"image-{index}")
        for index in range(spec.limits.max_items + 1)
    )
    controller = _ImageDirectoryController(records)

    observation = OSWorldArtifactEvidenceSource().capture(
        _TASK_ID,
        controller,
        guest_shared_dir="/home/user/shared",
    )

    assert len(observation.artifact_slots) == 1
    slot = observation.artifact_slots[0]
    assert (slot.status, slot.metric_scores) == ("schema_error", ())


def test_source_constructor_maps_unrepresentable_spec_limit_to_domain_error() -> None:
    """验证注入 spec 的极端整数不会把 Python 转换异常泄到 runtime。

    输入参数：
        无；把 getter timeout 替换为无法转成有限浮点的超大整数。
    输出返回值：
        无；source 构造阶段仅抛脱敏的 artifact evidence 领域错误。
    """

    source = OSWORLD_ARTIFACT_EVIDENCE_SPECS[_TASK_ID]
    unsafe_limits = replace(
        source.limits,
        getter_timeout_seconds=10**400,
    )
    unsafe_spec = replace(source, limits=unsafe_limits)

    with pytest.raises(OSWorldArtifactEvidenceError):
        OSWorldArtifactEvidenceSource(specs={_TASK_ID: unsafe_spec})


@pytest.mark.parametrize(
    "task_id",
    sorted(set(OSWORLD_ARTIFACT_EVIDENCE_SPECS) - {_TASK_ID, _BIBTEX_TASK_ID}),
)
def test_every_remaining_unimplemented_task_is_blocked_before_any_guest_io(
    task_id: str,
) -> None:
    """验证其余 14 条已冻结 spec 均不会冒充 production getter。

    输入参数：
        task_id：pytest 枚举的每一条尚未实现 production source 的任务。
    输出返回值：
        无；source 必须在调用 ``get_desktop_path`` 前显式失败，确保冻结
        metadata 不扩大运行能力，也不对 guest 产生无意义副作用。
    """

    class _CountingController:
        """只记录 guest home 读取次数的外部边界 fake。"""

        def __init__(self) -> None:
            """初始化零次调用计数。

            输入参数：无。
            输出返回值：无。
            """

            self.home_calls = 0

        def get_desktop_path(self) -> str:
            """记录不应发生的 guest I/O。

            输入参数：无。
            输出返回值：固定测试路径。
            """

            self.home_calls += 1
            return "/home/user/Desktop"

    controller = _CountingController()

    with pytest.raises(OSWorldArtifactEvidenceError):
        OSWorldArtifactEvidenceSource().capture(
            task_id,
            controller,
            guest_shared_dir="/home/user/shared",
        )

    assert controller.home_calls == 0


def test_capture_uses_frozen_shared_binding_without_rereading_desktop() -> None:
    """验证评价只使用 prepare 阶段冻结的 shared locator。

    输入参数：
        无；controller 的 Desktop 接口若被调用会立即失败。
    输出返回值：
        无；getter 必须收到显式冻结路径，证明 endpoint 漂移不会把同一
        Attempt 的评价切换到另一个 guest home。
    """

    spec = OSWORLD_ARTIFACT_EVIDENCE_SPECS[_TASK_ID]
    controller = _ImageDirectoryController(())

    def fail_if_reread() -> str:
        """拒绝 capture 阶段重新解析 Desktop。

        输入参数：无。
        输出返回值：永不返回；调用即表示 locator 绑定被破坏。
        """

        raise AssertionError("capture 不得重新读取 Desktop")

    controller.get_desktop_path = fail_if_reread  # type: ignore[method-assign]

    OSWorldArtifactEvidenceSource().capture(
        _TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/frozen/shared",
    )

    assert controller.collect_calls[0]["guest_directory"] == (
        "/srv/paraguibench-test/frozen/shared"
    )
    assert controller.collect_calls[0]["timeout_seconds"] == (
        spec.limits.getter_timeout_seconds
    )


def test_existing_empty_directory_is_available_and_scores_zero() -> None:
    """验证存在但为空的目录仍进入旧 metric，而不是伪装成 missing。

    输入参数：
        无；controller 用空 tuple 表示已安全打开但没有直接成员的目录。
    输出返回值：
        无；槽位状态为 ``available``，固定三山映射 metric 得分为 0。
    """

    controller = _ImageDirectoryController(())

    observation = OSWorldArtifactEvidenceSource().capture(
        _TASK_ID,
        controller,
        guest_shared_dir="/home/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert slot.status == "available"
    assert tuple((metric.metric_id, metric.score) for metric in slot.metric_scores) == (
        ("check_direct_json_object", 0.0),
    )


def test_absent_final_directory_maps_only_typed_missing_to_missing() -> None:
    """验证仅 controller 的 typed ENOENT 被投影为 missing。

    输入参数：
        无；fake getter 抛不携带路径的专用缺失异常。
    输出返回值：
        无；source 返回无 metric 的 ``missing`` 槽位，不把它误记为读取错误。
    """

    controller = _ImageDirectoryController(())

    def raise_missing(
        *args: object,
        **kwargs: object,
    ) -> tuple[tuple[str, str], ...]:
        """模拟最终目录 ENOENT 的受控 controller 边界。

        输入参数：
            args/kwargs：source 传入的固定 getter 参数，本测试无需读取。
        输出返回值：
            永不返回；总是抛专用缺失异常。
        """

        raise OSWorldGuestPathMissingError("guest artifact 目录缺失")

    controller.collect_image_pixel_hashes = raise_missing  # type: ignore[method-assign]

    observation = OSWorldArtifactEvidenceSource().capture(
        _TASK_ID,
        controller,
        guest_shared_dir="/home/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert (slot.status, slot.metric_scores) == ("missing", ())


def test_default_catalog_scores_all_three_canonical_mountains() -> None:
    """验证 production 默认 spec 的三条内联 gold 均参与评价。

    输入参数：
        无；使用审定 spec 内三张图片的像素摘要与各自允许名称。
    输出返回值：
        无；默认 catalog 必须得到完整 1 分，防止测试注入 spec 掩盖漏 gold。
    """

    controller = _ImageDirectoryController(
        (
            (
                "6ed4239ecc2be3ec15ad65a78c5c823b9004d640b8cc83a6a7af5930f354de91",
                "Everest",
            ),
            (
                "79f45d40d8413d4e81f1b9734ea39e58622cafd79e12bab32959643fc245147c",
                "Huashan",
            ),
            (
                "ec076282f61ba74642e94b5a6a1250c6988204d59d9b02936606b6b8ef1e4433",
                "Kilimanjaro",
            ),
        )
    )

    observation = OSWorldArtifactEvidenceSource().capture(
        _TASK_ID,
        controller,
        guest_shared_dir="/home/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert (slot.status, slot.metric_scores[0].score) == ("available", 1.0)


def test_duplicate_pixel_digest_uses_last_observed_member_name() -> None:
    """验证重复像素摘要按官方字典覆盖语义保留最后观察名称。

    输入参数：
        无；两张图片具有同一像素摘要、名称先错后对。
    输出返回值：
        无；最终 metric 为 1，证明未偷偷改成首条获胜或拒绝重复摘要。
    """

    _image_bytes, pixel_digest = _png_and_pixel_digest()
    spec = _spec_with_inline_pixel_digest(pixel_digest)
    controller = _ImageDirectoryController(
        ((pixel_digest, "Wrong Name"), (pixel_digest, "Mount Test"))
    )

    observation = OSWorldArtifactEvidenceSource(specs={_TASK_ID: spec}).capture(
        _TASK_ID,
        controller,
        guest_shared_dir="/home/user/shared",
    )

    assert observation.artifact_slots[0].metric_scores[0].score == 1.0


def test_getter_failure_and_malformed_record_have_distinct_error_statuses() -> None:
    """验证 I/O 失败与返回 schema 污染不会降级成 Agent 内容零分。

    输入参数：
        无；分别模拟普通 getter 异常和包含路径分隔符的非法成员名。
    输出返回值：
        无；两者映射为 ``read_error`` 与 ``schema_error``，且均无 metric。
    """

    failing = _ImageDirectoryController(())

    def raise_io_error(
        *args: object,
        **kwargs: object,
    ) -> tuple[tuple[str, str], ...]:
        """模拟非 ENOENT 的 controller 读取失败。

        输入参数：
            args/kwargs：source 的固定 getter 参数，本测试无需读取。
        输出返回值：
            永不返回；总是抛普通 runtime 异常。
        """

        raise RuntimeError("synthetic getter failure")

    failing.collect_image_pixel_hashes = raise_io_error  # type: ignore[method-assign]
    malformed = _ImageDirectoryController((("a" * 64, "nested/name"),))

    read_error = (
        OSWorldArtifactEvidenceSource()
        .capture(
            _TASK_ID,
            failing,
            guest_shared_dir="/home/user/shared",
        )
        .artifact_slots[0]
    )
    schema_error = (
        OSWorldArtifactEvidenceSource()
        .capture(
            _TASK_ID,
            malformed,
            guest_shared_dir="/home/user/shared",
        )
        .artifact_slots[0]
    )

    assert (read_error.status, read_error.metric_scores) == ("read_error", ())
    assert (schema_error.status, schema_error.metric_scores) == (
        "schema_error",
        (),
    )


def test_bibtex_capture_uses_desktop_actual_and_verified_gold() -> None:
    """验证 CombinationDocs-015 复现源路径与 ignore-blanks metric。

    输入参数：
        无；actual 与 gold 只在空白布局上不同，并分别由 guest getter 与
        evaluator-only resolver 提供。
    输出返回值：
        无；Desktop locator、资源上限、gold logical key/媒体与满分 observation
        全部匹配，且 observation 不含正文。
    """

    actual = b"@article{x,\n  title = {Pinned Gold}\n}\n"
    gold = b"@article{x, title = {Pinned Gold} }\n"
    controller = _SingleFileController(actual)
    resolver = _GoldResolver(gold)
    spec = OSWORLD_ARTIFACT_EVIDENCE_SPECS[_BIBTEX_TASK_ID]

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=resolver,
    ).capture(
        _BIBTEX_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert (slot.slot_id, slot.status) == ("bibtex_output", "available")
    assert tuple((metric.metric_id, metric.score) for metric in slot.metric_scores) == (
        ("compare_text_file", 1.0),
    )
    assert controller.collect_calls == [
        {
            "guest_path": ("/srv/paraguibench-test/user/Desktop/references.bib"),
            "max_bytes": spec.limits.max_text_bytes,
            "max_response_bytes": (4_096 + 4 * ((spec.limits.max_text_bytes + 2) // 3)),
            "timeout_seconds": spec.limits.getter_timeout_seconds,
        }
    ]
    assert resolver.open_calls == [
        {
            "logical_key": _BIBTEX_GOLD_KEY,
            "max_bytes": spec.limits.max_text_bytes,
            "expected_media_types": frozenset({"application/x-bibtex"}),
        }
    ]
    assert "Pinned Gold" not in repr(observation)


def test_bibtex_content_mismatch_is_available_zero_score() -> None:
    """验证可信 actual 与 gold 不等属于正常 Agent 评价失败。

    输入参数：
        无；构造两个 UTF-8 合法但内容不同的 BibTeX 文件。
    输出返回值：
        无；槽位仍是 available，metric 得 0，而不是 evaluator error。
    """

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=_GoldResolver(b"@article{gold}\n"),
    ).capture(
        _BIBTEX_TASK_ID,
        _SingleFileController(b"@article{actual}\n"),
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert (slot.status, slot.metric_scores[0].score) == ("available", 0.0)


def test_bibtex_missing_actual_is_missing_after_gold_is_verified() -> None:
    """验证 actual ENOENT 是零分语义，但不能掩盖未准备的 gold。

    输入参数：
        无；gold 可用，guest getter 抛专用缺失异常。
    输出返回值：
        无；resolver 先完成一次可信打开，随后槽位投影为 missing 且无 metric。
    """

    resolver = _GoldResolver(b"@article{gold}\n")
    controller = _SingleFileController(
        error=OSWorldGuestPathMissingError("synthetic missing"),
    )

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=resolver,
    ).capture(
        _BIBTEX_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert (slot.status, slot.metric_scores) == ("missing", ())
    assert len(resolver.open_calls) == 1
    assert len(controller.collect_calls) == 1


def test_bibtex_preflights_every_slot_gold_before_any_guest_io() -> None:
    """验证多槽位任务先完成全部 gold 门禁，再读取任一 actual。

    输入参数：
        无；从 canonical 015 spec 派生两个合法 BibTeX 槽位，resolver
            第一次打开成功、第二次打开失败。
    输出返回值：
        无；即使首槽位 gold 可用，第二槽位 gold 失败也必须使整个任务
        保持零次 guest/controller 调用，防止按槽位交错取证。
    """

    source_spec = OSWORLD_ARTIFACT_EVIDENCE_SPECS[_BIBTEX_TASK_ID]
    first_slot = source_spec.artifact_slots[0]
    second_slot = replace(
        first_slot,
        slot_id="bibtex_output_second",
        source_locator_relative_paths=("Desktop/references-second.bib",),
        locator_relative_paths=("Desktop/references-second.bib",),
    )
    unsigned_spec = replace(
        source_spec,
        artifact_slots=(first_slot, second_slot),
        evidence_spec_sha256="",
    )
    two_slot_spec = replace(
        unsigned_spec,
        evidence_spec_sha256=hashlib.sha256(
            canonical_artifact_evidence_spec_json(unsigned_spec).encode("utf-8")
        ).hexdigest(),
    )

    class _SecondGoldOpenFailsResolver:
        """第一次返回可信字节流，第二次以含秘密文本的异常失败。"""

        def __init__(self) -> None:
            """初始化 gold 打开调用记录。

            输入参数：无。
            输出返回值：无；调用记录初始为空。
            """

            self.open_calls: list[str] = []

        @contextmanager
        def open_verified(
            self,
            logical_key: str,
            *,
            max_bytes: int,
            expected_media_types: frozenset[str],
        ):
            """为首次调用交付 gold，第二次模拟可信门禁失败。

            输入参数：
                logical_key：当前槽位绑定的稳定 gold 身份。
                max_bytes/expected_media_types：source 传入的资源和媒体门禁。
            输出返回值：
                首次调用在 context 内交付只读 ``BytesIO``；第二次不返回。
            """

            assert max_bytes == source_spec.limits.max_text_bytes
            assert expected_media_types == frozenset({"application/x-bibtex"})
            self.open_calls.append(logical_key)
            if len(self.open_calls) == 2:
                raise RuntimeError("PRIVATE_SECOND_GOLD_FAILURE")
            with BytesIO(b"@article{gold}\n") as stream:
                yield stream

    resolver = _SecondGoldOpenFailsResolver()
    controller = _SingleFileController(b"@article{actual}\n")

    observation = OSWorldArtifactEvidenceSource(
        specs={_BIBTEX_TASK_ID: two_slot_spec},
        gold_resolver=resolver,
    ).capture(
        _BIBTEX_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    assert len(observation.artifact_slots) == 2
    assert len(resolver.open_calls) == 2
    assert controller.collect_calls == []


@pytest.mark.parametrize(
    ("resolver_error", "actual", "expected_status", "expected_guest_calls"),
    [
        (RuntimeError("gold unavailable"), b"valid", "read_error", 0),
        (None, b"\xff", "schema_error", 1),
    ],
)
def test_bibtex_gold_or_text_errors_are_evaluator_errors(
    resolver_error: Exception | None,
    actual: bytes,
    expected_status: str,
    expected_guest_calls: int,
) -> None:
    """验证 gold 基础设施与 UTF-8 解析失败不伪装成内容零分。

    输入参数：
        resolver_error：模拟 gold 未 provision/损坏；``None`` 表示可用。
        actual：guest getter 返回的 actual 原始字节。
        expected_status：脱敏槽位错误分类。
        expected_guest_calls：gold 门禁后允许发生的 guest getter 次数。
    输出返回值：
        无；两类 evaluator 故障均无 metric score，且 gold 失败发生在 guest
        actual 读取之前。
    """

    resolver = _GoldResolver(
        b"@article{gold}\n",
        error=resolver_error,
    )
    controller = _SingleFileController(actual)

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=resolver,
    ).capture(
        _BIBTEX_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert (slot.status, slot.metric_scores) == (expected_status, ())
    assert len(controller.collect_calls) == expected_guest_calls


@pytest.mark.parametrize("failing_boundary", ("gold", "guest"))
def test_bibtex_boundary_secrets_never_enter_observation_output(
    failing_boundary: str,
) -> None:
    """验证 gold 或 guest 边界异常中的秘密不会进入 observation。

    输入参数：
        failing_boundary：pytest 指定由 ``gold`` resolver 或 ``guest``
            controller 抛出包含正文、路径与标记的异常。
    输出返回值：
        无；槽位只公开 ``read_error``，observation 的 ``repr`` 与 ``str``
        均不得包含秘密标记、正文或路径片段。
    """

    secret_marker = "PRIVATE_BIBTEX_BOUNDARY_MARKER"
    secret_path = "/private/evaluator/cache/references.bib"
    secret_content = "@article{private-secret-content}"
    boundary_error = RuntimeError(f"{secret_marker}:{secret_path}:{secret_content}")
    resolver = _GoldResolver(
        b"@article{gold}\n",
        error=boundary_error if failing_boundary == "gold" else None,
    )
    controller = _SingleFileController(
        b"@article{actual}\n",
        error=boundary_error if failing_boundary == "guest" else None,
    )

    observation = OSWorldArtifactEvidenceSource(
        gold_resolver=resolver,
    ).capture(
        _BIBTEX_TASK_ID,
        controller,
        guest_shared_dir="/srv/paraguibench-test/user/shared",
    )

    slot = observation.artifact_slots[0]
    assert (slot.status, slot.metric_scores) == ("read_error", ())
    public_outputs = (
        repr(observation),
        str(observation),
        repr(slot),
        str(slot),
    )
    for output in public_outputs:
        assert secret_marker not in output
        assert secret_path not in output
        assert secret_content not in output


def test_bibtex_domain_error_does_not_echo_secret_binding() -> None:
    """验证 capture 向外抛出的领域错误不回显非法 guest binding。

    输入参数：
        无；将秘密标记放入末段不是 ``shared`` 的 guest 路径绑定。
    输出返回值：
        无；错误的 ``str``/``repr`` 均为固定脱敏文本，并且在 gold
        或 guest I/O 发生前失败。
    """

    secret_marker = "PRIVATE_GUEST_BINDING_MARKER"
    resolver = _GoldResolver(b"@article{gold}\n")
    controller = _SingleFileController(b"@article{actual}\n")

    with pytest.raises(OSWorldArtifactEvidenceError) as caught:
        OSWorldArtifactEvidenceSource(
            gold_resolver=resolver,
        ).capture(
            _BIBTEX_TASK_ID,
            controller,
            guest_shared_dir=f"/srv/{secret_marker}/not-shared",
        )

    assert secret_marker not in str(caught.value)
    assert secret_marker not in repr(caught.value)
    assert resolver.open_calls == []
    assert controller.collect_calls == []


def test_bibtex_source_without_gold_resolver_fails_before_guest_io() -> None:
    """验证 production 默认构造不会把缺失 gold resolver 降级为占位分数。

    输入参数：
        无；使用未注入 resolver 的 source 和带调用记录的 fake controller。
    输出返回值：
        无；capture 在 guest I/O 前抛领域错误，任务继续保持 blocked。
    """

    controller = _SingleFileController(b"private actual")

    with pytest.raises(OSWorldArtifactEvidenceError):
        OSWorldArtifactEvidenceSource().capture(
            _BIBTEX_TASK_ID,
            controller,
            guest_shared_dir="/srv/paraguibench-test/user/shared",
        )

    assert controller.collect_calls == []

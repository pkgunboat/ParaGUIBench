"""Excel-008 production artifact capture 到正式评价器的纵向测试。"""

from __future__ import annotations

from copy import copy
from dataclasses import replace
import hashlib
import json
import multiprocessing
import os
from pathlib import Path, PurePosixPath
import shutil
import stat
from typing import Any
import zipfile

import pytest

from paraguibench.agents import AgentRunResult
from paraguibench.benchmark import PreparedTask
from paraguibench.evaluation.pipeline_implicit import (
    HIDE_NA_ROWS_PROTOCOL_ID,
    HIDE_NA_ROWS_TASK_ID,
    HideNARowsObservation,
    evaluate_hide_na_rows,
)
from paraguibench.integrations.pipeline_implicit import (
    PipelineImplicitArtifactEvidenceError,
    PipelineImplicitArtifactEvidenceSource,
    PINNED_HIDE_NA_ROWS_BASELINE_SHA256,
    derive_hide_na_rows_baseline_sha256,
)
from paraguibench.integrations.pipeline_implicit import hide_na_rows_bridge
from paraguibench.runstore import (
    AttemptFailureStage,
    EvaluationOutcome,
    ExecutionOutcome,
    RunProvenanceStatus,
    RunStore,
)
from paraguibench.runtime.attempt_runner import AttemptRunner
from paraguibench.runtime.evaluators import PipelineImplicitTaskEvaluator
from tests.runstore._audit import (
    synthetic_run_version_vector,
    synthetic_task_audit,
)


_FIXTURE_ENVIRONMENT_VARIABLE = "PARAGUI_EXCEL008_FIXTURE_ROOT"
_TASK_UID = "1c73128f-a5ef-4a97-97ce-ef427d6d46b4"
_GUEST_SHARED_DIR = "/guest-home/shared"


def _consume_hide_na_rows_cpu(
    content: bytes,
) -> tuple[tuple[int, ...], str]:
    """持续占用 CPU，验证 Excel-008 worker 的 OS 硬限额。

    输入参数：
        content：故障注入用的有界字节，本函数不解析。
    输出返回值：
        不返回；子进程应被 ``RLIMIT_CPU`` 终止。
    """

    del content
    accumulator = 0
    while True:
        accumulator = (accumulator + 1) % 104729


def _attempt_hide_na_rows_filesystem_mutation(
    content: bytes,
) -> tuple[tuple[int, ...], str]:
    """在 Excel-008 真实 worker 内尝试指定文件系统变更。

    输入参数：
        content：JSON 字节，只含操作名、源路径和目标路径。
    输出返回值：
        仅当隔离边界错误放行时返回合法小投影；
        正确边界会先以 ``PermissionError`` 拒绝。
    """

    instruction = json.loads(content.decode("utf-8", errors="strict"))
    operation = instruction["operation"]
    source = instruction["source"]
    destination = instruction["destination"]
    if operation == "open-write":
        with open(source, mode="wb"):
            pass
    elif operation == "os-open-write":
        descriptor = os.open(source, os.O_WRONLY | os.O_TRUNC)
        os.close(descriptor)
    elif operation == "truncate":
        os.truncate(source, 0)
    elif operation == "delete":
        os.remove(source)
    elif operation == "rename":
        os.rename(source, destination)
    elif operation == "replace":
        os.replace(source, destination)
    elif operation == "mkdir":
        os.mkdir(destination)
    elif operation == "rmdir":
        os.rmdir(source)
    elif operation == "chmod":
        os.chmod(source, 0o600)
    elif operation == "link":
        os.link(source, destination)
    elif operation == "symlink":
        os.symlink(source, destination)
    elif operation == "mkfifo":
        os.mkfifo(destination, 0o600)
    elif operation == "mknod":
        os.mknod(destination, stat.S_IFREG | 0o600)
    elif operation == "utime":
        os.utime(source, ns=(1, 1))
    elif operation == "chown":
        os.chown(source, os.getuid(), os.getgid())
    else:
        raise AssertionError("未注册的文件系统故障注入操作")
    return (), "0" * 64


def _return_hide_na_rows_parser_projection(
    content: bytes,
) -> tuple[tuple[int, ...], str]:
    """返回可验证的小型 parser 投影以检查 JSON IPC。

    输入参数：
        content：有界测试字节，不包含工作簿私密内容。
    输出返回值：
        两个隐藏行号和固定小写 SHA-256。
    """

    del content
    return (8, 10), "0" * 64


def _create_readonly_hide_na_rows_cwd(tmp_path: Path, name: str) -> Path:
    """创建 Excel-008 parser 故障注入使用的只读空目录。

    输入参数：
        tmp_path：pytest 隔离根目录。
        name：当前测试唯一子目录名。
    输出返回值：
        已移除写权限的空目录路径。
    """

    sandbox_cwd = tmp_path / name
    sandbox_cwd.mkdir()
    sandbox_cwd.chmod(stat.S_IRUSR | stat.S_IXUSR)
    return sandbox_cwd


def _snapshot_hide_na_rows_test_tree(
    root: Path,
) -> tuple[tuple[object, ...], ...]:
    """快照故障注入树的身份、元数据和文件内容。

    输入参数：
        root：只含测试合成成员的可破坏根目录。
    输出返回值：
        按 UTF-8 路径排序的相对名、类型、权限、
        大小、修改时间和内容摘要元组。
    """

    records: list[tuple[object, ...]] = []
    for path in root.rglob("*"):
        metadata = path.lstat()
        payload_sha256 = (
            hashlib.sha256(path.read_bytes()).hexdigest()
            if path.is_file() and not path.is_symlink()
            else None
        )
        records.append(
            (
                path.relative_to(root).as_posix(),
                stat.S_IFMT(metadata.st_mode),
                stat.S_IMODE(metadata.st_mode),
                metadata.st_size,
                metadata.st_mtime_ns,
                payload_sha256,
            )
        )
    return tuple(sorted(records, key=lambda item: str(item[0]).encode("utf-8")))


class _DirectoryController:
    """把已核验的本地固定 revision 目录暴露为 guest seam。

    输入参数：
        root：包含 Excel-008 五个工作簿的目录。
    输出返回值：
        无；实例通过 production source 要求的窄接口返回
        稳定 manifest 和单文件字节。
    """

    def __init__(self, root: Path) -> None:
        self._root = root

    def collect_artifact_tree_manifest(
        self,
        guest_directory: str,
        **limits: Any,
    ) -> tuple[tuple[str, int, str], ...]:
        """返回按 UTF-8 字节序排列的完整常规文件树。

        输入参数：
            guest_directory：production source 冻结的 guest shared 根。
            limits：source 下发的资源与超时上限。
        输出返回值：
            ``(relative_path, size, sha256)`` 的有序不可变 tuple。
        """

        assert guest_directory == _GUEST_SHARED_DIR
        assert limits
        records = []
        for path in self._root.rglob("*"):
            if not path.is_file():
                continue
            payload = path.read_bytes()
            records.append(
                (
                    path.relative_to(self._root).as_posix(),
                    len(payload),
                    hashlib.sha256(payload).hexdigest(),
                )
            )
        return tuple(sorted(records, key=lambda item: item[0].encode("utf-8")))

    def collect_file_bytes(
        self,
        guest_path: str,
        **limits: Any,
    ) -> bytes:
        """按 production source 给出的 guest 路径返回真实字节。

        输入参数：
            guest_path：必须位于固定 guest shared 根下。
            limits：source 下发的单文件资源与超时上限。
        输出返回值：
            对应固定 revision 成员的原始字节。
        """

        assert limits
        guest_root = PurePosixPath(_GUEST_SHARED_DIR)
        relative_path = PurePosixPath(guest_path).relative_to(guest_root)
        return self._root.joinpath(*relative_path.parts).read_bytes()


class _ExcelObservationEnvironment:
    """在 AttemptRunner 生命周期内按需捕获 Excel-008 观测。

    输入参数：
        root：当前 Attempt 的 guest shared 工作簿闭集。
    输出返回值：
        无；对象通过 runtime 窄 seam 返回 production typed 观测。
    """

    def __init__(self, root: Path) -> None:
        self._controller = _DirectoryController(root)
        self.closed = False

    def start(self) -> None:
        """启动不占用外部资源的合成环境。

        输入参数：无。
        输出返回值：无。
        """

    def prepare(self, task: dict[str, Any]) -> None:
        """验证 AttemptRunner 传入的可信任务身份。

        输入参数：
            task：``PreparedTask`` 的 trusted projection。
        输出返回值：无；身份漂移时断言失败。
        """

        assert task["task_id"] == HIDE_NA_ROWS_TASK_ID

    def close(self) -> None:
        """标记 Attempt 环境已清理。

        输入参数：无。
        输出返回值：无。
        """

        self.closed = True

    def pipeline_implicit_observation(
        self,
        task_id: str,
        protocol_id: str,
    ) -> object:
        """从仍存活的 guest 文件闭集生成强类型观测。

        输入参数：
            task_id/protocol_id：runtime adapter 固定的任务与协议。
        输出返回值：
            manifest—nofollow—manifest 真实 capture 的 Excel-008 观测。
        """

        assert task_id == HIDE_NA_ROWS_TASK_ID
        assert protocol_id == HIDE_NA_ROWS_PROTOCOL_ID
        return PipelineImplicitArtifactEvidenceSource().capture(
            task_id,
            self._controller,
            guest_shared_dir=_GUEST_SHARED_DIR,
        )


class _SensitiveExcelAgent:
    """返回不得参与 Excel-008 评价或持久化的哨兵文本。"""

    def run(
        self,
        task_view: dict[str, Any],
        environment: object,
    ) -> AgentRunResult:
        """返回一步结束且含私密 final text 的合法 Agent 结果。

        输入参数：
            task_view：不含 gold 的 Agent 视图。
            environment：仍存活的工作簿环境；本 fake 不读取。
        输出返回值：
            含固定哨兵、一步和结束原因的 ``AgentRunResult``。
        """

        del environment
        assert task_view["task_id"] == HIDE_NA_ROWS_TASK_ID
        return AgentRunResult(
            final_output="PRIVATE-FINAL-TEXT-SENTINEL",
            step_count=1,
            termination="finished",
        )


def _prepared_excel_task() -> PreparedTask:
    """从 canonical JSON 构造 RunStore 纵向测试的三投影任务。

    输入参数：无。
    输出返回值：
        trusted task 保留正式身份，Agent view 仅含任务和指令，
        audit metadata 仅含 RunStore allowlist 字段。
    """

    repo_root = Path(__file__).resolve().parents[2]
    task = json.loads(
        (
            repo_root
            / "benchmark/tasks/Operation-FileOperate-BatchOperationExcel-008.json"
        ).read_text(encoding="utf-8")
    )
    return PreparedTask(
        trusted_task=task,
        agent_task={
            "task_id": HIDE_NA_ROWS_TASK_ID,
            "instruction": task["instruction"],
        },
        audit_metadata=synthetic_task_audit(
            HIDE_NA_ROWS_TASK_ID,
            task_uid=task["task_uid"],
            task_type=task["task_type"],
            task_source=task["task_source"],
            task_tag=task["task_tag"],
        ),
    )


def _fixed_revision_gold_fixture() -> Path:
    """返回固定 Lee revision 的五文件 gold 目录。

    输入参数：无。
    输出返回值：
        环境变量指向 fixture 存在时返回路径；未配置时
        跳过 download-only 纵向测试。
    """

    raw_path = os.environ.get(_FIXTURE_ENVIRONMENT_VARIABLE)
    if raw_path is None:
        pytest.skip(
            f"{_FIXTURE_ENVIRONMENT_VARIABLE} is required for download-only fixture"
        )
    fixture_path = Path(raw_path) / "answer_files" / _TASK_UID
    if not fixture_path.is_dir():
        pytest.fail("Excel-008 fixed-revision gold fixture is unavailable")
    return fixture_path


def _fixed_revision_input_fixture() -> Path:
    """返回固定 Lee revision 的五文件 input 目录。

    输入参数：无。
    输出返回值：
        与 gold 同一 fixture 根下的 benchmark input 路径。
    """

    gold_fixture = _fixed_revision_gold_fixture()
    input_fixture = gold_fixture.parents[1] / "benchmark_dataset" / _TASK_UID
    if not input_fixture.is_dir():
        pytest.fail("Excel-008 fixed-revision input fixture is unavailable")
    return input_fixture


def _copy_fixed_revision_gold_fixture(tmp_path: Path) -> Path:
    """把 download-only gold 树复制到当前测试隔离目录。

    输入参数：
        tmp_path：pytest 为当前测试创建的空临时根。
    输出返回值：
        可安全改动的独立五文件树。
    """

    target = tmp_path / "gold"
    shutil.copytree(_fixed_revision_gold_fixture(), target)
    return target


def _derive_kfc_semantic_variant_digests(
    tmp_path: Path,
    mutation: str,
) -> tuple[str, str]:
    """从同一真实 gold 生成对照/变体并返回受控语义摘要。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：当前 RED→GREEN 类别的已注册可见变更。
    输出返回值：
        对照工作簿和变体工作簿的受控语义 SHA-256。

    对照与变体都来自 Lee 固定 revision 的同一 KFC gold；
    字体/字号/颜色用例在两份中先建立相同页眉文本，
    确保测到的是样式字段本身，而非“从无到有”的文本差异。
    """

    import openpyxl

    source = _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    control_path = tmp_path / "control.xlsx"
    mutated_path = tmp_path / "mutated.xlsx"
    shutil.copy2(source, control_path)
    shutil.copy2(source, mutated_path)
    control = openpyxl.load_workbook(control_path)
    mutated = openpyxl.load_workbook(mutated_path)
    control_sheet = control.active
    mutated_sheet = mutated.active
    if mutation in {"header_font", "header_size", "header_color"}:
        control_sheet.oddHeader.left.text = "Visible header"
        mutated_sheet.oddHeader.left.text = "Visible header"
    if mutation == "header_flag":
        mutated_sheet.HeaderFooter.differentFirst = True
    elif mutation == "header_text":
        mutated_sheet.firstFooter.right.text = "Visible footer"
    elif mutation == "header_font":
        mutated_sheet.oddHeader.left.font = "Arial,Bold"
    elif mutation == "header_size":
        mutated_sheet.oddHeader.left.size = 18
    elif mutation == "header_color":
        mutated_sheet.oddHeader.left.color = "FF0000"
    elif mutation == "sheet_protection_flag":
        mutated_sheet.protection.sheet = True
    elif mutation == "sheet_protection_password":
        control_sheet.protection.sheet = True
        mutated_sheet.protection.sheet = True
        mutated_sheet.protection.password = "ABCD"
    elif mutation == "sheet_protection_hash":
        control_sheet.protection.sheet = True
        mutated_sheet.protection.sheet = True
        mutated_sheet.protection.algorithmName = "SHA-512"
        mutated_sheet.protection.hashValue = "QUJD"
        mutated_sheet.protection.saltValue = "REVG"
        mutated_sheet.protection.spinCount = 1
    elif mutation == "workbook_protection_flag":
        mutated.security.lockStructure = True
    elif mutation == "workbook_protection_hash":
        mutated.security.workbookAlgorithmName = "SHA-512"
        mutated.security.workbookHashValue = "QUJD"
        mutated.security.workbookSaltValue = "REVG"
        mutated.security.workbookSpinCount = 1
    elif mutation in {"defined_name", "defined_name_metadata"}:
        from openpyxl.workbook.defined_name import DefinedName

        reference = f"'{mutated_sheet.title}'!$A$1"
        if mutation == "defined_name":
            mutated.defined_names.add(DefinedName("VisibleRange", attr_text=reference))
        else:
            control.defined_names.add(DefinedName("VisibleRange", attr_text=reference))
            mutated.defined_names.add(
                DefinedName(
                    "VisibleRange",
                    attr_text=reference,
                    hidden=True,
                )
            )
    elif mutation == "print_titles":
        mutated_sheet.print_title_rows = "1:2"
        mutated_sheet.print_title_cols = "A:B"
    elif mutation == "print_area":
        mutated_sheet.print_area = "A1:B5"
    elif mutation == "page_first_number":
        mutated_sheet.page_setup.firstPageNumber = 2
        mutated_sheet.page_setup.useFirstPageNumber = True
    elif mutation == "page_physical_size":
        mutated_sheet.page_setup.paperHeight = "297mm"
        mutated_sheet.page_setup.paperWidth = "210mm"
    elif mutation == "page_printer_defaults":
        mutated_sheet.page_setup.usePrinterDefaults = True
    elif mutation == "page_comments_errors":
        mutated_sheet.page_setup.cellComments = "atEnd"
        mutated_sheet.page_setup.errors = "blank"
    elif mutation in {"row_break", "column_break"}:
        from openpyxl.worksheet.pagebreak import Break

        page_break = Break(
            id=5,
            min=0,
            max=16_383,
            man=True,
            pt=False,
        )
        if mutation == "row_break":
            mutated_sheet.row_breaks.append(page_break)
        else:
            mutated_sheet.col_breaks.append(page_break)
    elif mutation == "workbook_view_visibility":
        mutated.views[0].visibility = "hidden"
    elif mutation == "workbook_view_minimized":
        mutated.views[0].minimized = True
    elif mutation == "workbook_view_chrome":
        mutated.views[0].showHorizontalScroll = False
        mutated.views[0].showVerticalScroll = False
        mutated.views[0].showSheetTabs = False
    elif mutation == "workbook_view_sheet_indices":
        mutated.views[0].firstSheet = 1
        mutated.views[0].activeTab = 1
    elif mutation == "workbook_view_date_grouping":
        mutated.views[0].autoFilterDateGrouping = False
    elif mutation == "workbook_view_geometry":
        mutated.views[0].xWindow = 1
        mutated.views[0].tabRatio = 501
    elif mutation == "sheet_view_window_protection":
        mutated_sheet.sheet_view.windowProtection = True
    elif mutation == "sheet_view_ruler":
        mutated_sheet.sheet_view.showRuler = False
    elif mutation == "sheet_view_whitespace":
        mutated_sheet.sheet_view.showWhiteSpace = False
    elif mutation == "sheet_view_color_id":
        mutated_sheet.sheet_view.colorId = 7
    elif mutation == "sheet_view_zoom_to_fit":
        mutated_sheet.sheet_view.zoomToFit = True
    elif mutation == "sheet_view_workbook_id":
        mutated_sheet.sheet_view.workbookViewId = 1
    elif mutation == "sheet_format_custom_height":
        mutated_sheet.sheet_format.customHeight = True
    elif mutation == "sheet_format_row_outline":
        mutated_sheet.sheet_format.outlineLevelRow = 1
    elif mutation == "sheet_format_column_outline":
        mutated_sheet.sheet_format.outlineLevelCol = 1
    elif mutation == "sheet_properties_outline":
        mutated_sheet.sheet_properties.outlinePr.applyStyles = True
        mutated_sheet.sheet_properties.outlinePr.summaryBelow = False
        mutated_sheet.sheet_properties.outlinePr.summaryRight = False
        mutated_sheet.sheet_properties.outlinePr.showOutlineSymbols = False
    elif mutation == "sheet_properties_code_name":
        mutated_sheet.sheet_properties.codeName = "VisibleSheetCode"
    elif mutation == "sheet_properties_format_calculation":
        mutated_sheet.sheet_properties.enableFormatConditionsCalculation = False
    elif mutation == "sheet_properties_filter_mode":
        mutated_sheet.sheet_properties.filterMode = False
    elif mutation == "sheet_properties_published":
        mutated_sheet.sheet_properties.published = False
    elif mutation == "sheet_properties_sync":
        mutated_sheet.sheet_properties.syncHorizontal = True
        mutated_sheet.sheet_properties.syncVertical = True
        mutated_sheet.sheet_properties.syncRef = "A1"
    elif mutation == "sheet_properties_transition":
        mutated_sheet.sheet_properties.transitionEvaluation = True
        mutated_sheet.sheet_properties.transitionEntry = True
    elif mutation == "calculation_calc_id":
        mutated.calculation.calcId = 999_999
    elif mutation == "calculation_mode":
        mutated.calculation.calcMode = "manual"
    elif mutation == "calculation_full_on_load":
        mutated.calculation.fullCalcOnLoad = False
    elif mutation == "calculation_reference_mode":
        mutated.calculation.refMode = "R1C1"
    elif mutation == "calculation_iterate":
        mutated.calculation.iterate = True
    elif mutation == "calculation_iterate_count":
        mutated.calculation.iterateCount = 42
    elif mutation == "calculation_iterate_delta":
        mutated.calculation.iterateDelta = 0.5
    elif mutation == "calculation_full_precision":
        mutated.calculation.fullPrecision = False
    elif mutation == "calculation_completed":
        mutated.calculation.calcCompleted = False
    elif mutation == "calculation_on_save":
        mutated.calculation.calcOnSave = True
    elif mutation == "calculation_concurrent":
        mutated.calculation.concurrentCalc = False
    elif mutation == "calculation_concurrent_count":
        mutated.calculation.concurrentManualCount = 2
    elif mutation == "calculation_force_full":
        mutated.calculation.forceFullCalc = True
    elif mutation in {
        "cell_protection_locked",
        "cell_protection_hidden",
    }:
        protection = copy(mutated_sheet["A1"].protection)
        if mutation == "cell_protection_locked":
            protection.locked = False
        else:
            protection.hidden = True
        mutated_sheet["A1"].protection = protection
    elif mutation in {
        "font_charset",
        "font_family",
        "font_scheme",
    }:
        font = copy(mutated_sheet["A1"].font)
        if mutation == "font_charset":
            font.charset = 204
        elif mutation == "font_family":
            font.family = 4
        else:
            font.scheme = "major"
        mutated_sheet["A1"].font = font
    elif mutation == "print_options_gridlines_set":
        mutated_sheet.print_options.gridLinesSet = False
    elif mutation == "page_setup_relationship_id":
        mutated_sheet.page_setup.id = "rId99"
    else:
        pytest.fail("unregistered remaining Excel semantic mutation")
    control.save(control_path)
    control.close()
    mutated.save(mutated_path)
    mutated.close()
    if mutation == "sheet_format_column_outline":
        _inject_sheet_format_column_outline(mutated_path)
    if mutation == "page_setup_relationship_id":
        _inject_page_setup_relationship_id(mutated_path)
    return (
        derive_hide_na_rows_baseline_sha256(control_path.read_bytes()),
        derive_hide_na_rows_baseline_sha256(mutated_path.read_bytes()),
    )


def _inject_sheet_format_column_outline(workbook_path: Path) -> None:
    """在测试 XLSX 中注入 openpyxl 会在保存时重算掉的列 outline 摘要。

    输入参数：
        workbook_path：已完成正常 openpyxl round-trip 的隔离变体。
    输出返回值：
        无；仅将 ``sheet1.xml`` 的 ``sheetFormatPr`` 加入
        ``outlineLevelCol=1``，其他 ZIP 成员与元数据保持不变。

    openpyxl 会从 column dimensions 重算此摘要字段，直接设置
    Python 属性无法产生真实变体；本辅助函数仅用于验证
    parser 对已存在 OOXML 字段的观测能力。
    """

    rebuilt_path = workbook_path.with_suffix(".rebuilt.xlsx")
    with zipfile.ZipFile(workbook_path, "r") as source:
        with zipfile.ZipFile(rebuilt_path, "w") as rebuilt:
            for info in source.infolist():
                payload = source.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    marker = b"<sheetFormatPr "
                    assert marker in payload
                    payload = payload.replace(
                        marker,
                        b'<sheetFormatPr outlineLevelCol="1" ',
                        1,
                    )
                rebuilt.writestr(info, payload)
    os.replace(rebuilt_path, workbook_path)


def _inject_page_setup_relationship_id(workbook_path: Path) -> None:
    """在真实 sheet XML 中注入 openpyxl 保存时会丢弃的 r:id。

    输入参数：
        workbook_path：已完成正常 openpyxl round-trip 的变体。
    输出返回值：
        无；仅在 ``pageSetup`` 加入合法 Office relationship QName
        和第三值 ``rId99``。
    """

    member_name = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(workbook_path, "r") as archive:
        payload = archive.read(member_name)
    marker = b"<pageSetup "
    assert payload.count(marker) == 1
    replacement = (
        b'<pageSetup xmlns:r="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships" r:id="rId99" '
    )
    _replace_ooxml_member(
        workbook_path,
        member_name=member_name,
        payload=payload.replace(marker, replacement, 1),
    )


def _replace_ooxml_member(
    workbook_path: Path,
    *,
    member_name: str,
    payload: bytes,
) -> None:
    """在隔离的真实 XLSX 中精确替换一个 OOXML 成员。

    输入参数：
        workbook_path：当前测试的真实 gold 副本。
        member_name：已存在且要替换的规范 ZIP 成员名。
        payload：要写入该成员的有界原始字节。
    输出返回值：
        无；保留其他成员身份和 ZIP metadata，仅替换目标载荷。
    """

    rebuilt_path = workbook_path.with_suffix(".rebuilt.xlsx")
    replaced = False
    with zipfile.ZipFile(workbook_path, "r") as source:
        with zipfile.ZipFile(rebuilt_path, "w") as rebuilt:
            for info in source.infolist():
                member_payload = source.read(info.filename)
                if info.filename == member_name:
                    assert replaced is False
                    member_payload = payload
                    replaced = True
                rebuilt.writestr(info, member_payload)
    assert replaced is True
    os.replace(rebuilt_path, workbook_path)


def _inject_ooxml_fragment_before_closing_tag(
    workbook_path: Path,
    *,
    member_name: str,
    closing_tag: bytes,
    fragment: bytes,
) -> None:
    """在真实 OOXML 成员的唯一闭合标签前注入片段。

    输入参数：
        workbook_path：当前隔离的 XLSX 路径。
        member_name：要修改的已存在 XML part。
        closing_tag：该 part 预期唯一的根闭合标签。
        fragment：要注入的有界 XML 字节。
    输出返回值：
        无；转交给 ``_replace_ooxml_member`` 原子重建。
    """

    with zipfile.ZipFile(workbook_path, "r") as archive:
        payload = archive.read(member_name)
    assert payload.count(closing_tag) == 1
    mutated = payload.replace(closing_tag, fragment + closing_tag, 1)
    _replace_ooxml_member(
        workbook_path,
        member_name=member_name,
        payload=mutated,
    )


def _add_ooxml_member(
    workbook_path: Path,
    *,
    member_name: str,
    payload: bytes,
) -> None:
    """在真实 XLSX 副本中新增唯一 OOXML 成员。

    输入参数：
        workbook_path：当前隔离的工作簿路径。
        member_name：不得已存在的规范 ZIP 成员名。
        payload：要新增的有界原始字节。
    输出返回值：
        无；原成员不变，新成员仅出现一次。
    """

    rebuilt_path = workbook_path.with_suffix(".rebuilt.xlsx")
    with zipfile.ZipFile(workbook_path, "r") as source:
        assert member_name not in source.namelist()
        with zipfile.ZipFile(rebuilt_path, "w") as rebuilt:
            for info in source.infolist():
                rebuilt.writestr(info, source.read(info.filename))
            rebuilt.writestr(member_name, payload)
    os.replace(rebuilt_path, workbook_path)


def test_audited_baseline_rejects_unrecognized_drawingml_shape(
    tmp_path: Path,
) -> None:
    """验证 openpyxl 会丢弃的 DrawingML 形状仍破坏基线。

    输入参数：
        tmp_path：承载 Lee 固定 revision KFC gold 副本。
    输出返回值：
        无；在已存在的空 drawing part 注入红色矩形和文本后，
        受控 builder 必须返回不同摘要，不得依赖 openpyxl
        识别对象计数而 fail-open。
    """

    source = _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    control_path = tmp_path / "control.xlsx"
    mutated_path = tmp_path / "mutated.xlsx"
    shutil.copy2(source, control_path)
    shutil.copy2(source, mutated_path)
    drawing = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:twoCellAnchor><xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:to><xdr:col>2</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>4</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to><xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="Red Rectangle"/><xdr:cNvSpPr/></xdr:nvSpPr><xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:solidFill><a:srgbClr val="FF0000"/></a:solidFill></xdr:spPr><xdr:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>PRIVATE-DRAWING-TEXT-SENTINEL</a:t></a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>"""
    _replace_ooxml_member(
        mutated_path,
        member_name="xl/drawings/drawing1.xml",
        payload=drawing,
    )

    assert derive_hide_na_rows_baseline_sha256(
        control_path.read_bytes()
    ) != derive_hide_na_rows_baseline_sha256(mutated_path.read_bytes())


def test_audited_baseline_rejects_x14_sparkline_extension(
    tmp_path: Path,
) -> None:
    """验证 worksheet 扩展 namespace 不会被警告后静默丢弃。

    输入参数：
        tmp_path：承载 Lee 固定 revision KFC gold 副本。
    输出返回值：
        无；注入 x14 sparkline ``extLst`` 后语义摘要必须变化，
        即使 openpyxl 只产生“扩展将被移除”警告。
    """

    source = _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    control_path = tmp_path / "control.xlsx"
    mutated_path = tmp_path / "mutated.xlsx"
    shutil.copy2(source, control_path)
    shutil.copy2(source, mutated_path)
    extension = b"""<extLst><ext uri="{05C60535-1F16-4fd2-B633-F4F36F0B64E0}" xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main"><x14:sparklineGroups><x14:sparklineGroup displayEmptyCellsAs="gap"><x14:colorSeries rgb="FFFF0000"/><x14:sparklines><x14:sparkline><xm:f>Monthly Data!A1:A3</xm:f><xm:sqref>G4</xm:sqref></x14:sparkline></x14:sparklines></x14:sparklineGroup></x14:sparklineGroups></ext></extLst>"""
    _inject_ooxml_fragment_before_closing_tag(
        mutated_path,
        member_name="xl/worksheets/sheet1.xml",
        closing_tag=b"</worksheet>",
        fragment=extension,
    )

    assert derive_hide_na_rows_baseline_sha256(
        control_path.read_bytes()
    ) != derive_hide_na_rows_baseline_sha256(mutated_path.read_bytes())


def test_audited_builder_rejects_unregistered_internal_relationship(
    tmp_path: Path,
) -> None:
    """验证内部 relationship 也必须命中任务包图闭集。

    输入参数：
        tmp_path：承载 Lee 固定 revision KFC gold 副本。
    输出返回值：
        无；将已存在的 drawing relationship 改为内部
        diagramData 后，受控 builder 必须以脱敏固定错误拒绝，
        不得因 target 仍在 ZIP 内就放行。
    """

    source = _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    mutated_path = tmp_path / "mutated.xlsx"
    shutil.copy2(source, mutated_path)
    member_name = "xl/worksheets/_rels/sheet1.xml.rels"
    with zipfile.ZipFile(mutated_path, "r") as archive:
        relationships = archive.read(member_name)
    drawing_type = (
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
    )
    diagram_type = b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/diagramData"
    assert relationships.count(drawing_type) == 1
    _replace_ooxml_member(
        mutated_path,
        member_name=member_name,
        payload=relationships.replace(drawing_type, diagram_type, 1),
    )

    with pytest.raises(
        PipelineImplicitArtifactEvidenceError,
        match="^TYPED_OBSERVATION_INVALID$",
    ):
        derive_hide_na_rows_baseline_sha256(mutated_path.read_bytes())


def test_audited_builder_rejects_unregistered_non_core_part(
    tmp_path: Path,
) -> None:
    """验证无 relationship 的非核心 XML part 也不能漏过。

    输入参数：
        tmp_path：承载 Lee 固定 revision KFC gold 副本。
    输出返回值：
        无；新增一个可被通用 ``xml`` content type 覆盖、
        但不在 Excel-008 package 闭集内的 part 后，builder
        必须以固定脱敏错误拒绝。
    """

    source = _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    mutated_path = tmp_path / "mutated.xlsx"
    shutil.copy2(source, mutated_path)
    _add_ooxml_member(
        mutated_path,
        member_name="xl/customXml/item1.xml",
        payload=(
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<private xmlns="urn:private">PRIVATE-PART-SENTINEL</private>'
        ),
    )

    with pytest.raises(
        PipelineImplicitArtifactEvidenceError,
        match="^TYPED_OBSERVATION_INVALID$",
    ):
        derive_hide_na_rows_baseline_sha256(mutated_path.read_bytes())


def test_audited_builder_rejects_content_type_semantic_drift(
    tmp_path: Path,
) -> None:
    """验证 package content-type 映射属于精确闭集。

    输入参数：
        tmp_path：承载 Lee 固定 revision KFC gold 副本。
    输出返回值：
        无；仅把 drawing part 声明为普通 XML 后，受控
        builder 必须拒绝，不得由 openpyxl 的宽松加载决定。
    """

    source = _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    mutated_path = tmp_path / "mutated.xlsx"
    shutil.copy2(source, mutated_path)
    member_name = "[Content_Types].xml"
    with zipfile.ZipFile(mutated_path, "r") as archive:
        content_types = archive.read(member_name)
    drawing_type = b"application/vnd.openxmlformats-officedocument.drawing+xml"
    assert content_types.count(drawing_type) == 1
    _replace_ooxml_member(
        mutated_path,
        member_name=member_name,
        payload=content_types.replace(drawing_type, b"application/xml", 1),
    )

    with pytest.raises(
        PipelineImplicitArtifactEvidenceError,
        match="^TYPED_OBSERVATION_INVALID$",
    ):
        derive_hide_na_rows_baseline_sha256(mutated_path.read_bytes())


def test_audited_baseline_hashes_nonempty_custom_property_payload(
    tmp_path: Path,
) -> None:
    """验证已登记非核心 part 也不能只检查身份。

    输入参数：
        tmp_path：承载 Lee 固定 revision KFC gold 副本。
    输出返回值：
        无；把已审定空 ``custom.xml`` 替换为合法的非空
        custom property 后，固定 payload 语义摘要必须变化。
    """

    source = _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    control_path = tmp_path / "control.xlsx"
    mutated_path = tmp_path / "mutated.xlsx"
    shutil.copy2(source, control_path)
    shutil.copy2(source, mutated_path)
    custom_properties = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/custom-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="2" name="PrivateProperty"><vt:lpwstr>PRIVATE-CUSTOM-PROPERTY-SENTINEL</vt:lpwstr></property></Properties>"""
    _replace_ooxml_member(
        mutated_path,
        member_name="docProps/custom.xml",
        payload=custom_properties,
    )

    assert derive_hide_na_rows_baseline_sha256(
        control_path.read_bytes()
    ) != derive_hide_na_rows_baseline_sha256(mutated_path.read_bytes())


@pytest.mark.parametrize(
    "mutation",
    (
        "header_flag",
        "header_text",
        "header_font",
        "header_size",
        "header_color",
    ),
)
def test_audited_baseline_rejects_header_footer_semantic_mutations(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证页眉页脚标志、文本、字体、字号与颜色均进入基线。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：页眉页脚可见语义中的一类字段。
    输出返回值：
        无；变体摘要必须与对照摘要不同。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "sheet_protection_flag",
        "sheet_protection_password",
        "sheet_protection_hash",
        "workbook_protection_flag",
        "workbook_protection_hash",
    ),
)
def test_audited_baseline_rejects_full_protection_semantic_mutations(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证 sheet/workbook protection 的标志、密码与强哈希入基线。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：两级 protection 闭集中的代表性字段组。
    输出返回值：
        无；变体摘要必须与对照摘要不同。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "defined_name",
        "defined_name_metadata",
        "print_titles",
        "print_area",
    ),
)
def test_audited_baseline_rejects_defined_name_and_print_range_mutations(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证定义名全字段以及打印标题/区域进入基线。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：定义名身份、元数据、print titles 或 print area。
    输出返回值：
        无；变体摘要必须与对照摘要不同。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


def test_audited_baseline_sorts_defined_names_by_all_fields(
    tmp_path: Path,
) -> None:
    """验证定义名 XML/字典插入顺序不改变语义指纹。

    输入参数：
        tmp_path：创建两份只有定义名插入顺序不同的真实 gold 变体。
    输出返回值：
        无；两份变体的受控摘要必须相等，证明排序键
        来自全字段内容而非字典迭代顺序。
    """

    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName

    source = _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    forward_path = tmp_path / "defined-forward.xlsx"
    reverse_path = tmp_path / "defined-reverse.xlsx"
    shutil.copy2(source, forward_path)
    shutil.copy2(source, reverse_path)
    forward = openpyxl.load_workbook(forward_path)
    reverse = openpyxl.load_workbook(reverse_path)
    reference = f"'{forward.active.title}'!$A$1"
    first = DefinedName("AlphaRange", attr_text=reference, hidden=False)
    second = DefinedName("ZuluRange", attr_text=reference, hidden=True)
    forward.defined_names.add(first)
    forward.defined_names.add(second)
    reverse.defined_names.add(
        DefinedName("ZuluRange", attr_text=reference, hidden=True)
    )
    reverse.defined_names.add(
        DefinedName("AlphaRange", attr_text=reference, hidden=False)
    )
    forward.save(forward_path)
    forward.close()
    reverse.save(reverse_path)
    reverse.close()

    assert derive_hide_na_rows_baseline_sha256(
        forward_path.read_bytes()
    ) == derive_hide_na_rows_baseline_sha256(reverse_path.read_bytes())


@pytest.mark.parametrize(
    "mutation",
    (
        "page_first_number",
        "page_physical_size",
        "page_printer_defaults",
        "page_comments_errors",
        "row_break",
        "column_break",
    ),
)
def test_audited_baseline_rejects_page_setup_and_break_mutations(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证页码、物理纸张、打印策略与手动分页入基线。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：新增 page setup 字段组或 row/column break。
    输出返回值：
        无；变体摘要必须与对照摘要不同。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "workbook_view_visibility",
        "workbook_view_minimized",
        "workbook_view_chrome",
        "workbook_view_sheet_indices",
        "workbook_view_date_grouping",
        "workbook_view_geometry",
    ),
)
def test_audited_baseline_rejects_workbook_view_mutations(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证工作簿窗口可见状态及已审定几何第三值被拒绝。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：可见性、最小化、滚动条/标签、sheet 索引、
            日期分组或窗口几何之一。
    输出返回值：
        无；变体摘要必须与对照摘要不同。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "sheet_view_window_protection",
        "sheet_view_ruler",
        "sheet_view_whitespace",
        "sheet_view_color_id",
        "sheet_view_zoom_to_fit",
        "sheet_view_workbook_id",
    ),
)
def test_audited_baseline_rejects_remaining_sheet_view_mutations(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证剩余 sheet view 可见字段与窗口绑定进入基线。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：windowProtection、ruler、whitespace、colorId、
            zoomToFit 或 workbookViewId 之一。
    输出返回值：
        无；变体摘要必须与对照摘要不同。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "sheet_format_custom_height",
        "sheet_format_row_outline",
        "sheet_format_column_outline",
    ),
)
def test_audited_baseline_rejects_remaining_sheet_format_mutations(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证 sheet format 的 customHeight 与行/列 outline level 进入基线。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：customHeight、outlineLevelRow 或 outlineLevelCol。
    输出返回值：
        无；变体摘要必须与对照摘要不同。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "sheet_properties_outline",
        "sheet_properties_code_name",
        "sheet_properties_format_calculation",
        "sheet_properties_filter_mode",
        "sheet_properties_published",
        "sheet_properties_sync",
        "sheet_properties_transition",
    ),
)
def test_audited_baseline_rejects_sheet_properties_mutations(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证 outlinePr 与 worksheet properties 全字段进入基线。

    输入参数：
        tmp_path：pytest 隔离目录。
        mutation：outline、codeName、格式计算、filter mode、published、
            sync 或 transition 字段组。
    输出返回值：
        无；变体摘要必须与对照摘要不同。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "calculation_calc_id",
        "calculation_mode",
        "calculation_full_on_load",
        "calculation_reference_mode",
        "calculation_iterate",
        "calculation_iterate_count",
        "calculation_iterate_delta",
        "calculation_full_precision",
        "calculation_completed",
        "calculation_on_save",
        "calculation_concurrent",
        "calculation_concurrent_count",
        "calculation_force_full",
    ),
)
def test_audited_baseline_rejects_each_workbook_calculation_field(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证 workbook calculation 序列化全字段进入基线。

    输入参数：
        tmp_path：承载同一真实 KFC gold 的对照与变体。
        mutation：``CalcProperties.__attrs__`` 中当前字段的
            第三语义值。
    输出返回值：
        无；13 个已审定字段任一变化都必须改变摘要。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "cell_protection_locked",
        "cell_protection_hidden",
        "font_charset",
        "font_family",
        "font_scheme",
    ),
)
def test_audited_baseline_rejects_cell_protection_and_font_metadata(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证单元格 protection 与剩余字体字段进入基线。

    输入参数：
        tmp_path：承载同一真实 KFC gold 的对照与变体。
        mutation：locked/hidden 或 charset/family/scheme 的第三值。
    输出返回值：
        无；任一变体都必须改变语义摘要。
    """

    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "print_options_gridlines_set",
        "page_setup_relationship_id",
    ),
)
def test_audited_baseline_rejects_remaining_print_and_page_fields(
    tmp_path: Path,
    mutation: str,
) -> None:
    """验证 PrintOptions/PageSetup 序列化字段闭集无漏项。

    输入参数：
        tmp_path：承载同一真实 KFC gold 的对照与变体。
        mutation：``gridLinesSet`` 或 ``PrintPageSetup.id`` 第三值。
    输出返回值：
        无；``gridLinesSet`` 第三值改变摘要；无已审定
        printer-settings part 的 ``r:id`` 以固定脱敏错误失败关闭。
    """

    if mutation == "page_setup_relationship_id":
        with pytest.raises(
            PipelineImplicitArtifactEvidenceError,
            match="^TYPED_OBSERVATION_INVALID$",
        ):
            _derive_kfc_semantic_variant_digests(tmp_path, mutation)
        return
    control_digest, mutated_digest = _derive_kfc_semantic_variant_digests(
        tmp_path,
        mutation,
    )

    assert mutated_digest != control_digest


@pytest.mark.parametrize(
    "object_name",
    ("print_options", "page_setup"),
)
def test_audited_page_types_fail_closed_on_dependency_field_drift(
    object_name: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 openpyxl 新增序列化字段时不会被静默忽略。

    输入参数：
        object_name：真实 gold 上的 PrintOptions 或 PrintPageSetup。
        monkeypatch：仅在当前进程模拟依赖版本新增字段。
    输出返回值：
        无；全字段投影在依赖闭集漂移时抛出固定
        ``ValueError``，而非继续生成不完整摘要。
    """

    import openpyxl

    workbook = openpyxl.load_workbook(
        _fixed_revision_gold_fixture() / "KFC_Monthly_Data.xlsx"
    )
    worksheet = workbook.active
    instance = getattr(worksheet, object_name)
    original_fields = tuple(instance.__attrs__)
    monkeypatch.setattr(
        type(instance),
        "__attrs__",
        original_fields + ("privateUnknown",),
    )
    try:
        with pytest.raises(ValueError, match="字段闭集漂移"):
            hide_na_rows_bridge._normalize_page_setup(worksheet)
    finally:
        workbook.close()


def test_production_capture_builds_typed_gold_observation() -> None:
    """验证 generic capture 会自行生产 typed 证据并通过正式评价。

    输入参数：无；使用固定 revision 的真实五文件 gold。
    输出返回值：
        无；不注入 typed fake，production source 的返回值
        必须是可获得满分的 ``HideNARowsObservation``。
    """

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        HIDE_NA_ROWS_TASK_ID,
        _DirectoryController(_fixed_revision_gold_fixture()),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )

    assert isinstance(observation, HideNARowsObservation)
    result = evaluate_hide_na_rows(observation)
    assert result.passed is True
    assert result.score == 1.0
    assert result.expected_document_count == 5


def test_audited_builder_locks_equal_input_and_gold_baselines() -> None:
    """验证受控 builder 从真实固定资产重算五对相等摘要。

    输入参数：无；使用 Lee 固定 revision 的 input/gold 闭集。
    输出返回值：
        无；每对 input/gold 在排除任务允许的 hidden 状态后
        必须相等，且精确命中 production 锁定摘要。
    """

    input_fixture = _fixed_revision_input_fixture()
    gold_fixture = _fixed_revision_gold_fixture()
    observed: dict[str, str] = {}
    for name in sorted(PINNED_HIDE_NA_ROWS_BASELINE_SHA256):
        input_digest = derive_hide_na_rows_baseline_sha256(
            (input_fixture / name).read_bytes()
        )
        gold_digest = derive_hide_na_rows_baseline_sha256(
            (gold_fixture / name).read_bytes()
        )
        assert input_digest == gold_digest
        observed[name] = input_digest

    assert observed == dict(PINNED_HIDE_NA_ROWS_BASELINE_SHA256)


@pytest.mark.parametrize(
    "visible_mutation",
    (
        "conditional_formatting",
        "view_gridlines",
        "view_zoom",
        "view_headers",
        "view_zeros",
        "view_rtl",
        "font_shadow",
        "font_outline",
        "font_vertical_alignment",
        "workbook_theme",
        "tab_color",
        "auto_filter",
        "page_orientation",
    ),
)
def test_audited_baseline_rejects_each_visible_semantic_third_value(
    tmp_path: Path,
    visible_mutation: str,
) -> None:
    """验证 input/gold 已审定默认差异外的第三值都使基线失效。

    输入参数：
        tmp_path：pytest 隔离目录。
        visible_mutation：条件格式、sheet view、字体、theme、
            tab color、auto-filter 或页面方向中的一项。
    输出返回值：
        无；无操作的 openpyxl round-trip 仍命中锁定摘要，
        而任一列出的第三值都产生不同摘要并在正式评价中
        只计一个 ``BASELINE_CONTENT_CHANGED``。
    """

    import openpyxl
    from openpyxl.formatting.rule import FormulaRule

    control_fixture = tmp_path / "control"
    mutated_fixture = tmp_path / "mutated"
    shutil.copytree(_fixed_revision_gold_fixture(), control_fixture)
    shutil.copytree(_fixed_revision_gold_fixture(), mutated_fixture)
    name = "KFC_Monthly_Data.xlsx"

    control_path = control_fixture / name
    control_workbook = openpyxl.load_workbook(control_path)
    control_workbook.save(control_path)
    control_workbook.close()
    assert (
        derive_hide_na_rows_baseline_sha256(control_path.read_bytes())
        == PINNED_HIDE_NA_ROWS_BASELINE_SHA256[name]
    )

    target_path = mutated_fixture / name
    workbook = openpyxl.load_workbook(target_path)
    worksheet = workbook.active
    if visible_mutation == "conditional_formatting":
        worksheet.conditional_formatting.add(
            "A1",
            FormulaRule(formula=["1=1"]),
        )
    elif visible_mutation == "view_gridlines":
        worksheet.sheet_view.showGridLines = False
    elif visible_mutation == "view_zoom":
        worksheet.sheet_view.zoomScale = 125
    elif visible_mutation == "view_headers":
        worksheet.sheet_view.showRowColHeaders = False
    elif visible_mutation == "view_zeros":
        worksheet.sheet_view.showZeros = False
    elif visible_mutation == "view_rtl":
        worksheet.sheet_view.rightToLeft = True
    elif visible_mutation in {
        "font_shadow",
        "font_outline",
        "font_vertical_alignment",
    }:
        font = copy(worksheet["A1"].font)
        if visible_mutation == "font_shadow":
            font.shadow = True
        elif visible_mutation == "font_outline":
            font.outline = True
        else:
            font.vertAlign = "superscript"
        worksheet["A1"].font = font
    elif visible_mutation == "workbook_theme":
        input_workbook = openpyxl.load_workbook(_fixed_revision_input_fixture() / name)
        verified_theme = input_workbook.loaded_theme
        input_workbook.close()
        assert isinstance(verified_theme, bytes)
        modified_theme = verified_theme.replace(
            b'lastClr="000000"',
            b'lastClr="FF0000"',
            1,
        )
        assert modified_theme != verified_theme
        workbook.loaded_theme = modified_theme
    elif visible_mutation == "tab_color":
        worksheet.sheet_properties.tabColor = "FFFF00FF"
    elif visible_mutation == "auto_filter":
        worksheet.auto_filter.ref = "A1:A2"
    elif visible_mutation == "page_orientation":
        worksheet.page_setup.orientation = "landscape"
    else:
        pytest.fail("unregistered visible mutation")
    workbook.save(target_path)
    workbook.close()

    mutated_digest = derive_hide_na_rows_baseline_sha256(target_path.read_bytes())
    assert mutated_digest != PINNED_HIDE_NA_ROWS_BASELINE_SHA256[name]

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        HIDE_NA_ROWS_TASK_ID,
        _DirectoryController(mutated_fixture),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    assert isinstance(observation, HideNARowsObservation)
    result = evaluate_hide_na_rows(observation)
    assert result.passed is False
    assert result.score == 0.8
    assert result.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert result.mutated_document_count == 1


def test_pinned_input_is_baseline_authority_but_noop_fails() -> None:
    """验证原 input 内容基线有效，但未隐藏 N/A 行不通过。

    输入参数：无；使用固定 revision 的真实五文件 input。
    输出返回值：
        无；五个工作簿都必须命中内容基线，但因缺失
        固定八个隐藏行而得到 FAIL。
    """

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        HIDE_NA_ROWS_TASK_ID,
        _DirectoryController(_fixed_revision_input_fixture()),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )

    assert isinstance(observation, HideNARowsObservation)
    result = evaluate_hide_na_rows(observation)
    assert result.passed is False
    assert result.reason_codes == ("MISSING_HIDDEN_ROW",)
    assert result.missing_hidden_row_count == 8
    assert result.mutated_document_count == 0


def test_visible_column_width_mutation_fails_baseline(
    tmp_path: Path,
) -> None:
    """验证非任务所需的可见列宽改动会破坏语义基线。

    输入参数：
        tmp_path：pytest 隔离目录，用于修改 gold 副本。
    输出返回值：
        无；隐藏行仍正确，但评价必须以
        ``BASELINE_CONTENT_CHANGED`` 失败。
    """

    import openpyxl

    fixture = _copy_fixed_revision_gold_fixture(tmp_path)
    target = fixture / "KFC_Monthly_Data.xlsx"
    workbook = openpyxl.load_workbook(target)
    workbook.active.column_dimensions["A"].width = 21.0
    workbook.save(target)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        HIDE_NA_ROWS_TASK_ID,
        _DirectoryController(fixture),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )

    assert isinstance(observation, HideNARowsObservation)
    result = evaluate_hide_na_rows(observation)
    assert result.passed is False
    assert result.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert result.mutated_document_count == 1


def test_missing_and_extra_files_are_task_failures_not_evaluator_errors(
    tmp_path: Path,
) -> None:
    """验证精确五文件闭集的缺失与额外成员都计为 FAIL。

    输入参数：
        tmp_path：pytest 隔离目录，用于构造副作用。
    输出返回值：
        无；typed bridge 不应因 Agent 少产或多产文件而
        把本应评分的结果误分类为 evaluator ERROR。
    """

    fixture = _copy_fixed_revision_gold_fixture(tmp_path)
    (fixture / "Mixue_Monthly_Data.xlsx").unlink()
    (fixture / "private-extra.txt").write_bytes(b"not an xlsx")

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        HIDE_NA_ROWS_TASK_ID,
        _DirectoryController(fixture),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )

    assert isinstance(observation, HideNARowsObservation)
    result = evaluate_hide_na_rows(observation)
    assert result.passed is False
    assert result.reason_codes == (
        "MISSING_DOCUMENT",
        "UNEXPECTED_DOCUMENT",
    )
    assert result.evaluated_document_count == 4
    assert result.unexpected_document_count == 1
    assert "private-extra" not in repr(observation)


def test_corrupt_expected_xlsx_is_redacted_task_failure(
    tmp_path: Path,
) -> None:
    """验证恶意或损坏的期望 XLSX 不进入 Office parser 可信域。

    输入参数：
        tmp_path：pytest 隔离目录，用于破坏一份 gold 副本。
    输出返回值：
        无；该文件以固定分母中的零分占位表示，
        评价结果只含固定 reason code 和计数。
    """

    fixture = _copy_fixed_revision_gold_fixture(tmp_path)
    target = fixture / "KFC_Monthly_Data.xlsx"
    target.write_bytes(b"PRIVATE-NOT-AN-XLSX")

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        HIDE_NA_ROWS_TASK_ID,
        _DirectoryController(fixture),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )

    assert isinstance(observation, HideNARowsObservation)
    result = evaluate_hide_na_rows(observation)
    assert result.passed is False
    assert result.reason_codes == (
        "MISSING_HIDDEN_ROW",
        "BASELINE_CONTENT_CHANGED",
    )
    assert result.missing_hidden_row_count == 2
    assert result.mutated_document_count == 1
    rendered = repr(result)
    assert "KFC_Monthly_Data.xlsx" not in rendered
    assert "PRIVATE-NOT-AN-XLSX" not in rendered


def test_hide_na_rows_parser_cleanup_failure_is_fixed_redacted_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 Excel-008 parser sandbox 清理 OSError 不泄露路径或底层文本。

    输入参数：
        tmp_path：创建由 pytest 回收的合成 sandbox。
        monkeypatch：使 parser 返回合法小投影，cleanup 抛出
            包含私密哨兵的 ``OSError``。
    输出返回值：
        无；入口仅抛固定 ``PARSER_SANDBOX_FAILED`` 内部错误，
        异常表示中不含绝对路径或哨兵。
    """

    sandbox_path = tmp_path / "PRIVATE-EXCEL-CLEANUP-PATH"
    sandbox_path.mkdir()

    class _FailingTemporaryDirectory:
        """提供合法目录身份并在 cleanup 阶段注入故障。"""

        name = os.fspath(sandbox_path)

        def cleanup(self) -> None:
            """抛出含私密文本的合成清理异常。

            输入参数：无。
            输出返回值：不返回，始终抛出 ``OSError``。
            """

            raise OSError("PRIVATE-EXCEL-CLEANUP-SENTINEL")

    monkeypatch.setattr(
        hide_na_rows_bridge.tempfile,
        "TemporaryDirectory",
        lambda prefix: _FailingTemporaryDirectory(),
    )
    monkeypatch.setattr(
        hide_na_rows_bridge,
        "_run_parser_process",
        lambda content, **kwargs: ((), "0" * 64),
    )

    with pytest.raises(
        hide_na_rows_bridge._WorkbookParserInternalError,
        match="^PARSER_SANDBOX_FAILED$",
    ) as captured:
        hide_na_rows_bridge._parse_xlsx_controlled(b"bounded")

    assert "PRIVATE" not in repr(captured.value)
    assert os.fspath(sandbox_path) not in repr(captured.value)


def test_hide_na_rows_parser_wall_timeout_reaps_worker(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 Excel-008 parser wall-clock 超时后子进程已回收。

    输入参数：
        monkeypatch：把当前任务的 wall 上限收紧到 1 微秒。
    输出返回值：
        无；真实 gold 因资源拒绝按固定五文档分母零分，
        且不存在活动的 Excel-008 parser 子进程。
    """

    monkeypatch.setattr(
        hide_na_rows_bridge,
        "_PARSER_WALL_TIMEOUT_SECONDS",
        0.000001,
    )
    observation = PipelineImplicitArtifactEvidenceSource().capture(
        HIDE_NA_ROWS_TASK_ID,
        _DirectoryController(_fixed_revision_gold_fixture()),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_hide_na_rows(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.0
    assert evaluation.expected_document_count == 5
    assert all(
        child.name != "paraguibench-hide-na-rows-parser"
        for child in multiprocessing.active_children()
    )


def test_hide_na_rows_parser_rss_budget_reaps_worker(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 Excel-008 parser RSS 软监控超限后失败关闭。

    输入参数：
        monkeypatch：将资源上限收紧到 1 字节，并固定
            父进程 RSS 采样为 2 字节。
    输出返回值：
        无；未信任 parser 不得返回部分成功，且必须完整回收。
    """

    monkeypatch.setattr(hide_na_rows_bridge, "_PARSER_RSS_LIMIT_BYTES", 1)
    monkeypatch.setattr(
        hide_na_rows_bridge,
        "_PARSER_POLL_INTERVAL_SECONDS",
        0.0001,
    )
    monkeypatch.setattr(
        hide_na_rows_bridge,
        "_parser_resident_bytes",
        lambda pid: 2,
    )
    observation = PipelineImplicitArtifactEvidenceSource().capture(
        HIDE_NA_ROWS_TASK_ID,
        _DirectoryController(_fixed_revision_gold_fixture()),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_hide_na_rows(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.0
    assert all(
        child.name != "paraguibench-hide-na-rows-parser"
        for child in multiprocessing.active_children()
    )


def test_hide_na_rows_parser_cpu_limit_stops_busy_materializer(
    tmp_path: Path,
) -> None:
    """验证 Excel-008 worker 的 CPU 硬限额独立于 wall 监控生效。

    输入参数：
        tmp_path：为真实 spawn worker 提供只读空工作目录。
    输出返回值：
        无；无限纯 CPU materializer 必须在 5 秒内被 OS 信号终止，
        且不留活动子进程。
    """

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    sandbox_cwd = _create_readonly_hide_na_rows_cwd(
        tmp_path,
        "hide-na-rows-parser-cpu-cwd",
    )
    process = context.Process(
        target=hide_na_rows_bridge._workbook_parse_worker,
        args=(
            sender,
            b"bounded",
            _consume_hide_na_rows_cpu,
            os.fspath(sandbox_cwd),
        ),
        daemon=True,
        name="paraguibench-hide-na-rows-parser-cpu-test",
    )
    process.start()
    sender.close()
    process.join(timeout=5.0)
    receiver.close()
    if process.is_alive():
        process.terminate()
        process.join(timeout=1.0)
        pytest.fail("Excel-008 parser CPU hard limit did not terminate worker")

    assert isinstance(process.exitcode, int)
    assert process.exitcode < 0
    process.close()
    sandbox_cwd.chmod(stat.S_IRWXU)
    assert all(
        child.name != "paraguibench-hide-na-rows-parser-cpu-test"
        for child in multiprocessing.active_children()
    )


@pytest.mark.parametrize(
    "operation",
    (
        "open-write",
        "os-open-write",
        "truncate",
        "delete",
        "rename",
        "replace",
        "mkdir",
        "rmdir",
        "chmod",
        "link",
        "symlink",
        "mkfifo",
        "mknod",
        "utime",
        "chown",
    ),
)
def test_hide_na_rows_parser_rejects_python_filesystem_mutation(
    tmp_path: Path,
    operation: str,
) -> None:
    """逐项验证 Excel-008 真实 worker 的 Python 文件系统禁写边界。

    输入参数：
        tmp_path：只供本测试创建可破坏合成树。
        operation：写打开、截断、删除、重命名/替换、建删目录、
            改权限/时间/属主、硬/软链接、FIFO 或 node 之一。
    输出返回值：
        无；worker 只返回固定 ``rejected``，测试树前后完全一致。
    """

    protected_root = tmp_path / "protected"
    protected_root.mkdir()
    source = protected_root / "source.bin"
    source.write_bytes(b"PRIVATE-SYNTHETIC-CONTENT")
    source.chmod(0o640)
    empty_directory = protected_root / "empty"
    empty_directory.mkdir()
    destination = protected_root / "destination"
    operation_source = empty_directory if operation == "rmdir" else source
    before = _snapshot_hide_na_rows_test_tree(protected_root)
    sandbox_cwd = _create_readonly_hide_na_rows_cwd(
        tmp_path,
        "hide-na-rows-parser-write-cwd",
    )
    payload = json.dumps(
        {
            "operation": operation,
            "source": os.fspath(operation_source),
            "destination": os.fspath(destination),
        }
    ).encode("utf-8")

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    process = context.Process(
        target=hide_na_rows_bridge._workbook_parse_worker,
        args=(
            sender,
            payload,
            _attempt_hide_na_rows_filesystem_mutation,
            os.fspath(sandbox_cwd),
        ),
        daemon=True,
        name="paraguibench-hide-na-rows-parser-filesystem-test",
    )
    process.start()
    sender.close()
    assert receiver.poll(5.0)
    message = hide_na_rows_bridge._decode_parser_message(receiver.recv_bytes(64 * 1024))
    receiver.close()
    process.join(timeout=5.0)

    assert process.is_alive() is False
    assert process.exitcode == 0
    process.close()
    assert message == ("rejected",)
    assert _snapshot_hide_na_rows_test_tree(protected_root) == before
    sandbox_cwd.chmod(stat.S_IRWXU)


def test_hide_na_rows_parser_ipc_is_bounded_json_not_pickle(
    tmp_path: Path,
) -> None:
    """验证 Excel-008 不可信 worker 只向父进程发送有界 JSON。

    输入参数：
        tmp_path：为真实 worker 提供只读空工作目录。
    输出返回值：
        无；原始 frame 不超过 64 KiB，可严格解码为 JSON 列表，
        再经父进程结构验证，全链不使用 pickle。
    """

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    sandbox_cwd = _create_readonly_hide_na_rows_cwd(
        tmp_path,
        "hide-na-rows-parser-json-cwd",
    )
    process = context.Process(
        target=hide_na_rows_bridge._workbook_parse_worker,
        args=(
            sender,
            b"bounded",
            _return_hide_na_rows_parser_projection,
            os.fspath(sandbox_cwd),
        ),
        daemon=True,
        name="paraguibench-hide-na-rows-parser-json-test",
    )
    process.start()
    sender.close()
    assert receiver.poll(5.0)
    raw_message = receiver.recv_bytes(64 * 1024)
    receiver.close()
    process.join(timeout=5.0)

    assert process.is_alive() is False
    assert process.exitcode == 0
    process.close()
    sandbox_cwd.chmod(stat.S_IRWXU)
    assert len(raw_message) <= 64 * 1024
    assert json.loads(raw_message.decode("utf-8", errors="strict")) == [
        "ok",
        [8, 10],
        "0" * 64,
    ]
    decoded = hide_na_rows_bridge._decode_parser_message(raw_message)
    assert hide_na_rows_bridge._validate_parser_message(decoded) == (
        (8, 10),
        "0" * 64,
    )


def test_hide_na_rows_worker_holds_exit_until_parent_rss_ack(
    tmp_path: Path,
) -> None:
    """验证快速 JSON frame 发送后 worker 仍留存到父进程确认。

    输入参数：
        tmp_path：为真实 spawn worker 提供只读空工作目录。
    输出返回值：
        无；父进程收到 frame 后、发出 RSS 采样确认前，
        worker 必须仍存活，从而消除 macOS 进程退出与采样的竞态。
    """

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    parent_rss_ack = context.Event()
    sandbox_cwd = _create_readonly_hide_na_rows_cwd(
        tmp_path,
        "hide-na-rows-parser-ack-cwd",
    )
    process = context.Process(
        target=hide_na_rows_bridge._workbook_parse_worker,
        args=(
            sender,
            b"bounded",
            _return_hide_na_rows_parser_projection,
            os.fspath(sandbox_cwd),
            parent_rss_ack,
        ),
        daemon=True,
        name="paraguibench-hide-na-rows-parser-ack-test",
    )
    process.start()
    sender.close()
    assert receiver.poll(5.0)
    raw_message = receiver.recv_bytes(64 * 1024)
    assert process.is_alive() is True
    parent_rss_ack.set()
    receiver.close()
    process.join(timeout=5.0)

    assert process.is_alive() is False
    assert process.exitcode == 0
    process.close()
    sandbox_cwd.chmod(stat.S_IRWXU)
    assert json.loads(raw_message.decode("utf-8", errors="strict"))[0] == "ok"


def test_attempt_runner_persists_only_redacted_excel_counts(
    tmp_path: Path,
) -> None:
    """验证 Excel 文件名、行列、单元格和 Agent final text 不落盘。

    输入参数：
        tmp_path：pytest 隔离目录，承载修改后的 guest 闭集和 RunStore。
    输出返回值：
        无；真实 capture→runtime adapter→AttemptRunner→RunStore
        产生 0.8 的普通 FAIL，持久化仅保留协议、原因码和计数。
    """

    import openpyxl

    fixture_path = tmp_path / "guest" / _TASK_UID
    shutil.copytree(_fixed_revision_gold_fixture(), fixture_path)
    target_path = fixture_path / "KFC_Monthly_Data.xlsx"
    workbook = openpyxl.load_workbook(target_path)
    workbook.active["F4"] = "PRIVATE-CELL-VALUE-SENTINEL"
    workbook.save(target_path)
    workbook.close()
    extra_workbook = openpyxl.Workbook()
    extra_workbook.active["A1"] = "PRIVATE-FILE-CONTENT-SENTINEL"
    extra_workbook.save(fixture_path / "PRIVATE-PATH-SENTINEL.xlsx")
    extra_workbook.close()

    prepared = _prepared_excel_task()
    store_root = tmp_path / "runstore"
    store = RunStore(store_root)
    version_vector = replace(
        synthetic_run_version_vector(),
        evaluation_protocol=HIDE_NA_ROWS_PROTOCOL_ID,
    )
    store.start_run(
        run_id="run-excel008-privacy",
        run_record={"environment_id": "synthetic-osworld"},
        version_vector=version_vector,
    )
    attempt = store.start_attempt(
        run_id="run-excel008-privacy",
        task_id=HIDE_NA_ROWS_TASK_ID,
        attempt_id="attempt-001",
        task_record=prepared.audit_metadata,
    )
    environment = _ExcelObservationEnvironment(fixture_path)

    result = AttemptRunner(store).run(
        attempt=attempt,
        prepared_task=prepared,
        environment=environment,
        agent=_SensitiveExcelAgent(),
        evaluator=PipelineImplicitTaskEvaluator(
            task_id=HIDE_NA_ROWS_TASK_ID,
            evaluation_protocol=HIDE_NA_ROWS_PROTOCOL_ID,
        ),
    )

    assert result.execution_outcome is ExecutionOutcome.SUCCEEDED
    assert result.evaluation_outcome is EvaluationOutcome.FAILED
    assert result.score == pytest.approx(0.8, abs=1e-4)
    assert environment.closed is True
    inspection = store.inspect_attempt(
        run_id="run-excel008-privacy",
        task_id=HIDE_NA_ROWS_TASK_ID,
        attempt_id="attempt-001",
    )
    assert inspection.execution_outcome is ExecutionOutcome.SUCCEEDED
    assert inspection.evaluation_outcome is EvaluationOutcome.FAILED
    assert inspection.score == pytest.approx(0.8, abs=1e-4)
    assert inspection.failure_stage is AttemptFailureStage.NOT_FAILED
    assert inspection.provenance_status is RunProvenanceStatus.VERSIONED
    assert inspection.version_vector == version_vector
    persisted = b"\n".join(
        path.read_bytes() for path in store_root.rglob("*") if path.is_file()
    )
    for sentinel in (
        b"PRIVATE-CELL-VALUE-SENTINEL",
        b"PRIVATE-PATH-SENTINEL",
        b"PRIVATE-FILE-CONTENT-SENTINEL",
        b"PRIVATE-FINAL-TEXT-SENTINEL",
        b"KFC_Monthly_Data.xlsx",
        b"F4",
    ):
        assert sentinel not in persisted
    for safe_field in (
        HIDE_NA_ROWS_PROTOCOL_ID.encode("utf-8"),
        b"reason_codes",
        b"unexpected_document_count",
        b"mutated_document_count",
        b"matched_hidden_row_count",
    ):
        assert safe_field in persisted

"""SearchAndWrite-008 production artifact capture 到正式评价器的纵向测试。"""

from __future__ import annotations

from copy import copy
import hashlib
from io import BytesIO
import json
import multiprocessing
import os
from pathlib import Path, PurePosixPath
import shutil
import stat
import sys
from typing import Any
import zipfile

import pytest

from paraguibench.agents import AgentRunResult
from paraguibench.benchmark import PreparedTask
from paraguibench.evaluation.pipeline_implicit import (
    SEARCHWRITE_XLSX_PROTOCOL_ID,
    SEARCHWRITE_XLSX_TASK_ID,
    evaluate_searchwrite_xlsx,
)
from paraguibench.integrations.pipeline_implicit import (
    PipelineImplicitArtifactEvidenceSource,
)
from paraguibench.integrations.pipeline_implicit import searchwrite_bridge
from paraguibench.integrations.pipeline_implicit.searchwrite_bridge import (
    _WorkbookParserInternalError,
    _decode_parser_message,
    _normalize_cell_visible_style,
    _preflight_xlsx_bytes,
    _validate_parser_message,
    _workbook_parse_worker,
)
from paraguibench.integrations.pipeline_implicit.artifact_evidence import (
    PipelineImplicitArtifactEvidenceError,
)
from paraguibench.runstore import EvaluationOutcome, RunStore
from paraguibench.runtime.attempt_runner import AttemptRunner
from paraguibench.runtime.evaluators import PipelineImplicitTaskEvaluator
from tests.runstore._audit import (
    synthetic_run_version_vector,
    synthetic_task_audit,
)


_FIXTURE_ENVIRONMENT_VARIABLE = "PARAGUI_SEARCHWRITE008_FIXTURE_ROOT"
_TASK_UID = "65a4848d-b4b2-4173-8308-a0213fdafbd0"
_GUEST_SHARED_DIR = "/guest-home/shared"
_PINNED_INPUT_SHA256 = {
    "UK_Universities_Group1.xlsx": (
        "df08dc5e24d04a9587c21154b363511e01bc2ec18e9411d179e29e9231188e27"
    ),
    "UK_Universities_Group2.xlsx": (
        "7936c66869e26be9e787e703e801c74b7034afd22f934ca3b166a3d4b021caaa"
    ),
}
_PINNED_GOLD_SHA256 = {
    "UK_Universities_Group1.xlsx": (
        "0170c5dab6a6062c610517b297708ad496a8bfa53699915ad6c3ff3948bf81cd"
    ),
    "UK_Universities_Group2.xlsx": (
        "b19a72eb28ad9a55ed956247dd8fb97f59ec5ede751ece25ac963614631ef257"
    ),
}


def _raise_private_unknown_parser_error(
    content: bytes,
    *,
    coordinates: tuple[str, ...],
    expected_baseline_sha256: str,
) -> tuple[tuple[tuple[str, object], ...], bool]:
    """在子进程边界测试中触发未知 parser 故障。

    输入参数：
        content/coordinates/expected_baseline_sha256：与生产 materializer
            相同的已受控调用签名；测试不消费其内容。
    输出返回值：
        不返回；向子进程 stderr 写入私密哨兵后抛出不在
        已知输入错误闭集内的 ``RuntimeError``。
    """

    del content, coordinates, expected_baseline_sha256
    print("PRIVATE-UNKNOWN-PARSER-TRACEBACK", file=sys.stderr)
    raise RuntimeError("PRIVATE-UNKNOWN-PARSER-VALUE")


def _consume_parser_cpu_until_limited(
    content: bytes,
    *,
    coordinates: tuple[str, ...],
    expected_baseline_sha256: str,
) -> tuple[tuple[tuple[str, object], ...], bool]:
    """持续占用 CPU，用于验证 parser worker 的 OS 硬限额。

    输入参数：
        content/coordinates/expected_baseline_sha256：与生产 materializer
            一致的可拾取签名；测试不读取其内容。
    输出返回值：
        不返回；子进程应由 2 秒 ``RLIMIT_CPU`` 终止。
    """

    del content, coordinates, expected_baseline_sha256
    accumulator = 0
    while True:
        accumulator = (accumulator + 1) % 104729


def _attempt_forbidden_filesystem_mutation(
    content: bytes,
    *,
    coordinates: tuple[str, ...],
    expected_baseline_sha256: str,
) -> tuple[tuple[tuple[str, object], ...], bool]:
    """在真实 parser 子进程中尝试测试指定的文件系统变更。

    输入参数：
        content：只由本测试生成的 JSON，包含操作名称和隔离临时路径。
        coordinates/expected_baseline_sha256：与生产 materializer
            一致的调用参数，本故障注入器不读取。
    输出返回值：
        仅当隔离边界错误地允许操作时返回空投影；正确边界应先抛出
        固定 ``PermissionError``，worker 将其归类为输入拒绝。
    """

    del coordinates, expected_baseline_sha256
    instruction = json.loads(content.decode("utf-8", errors="strict"))
    operation = instruction["operation"]
    source = instruction["source"]
    destination = instruction["destination"]
    if operation == "truncate":
        os.truncate(source, 0)
    elif operation == "delete":
        os.remove(source)
    elif operation == "rename":
        os.rename(source, destination)
    elif operation == "mkdir":
        os.mkdir(destination)
    elif operation == "rmdir":
        os.rmdir(source)
    elif operation == "chmod":
        os.chmod(source, 0o600)
    elif operation == "link":
        os.link(source, destination)
    elif operation == "symlink":
        os.symlink(source, destination)
    elif operation == "open-write":
        with open(source, mode="wb"):
            pass
    elif operation == "os-open-write":
        descriptor = os.open(source, os.O_WRONLY | os.O_TRUNC)
        os.close(descriptor)
    elif operation == "replace":
        os.replace(source, destination)
    elif operation == "mkfifo":
        os.mkfifo(destination, 0o600)
    elif operation == "mknod":
        os.mknod(destination, stat.S_IFREG | 0o600)
    elif operation == "utime":
        os.utime(source, ns=(1, 1))
    elif operation == "chown":
        os.chown(source, os.getuid(), os.getgid())
    else:
        raise AssertionError("未知文件系统故障注入操作")
    return (), True


def _verify_parser_readonly_empty_cwd(
    content: bytes,
    *,
    coordinates: tuple[str, ...],
    expected_baseline_sha256: str,
) -> tuple[tuple[tuple[str, object], ...], bool]:
    """在真实 worker 内验证当前目录身份、空目录和只读权限。

    输入参数：
        content：父测试传入的预期绝对 cwd UTF-8 字节。
        coordinates/expected_baseline_sha256：生产 materializer 兼容参数，
            本测试 seam 不消费。
    输出返回值：
        三项条件均成立时返回空投影；否则抛出已知 ``ValueError``，
        且不把实际路径送回父进程。
    """

    del coordinates, expected_baseline_sha256
    expected_cwd = content.decode("utf-8", errors="strict")
    actual_cwd = os.getcwd()
    mode = stat.S_IMODE(os.stat(actual_cwd, follow_symlinks=False).st_mode)
    if (
        actual_cwd != expected_cwd
        or os.listdir(actual_cwd)
        or mode & (stat.S_IWUSR | stat.S_IWGRP | stat.S_IWOTH)
    ):
        raise ValueError("parser cwd 未正确隔离")
    return (), True


def _snapshot_test_tree(root: Path) -> tuple[tuple[object, ...], ...]:
    """生成隔离测试树的内容、类型和权限快照。

    输入参数：
        root：仅由当前 pytest ``tmp_path`` 创建的合成目录。
    输出返回值：
        按相对路径排序的不可变元组；常规文件仅记录 SHA-256，
        符号链接仅记录目标，避免测试失败信息回显原始内容。
    """

    records: list[tuple[object, ...]] = []
    for path in sorted(root.rglob("*")):
        metadata = path.lstat()
        relative_path = path.relative_to(root).as_posix()
        mode = stat.S_IMODE(metadata.st_mode)
        ownership_and_time = (
            metadata.st_uid,
            metadata.st_gid,
            metadata.st_mtime_ns,
        )
        if path.is_symlink():
            records.append(
                (
                    relative_path,
                    "symlink",
                    mode,
                    ownership_and_time,
                    os.readlink(path),
                )
            )
        elif path.is_file():
            records.append(
                (
                    relative_path,
                    "file",
                    mode,
                    ownership_and_time,
                    hashlib.sha256(path.read_bytes()).hexdigest(),
                )
            )
        elif path.is_dir():
            records.append((relative_path, "directory", mode, ownership_and_time))
        else:
            records.append((relative_path, "other", mode, ownership_and_time))
    return tuple(records)


def _create_readonly_test_parser_cwd(tmp_path: Path, name: str) -> Path:
    """创建真实 worker 测试使用的只读空工作目录。

    输入参数：
        tmp_path：pytest 隔离根目录。
        name：当前测试内唯一、不会持久化的目录名。
    输出返回值：
        已确认空且仅保留 owner 读/执行权限的绝对目录路径。
    """

    sandbox_cwd = tmp_path / name
    sandbox_cwd.mkdir()
    sandbox_cwd.chmod(stat.S_IRUSR | stat.S_IXUSR)
    return sandbox_cwd


class _DirectoryController:
    """把已核验的本地固定 revision 目录暴露为 guest controller seam。

    输入参数：
        root：包含 SearchAndWrite-008 两个工作簿的目录。
    输出返回值：
        无；实例通过 production source 的两个窄接口返回
        稳定 manifest 和单文件字节。
    """

    def __init__(self, root: Path) -> None:
        self._root = root

    def collect_artifact_tree_manifest(
        self,
        guest_directory: str,
        **limits: Any,
    ) -> tuple[tuple[str, int, str], ...]:
        """返回按 UTF-8 字节序排列的完整常规文件树。

        输入参数：
            guest_directory：production source 冻结的 guest shared 根。
            limits：source 下发的资源与超时上限。
        输出返回值：
            ``(relative_path, size, sha256)`` 的有序不可变 tuple。
        """

        assert guest_directory == _GUEST_SHARED_DIR
        assert limits
        records = []
        for path in self._root.rglob("*"):
            if not path.is_file():
                continue
            payload = path.read_bytes()
            records.append(
                (
                    path.relative_to(self._root).as_posix(),
                    len(payload),
                    hashlib.sha256(payload).hexdigest(),
                )
            )
        return tuple(sorted(records, key=lambda item: item[0].encode("utf-8")))

    def collect_file_bytes(
        self,
        guest_path: str,
        **limits: Any,
    ) -> bytes:
        """按 production source 给出的 guest 路径返回真实字节。

        输入参数：
            guest_path：必须位于测试冻结的合成 guest 根下。
            limits：source 下发的单文件资源与超时上限。
        输出返回值：
            对应固定 revision 成员的原始字节。
        """

        assert limits
        guest_root = PurePosixPath(_GUEST_SHARED_DIR)
        relative_path = PurePosixPath(guest_path).relative_to(guest_root)
        return self._root.joinpath(*relative_path.parts).read_bytes()


class _SearchWriteObservationEnvironment:
    """在 AttemptRunner 生命周期中按需捕获真实 XLSX 观测。

    输入参数：
        root：当前测试 Attempt 的 guest shared 工作簿闭集。
    输出返回值：
        无；对象通过 runtime 窄 seam 返回 production capture 观测。
    """

    def __init__(self, root: Path) -> None:
        self._controller = _DirectoryController(root)
        self.closed = False

    def start(self) -> None:
        """启动无外部资源的合成环境。

        输入参数：无。
        输出返回值：无。
        """

    def prepare(self, task: dict[str, Any]) -> None:
        """验证 AttemptRunner 传入的可信任务身份。

        输入参数：
            task：PreparedTask 的 trusted projection。
        输出返回值：无；身份漂移时断言失败。
        """

        assert task["task_id"] == SEARCHWRITE_XLSX_TASK_ID

    def close(self) -> None:
        """标记 Attempt 环境已清理。

        输入参数：无。
        输出返回值：无。
        """

        self.closed = True

    def pipeline_implicit_observation(
        self,
        task_id: str,
        protocol_id: str,
    ) -> object:
        """从仍存活的 guest 文件闭集生成强类型观测。

        输入参数：
            task_id/protocol_id：runtime adapter 固定的任务与协议。
        输出返回值：
            manifest—nofollow—manifest 生产 capture 的 SearchWrite observation。
        """

        assert task_id == SEARCHWRITE_XLSX_TASK_ID
        assert protocol_id == SEARCHWRITE_XLSX_PROTOCOL_ID
        return PipelineImplicitArtifactEvidenceSource().capture(
            task_id,
            self._controller,
            guest_shared_dir=_GUEST_SHARED_DIR,
        )


class _SensitiveSearchWriteAgent:
    """返回不得参与 SearchWrite 评价或持久化的哨兵文本。"""

    def run(
        self,
        task_view: dict[str, Any],
        environment: object,
    ) -> AgentRunResult:
        """返回一步结束且含私密 final text 的合法 Agent 结果。

        输入参数：
            task_view：不含 gold 的 Agent 视图。
            environment：仍存活的工作簿环境；Agent fake 不读取。
        输出返回值：含固定哨兵、一步和结束原因的结果。
        """

        del environment
        assert task_view["task_id"] == SEARCHWRITE_XLSX_TASK_ID
        return AgentRunResult(
            final_output="PRIVATE-FINAL-TEXT-SENTINEL",
            step_count=1,
            termination="finished",
        )


def _fixed_revision_fixture(role: str) -> Path:
    """返回显式环境变量下的 input 或 gold 固定树。

    输入参数：
        role：``benchmark_dataset`` 或 ``answer_files``。
    输出返回值：
        Lee ``13bf942d…`` 下 SearchAndWrite-008 的两文件目录。
    """

    raw_path = os.environ.get(_FIXTURE_ENVIRONMENT_VARIABLE)
    if raw_path is None:
        pytest.skip(
            f"{_FIXTURE_ENVIRONMENT_VARIABLE} is required for download-only fixture"
        )
    fixture_path = Path(raw_path) / role / _TASK_UID
    if not fixture_path.is_dir():
        pytest.fail("SearchAndWrite-008 fixed-revision fixture is unavailable")
    return fixture_path


def _prepared_searchwrite_task() -> PreparedTask:
    """从 canonical JSON 构造 RunStore 测试使用的三投影任务。

    输入参数：无。
    输出返回值：
        trusted task 保留正式身份，Agent view 仅含任务和指令，
        audit metadata 仅含 RunStore allowlist 字段的 ``PreparedTask``。
    """

    repo_root = Path(__file__).resolve().parents[2]
    task = json.loads(
        (
            repo_root / "benchmark/tasks/Operation-FileOperate-SearchAndWrite-008.json"
        ).read_text(encoding="utf-8")
    )
    return PreparedTask(
        trusted_task=task,
        agent_task={
            "task_id": SEARCHWRITE_XLSX_TASK_ID,
            "instruction": task["instruction"],
        },
        audit_metadata=synthetic_task_audit(
            SEARCHWRITE_XLSX_TASK_ID,
            task_uid=task["task_uid"],
            task_type=task["task_type"],
            task_source=task["task_source"],
            task_tag=task["task_tag"],
        ),
    )


def _minimal_xlsx_with_cell_elements(cell_count: int) -> bytes:
    """构造仅用于 parser 前结构资源门的最小 OOXML ZIP。

    输入参数：
        cell_count：首个 worksheet XML 中重复的 ``c`` 结构数。
    输出返回值：
        含 Content Types、workbook 和 worksheet 三个 XML 成员的字节。
        它不作为可执行工作簿，只验证 openpyxl 之前的门。
    """

    if not isinstance(cell_count, int) or isinstance(cell_count, bool):
        raise TypeError("cell_count 必须是整数")
    worksheet = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main"><sheetData><row r="1">'
        + b'<c r="A1"/>' * cell_count
        + b"</row></sheetData></worksheet>"
    )
    stream = BytesIO()
    with zipfile.ZipFile(stream, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            b'<Types xmlns="http://schemas.openxmlformats.org/'
            b'package/2006/content-types"/>',
        )
        archive.writestr(
            "xl/workbook.xml",
            b'<workbook xmlns="http://schemas.openxmlformats.org/'
            b'spreadsheetml/2006/main"/>',
        )
        archive.writestr("xl/worksheets/sheet1.xml", worksheet)
    return stream.getvalue()


def test_preflight_stream_rejects_excessive_cell_structure() -> None:
    """验证过多 worksheet 单元格元素在 Office parser 前被拒绝。

    输入参数：无；使用体积很小但含 1025 个 ``c``
        元素的合成 worksheet XML。
    输出返回值：
        无；流式预检必须直接拒绝，不依赖 openpyxl 物化。
    """

    with pytest.raises(ValueError):
        _preflight_xlsx_bytes(_minimal_xlsx_with_cell_elements(1025))


def test_office_parser_runs_outside_parent_process(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 openpyxl 完整物化不在 AttemptRunner 所在进程执行。

    输入参数：
        monkeypatch：只在当前 pytest/父进程替换三方 parser
            边界；spawn 的受控子进程从安装环境重新导入。
    输出返回值：
        无；真实固定 gold 仍必须满分，证明父进程的
        ``load_workbook`` 从未被调用。
    """

    try:
        import openpyxl
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")

    def reject_parent_materialization(*args: object, **kwargs: object) -> object:
        """若三方 parser 在父进程被调用则立即失败。

        输入参数：args/kwargs 为未使用的 openpyxl 调用参数。
        输出返回值：不返回，始终抛出合成断言。
        """

        del args, kwargs
        raise AssertionError("parent Office parser must remain unused")

    monkeypatch.setattr(openpyxl, "load_workbook", reject_parent_materialization)
    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(_fixed_revision_fixture("answer_files")),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is True
    assert evaluation.score == 1.0


def test_parser_wall_timeout_joins_and_closes_spawned_process(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 wall-clock 超时后 parser 子进程已终止并回收。

    输入参数：
        monkeypatch：把本次测试的固定 parser 时限收紧至
            1 微秒，确保 spawn 无法在截止前物化真实工作簿。
    输出返回值：
        无；任务按固定分母失败，且 multiprocessing 不存在
        名为 ``paraguibench-searchwrite-parser`` 的活动子进程。
    """

    monkeypatch.setattr(
        searchwrite_bridge,
        "_PARSER_WALL_TIMEOUT_SECONDS",
        0.000001,
    )
    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(_fixed_revision_fixture("answer_files")),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.0
    assert evaluation.missing_cell_count == 9
    assert all(
        child.name != "paraguibench-searchwrite-parser"
        for child in multiprocessing.active_children()
    )


def test_parser_rss_budget_terminates_and_reaps_worker(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 RSS 超过固定内存预算时失败关闭且无子进程泄漏。

    输入参数：
        monkeypatch：本次测试把 256 MiB 生产预算收紧到
            1 字节，使任何 spawn 解释器在 openpyxl 前就超限。
    输出返回值：
        无；评价为固定分母失败，且命名 parser 子进程已回收。
    """

    monkeypatch.setattr(searchwrite_bridge, "_PARSER_RSS_LIMIT_BYTES", 1)
    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(_fixed_revision_fixture("answer_files")),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.0
    assert all(
        child.name != "paraguibench-searchwrite-parser"
        for child in multiprocessing.active_children()
    )


def test_parser_checks_final_rss_before_receiving_ready_frame(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 frame 就绪后仍先执行一次最终 RSS 软监控。

    输入参数：
        monkeypatch：把最终 RSS 读数固定为超过协议预算。
    输出返回值：
        helper 在调用 ``recv_bytes`` 前以资源拒绝退出，证明快速发送
        frame 不能绕过父进程最后一次采样；该采样不冒充 OS 硬限制。
    """

    class _ReadyReceiver:
        """记录测试是否错误地先消费了 parser frame。"""

        receive_count = 0

        def recv_bytes(self, maximum: int) -> bytes:
            """记录接收调用并返回一个固定小 frame。

            输入参数：maximum 为生产协议的 frame 字节上限。
            输出返回值：固定 JSON 状态字节。
            """

            assert maximum == 64 * 1024
            self.receive_count += 1
            return b'["rejected"]'

    class _LiveProcess:
        """提供固定正 pid 的最小进程监控 seam。"""

        pid = 4242

    receiver = _ReadyReceiver()
    monkeypatch.setattr(
        searchwrite_bridge,
        "_parser_resident_bytes",
        lambda pid: searchwrite_bridge._PARSER_RSS_LIMIT_BYTES + 1,
    )

    with pytest.raises(
        searchwrite_bridge._WorkbookParseRejected,
        match="^PARSER_RESOURCE_LIMIT$",
    ):
        searchwrite_bridge._receive_parser_frame_after_final_rss_check(
            receiver,
            _LiveProcess(),
        )

    assert receiver.receive_count == 0


def test_parser_sandbox_cleanup_failure_is_redacted_internal_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证父进程临时目录清理故障不会泄露路径或降为任务失败。

    输入参数：
        tmp_path：创建可安全保留到 pytest 回收的合成 sandbox。
        monkeypatch：令 parser 返回合法小结果、cleanup 抛含哨兵 OSError。
    输出返回值：
        受控入口只抛固定 ``PARSER_SANDBOX_FAILED`` 内部错误，异常表示
        不含清理器的私密路径/文本。
    """

    sandbox_path = tmp_path / "PRIVATE-CLEANUP-PATH"
    sandbox_path.mkdir()

    class _FailingTemporaryDirectory:
        """提供有效目录身份并在 cleanup 阶段注入未知故障。"""

        name = os.fspath(sandbox_path)

        def cleanup(self) -> None:
            """模拟含私密信息的底层清理异常。

            输入参数：无。
            输出返回值：不返回，始终抛出测试 OSError。
            """

            raise OSError("PRIVATE-CLEANUP-SENTINEL")

    monkeypatch.setattr(
        searchwrite_bridge.tempfile,
        "TemporaryDirectory",
        lambda prefix: _FailingTemporaryDirectory(),
    )
    monkeypatch.setattr(
        searchwrite_bridge,
        "_run_parser_process",
        lambda content, **kwargs: ((), True),
    )

    with pytest.raises(
        _WorkbookParserInternalError,
        match="^PARSER_SANDBOX_FAILED$",
    ) as error:
        searchwrite_bridge._parse_xlsx_controlled(
            b"bounded",
            coordinates=("A1",),
            expected_baseline_sha256="0" * 64,
        )

    assert "PRIVATE" not in repr(error.value)


def test_parser_worker_enters_readonly_empty_cwd(tmp_path: Path) -> None:
    """验证真实 spawn worker 切换到父进程提供的只读空 cwd。

    输入参数：
        tmp_path：创建与仓库、资产完全隔离的合成工作目录。
    输出返回值：
        worker 内部验证目录身份、空状态和权限后只返回固定 ``ok``，
        子进程被完整 join/close，测试结束前恢复清理权限。
    """

    sandbox_cwd = _create_readonly_test_parser_cwd(tmp_path, "parser-cwd-contract")
    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    process = context.Process(
        target=_workbook_parse_worker,
        args=(
            sender,
            os.fspath(sandbox_cwd).encode("utf-8"),
            ("A1",),
            "0" * 64,
            _verify_parser_readonly_empty_cwd,
            os.fspath(sandbox_cwd),
        ),
        daemon=True,
        name="paraguibench-searchwrite-parser-cwd-test",
    )
    process.start()
    sender.close()
    assert receiver.poll(5.0)
    message = _decode_parser_message(receiver.recv_bytes(64 * 1024))
    receiver.close()
    process.join(timeout=5.0)

    assert process.is_alive() is False
    assert process.exitcode == 0
    process.close()
    sandbox_cwd.chmod(stat.S_IRWXU)
    assert message == ("ok", (), True)
    assert all(
        child.name != "paraguibench-searchwrite-parser-cwd-test"
        for child in multiprocessing.active_children()
    )


@pytest.mark.parametrize(
    "operation",
    (
        "open-write",
        "os-open-write",
        "truncate",
        "delete",
        "rename",
        "replace",
        "mkdir",
        "rmdir",
        "chmod",
        "link",
        "symlink",
        "mkfifo",
        "mknod",
        "utime",
        "chown",
    ),
)
def test_parser_worker_rejects_python_filesystem_mutation(
    tmp_path: Path,
    operation: str,
) -> None:
    """逐项验证真实 worker 的 Python 层文件系统写隔离。

    输入参数：
        tmp_path：仅供本测试创建可破坏合成树，不触碰仓库或真实资产。
        operation：固定参数矩阵中的 builtin/os.open 写开关、截断、
            删除、重命名/替换、建删目录、改权限、硬链接或符号链接。
            还覆盖 CPython 当前不发可用 audit event 的 FIFO/node 创建。
    输出返回值：
        worker 只返回固定 ``rejected``，测试树的成员、内容和权限
        与运行前完全相同，且子进程已经回收。
    """

    protected_root = tmp_path / "protected"
    protected_root.mkdir()
    source = protected_root / "source.bin"
    source.write_bytes(b"PRIVATE-SYNTHETIC-CONTENT")
    source.chmod(0o640)
    empty_directory = protected_root / "empty"
    empty_directory.mkdir()
    destination = protected_root / "destination"
    operation_source = empty_directory if operation == "rmdir" else source
    before = _snapshot_test_tree(protected_root)
    sandbox_cwd = _create_readonly_test_parser_cwd(tmp_path, "parser-cwd")
    payload = json.dumps(
        {
            "operation": operation,
            "source": os.fspath(operation_source),
            "destination": os.fspath(destination),
        }
    ).encode("utf-8")

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    process = context.Process(
        target=_workbook_parse_worker,
        args=(
            sender,
            payload,
            ("A1",),
            "0" * 64,
            _attempt_forbidden_filesystem_mutation,
            os.fspath(sandbox_cwd),
        ),
        daemon=True,
        name="paraguibench-searchwrite-parser-filesystem-test",
    )
    process.start()
    sender.close()
    assert receiver.poll(5.0)
    message = _decode_parser_message(receiver.recv_bytes(64 * 1024))
    receiver.close()
    process.join(timeout=5.0)

    assert process.is_alive() is False
    assert process.exitcode == 0
    process.close()
    assert message == ("rejected",)
    assert _snapshot_test_tree(protected_root) == before
    sandbox_cwd.chmod(stat.S_IRWXU)


def test_parser_worker_cpu_limit_stops_busy_materializer(
    tmp_path: Path,
) -> None:
    """验证子进程 CPU 硬限额不依赖 wall-clock 监控才生效。

    输入参数：
        tmp_path：为真实 worker 提供隔离的只读空工作目录。
        测试向该 worker 注入无限纯 CPU materializer。
    输出返回值：
        无；worker 必须在 5 秒 join 窗口内被 OS 信号终止，
        退出码为负，且不留活动子进程。
    """

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    sandbox_cwd = _create_readonly_test_parser_cwd(tmp_path, "parser-cpu-cwd")
    process = context.Process(
        target=_workbook_parse_worker,
        args=(
            sender,
            b"bounded",
            ("A1",),
            "0" * 64,
            _consume_parser_cpu_until_limited,
            os.fspath(sandbox_cwd),
        ),
        daemon=True,
        name="paraguibench-searchwrite-parser-cpu-test",
    )
    process.start()
    sender.close()
    process.join(timeout=5.0)
    receiver.close()

    assert process.is_alive() is False
    assert isinstance(process.exitcode, int)
    assert process.exitcode < 0
    process.close()
    sandbox_cwd.chmod(stat.S_IRWXU)
    assert all(
        child.name != "paraguibench-searchwrite-parser-cpu-test"
        for child in multiprocessing.active_children()
    )


def test_unknown_parser_exception_is_redacted_and_classified_internal(
    capsys: pytest.CaptureFixture[str],
    tmp_path: Path,
) -> None:
    """验证未知 parser 异常不泄露且不被当作 Agent 零分。

    输入参数：
        capsys：pytest 父进程输出捕获器，用于确认子进程
            继承的 stdout/stderr 中不存在私密 traceback。
        tmp_path：为真实 worker 提供隔离的只读空工作目录。
    输出返回值：
        无；worker 只返回固定 ``internal_error``，父层将它
        分类为 evaluator 内部错误，而不是可评价的工作簿失败。
    """

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    sandbox_cwd = _create_readonly_test_parser_cwd(tmp_path, "parser-error-cwd")
    process = context.Process(
        target=_workbook_parse_worker,
        args=(
            sender,
            b"bounded",
            ("A1",),
            "0" * 64,
            _raise_private_unknown_parser_error,
            os.fspath(sandbox_cwd),
        ),
        daemon=True,
        name="paraguibench-searchwrite-parser-error-test",
    )
    process.start()
    sender.close()
    assert receiver.poll(5.0)
    raw_message = receiver.recv_bytes(64 * 1024)
    message = _decode_parser_message(raw_message)
    receiver.close()
    process.join(timeout=5.0)
    assert process.is_alive() is False
    assert process.exitcode == 0
    process.close()
    sandbox_cwd.chmod(stat.S_IRWXU)

    assert message == ("internal_error",)
    with pytest.raises(_WorkbookParserInternalError):
        _validate_parser_message(message, coordinates=("A1",))
    captured = capsys.readouterr()
    assert "PRIVATE-UNKNOWN-PARSER" not in captured.out
    assert "PRIVATE-UNKNOWN-PARSER" not in captured.err


def test_parser_ipc_uses_bounded_json_instead_of_pickle(
    tmp_path: Path,
) -> None:
    """验证不可信 parser 子进程不向父进程发送 pickle。

    输入参数：
        tmp_path：为真实 worker 提供隔离的只读空工作目录；
            测试复用未知 parser 故障注入产生固定状态。
    输出返回值：
        无；单向 pipe 的原始 frame 必须不超过 64 KiB，
        可以严格解码为 JSON 列表 ``["internal_error"]``。
    """

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    sandbox_cwd = _create_readonly_test_parser_cwd(tmp_path, "parser-json-cwd")
    process = context.Process(
        target=_workbook_parse_worker,
        args=(
            sender,
            b"bounded",
            ("A1",),
            "0" * 64,
            _raise_private_unknown_parser_error,
            os.fspath(sandbox_cwd),
        ),
        daemon=True,
        name="paraguibench-searchwrite-parser-json-test",
    )
    process.start()
    sender.close()
    assert receiver.poll(5.0)
    raw_message = receiver.recv_bytes(64 * 1024)
    receiver.close()
    process.join(timeout=5.0)
    assert process.is_alive() is False
    process.close()
    sandbox_cwd.chmod(stat.S_IRWXU)

    assert len(raw_message) <= 64 * 1024
    assert json.loads(raw_message.decode("utf-8", errors="strict")) == [
        "internal_error"
    ]


def test_production_capture_feeds_fixed_revision_gold_to_formal_evaluator() -> None:
    """验证固定 revision 的两个 gold XLSX 可直接满分评价。

    输入参数：
        无；真实字节由 production manifest—nofollow—manifest
        capture 读取，不注入 typed fake。
    输出返回值：
        无；source 必须产生正式 evaluator 可直接消费的
        SearchWrite typed observation，且不依赖 Agent final text。
    """

    fixture_path = _fixed_revision_fixture("answer_files")
    assert sorted(path.name for path in fixture_path.iterdir()) == [
        "UK_Universities_Group1.xlsx",
        "UK_Universities_Group2.xlsx",
    ]

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is True
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ()
    assert evaluation.expected_cell_count == 9


def test_pinned_input_is_baseline_authority_but_not_answer_authority() -> None:
    """验证固定 input/gold 两套字节在协议中承担不同权威。

    输入参数：
        无；从 Lee ``13bf942d…`` 真实下载树读取两个
        input 模板和两个 gold 答案。
    输出返回值：
        无；四个文件 SHA-256 必须匹配固定 revision。input
        的非目标单元格/sheet 结构通过基线校验，但因九个
        目标格仍空而得零分；目标值权威只来自 pinned gold。
    """

    input_path = _fixed_revision_fixture("benchmark_dataset")
    gold_path = _fixed_revision_fixture("answer_files")
    for directory, expected_sha256 in (
        (input_path, _PINNED_INPUT_SHA256),
        (gold_path, _PINNED_GOLD_SHA256),
    ):
        actual = {
            path.name: hashlib.sha256(path.read_bytes()).hexdigest()
            for path in directory.iterdir()
            if path.is_file()
        }
        assert actual == expected_sha256

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(input_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.0
    assert evaluation.reason_codes == ("MISSING_CELL",)
    assert evaluation.expected_cell_count == 9
    assert evaluation.missing_cell_count == 9
    assert evaluation.mutated_document_count == 0


def test_production_capture_detects_prefilled_baseline_mutation(
    tmp_path: Path,
) -> None:
    """验证目标九格全对也不能遮蔽预填内容变更。

    输入参数：
        tmp_path：pytest 隔离目录；测试复制固定 gold 后
            仅改动 Group1 中一个原本预填的单元格。
    输出返回值：
        无；正式 evaluator 必须保留九格满分分母，但以
        ``BASELINE_CONTENT_CHANGED`` 拒绝修改后工作簿。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook_path = fixture_path / "UK_Universities_Group1.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook.worksheets[0]["A4"] = "PRIVATE MUTATION"
    workbook.save(workbook_path)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert evaluation.expected_cell_count == 9
    assert evaluation.matched_cell_count == 9


def test_production_capture_detects_visible_row_height_mutation(
    tmp_path: Path,
) -> None:
    """验证九格内容正确时仍不允许改动可见行高。

    输入参数：
        tmp_path：pytest 隔离目录；测试只改动 Group1
            首表的显式行高，不改九个答案单元格。
    输出返回值：
        无；评价保留九格满分，但以基线变更拒绝该工作簿。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook_path = fixture_path / "UK_Universities_Group1.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook.worksheets[0].row_dimensions[4].height = 31.25
    workbook.save(workbook_path)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert evaluation.matched_cell_count == 9


def test_production_capture_detects_visible_column_width_mutation(
    tmp_path: Path,
) -> None:
    """验证九格内容正确时仍不允许改动可见列宽。

    输入参数：
        tmp_path：pytest 隔离目录；测试只改动 Group2
            首表 A 列宽度，不改九个答案单元格。
    输出返回值：
        无；评价保留九格满分，但以基线变更拒绝该工作簿。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook_path = fixture_path / "UK_Universities_Group2.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook.worksheets[0].column_dimensions["A"].width = 80.25
    workbook.save(workbook_path)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert evaluation.matched_cell_count == 9


def test_production_capture_detects_target_cell_style_mutation(
    tmp_path: Path,
) -> None:
    """验证目标格只允许改内容，不允许改动其可见样式。

    输入参数：
        tmp_path：pytest 隔离目录；测试复制真实 gold 后只改一个
            目标格的填充色，不改变九格答案内容。
    输出返回值：
        正式评价保留九格满分，但基线闭集报告文档结构变更。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook_path = fixture_path / "UK_Universities_Group1.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook.worksheets[0]["C6"].fill = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF00",
    )
    workbook.save(workbook_path)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert evaluation.matched_cell_count == 9


def test_production_capture_rejects_theme_color_canonicalization_bypass(
    tmp_path: Path,
) -> None:
    """验证修改 theme 后不能把 theme-1 伪装成固定黑色样式。

    输入参数：
        tmp_path：pytest 隔离 gold 副本；测试从同一固定 revision 的
            input 取得已核验 theme，再只改变其默认深色并让目标格引用。
    输出返回值：
        九格内容仍正确，但 theme 与样式组合必须破坏基线，防止仅按
        ``theme=1`` 编号归一化而忽略实际渲染颜色。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
        from openpyxl.styles import Color
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    input_workbook = openpyxl.load_workbook(
        _fixed_revision_fixture("benchmark_dataset") / "UK_Universities_Group1.xlsx"
    )
    original_theme = input_workbook.loaded_theme
    input_workbook.close()
    assert isinstance(original_theme, bytes)
    modified_theme = original_theme.replace(
        b'lastClr="000000"',
        b'lastClr="FF0000"',
        1,
    )
    assert modified_theme != original_theme

    workbook_path = fixture_path / "UK_Universities_Group1.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook.loaded_theme = modified_theme
    target_cell = workbook.worksheets[0]["C6"]
    target_font = copy(target_cell.font)
    target_font.color = Color(theme=1)
    target_cell.font = target_font
    workbook.save(workbook_path)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert evaluation.matched_cell_count == 9


@pytest.mark.parametrize("font_metadata", ("scheme", "charset"))
def test_cell_style_normalizer_rejects_unverified_font_metadata(
    font_metadata: str,
) -> None:
    """直接验证字体规范化器不忽略未核验 metadata。

    输入参数：
        font_metadata：固定选择 theme scheme 或 charset 字段。
    输出返回值：
        真实 gold 目标格的规范化样式必须在单字段变更后立即不同，
        不依赖 save/reload 带来的其他 OOXML 差异才发现变更。
    """

    try:
        import openpyxl
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook = openpyxl.load_workbook(
        _fixed_revision_fixture("answer_files") / "UK_Universities_Group1.xlsx"
    )
    target_cell = workbook.worksheets[0]["C6"]
    before = _normalize_cell_visible_style(target_cell)
    target_font = copy(target_cell.font)
    if font_metadata == "scheme":
        target_font.scheme = "major"
    elif font_metadata == "charset":
        target_font.charset = 2
    else:
        raise AssertionError("未知字体 metadata 测试类型")
    target_cell.font = target_font
    after = _normalize_cell_visible_style(target_cell)
    workbook.close()

    assert after != before


@pytest.mark.parametrize("color_component", ("font", "fill"))
def test_cell_style_normalizer_distinguishes_rgb_colors(
    color_component: str,
) -> None:
    """验证 openpyxl 未设置 auto descriptor 不会吞掉 RGB 颜色。

    输入参数：
        color_component：固定选择真实表头的字体色或 solid 填充前景色。
    输出返回值：
        只改对应 RGB 色时规范化结果必须立即不同，避免把 openpyxl
        未设置 ``auto`` 时的 descriptor 提示字符串误当成 ``True``。
    """

    try:
        import openpyxl
        from openpyxl.styles import Color, PatternFill
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook = openpyxl.load_workbook(
        _fixed_revision_fixture("answer_files") / "UK_Universities_Group1.xlsx"
    )
    header_cell = workbook.worksheets[0]["A3"]
    before = _normalize_cell_visible_style(header_cell)
    if color_component == "font":
        header_font = copy(header_cell.font)
        header_font.color = Color(rgb="FF000000")
        header_cell.font = header_font
    elif color_component == "fill":
        header_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="FFFFFF00",
        )
    else:
        raise AssertionError("未知颜色组件测试类型")
    after = _normalize_cell_visible_style(header_cell)
    workbook.close()

    assert after != before


@pytest.mark.parametrize("font_metadata", ("scheme", "charset"))
def test_production_capture_detects_target_font_metadata_mutation(
    tmp_path: Path,
    font_metadata: str,
) -> None:
    """验证只合并真实 input/gold 的字体序列化等价项。

    输入参数：
        tmp_path：pytest 隔离的真实 gold 副本。
        font_metadata：固定选择 theme scheme 或 charset 字段。
    输出返回值：
        任意未核验的 scheme/charset 组合均破坏目标格样式基线，
        不能因兼容 LibreOffice 的已知差异而被宽泛忽略。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook_path = fixture_path / "UK_Universities_Group1.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    target_cell = workbook.worksheets[0]["C6"]
    target_font = copy(target_cell.font)
    if font_metadata == "scheme":
        target_font.scheme = "major"
    elif font_metadata == "charset":
        target_font.charset = 2
    else:
        raise AssertionError("未知字体 metadata 测试类型")
    target_cell.font = target_font
    workbook.save(workbook_path)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert evaluation.matched_cell_count == 9


@pytest.mark.parametrize(
    "visible_object_kind",
    (
        "comment",
        "table",
        "chart",
        "image",
        "hyperlink",
        "data-validation",
        "conditional-formatting",
    ),
)
def test_production_capture_detects_added_visible_object(
    tmp_path: Path,
    visible_object_kind: str,
) -> None:
    """逐项验证新增批注、表格、图表或图片破坏基线闭集。

    输入参数：
        tmp_path：pytest 隔离目录及合成图片文件位置。
        visible_object_kind：固定参数矩阵中的 openpyxl 可见对象类型。
    输出返回值：
        九格答案内容保持正确，但正式评价必须以基线变更失败，
        observation/evaluation 表示均不含合成私密文本。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
        from openpyxl.chart import BarChart, Reference
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.drawing.image import Image as WorksheetImage
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.worksheet.hyperlink import Hyperlink
        from openpyxl.worksheet.table import Table
        from PIL import Image as PillowImage
    except ImportError:
        pytest.fail("artifact optional spreadsheet dependencies unavailable")
    workbook_path = fixture_path / "UK_Universities_Group1.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook.worksheets[0]
    if visible_object_kind == "comment":
        worksheet["A4"].comment = Comment(
            "PRIVATE-VISIBLE-OBJECT-SENTINEL",
            "PRIVATE-AUTHOR",
        )
    elif visible_object_kind == "table":
        worksheet.add_table(Table(displayName="AddedTable", ref="A4:A8"))
    elif visible_object_kind == "chart":
        chart = BarChart()
        chart.add_data(
            Reference(
                worksheet,
                min_col=4,
                min_row=4,
                max_row=8,
            ),
            titles_from_data=True,
        )
        worksheet.add_chart(chart, "F2")
    elif visible_object_kind == "image":
        image_path = tmp_path / "pixel.png"
        PillowImage.new("RGB", (1, 1), color=(255, 0, 0)).save(image_path)
        worksheet.add_image(WorksheetImage(image_path), "F2")
    elif visible_object_kind == "hyperlink":
        worksheet["A4"].hyperlink = Hyperlink(
            ref="A4",
            location="'UK Universities'!A4",
        )
    elif visible_object_kind == "data-validation":
        validation = DataValidation(
            type="whole",
            operator="between",
            formula1="1",
            formula2="10",
        )
        validation.add("D4:D8")
        worksheet.add_data_validation(validation)
    elif visible_object_kind == "conditional-formatting":
        worksheet.conditional_formatting.add(
            "D4:D8",
            CellIsRule(operator="greaterThan", formula=["0"]),
        )
    else:
        raise AssertionError("未知可见对象测试类型")
    workbook.save(workbook_path)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ("BASELINE_CONTENT_CHANGED",)
    assert evaluation.matched_cell_count == 9
    assert "PRIVATE-VISIBLE-OBJECT" not in repr(observation)
    assert "PRIVATE-VISIBLE-OBJECT" not in repr(evaluation)


def test_corrupt_expected_workbook_keeps_its_cells_in_fixed_denominator(
    tmp_path: Path,
) -> None:
    """验证单个损坏工作簿不会被从九格分母删除。

    输入参数：
        tmp_path：pytest 隔离目录；Group1 保持 gold，Group2
            替换为无法解析且含私密占位符的字节。
    输出返回值：
        无；production source 仍应产生正式 typed observation，
        Group2 的五格全部作为缺失计分，任何结果不回显原字节。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    private_payload = b"PRIVATE corrupt workbook payload"
    (fixture_path / "UK_Universities_Group2.xlsx").write_bytes(private_payload)

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.4444
    assert evaluation.reason_codes == (
        "MISSING_CELL",
        "BASELINE_CONTENT_CHANGED",
    )
    assert evaluation.expected_cell_count == 9
    assert evaluation.matched_cell_count == 4
    assert evaluation.missing_cell_count == 5
    assert evaluation.evaluated_document_count == 2
    assert "PRIVATE" not in repr(evaluation)


def test_missing_expected_workbook_keeps_its_cells_in_fixed_denominator(
    tmp_path: Path,
) -> None:
    """验证整个期望工作簿缺失也不会缩小九格分母。

    输入参数：
        tmp_path：pytest 隔离目录；从真实 gold 副本删除
            Group2，Group1 保持完全正确。
    输出返回值：
        无；正式 evaluator 必须显式报告缺文档与缺五格，
        分数只为 Group1 的 ``4/9``。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    (fixture_path / "UK_Universities_Group2.xlsx").unlink()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.4444
    assert evaluation.reason_codes == (
        "MISSING_DOCUMENT",
        "MISSING_CELL",
    )
    assert evaluation.expected_document_count == 2
    assert evaluation.evaluated_document_count == 1
    assert evaluation.expected_cell_count == 9
    assert evaluation.matched_cell_count == 4
    assert evaluation.missing_cell_count == 5


def test_extra_regular_file_fails_the_exact_two_workbook_closed_set(
    tmp_path: Path,
) -> None:
    """验证九格全对时任何额外常规文件仍使任务失败。

    输入参数：
        tmp_path：pytest 隔离目录；在真实两个 gold 工作簿
            之外添加名称和内容都不应外泄的文件。
    输出返回值：
        无；bridge 必须以脱敏逻辑身份保留额外文件事实，
        正式 evaluator 返回 ``UNEXPECTED_DOCUMENT``。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    (fixture_path / "PRIVATE-EXTRA.xlsx").write_bytes(b"PRIVATE unexpected bytes")

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 1.0
    assert evaluation.reason_codes == ("UNEXPECTED_DOCUMENT",)
    assert evaluation.expected_cell_count == 9
    assert evaluation.matched_cell_count == 9
    assert evaluation.unexpected_document_count == 1
    assert "PRIVATE" not in repr(evaluation)


def test_attempt_runner_persists_only_redacted_searchwrite_counts(
    tmp_path: Path,
) -> None:
    """验证 XLSX 路径、单元格值与 Agent final text 不进入 RunStore。

    输入参数：
        tmp_path：pytest 隔离目录；同时承载修改后的 guest
            shared fixture 和任务级 RunStore。
    输出返回值：
        无；真实 production capture、runtime adapter、AttemptRunner 与
        RunStore 必须得到普通 FAIL 结果，所有持久化文件仅含
        协议、原因码和计数，不含四类私密哨兵。
    """

    fixture_path = tmp_path / "guest" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook_path = fixture_path / "UK_Universities_Group1.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook.worksheets[0]["D6"] = "PRIVATE-CELL-VALUE-SENTINEL"
    workbook.save(workbook_path)
    workbook.close()
    (fixture_path / "PRIVATE-PATH-SENTINEL.xlsx").write_bytes(
        b"PRIVATE-FILE-CONTENT-SENTINEL"
    )

    prepared = _prepared_searchwrite_task()
    store_root = tmp_path / "runstore"
    store = RunStore(store_root)
    store.start_run(
        run_id="run-searchwrite-privacy",
        run_record={"environment_id": "synthetic-osworld"},
        version_vector=synthetic_run_version_vector(),
    )
    attempt = store.start_attempt(
        run_id="run-searchwrite-privacy",
        task_id=SEARCHWRITE_XLSX_TASK_ID,
        attempt_id="attempt-001",
        task_record=prepared.audit_metadata,
    )
    environment = _SearchWriteObservationEnvironment(fixture_path)

    result = AttemptRunner(store).run(
        attempt=attempt,
        prepared_task=prepared,
        environment=environment,
        agent=_SensitiveSearchWriteAgent(),
        evaluator=PipelineImplicitTaskEvaluator(
            task_id=SEARCHWRITE_XLSX_TASK_ID,
            evaluation_protocol=SEARCHWRITE_XLSX_PROTOCOL_ID,
        ),
    )

    assert result.evaluation_outcome is EvaluationOutcome.FAILED
    assert result.score == pytest.approx(8 / 9, abs=1e-4)
    assert environment.closed is True
    persisted = b"\n".join(
        path.read_bytes() for path in store_root.rglob("*") if path.is_file()
    )
    for sentinel in (
        b"PRIVATE-CELL-VALUE-SENTINEL",
        b"PRIVATE-PATH-SENTINEL",
        b"PRIVATE-FILE-CONTENT-SENTINEL",
        b"PRIVATE-FINAL-TEXT-SENTINEL",
        b"D6",
    ):
        assert sentinel not in persisted
    for safe_field in (
        b"paraguibench.operation.searchwrite-xlsx.v1",
        b"reason_codes",
        b"unexpected_document_count",
        b"mismatched_cell_count",
    ):
        assert safe_field in persisted


def test_unknown_parser_failure_persists_error_and_null_score(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证 parser 内部故障不被静默降级为 Agent 任务失败。

    输入参数：
        tmp_path：隔离 RunStore 与真实 fixed-revision fixture。
        monkeypatch：在受控子进程边界向父层注入一个
            含私密哨兵的未知内部故障。
    输出返回值：
        无；AttemptRunner 必须保留 execution SUCCEEDED，将 evaluation
        记为 ERROR/score null，持久化数据不含内部哨兵或 final text。
    """

    def raise_internal_parser_error(
        content: bytes,
        *,
        coordinates: tuple[str, ...],
        expected_baseline_sha256: str,
    ) -> tuple[tuple[tuple[str, object], ...], bool]:
        """模拟 worker 已将未知异常分类为父层内部故障。

        输入参数：content/coordinates/指纹为已受控调用参数。
        输出返回值：不返回；抛出含哨兵的内部错误。
        """

        del content, coordinates, expected_baseline_sha256
        raise _WorkbookParserInternalError("PRIVATE-PARSER-INTERNAL-SENTINEL")

    monkeypatch.setattr(
        searchwrite_bridge,
        "_parse_xlsx_controlled",
        raise_internal_parser_error,
    )
    prepared = _prepared_searchwrite_task()
    store_root = tmp_path / "runstore"
    store = RunStore(store_root)
    store.start_run(
        run_id="run-searchwrite-parser-error",
        run_record={"environment_id": "synthetic-osworld"},
        version_vector=synthetic_run_version_vector(),
    )
    attempt = store.start_attempt(
        run_id="run-searchwrite-parser-error",
        task_id=SEARCHWRITE_XLSX_TASK_ID,
        attempt_id="attempt-001",
        task_record=prepared.audit_metadata,
    )

    with pytest.raises(
        PipelineImplicitArtifactEvidenceError,
        match="^TYPED_OBSERVATION_INVALID$",
    ):
        AttemptRunner(store).run(
            attempt=attempt,
            prepared_task=prepared,
            environment=_SearchWriteObservationEnvironment(
                _fixed_revision_fixture("answer_files")
            ),
            agent=_SensitiveSearchWriteAgent(),
            evaluator=PipelineImplicitTaskEvaluator(
                task_id=SEARCHWRITE_XLSX_TASK_ID,
                evaluation_protocol=SEARCHWRITE_XLSX_PROTOCOL_ID,
            ),
        )

    summary = json.loads((attempt.path / "summary.json").read_text(encoding="utf-8"))
    assert summary["execution"]["outcome"] == "SUCCEEDED"
    assert summary["evaluation"]["outcome"] == "ERROR"
    assert summary["evaluation"]["score"] is None
    persisted = b"\n".join(
        path.read_bytes() for path in store_root.rglob("*") if path.is_file()
    )
    assert b"PRIVATE-PARSER-INTERNAL-SENTINEL" not in persisted
    assert b"PRIVATE-FINAL-TEXT-SENTINEL" not in persisted


def test_ooxml_casefold_member_collision_is_rejected_before_semantic_use(
    tmp_path: Path,
) -> None:
    """验证 OOXML 内部大小写折叠成员不能绕过解析门。

    输入参数：
        tmp_path：pytest 隔离目录；在 Group2 gold ZIP 中添加
            与 ``xl/workbook.xml`` 便携大小写折叠的成员。
    输出返回值：
        无；即使 Office parser 会忽略该成员，bridge 也必须
        把 Group2 作为损坏期望文件计入九格固定分母。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    workbook_path = fixture_path / "UK_Universities_Group2.xlsx"
    with zipfile.ZipFile(workbook_path, mode="a") as archive:
        archive.writestr("XL/WORKBOOK.XML", b"<PRIVATE-ignored/>")

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.4444
    assert evaluation.reason_codes == (
        "MISSING_CELL",
        "BASELINE_CONTENT_CHANGED",
    )
    assert evaluation.matched_cell_count == 4
    assert evaluation.missing_cell_count == 5
    assert "PRIVATE" not in repr(evaluation)


def test_semantic_projection_rejects_oversized_sheet_dimension(
    tmp_path: Path,
) -> None:
    """验证小 ZIP 中的超大 worksheet dimension 不会触发无界迭代。

    输入参数：
        tmp_path：pytest 隔离目录；在 Group2 gold 中增加
            一个远超任务语义上限的非空远端单元格。
    输出返回值：
        无；bridge 必须在 ``iter_rows`` 前把 Group2 转换为
        零分占位，而不是仅保留五个目标值并只报基线变更。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    try:
        import openpyxl
    except ImportError:
        pytest.fail("artifact optional dependency openpyxl is unavailable")
    workbook_path = fixture_path / "UK_Universities_Group2.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook.worksheets[0]["A4097"] = "PRIVATE far cell"
    workbook.save(workbook_path)
    workbook.close()

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.4444
    assert evaluation.reason_codes == (
        "MISSING_CELL",
        "BASELINE_CONTENT_CHANGED",
    )
    assert evaluation.matched_cell_count == 4
    assert evaluation.missing_cell_count == 5


def test_ooxml_external_relationship_is_rejected_before_office_parser(
    tmp_path: Path,
) -> None:
    """验证协议不需要的 OOXML 外部 relationship 失败关闭。

    输入参数：
        tmp_path：pytest 隔离目录；在 Group2 gold 中增加
            Office parser 会忽略的外部链接 relationship。
    输出返回值：
        无；bridge 必须在语义解析前拒绝该文件，并用
        Group2 的五格零分占位保持固定分母。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    workbook_path = fixture_path / "UK_Universities_Group2.xlsx"
    relationship = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/'
        b'package/2006/relationships">'
        b'<Relationship Id="private" Type="http://example.invalid/private" '
        b'Target="https://example.invalid/private" TargetMode="External"/>'
        b"</Relationships>"
    )
    with zipfile.ZipFile(workbook_path, mode="a") as archive:
        archive.writestr("xl/_rels/private.xml.rels", relationship)

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.4444
    assert evaluation.reason_codes == (
        "MISSING_CELL",
        "BASELINE_CONTENT_CHANGED",
    )
    assert evaluation.matched_cell_count == 4
    assert evaluation.missing_cell_count == 5
    assert "private" not in repr(evaluation).lower()


def test_ooxml_utf16_active_entity_declaration_is_rejected(
    tmp_path: Path,
) -> None:
    """验证 UTF-16 编码的 DTD/entity 声明不能绕过 XML 门。

    输入参数：
        tmp_path：pytest 隔离目录；在 Group2 gold ZIP 中添加
            一个含主动实体声明的 UTF-16 XML 成员。
    输出返回值：
        无；该工作簿必须在 openpyxl 前被转为零分占位，
        评价结果不得回显实体或成员内容。
    """

    fixture_path = tmp_path / "answer_files" / _TASK_UID
    shutil.copytree(_fixed_revision_fixture("answer_files"), fixture_path)
    workbook_path = fixture_path / "UK_Universities_Group2.xlsx"
    active_xml = (
        '<?xml version="1.0" encoding="UTF-16"?>'
        '<!DOCTYPE x [<!ENTITY private "PRIVATE">]><x>&private;</x>'
    ).encode("utf-16")
    with zipfile.ZipFile(workbook_path, mode="a") as archive:
        archive.writestr("xl/private.xml", active_xml)

    observation = PipelineImplicitArtifactEvidenceSource().capture(
        SEARCHWRITE_XLSX_TASK_ID,
        _DirectoryController(fixture_path),
        guest_shared_dir=_GUEST_SHARED_DIR,
    )
    evaluation = evaluate_searchwrite_xlsx(observation)

    assert evaluation.passed is False
    assert evaluation.score == 0.4444
    assert evaluation.reason_codes == (
        "MISSING_CELL",
        "BASELINE_CONTENT_CHANGED",
    )
    assert evaluation.matched_cell_count == 4
    assert evaluation.missing_cell_count == 5
    assert "PRIVATE" not in repr(evaluation)
